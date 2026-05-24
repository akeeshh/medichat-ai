import streamlit as st
from groq import Groq
try:
    from anthropic import Anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
from sentence_transformers import SentenceTransformer
from datasets import load_dataset
import faiss
import os
import base64
import hashlib
import html
from PIL import Image, ImageEnhance, ImageOps
import io
import re
import time
import difflib
from datetime import datetime, timedelta, date as _date
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None
from fpdf import FPDF

# Firebase for cross-session analytics (optional - fails gracefully if missing)
try:
    import firebase_admin
    from firebase_admin import credentials, firestore
    FIREBASE_AVAILABLE = True
except ImportError:
    FIREBASE_AVAILABLE = False

def _safe_int_env(name, default):
    try:
        return int(os.environ.get(name, str(default)))
    except Exception:
        return default

def _safe_secret(name, default=None):
    try:
        return st.secrets[name]
    except Exception:
        return default

APP_TITLE = "MediChat Ai"
APP_SUBTITLE = "Your Ai Health Assistant"
APP_VERSION_LABEL = APP_TITLE

def _resolve_asset_path(filename):
    """Locate an asset by filename, enforcing an absolute path relative to this script.
    
    This guarantees reliability across local environments and Streamlit Cloud container restarts.
    """
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base_dir = os.getcwd()
        
    candidates = [
        os.path.join(base_dir, "assets", filename),
        os.path.join(base_dir, filename),
        os.path.join("assets", filename)
    ]
    
    for path in candidates:
        if os.path.exists(path):
            return path
            
    # Fallback to prevent unhandled runtime errors if file is created dynamically
    return os.path.join(base_dir, "assets", filename)

@st.cache_data(show_spinner=False)
def _load_asset_data_uri_cached(path, mtime):
    """Cache key is (path, mtime) so the cache invalidates automatically when
    the file changes on disk — no stale None after a missing-asset boot."""
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("ascii")
    ext = os.path.splitext(path)[1].lstrip(".").lower() or "png"
    if ext == "jpg":
        ext = "jpeg"
    return "data:image/" + ext + ";base64," + b64

def load_asset_data_uri(filename):
    """Return an asset as a base64 data URI, or None if missing/unreadable."""
    path = _resolve_asset_path(filename)
    if not path:
        return None
    try:
        return _load_asset_data_uri_cached(path, os.path.getmtime(path))
    except Exception:
        return None

def get_brand_logo_data_uri():
    """Backwards-compatible accessor for the MediChat brand logo."""
    return load_asset_data_uri("MediChat logo.png")
MEDICAL_REFERENCE_TARGET = max(1000, _safe_int_env("MEDICHAT_REFERENCE_TARGET", 5000))
PRIVACY_POLICY_URL = _safe_secret(
    "PRIVACY_POLICY_URL",
    os.environ.get("PRIVACY_POLICY_URL", "?mode=privacy"),
)

st.set_page_config(
    page_title=APP_VERSION_LABEL,
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Firebase Initialization (cross-session analytics) ────────────────
@st.cache_resource
def init_firebase():
    """Initialize Firebase Admin SDK. Returns Firestore client or None if unavailable."""
    if not FIREBASE_AVAILABLE:
        return None
    try:
        firebase_config = _safe_secret("firebase", {})
        if not firebase_config or not firebase_config.get("project_id"):
            return None
        if not firebase_admin._apps:
            cred = credentials.Certificate(dict(firebase_config))
            firebase_admin.initialize_app(cred)
        return firestore.client()
    except Exception as e:
        print("Firebase init failed:", e)
        return None

firestore_db = init_firebase()
FIREBASE_ACTIVE = firestore_db is not None

# ── Persistent Patient Profiles (email + PIN, Firestore-backed) ──────
PROFILE_SALT = _safe_secret("PROFILE_SALT", os.environ.get("PROFILE_SALT", "medichat-default-change-me"))

def hash_email(email):
    return hashlib.sha256((email.lower().strip() + PROFILE_SALT).encode()).hexdigest()

def hash_pin(pin, email_hash):
    return hashlib.sha256((str(pin) + email_hash + PROFILE_SALT).encode()).hexdigest()

def get_profile(email_hash):
    if not FIREBASE_ACTIVE:
        return None
    try:
        doc = firestore_db.collection("medichat_profiles").document(email_hash).get()
        return doc.to_dict() if doc.exists else None
    except Exception as e:
        print("Profile fetch failed:", e)
        return None

def create_profile(email, pin, name=""):
    if not FIREBASE_ACTIVE:
        return None
    eh = hash_email(email)
    profile = {
        "email_hash": eh,
        "pin_hash": hash_pin(pin, eh),
        "name": (name or "").strip()[:30],
        "patient_memory": {"symptoms": [], "conditions": [], "medications": []},
        "language": "English",
        "created_at": firestore.SERVER_TIMESTAMP,
        "last_visit": firestore.SERVER_TIMESTAMP,
        "visit_count": 1,
    }
    try:
        firestore_db.collection("medichat_profiles").document(eh).set(profile)
        profile["email_hash"] = eh
        return profile
    except Exception as e:
        print("Profile create failed:", e)
        return None

def authenticate_profile(email, pin):
    eh = hash_email(email)
    profile = get_profile(eh)
    if profile is None:
        return None, "not_found"
    if profile.get("pin_hash") != hash_pin(pin, eh):
        return None, "wrong_pin"
    try:
        firestore_db.collection("medichat_profiles").document(eh).update({
            "last_visit": firestore.SERVER_TIMESTAMP,
            "visit_count": firestore.Increment(1),
        })
    except Exception as e:
        print("Profile visit-count update failed:", e)
    profile["email_hash"] = eh
    return profile, "ok"

def persist_profile_state(email_hash, patient_memory=None, name=None, language=None, messages=None):
    if not FIREBASE_ACTIVE or not email_hash:
        return
    update = {"last_visit": firestore.SERVER_TIMESTAMP}
    if patient_memory is not None:
        update["patient_memory"] = patient_memory
    if name is not None:
        update["name"] = (name or "").strip()[:30]
    if language is not None:
        update["language"] = language
    if messages is not None:
        # Trim to last 60 messages, strip non-serialisable fields, cap content length
        trimmed = []
        for m in messages[-60:]:
            if not isinstance(m, dict):
                continue
            trimmed.append({
                "role": m.get("role", ""),
                "type": m.get("type", "text"),
                "content": (m.get("content", "") or "")[:4000],
                "sources": m.get("sources", []),
                "confidence": m.get("confidence", ""),
                "confidence_pct": m.get("confidence_pct", 0),
                "engine": m.get("engine", ""),
            })
        update["messages"] = trimmed
    try:
        firestore_db.collection("medichat_profiles").document(email_hash).update(update)
    except Exception as e:
        print("Profile state persist failed:", e)

# ── Per-Chat History (each conversation is its own Firestore doc) ────
def _trim_messages_for_storage(messages):
    trimmed = []
    for m in (messages or [])[-80:]:
        if not isinstance(m, dict):
            continue
        trimmed.append({
            "role": m.get("role", ""),
            "type": m.get("type", "text"),
            "content": (m.get("content", "") or "")[:4000],
            "sources": m.get("sources", []),
            "confidence": m.get("confidence", ""),
            "confidence_pct": m.get("confidence_pct", 0),
            "engine": m.get("engine", ""),
        })
    return trimmed

def derive_chat_title(messages):
    for m in messages or []:
        if m.get("role") == "user" and m.get("content"):
            t = (m["content"] or "").strip().replace("\n", " ")
            return (t[:50] + "…") if len(t) > 50 else t
    return "New chat"

def generate_ai_chat_title(messages):
    """Use Claude to summarise a chat into a 3-6 word clinical title.
    One call per chat. Falls back to derive_chat_title on any failure."""
    if not CLAUDE_ACTIVE or not messages:
        return derive_chat_title(messages)
    try:
        lines = []
        for m in messages[:8]:
            role = "Patient" if m.get("role") == "user" else "MediChat"
            content = (m.get("content", "") or "")[:280]
            if content:
                lines.append(role + ": " + content)
        transcript = "\n".join(lines)
        resp = anthropic_client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=24,
            system=(
                "You are a clinical scribe. Read the chat and output a 3-6 word title "
                "summarising the patient's chief concern. Output the title and nothing else: "
                "no quotes, no punctuation at the end, no preamble. Examples: "
                "Migraine workup, Type 2 diabetes review, Persistent dry cough, "
                "Lower back pain after lifting."
            ),
            messages=[{"role": "user", "content": transcript}],
            temperature=0.3,
        )
        title = (resp.content[0].text or "").strip().strip('"\'').rstrip(".")[:60]
        return title or derive_chat_title(messages)
    except Exception as e:
        print("AI title generation failed:", e)
        return derive_chat_title(messages)

def list_conversations(email_hash, limit=30):
    if not FIREBASE_ACTIVE or not email_hash:
        return []
    try:
        docs = (firestore_db.collection("medichat_profiles")
                .document(email_hash)
                .collection("conversations")
                .order_by("last_updated", direction=firestore.Query.DESCENDING)
                .limit(limit).stream())
        out = []
        for d in docs:
            data = d.to_dict() or {}
            out.append({
                "id": d.id,
                "title": data.get("title", "Chat"),
                "message_count": data.get("message_count", 0),
                "last_updated": data.get("last_updated"),
                "first_user_msg": data.get("first_user_msg", ""),
                "last_assistant_msg": data.get("last_assistant_msg", ""),
            })
        return out
    except Exception as e:
        print("list_conversations failed:", e)
        return []

def load_conversation(email_hash, conv_id):
    if not FIREBASE_ACTIVE or not email_hash or not conv_id:
        return None
    try:
        doc = (firestore_db.collection("medichat_profiles")
               .document(email_hash)
               .collection("conversations")
               .document(conv_id).get())
        if not doc.exists:
            return None
        return doc.to_dict()
    except Exception as e:
        print("load_conversation failed:", e)
        return None

def save_conversation(email_hash, conv_id, messages):
    if not FIREBASE_ACTIVE or not email_hash:
        return None
    trimmed = _trim_messages_for_storage(messages)
    msg_count = len(trimmed)
    # Lightweight summary fields for cross-chat context (avoids re-reading the full doc).
    _first_user = next((m.get("content", "") for m in trimmed if m.get("role") == "user"), "")
    _last_asst = next((m.get("content", "") for m in reversed(trimmed) if m.get("role") == "assistant"), "")
    payload = {
        "messages": trimmed,
        "message_count": msg_count,
        "first_user_msg": (_first_user or "")[:240],
        "last_assistant_msg": (_last_asst or "")[:320],
        "last_updated": firestore.SERVER_TIMESTAMP,
    }
    # Title strategy: cheap fallback on first save, AI upgrade once at 4 messages.
    if not conv_id:
        payload["title"] = derive_chat_title(trimmed)
    elif msg_count == 4:
        payload["title"] = generate_ai_chat_title(trimmed)
    try:
        coll = firestore_db.collection("medichat_profiles").document(email_hash).collection("conversations")
        if conv_id:
            coll.document(conv_id).set(payload, merge=True)
            return conv_id
        # New conversation: include created_at
        payload["created_at"] = firestore.SERVER_TIMESTAMP
        ref = coll.add(payload)
        return ref[1].id if isinstance(ref, tuple) else ref.id
    except Exception as e:
        print("save_conversation failed:", e)
        return None

def delete_conversation(email_hash, conv_id):
    if not FIREBASE_ACTIVE or not email_hash or not conv_id:
        return False
    try:
        (firestore_db.collection("medichat_profiles")
         .document(email_hash)
         .collection("conversations")
         .document(conv_id).delete())
        return True
    except Exception as e:
        print("delete_conversation failed:", e)
        return False

# ── Generic per-user data store (Firestore for auth, session for guest) ─
import uuid as _uuid

GUEST_DATA_KEY = "guest_user_data"

def _ensure_guest_store():
    if GUEST_DATA_KEY not in st.session_state:
        st.session_state[GUEST_DATA_KEY] = {
            "medications": [],
            "appointments": [],
            "health_records": [],
            "daily_metrics": {},
        }
    return st.session_state[GUEST_DATA_KEY]

def _today_key():
    return datetime.now().strftime("%Y-%m-%d")

def get_user_doc():
    """Read fresh profile doc for the signed-in user; returns {} if guest/none."""
    if not (st.session_state.get("is_authenticated") and st.session_state.get("user_email_hash") and FIREBASE_ACTIVE):
        return None
    try:
        snap = firestore_db.collection("medichat_profiles").document(st.session_state.user_email_hash).get()
        return snap.to_dict() or {} if snap.exists else {}
    except Exception as e:
        print("get_user_doc failed:", e)
        return {}

def update_user_doc(updates):
    """Patch the signed-in user's profile doc."""
    if not (st.session_state.get("is_authenticated") and st.session_state.get("user_email_hash") and FIREBASE_ACTIVE):
        return False
    try:
        firestore_db.collection("medichat_profiles").document(st.session_state.user_email_hash).set(updates, merge=True)
        return True
    except Exception as e:
        print("update_user_doc failed:", e)
        return False

# ── Medications ──────────────────────────────────────────────────────
def list_medications():
    if st.session_state.get("is_authenticated"):
        return (get_user_doc() or {}).get("medications", []) or []
    return _ensure_guest_store()["medications"]

def add_medication(name, dose, frequency, time_of_day, notes=""):
    name = (name or "").strip()
    if not name:
        return False
    entry = {
        "id": str(_uuid.uuid4())[:12],
        "name": name[:80],
        "dose": (dose or "").strip()[:40],
        "frequency": (frequency or "Once daily")[:30],
        "time_of_day": (time_of_day or "")[:30],
        "notes": (notes or "").strip()[:240],
        "added_at": datetime.now().isoformat(timespec="seconds"),
    }
    if st.session_state.get("is_authenticated"):
        current = list_medications()
        update_user_doc({"medications": current + [entry]})
    else:
        _ensure_guest_store()["medications"].append(entry)
    return True

def delete_medication(med_id):
    if st.session_state.get("is_authenticated"):
        current = [m for m in list_medications() if m.get("id") != med_id]
        update_user_doc({"medications": current})
    else:
        store = _ensure_guest_store()
        store["medications"] = [m for m in store["medications"] if m.get("id") != med_id]
    return True

# ── Appointments ─────────────────────────────────────────────────────
def list_appointments():
    if st.session_state.get("is_authenticated"):
        return (get_user_doc() or {}).get("appointments", []) or []
    return _ensure_guest_store()["appointments"]

def add_appointment(title, date_iso, doctor, location, notes=""):
    title = (title or "").strip()
    if not title or not date_iso:
        return False
    entry = {
        "id": str(_uuid.uuid4())[:12],
        "title": title[:80],
        "date": date_iso,
        "doctor": (doctor or "").strip()[:60],
        "location": (location or "").strip()[:80],
        "notes": (notes or "").strip()[:240],
        "added_at": datetime.now().isoformat(timespec="seconds"),
    }
    if st.session_state.get("is_authenticated"):
        current = list_appointments()
        update_user_doc({"appointments": current + [entry]})
    else:
        _ensure_guest_store()["appointments"].append(entry)
    return True

def delete_appointment(appt_id):
    if st.session_state.get("is_authenticated"):
        current = [a for a in list_appointments() if a.get("id") != appt_id]
        update_user_doc({"appointments": current})
    else:
        store = _ensure_guest_store()
        store["appointments"] = [a for a in store["appointments"] if a.get("id") != appt_id]
    return True

# ── Health Records (file metadata only — file content not stored to keep doc <1MB) ──
def list_health_records():
    if st.session_state.get("is_authenticated"):
        return (get_user_doc() or {}).get("health_records", []) or []
    return _ensure_guest_store()["health_records"]

def add_health_record(name, file_type, size_bytes, summary=""):
    name = (name or "").strip()
    if not name:
        return False
    entry = {
        "id": str(_uuid.uuid4())[:12],
        "name": name[:120],
        "file_type": (file_type or "")[:24],
        "size_bytes": int(size_bytes or 0),
        "summary": (summary or "").strip()[:1500],
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
    }
    if st.session_state.get("is_authenticated"):
        current = list_health_records()
        update_user_doc({"health_records": current + [entry]})
    else:
        _ensure_guest_store()["health_records"].append(entry)
    return True

def delete_health_record(rec_id):
    if st.session_state.get("is_authenticated"):
        current = [r for r in list_health_records() if r.get("id") != rec_id]
        update_user_doc({"health_records": current})
    else:
        store = _ensure_guest_store()
        store["health_records"] = [r for r in store["health_records"] if r.get("id") != rec_id]
    return True

# ── Daily Metrics (water, sleep) ─────────────────────────────────────
DAILY_METRIC_DEFAULTS = {
    "water_glasses": 0,
    "sleep_hours": None,
    "mood": None,
    "steps": None,
    "heart_rate_resting": None,
}

def get_daily_metrics(date_key=None):
    date_key = date_key or _today_key()
    if st.session_state.get("is_authenticated"):
        all_dm = (get_user_doc() or {}).get("daily_metrics", {}) or {}
    else:
        all_dm = _ensure_guest_store()["daily_metrics"]
    # Merge with defaults so missing keys are explicit None rather than KeyError-prone.
    raw = all_dm.get(date_key, {})
    return {**DAILY_METRIC_DEFAULTS, **raw}

def heart_rate_status(bpm):
    """Return (label, css_status_class) for a resting heart rate, or (None, None)."""
    if bpm is None:
        return (None, None)
    try:
        v = int(bpm)
    except Exception:
        return (None, None)
    if 60 <= v <= 100:
        return ("Normal", "md-status-good")
    if v < 60:
        return ("Low", "md-status-info")
    return ("High", "md-status-warn")

def sleep_status(hours):
    if hours is None:
        return (None, None)
    try:
        v = float(hours)
    except Exception:
        return (None, None)
    if 7 <= v <= 9:
        return ("Good", "md-status-good")
    if 6 <= v < 7 or 9 < v <= 10:
        return ("Fair", "md-status-info")
    return ("Low" if v < 6 else "High", "md-status-warn")

def update_daily_metric(field, value, date_key=None):
    date_key = date_key or _today_key()
    if st.session_state.get("is_authenticated"):
        doc = get_user_doc() or {}
        all_dm = doc.get("daily_metrics", {}) or {}
        day = all_dm.get(date_key, {})
        day[field] = value
        all_dm[date_key] = day
        update_user_doc({"daily_metrics": all_dm})
    else:
        store = _ensure_guest_store()
        day = store["daily_metrics"].get(date_key, {})
        day[field] = value
        store["daily_metrics"][date_key] = day

def get_metrics_history(days=7):
    """Return list of (date_str, metrics_dict) for last N days."""
    if st.session_state.get("is_authenticated"):
        all_dm = (get_user_doc() or {}).get("daily_metrics", {}) or {}
    else:
        all_dm = _ensure_guest_store()["daily_metrics"]
    out = []
    for i in range(days - 1, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        out.append((d, {**DAILY_METRIC_DEFAULTS, **all_dm.get(d, {})}))
    return out

def log_query_to_firestore(query_data):
    """Write anonymised query metadata to Firestore. Silent failure if Firebase unavailable."""
    if not FIREBASE_ACTIVE:
        return
    try:
        safe_data = {
            "query_word_count": len(query_data.get("query", "").split()),
            "confidence": query_data.get("confidence", "unknown"),
            "confidence_pct": query_data.get("confidence_pct", 0),
            "sources": query_data.get("sources", []),
            "response_time": query_data.get("response_time", 0),
            "language": query_data.get("language", "English"),
            "mode": query_data.get("mode", "free_chat"),
            "emergency_triggered": query_data.get("emergency_triggered", False),
            "drug_alerts": query_data.get("drug_alerts", 0),
            "timestamp": firestore.SERVER_TIMESTAMP,
        }
        firestore_db.collection("medichat_queries").add(safe_data)
    except Exception as e:
        print("Firestore write failed:", e)

def fetch_all_queries_from_firestore(limit=500):
    """Retrieve all anonymised query logs for admin dashboard."""
    if not FIREBASE_ACTIVE:
        return []
    try:
        docs = firestore_db.collection("medichat_queries").order_by("timestamp", direction=firestore.Query.DESCENDING).limit(limit).stream()
        return [doc.to_dict() for doc in docs]
    except Exception as e:
        print("Firestore read failed:", e)
        return []

def ui_escape(value):
    """Escape dynamic text before placing it inside custom HTML."""
    return html.escape(str(value or ""), quote=True)

def ui_text(value, max_chars=None):
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if max_chars and len(text) > max_chars:
        text = text[:max_chars - 3].rstrip() + "..."
    return ui_escape(text)

def ui_lines(value):
    return ui_escape(value).replace("\n", "<br>")

def get_user_local_now():
    tz_name = ""
    try:
        tz_name = getattr(st.context, "timezone", "") or ""
    except Exception:
        tz_name = ""
    if tz_name and ZoneInfo is not None:
        try:
            return datetime.now(ZoneInfo(tz_name))
        except Exception:
            pass
    return datetime.now()

def _msg_now_ts():
    """Return the current user-local time as a short '10:21 AM' string,
    used as a per-message timestamp for new conversations going forward.
    Old messages stored without a ts will simply skip the timestamp line."""
    try:
        return get_user_local_now().strftime("%-I:%M %p")
    except Exception:
        try:
            return get_user_local_now().strftime("%I:%M %p").lstrip("0")
        except Exception:
            return ""

def reset_prescription_reader_state():
    st.session_state.rx_reader_result = None
    st.session_state.rx_uploader_key = st.session_state.get("rx_uploader_key", 0) + 1

def start_new_chat_session():
    st.session_state.current_conversation_id = ""
    st.session_state.messages = []
    st.session_state.qcount = 0
    st.session_state.feedback = {}
    st.session_state.last_sources = []
    st.session_state.last_pdf_context = ""
    st.session_state.last_image_context = ""
    st.session_state.emergency_detected = False
    st.session_state.triage_assessment = None
    st.session_state.pending_user_input = ""
    st.session_state.chat_input_key = st.session_state.get("chat_input_key", 0) + 1
    st.session_state.uploader_key = st.session_state.get("uploader_key", 0) + 1
    st.session_state.home_show_vision_upload = False
    st.session_state.home_show_voice = False
    st.session_state.voice_audio_key = st.session_state.get("voice_audio_key", 0) + 1
    st.session_state.assessment_stage = 0
    st.session_state.assessment_data = {}
    st.session_state.assessment_complete = False
    st.session_state.assessment_report = None
    st.session_state.assessment_parsed = None
    st.session_state.mode = "chat"
    reset_prescription_reader_state()

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=DM+Serif+Display:ital@0;1&display=swap');

    :root {
        --clinical-900: #0c2d48;
        --clinical-800: #144272;
        --clinical-700: #1a5b8a;
        --clinical-600: #2176ae;
        --clinical-500: #2a8fc5;
        --clinical-400: #4da8d6;
        --clinical-300: #7ec3e6;
        --clinical-200: #b0daf2;
        --clinical-100: #d6edf9;
        --clinical-50: #edf6fc;

        --neutral-900: #1a1d21;
        --neutral-800: #2d3238;
        --neutral-700: #4a5058;
        --neutral-600: #6b7280;
        --neutral-500: #9ca3af;
        --neutral-400: #cbd5e1;
        --neutral-300: #e2e8f0;
        --neutral-200: #f1f5f9;
        --neutral-100: #f8fafc;

        --accent-rose: #c4766a;
        --accent-amber: #d69e2e;
        --accent-lavender: #8b7aa8;
    }

    * { font-family: 'Inter', sans-serif; }

    .stApp {
        background:
            radial-gradient(ellipse 80% 60% at top left, rgba(42, 143, 197, 0.08) 0%, transparent 50%),
            radial-gradient(ellipse 70% 50% at bottom right, rgba(139, 122, 168, 0.05) 0%, transparent 50%),
            linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
        min-height: 100vh;
    }

    .main .block-container {
        padding: 1rem 1.5rem 2rem 1.5rem;
        max-width: 800px;
    }

    /* ── Header ──────────────────────────────────────────────────── */
    .header-card {
        background: white;
        border-radius: 16px;
        padding: 1rem 1.5rem;
        margin-bottom: 0.8rem;
        box-shadow:
            0 1px 2px rgba(12, 45, 72, 0.04),
            0 4px 16px rgba(12, 45, 72, 0.06);
        position: relative;
        overflow: hidden;
        border: 1px solid var(--clinical-100);
    }

    .header-card::before {
        content: "";
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, var(--clinical-500), var(--clinical-300), var(--clinical-500));
        background-size: 200% 100%;
        animation: shimmer 6s ease-in-out infinite;
    }

    @keyframes shimmer {
        0%, 100% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
    }

    .header-brand {
        display: flex;
        align-items: center;
        gap: 0.85rem;
        margin-bottom: 0;
    }

    .header-logo {
        width: 42px;
        height: 42px;
        border-radius: 12px;
        background: linear-gradient(135deg, var(--clinical-500), var(--clinical-700));
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.3rem;
        color: white;
        box-shadow: 0 3px 10px rgba(42, 143, 197, 0.25);
        flex-shrink: 0;
    }

    .header-title {
        font-family: 'Inter', sans-serif;
        font-size: 1.5rem;
        font-weight: 700;
        color: var(--clinical-900);
        margin: 0;
        letter-spacing: -0.01em;
        line-height: 1.2;
    }

    .header-subtitle {
        color: var(--neutral-600);
        font-size: 0.78rem;
        margin: 0.15rem 0 0 0;
        font-weight: 400;
        line-height: 1.4;
    }

    /* ── Trust Strip ─────────────────────────────────────────────── */
    .trust-strip {
        display: flex;
        gap: 0.5rem;
        margin: 0.8rem 0 1rem 0;
        flex-wrap: wrap;
        justify-content: center;
    }

    .trust-pill {
        background: white;
        border: 1px solid var(--clinical-100);
        border-radius: 100px;
        padding: 0.35rem 0.85rem;
        font-size: 0.7rem;
        font-weight: 500;
        color: var(--clinical-700);
        display: inline-flex;
        align-items: center;
        gap: 0.35rem;
        box-shadow: 0 1px 2px rgba(12, 45, 72, 0.03);
        transition: all 0.2s ease;
    }

    .trust-pill:hover {
        border-color: var(--clinical-300);
        transform: translateY(-1px);
    }

    .trust-pill-icon {
        width: 14px;
        height: 14px;
        border-radius: 50%;
        background: var(--clinical-500);
        display: inline-flex;
        align-items: center;
        justify-content: center;
        color: white;
        font-size: 0.5rem;
        font-weight: 700;
    }

    .stats-row { display: flex; gap: 0.5rem; margin-bottom: 0.8rem; flex-wrap: wrap; justify-content: center; }
    .stat-pill { background: white; border: 1px solid var(--clinical-100); border-radius: 100px; padding: 0.35rem 0.85rem; font-size: 0.7rem; font-weight: 500; color: var(--clinical-700); }
    .stat-pill.green { color: var(--clinical-700); border-color: var(--clinical-300); background: var(--clinical-50); }
    .stat-pill.blue { color: #3a6b8f; border-color: #bed2e0; background: #eff5fa; }
    .stat-pill.purple { color: var(--accent-lavender); border-color: #d4c9e3; background: #f5f0fa; }
    .stat-pill.orange { color: var(--accent-amber); border-color: #e8cf9e; background: #fbf5e7; }

    /* ── Disclaimer ──────────────────────────────────────────────── */
    .disclaimer {
        display: none;
    }

    .disclaimer-mini {
        font-size: 0.68rem;
        color: var(--neutral-500);
        text-align: center;
        padding: 0.3rem 0;
        margin-top: 0.5rem;
        opacity: 0.8;
    }
    .disclaimer-mini-red {
        color: #a85c50;
    }

    /* ── Emergency Banner ────────────────────────────────────────── */
    .emergency-banner {
        background: linear-gradient(135deg, #dc2626, #991b1b);
        color: white;
        border-radius: 14px;
        padding: 1rem 1.3rem;
        margin-bottom: 1rem;
        box-shadow: 0 6px 20px rgba(220, 38, 38, 0.2);
        animation: pulse 2s ease-in-out infinite;
    }
    @keyframes pulse {
        0%, 100% { box-shadow: 0 6px 20px rgba(220, 38, 38, 0.2); }
        50% { box-shadow: 0 6px 30px rgba(220, 38, 38, 0.45); }
    }
    .emergency-title { font-size: 1.05rem; font-weight: 700; margin-bottom: 0.3rem; display: flex; align-items: center; gap: 0.5rem; }
    .emergency-text { font-size: 0.82rem; line-height: 1.5; margin-bottom: 0.5rem; opacity: 0.95; }
    .emergency-number {
        background: white; color: #991b1b;
        padding: 0.5rem 1.1rem; border-radius: 10px;
        font-size: 1.1rem; font-weight: 700;
        display: inline-block; margin-top: 0.15rem;
        letter-spacing: 0.04em;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }

    /* ── Mode Buttons ────────────────────────────────────────────── */
    .stButton > button {
        font-family: 'Inter', sans-serif;
        font-weight: 700 !important;
        border-radius: 12px !important;
        border: 1.5px solid var(--clinical-300) !important;
        background: white !important;
        color: #1a1d21 !important;
        padding: 0 1rem !important;
        height: 44px !important;
        min-height: 44px !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 1px 3px rgba(12, 45, 72, 0.08) !important;
        font-size: 0.85rem !important;
    }
    .stButton > button:hover {
        border-color: var(--clinical-500) !important;
        background: var(--clinical-50) !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(42, 143, 197, 0.15) !important;
    }

    /* ── Primary Send button (chat form only — uses type="primary") ── */
    .stForm [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
    .stForm [data-testid="stFormSubmitButton"] > button[kind="primary"] {
        background: linear-gradient(135deg, var(--clinical-600), var(--clinical-800)) !important;
        color: white !important;
        border: 1px solid var(--clinical-700) !important;
        font-weight: 600 !important;
        height: 44px !important;
        min-height: 44px !important;
        border-radius: 14px !important;
        box-shadow: 0 4px 12px rgba(12, 45, 72, 0.2) !important;
    }
    .stForm [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"]:hover,
    .stForm [data-testid="stFormSubmitButton"] > button[kind="primary"]:hover {
        background: linear-gradient(135deg, var(--clinical-800), var(--clinical-900)) !important;
        border-color: var(--clinical-900) !important;
        box-shadow: 0 6px 18px rgba(12, 45, 72, 0.3) !important;
        transform: translateY(-1px);
    }

    /* Secondary form buttons (Save / Skip / Next / Cancel) — visible dark text */
    .stForm [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"],
    .stForm [data-testid="stFormSubmitButton"] > button[kind="secondary"] {
        background: white !important;
        color: var(--clinical-900) !important;
        border: 1px solid var(--clinical-300) !important;
        font-weight: 600 !important;
        height: 44px !important;
        min-height: 44px !important;
        border-radius: 14px !important;
    }
    .stForm [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"]:hover,
    .stForm [data-testid="stFormSubmitButton"] > button[kind="secondary"]:hover {
        background: var(--clinical-50) !important;
        border-color: var(--clinical-500) !important;
        color: var(--clinical-900) !important;
    }


    /* Match input height to buttons */
    .stForm .stTextInput > div > div > input {
        height: 44px !important;
        padding: 0 1.1rem !important;
    }

    /* ── Welcome Card ────────────────────────────────────────────── */
    .welcome-card {
        background: white;
        border-radius: 18px;
        padding: 2rem 2rem;
        text-align: center;
        box-shadow: 0 4px 20px rgba(12, 45, 72, 0.06);
        margin: 0.5rem 0 1rem 0;
        border: 1px solid var(--clinical-100);
        animation: fadeInUp 0.5s ease-out;
    }

    .hero-wrap {
        background: linear-gradient(135deg, #f0f7fc 0%, #e3f0f9 100%);
        border-radius: 18px;
        padding: 1.4rem 1.6rem 1.3rem 1.6rem;
        margin: 0 0 0.8rem 0;
        position: relative;
        overflow: hidden;
        border: 1px solid var(--clinical-100);
        animation: fadeInUp 0.6s ease-out;
    }
    .hero-wrap::before {
        content: "";
        position: absolute;
        top: -60px; right: -60px;
        width: 180px; height: 180px;
        background: radial-gradient(circle, rgba(42, 143, 197, 0.08), transparent 70%);
        border-radius: 50%;
        z-index: 0;
    }
    .hero-eyebrow {
        display: inline-flex;
        align-items: center;
        gap: 0.35rem;
        background: white;
        border: 1px solid var(--clinical-100);
        color: var(--clinical-700);
        font-size: 0.62rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        padding: 0.25rem 0.7rem;
        border-radius: 100px;
        margin-bottom: 0.6rem;
        position: relative;
        z-index: 1;
    }
    .hero-eyebrow::before {
        content: "●";
        color: #22c55e;
        font-size: 0.45rem;
        animation: pulse 2s infinite;
    }
    @keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.4; }
    }
    .hero-title {
        font-family: 'Inter', sans-serif;
        font-size: 1.45rem;
        font-weight: 700;
        color: var(--clinical-900);
        line-height: 1.2;
        letter-spacing: -0.01em;
        margin-bottom: 0.4rem;
        position: relative;
        z-index: 1;
    }
    .hero-subtitle {
        color: var(--neutral-600);
        font-size: 0.82rem;
        line-height: 1.5;
        max-width: 560px;
        margin-bottom: 0.8rem;
        position: relative;
        z-index: 1;
    }
    .hero-trust-row {
        display: flex;
        gap: 0.3rem;
        flex-wrap: wrap;
        margin-bottom: 0;
        position: relative;
        z-index: 1;
    }
    .trust-pill {
        display: inline-flex;
        align-items: center;
        gap: 0.3rem;
        background: white;
        border: 1px solid var(--clinical-100);
        color: var(--clinical-700);
        font-size: 0.68rem;
        font-weight: 500;
        padding: 0.3rem 0.65rem;
        border-radius: 100px;
        box-shadow: 0 1px 3px rgba(12, 45, 72, 0.04);
    }
    .trust-icon {
        width: 12px;
        height: 12px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 0.75rem;
    }

    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }

    .welcome-title {
        font-family: 'Inter', sans-serif;
        font-size: 1.5rem;
        font-weight: 700;
        color: var(--clinical-900);
        margin: 0.6rem 0 0.5rem 0;
        letter-spacing: -0.01em;
    }
    .welcome-text {
        color: var(--neutral-600);
        font-size: 0.9rem;
        line-height: 1.6;
        margin-bottom: 1.2rem;
    }

    .chip-row {
        display: flex;
        gap: 0.4rem;
        flex-wrap: wrap;
        justify-content: center;
        margin-top: 0.8rem;
    }
    .chip {
        background: var(--clinical-50);
        border: 1px solid var(--clinical-100);
        color: var(--clinical-700);
        padding: 0.35rem 0.85rem;
        border-radius: 100px;
        font-size: 0.72rem;
        font-weight: 500;
        transition: all 0.2s ease;
    }
    .chip:hover {
        background: var(--clinical-100);
        border-color: var(--clinical-300);
        transform: translateY(-1px);
    }

    /* ── Memory Card (Apple-style) ───────────────────────────────────
       Soft white card with subtle indigo tint, brain icon, and each
       remembered item rendered as its own coloured chip per category. */
    .memory-card {
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%);
        border: 1px solid #e6ecf6;
        border-radius: 18px;
        padding: 1rem 1.15rem 1.05rem 1.15rem;
        margin-bottom: 1.4rem;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.95) inset,
            0 6px 18px rgba(15, 23, 42, 0.04);
        animation: fadeIn 0.4s ease;
    }
    .memory-head {
        display: flex;
        align-items: center;
        gap: 0.55rem;
        margin-bottom: 0.7rem;
    }
    .memory-icon {
        width: 30px;
        height: 30px;
        border-radius: 9px;
        display: flex;
        align-items: center;
        justify-content: center;
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.14), rgba(139, 92, 246, 0.12));
        color: #4f46e5;
        flex-shrink: 0;
    }
    .memory-icon .material-symbols-rounded {
        font-size: 1.05rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
    }
    .memory-title {
        font-size: 0.92rem;
        font-weight: 700;
        color: #0f172a;
        letter-spacing: -0.01em;
        line-height: 1.2;
        margin: 0;
        display: block;
    }
    .memory-section {
        display: grid;
        grid-template-columns: 88px minmax(0, 1fr);
        align-items: center;
        gap: 0.65rem;
        padding: 0.4rem 0;
    }
    .memory-section + .memory-section {
        border-top: 1px solid #eef1f8;
    }
    .memory-label {
        font-size: 0.66rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #94a3b8;
    }
    .memory-chips {
        display: flex;
        flex-wrap: wrap;
        gap: 0.35rem;
    }
    .memory-chip {
        display: inline-flex;
        align-items: center;
        font-size: 0.74rem;
        font-weight: 600;
        line-height: 1;
        padding: 0.35rem 0.65rem;
        border-radius: 999px;
        background: #f7f9ff;
        border: 1px solid #e6ecf6;
        color: #475569;
        white-space: nowrap;
        max-width: 100%;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    /* Per-category accent tints (each section gets its own colour family
       so symptoms ≠ conditions ≠ medications at a glance). */
    .memory-label-rose   { color: #be123c; }
    .memory-label-indigo { color: #4f46e5; }
    .memory-label-violet { color: #7c3aed; }
    .memory-chip-rose   { background: #fff1f2; border-color: #ffd5dd; color: #9f1239; }
    .memory-chip-indigo { background: #eef2ff; border-color: #c7d2fe; color: #4338ca; }
    .memory-chip-violet { background: #f5f0ff; border-color: #ddd2fc; color: #6d28d9; }

    @keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }

    /* ── Chat Messages ───────────────────────────────────────────── */
    .bot-label {
        font-size: 0.65rem;
        color: var(--neutral-500);
        font-weight: 600;
        margin-left: 50px;
        margin-bottom: 0.25rem;
        margin-top: 0.7rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }

    .bot-wrap, .user-wrap {
        display: flex;
        align-items: flex-start;
        gap: 0.65rem;
        margin-bottom: 0.7rem;
        animation: messageSlideIn 0.35s ease-out;
    }

    .user-wrap {
        justify-content: flex-end;
    }

    @keyframes messageSlideIn {
        from { opacity: 0; transform: translateY(8px); }
        to { opacity: 1; transform: translateY(0); }
    }

    .av {
        width: 38px;
        height: 38px;
        border-radius: 12px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 600;
        font-size: 0.9rem;
        flex-shrink: 0;
        box-shadow: 0 2px 6px rgba(12, 45, 72, 0.1);
    }

    .av-bot {
        background: linear-gradient(135deg, var(--clinical-500), var(--clinical-700));
        color: white;
    }

    .av-user {
        background: linear-gradient(135deg, var(--clinical-100), var(--clinical-200));
        color: var(--clinical-900);
    }

    .bot-bubble {
        background: white;
        color: var(--neutral-800);
        padding: 0.9rem 1.15rem;
        border-radius: 4px 16px 16px 16px;
        max-width: 85%;
        font-size: 0.9rem;
        line-height: 1.6;
        box-shadow: 0 1px 2px rgba(12, 45, 72, 0.04), 0 4px 16px rgba(12, 45, 72, 0.04);
        border: 1px solid var(--clinical-100);
        position: relative;
    }

    .bot-bubble::before {
        content: "";
        position: absolute;
        left: 0; top: 0; bottom: 0;
        width: 3px;
        background: linear-gradient(180deg, var(--clinical-500), var(--clinical-300));
        border-radius: 4px 0 0 0;
    }

    .bot-bubble .md-h {
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        color: var(--clinical-900);
        margin: 0.5rem 0 0.35rem 0;
        line-height: 1.25;
        letter-spacing: -0.01em;
    }
    .bot-bubble h3.md-h { font-size: 1.1rem; }
    .bot-bubble h4.md-h { font-size: 1rem; }
    .bot-bubble h5.md-h { font-size: 0.95rem; font-weight: 600; color: var(--clinical-700); }
    .bot-bubble h6.md-h { font-size: 0.9rem; font-weight: 600; color: var(--clinical-700); }
    .bot-bubble .md-h:first-child { margin-top: 0; }
    .bot-bubble .md-p { margin: 0.45rem 0; }
    .bot-bubble .md-p:first-child { margin-top: 0; }
    .bot-bubble .md-p:last-child { margin-bottom: 0; }
    .bot-bubble .md-ul, .bot-bubble .md-ol {
        margin: 0.45rem 0 0.5rem 0;
        padding-left: 1.3rem;
    }
    .bot-bubble .md-ul li, .bot-bubble .md-ol li {
        margin: 0.2rem 0;
        line-height: 1.5;
    }
    .bot-bubble .md-ul li::marker { color: var(--clinical-500); }
    .bot-bubble .md-ol li::marker { color: var(--clinical-500); font-weight: 600; }
    .bot-bubble strong { color: var(--clinical-900); font-weight: 600; }
    .bot-bubble em { font-style: italic; color: var(--clinical-700); }
    .bot-bubble .md-hr { border: none; border-top: 1px solid var(--clinical-100); margin: 0.7rem 0; }
    .bot-bubble .md-code {
        background: var(--clinical-50);
        color: var(--clinical-700);
        padding: 0.1rem 0.35rem;
        border-radius: 5px;
        font-family: 'SF Mono', Monaco, Consolas, monospace;
        font-size: 0.85em;
    }

    .user-bubble {
        background: linear-gradient(135deg, #e0e7ff, #e0f2fe) !important;
        color: #1f2937 !important;
        padding: 0.8rem 1.1rem;
        border-radius: 16px 4px 16px 16px;
        max-width: 80%;
        font-size: 0.9rem;
        line-height: 1.5;
        border: 1px solid #c7d2fe !important;
        box-shadow: 0 4px 12px rgba(99, 102, 241, 0.05) !important;
    }

    /* ── Thinking Indicator ──────────────────────────────────────── */
    .thinking-indicator {
        display: flex;
        align-items: center;
        gap: 0.35rem;
        padding: 0.7rem 0.9rem;
        background: white;
        border-radius: 4px 16px 16px 16px;
        border-left: 3px solid var(--clinical-500);
        max-width: 160px;
        box-shadow: 0 1px 2px rgba(12, 45, 72, 0.04);
    }
    .thinking-dot {
        width: 7px;
        height: 7px;
        border-radius: 50%;
        background: var(--clinical-500);
        animation: bounce 1.4s infinite ease-in-out;
    }
    .thinking-dot:nth-child(1) { animation-delay: -0.32s; }
    .thinking-dot:nth-child(2) { animation-delay: -0.16s; }
    @keyframes bounce {
        0%, 80%, 100% { transform: scale(0.6); opacity: 0.4; }
        40% { transform: scale(1); opacity: 1; }
    }
    .thinking-label {
        font-size: 0.75rem;
        color: var(--neutral-500);
        margin-left: 0.25rem;
        font-style: italic;
    }

    /* ── Streaming Cursor ────────────────────────────────────────── */
    .stream-cursor {
        display: inline-block;
        width: 2px;
        height: 1em;
        background: var(--clinical-500);
        margin-left: 2px;
        vertical-align: text-bottom;
        animation: blink 1s step-end infinite;
    }
    @keyframes blink {
        0%, 50% { opacity: 1; }
        51%, 100% { opacity: 0; }
    }

    /* ── Source & Confidence Rows ────────────────────────────────── */
    .source-row {
        margin-left: 50px;
        margin-bottom: 0.25rem;
        font-size: 0.7rem;
        color: var(--neutral-500);
    }

    .engine-badge {
        display: inline-flex;
        align-items: center;
        gap: 0.25rem;
        font-size: 0.62rem;
        font-weight: 600;
        padding: 0.12rem 0.5rem;
        border-radius: 100px;
        letter-spacing: 0.02em;
        margin-right: 0.35rem;
    }
    .engine-claude {
        background: linear-gradient(135deg, #fef2e8, #fdeede);
        color: #a8521a;
        border: 1px solid #f5c4a1;
    }
    .engine-groq {
        background: var(--clinical-50);
        color: var(--clinical-700);
        border: 1px solid var(--clinical-300);
    }
    .engine-vision {
        background: #f3e8ff;
        color: #6b21a8;
        border: 1px solid #d8b4fe;
    }
    .engine-badge::before { content: "●"; font-size: 0.45rem; }
    .source-tag {
        display: inline-block;
        background: var(--clinical-50);
        border: 1px solid var(--clinical-100);
        color: var(--clinical-700);
        font-size: 0.65rem;
        font-weight: 500;
        padding: 0.15rem 0.6rem;
        border-radius: 100px;
        margin-right: 0.25rem;
        margin-top: 0.25rem;
    }
    .confidence-row {
        margin-left: 50px;
        margin-bottom: 0.7rem;
        display: flex;
        align-items: center;
        gap: 0.45rem;
        font-size: 0.65rem;
    }
    .confidence-pill {
        padding: 0.12rem 0.65rem;
        border-radius: 100px;
        font-weight: 600;
        letter-spacing: 0.03em;
        font-size: 0.63rem;
    }
    .conf-high { background: var(--clinical-50); color: var(--clinical-700); border: 1px solid var(--clinical-300); }
    .conf-medium { background: #fbf5e7; color: #7a5d1a; border: 1px solid #e8cf9e; }
    .conf-low { background: #fdf2f1; color: #8f3f34; border: 1px solid #e3bfb8; }
    .confidence-bar {
        display: inline-block;
        width: 75px;
        height: 5px;
        background: var(--clinical-100);
        border-radius: 100px;
        overflow: hidden;
    }
    .confidence-fill {
        display: block;
        height: 100%;
        border-radius: 100px;
    }

    /* ── Input ───────────────────────────────────────────────────── */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea {
        border-radius: 12px !important;
        border: 1.5px solid var(--clinical-100) !important;
        padding: 0.75rem 1rem !important;
        font-size: 0.9rem !important;
        background: white !important;
        transition: all 0.2s ease !important;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: var(--clinical-500) !important;
        box-shadow: 0 0 0 3px rgba(42, 143, 197, 0.15) !important;
    }

    /* ── Attach Uploader: styled to look like a clean Attach button ── */
    [data-testid="stFileUploader"] > label {
        display: none !important;
    }
    [data-testid="stFileUploader"] section {
        padding: 0 !important;
        border: 1px solid var(--clinical-200) !important;
        border-radius: 14px !important;
        background: white !important;
        box-shadow: 0 1px 2px rgba(12, 45, 72, 0.04) !important;
        transition: all 0.2s ease !important;
        min-height: 44px !important;
        height: 44px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        cursor: pointer !important;
    }
    [data-testid="stFileUploader"] section:hover {
        border-color: var(--clinical-400) !important;
        background: var(--clinical-50) !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(42, 143, 197, 0.12) !important;
    }

    [data-testid="stFileUploaderDropzone"] {
        padding: 0 1rem !important;
        min-height: 44px !important;
        height: 44px !important;
        background: transparent !important;
        border: none !important;
        border-radius: 14px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 100% !important;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] {
        position: relative !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        margin: 0 !important;
        padding: 0 !important;
        width: 100% !important;
        height: 100% !important;
    }
    /* Hide Streamlit's original dropzone children */
    [data-testid="stFileUploaderDropzoneInstructions"] > * {
        visibility: hidden !important;
        position: absolute !important;
        pointer-events: none !important;
    }
    /* Inject our own visible label */
    [data-testid="stFileUploaderDropzoneInstructions"]::after {
        content: "📎  Attach image or PDF";
        position: absolute;
        inset: 0;
        display: flex;
        align-items: center;
        justify-content: center;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        color: var(--clinical-900) !important;
        letter-spacing: 0.01em;
        pointer-events: none;
    }
    [data-testid="stFileUploaderDropzone"] button {
        display: none !important;
    }
    /* Show filename neatly once a file is selected */
    [data-testid="stFileUploaderFile"] {
        padding: 0.5rem 0.8rem !important;
        background: var(--clinical-50) !important;
        border-radius: 12px !important;
        margin-top: 0.4rem !important;
    }


    /* ── Image Tag ───────────────────────────────────────────────── */
    .image-tag {
        display: inline-block;
        background: #f5f0fa;
        color: var(--accent-lavender);
        border: 1px solid #d4c9e3;
        padding: 0.25rem 0.75rem;
        border-radius: 100px;
        font-size: 0.7rem;
        font-weight: 500;
        margin-bottom: 0.4rem;
    }

    /* ── Name Welcome ────────────────────────────────────────────── */
    .name-welcome {
        background: linear-gradient(135deg, var(--clinical-50), #e3f0f9);
        border: 1px solid var(--clinical-100);
        border-radius: 12px;
        padding: 0.8rem 1.1rem;
        margin-bottom: 0.8rem;
        animation: fadeInUp 0.4s ease;
    }
    .name-welcome-text {
        font-size: 0.88rem;
        color: var(--clinical-900);
        font-weight: 400;
    }

    /* ── Suggested Follow-ups ────────────────────────────────────── */
    .suggestion-row {
        margin-left: 50px;
        margin-bottom: 0.8rem;
        display: flex;
        gap: 0.35rem;
        flex-wrap: wrap;
    }
    .suggestion-label {
        font-size: 0.65rem;
        color: var(--neutral-500);
        margin-left: 50px;
        margin-bottom: 0.35rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }

    /* ── Sidebar Styling ─────────────────────────────────────────── */
    [data-testid="stSidebar"] {
        background: white !important;
        border-right: 1px solid var(--clinical-100);
    }
    [data-testid="stSidebar"] .stMarkdown { color: var(--clinical-900); }
    .sb-title { font-size: 0.65rem; font-weight: 700; color: var(--neutral-500); text-transform: uppercase; letter-spacing: 0.12em; margin: 0.7rem 0 0.45rem 0; }
    .sb-footer { font-size: 0.62rem; color: var(--neutral-500) !important; text-align: center; padding-top: 0.8rem; border-top: 1px solid var(--clinical-50); line-height: 1.6; }

    /* ── Symptom Assessment Card ─────────────────────────────────── */
    .assessment-card {
        background: white;
        border-radius: 18px;
        padding: 1.4rem 1.6rem 1.5rem 1.6rem;
        margin-bottom: 1.2rem;
        border: 1px solid var(--clinical-100);
        box-shadow: 0 1px 3px rgba(12, 45, 72, 0.04), 0 8px 28px rgba(12, 45, 72, 0.05);
        animation: fadeInUp 0.4s ease;
    }
    .assessment-title {
        font-family: 'DM Serif Display', serif;
        font-size: 1.45rem;
        font-weight: 400;
        color: var(--clinical-900);
        letter-spacing: -0.01em;
        margin-bottom: 0.35rem;
        line-height: 1.2;
    }
    .assessment-subtitle {
        color: var(--neutral-600, #6b6660);
        font-size: 0.88rem;
        line-height: 1.55;
        margin-bottom: 1rem;
    }
    .progress-label {
        display: flex;
        align-items: center;
        justify-content: space-between;
        font-size: 0.78rem;
        font-weight: 600;
        color: var(--clinical-700);
        margin-bottom: 0.45rem;
        letter-spacing: 0.02em;
    }
    .progress-label span:first-child {
        color: var(--clinical-900);
    }
    .progress-label span:last-child {
        color: var(--clinical-600);
        font-weight: 700;
    }
    .progress-bar-wrap {
        width: 100%;
        height: 8px;
        background: var(--clinical-50);
        border: 1px solid var(--clinical-100);
        border-radius: 100px;
        overflow: hidden;
    }
    .progress-bar-fill {
        height: 100%;
        background: linear-gradient(90deg, var(--clinical-500), var(--clinical-700));
        border-radius: 100px;
        transition: width 0.4s ease;
    }
    .question-bubble {
        background: var(--clinical-50);
        border: 1px solid var(--clinical-100);
        border-left: 4px solid var(--clinical-600);
        border-radius: 14px;
        padding: 1rem 1.2rem;
        margin: 0.8rem 0 1rem 0;
        font-size: 1rem;
        font-weight: 500;
        color: var(--clinical-900);
        line-height: 1.5;
    }

    /* ── Triage Tier Badge ──────────────────────────────────────── */
    .triage-card {
        border-radius: 16px;
        padding: 1rem 1.2rem;
        margin: 0.5rem 0 1rem 0;
        color: white;
        box-shadow: 0 6px 20px rgba(12, 45, 72, 0.18);
        animation: fadeInUp 0.35s ease;
    }
    .triage-head {
        display: flex;
        align-items: center;
        gap: 0.6rem;
        margin-bottom: 0.4rem;
    }
    .triage-icon { font-size: 1.4rem; line-height: 1; }
    .triage-tier-num {
        background: rgba(255,255,255,0.22);
        padding: 0.18rem 0.55rem;
        border-radius: 100px;
        font-size: 0.68rem;
        font-weight: 700;
        letter-spacing: 0.06em;
    }
    .triage-label {
        font-size: 1.05rem;
        font-weight: 700;
        letter-spacing: -0.01em;
    }
    .triage-step {
        font-size: 0.86rem;
        line-height: 1.55;
        opacity: 0.95;
        margin-bottom: 0.45rem;
    }
    .triage-reasons {
        display: flex;
        flex-wrap: wrap;
        gap: 0.3rem;
        margin-top: 0.35rem;
    }
    .triage-reason-pill {
        background: rgba(255,255,255,0.18);
        border: 1px solid rgba(255,255,255,0.28);
        color: white;
        font-size: 0.7rem;
        font-weight: 500;
        padding: 0.18rem 0.6rem;
        border-radius: 100px;
    }

    /* ── Health Profile Sidebar Card ─────────────────────────────── */
    .hp-card {
        background: linear-gradient(135deg, #edf6fc, #d6edf9);
        border: 1px solid #b0daf2;
        border-radius: 12px;
        padding: 0.7rem 0.85rem;
        margin-bottom: 0.55rem;
    }
    .hp-section-title {
        font-size: 0.62rem;
        font-weight: 700;
        color: #1a5b8a;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin-bottom: 0.3rem;
    }
    .hp-chip-row {
        display: flex;
        flex-wrap: wrap;
        gap: 0.25rem;
    }
    .hp-chip {
        background: white;
        border: 1px solid #b0daf2;
        color: #144272;
        font-size: 0.7rem;
        font-weight: 500;
        padding: 0.18rem 0.55rem;
        border-radius: 100px;
    }
    .hp-empty {
        font-size: 0.7rem;
        color: #64748b;
        font-style: italic;
    }

    /* ── Hide Streamlit Branding ─────────────────────────────────── */
    footer { visibility: hidden; }
    [data-testid="stHeader"] { background: transparent; }

    .bot-bubble a[class*="anchor"],
    .bot-bubble svg[class*="anchor"],
    .bot-bubble a[href^="#"]:not([href*="://"]) { display: none !important; }
    [data-testid="stMarkdownContainer"] a.heading-anchor-icon,
    [data-testid="stMarkdownContainer"] a[href^="#"] svg { display: none !important; }

    [data-testid="InputInstructions"] { display: none !important; }
    .stForm [data-testid="stFormSubmitButton"] + small { display: none !important; }
    .stForm small { display: none !important; }

    [data-testid="stFileUploader"] [data-testid="stAlert"] {
        margin-top: 0.5rem !important;
        position: relative !important;
        z-index: 5 !important;
        display: block !important;
    }

</style>

""", unsafe_allow_html=True)

# ── Dashboard Reskin (overrides above via cascade) ────────────────────
st.markdown("""
<style>
:root {
    --md-bg: #f7f8fb;
    --md-surface: #ffffff;
    --md-border: #eef0f4;
    --md-border-strong: #e2e6ee;
    --md-text-1: #0f172a;
    --md-text-2: #475569;
    --md-text-3: #94a3b8;
    --md-brand-1: #06b6d4;
    --md-brand-2: #0891b2;
    --md-brand-3: #155e75;
    --md-accent-blue: #3b82f6;
    --md-accent-violet: #8b5cf6;
    --md-accent-pink: #ec4899;
    --md-accent-green: #10b981;
    --md-accent-amber: #f59e0b;
    --md-accent-red: #ef4444;
    --md-soft-blue: #eff6ff;
    --md-soft-violet: #f5f3ff;
    --md-soft-pink: #fdf2f8;
    --md-soft-green: #ecfdf5;
    --md-soft-amber: #fffbeb;
    --md-shadow-sm: 0 1px 2px rgba(15,23,42,0.04);
    --md-shadow-md: 0 4px 14px rgba(15,23,42,0.06);
    --md-shadow-lg: 0 12px 32px rgba(15,23,42,0.08);
}

.stApp {
    background:
        radial-gradient(ellipse 60% 40% at top right, rgba(6, 182, 212, 0.06), transparent 60%),
        radial-gradient(ellipse 50% 30% at bottom left, rgba(139, 92, 246, 0.05), transparent 60%),
        var(--md-bg) !important;
}
.main .block-container {
    max-width: 1280px !important;
    padding: 1.4rem 1.6rem 2rem 1.6rem !important;
}

/* Remove leaked sidebar collapse glyph container ("keyboard_double_arrow_left"). */
[data-testid="stSidebar"] > div > div:first-child {
    display: none !important;
}
* { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important; }

/* Compliance / status strip */
@keyframes mdPulse {
    0% { box-shadow: 0 0 0 0 rgba(16,185,129,0.5); }
    70% { box-shadow: 0 0 0 8px rgba(16,185,129,0); }
    100% { box-shadow: 0 0 0 0 rgba(16,185,129,0); }
}

/* Greeting */
.md-greet-wrap { margin-bottom: 1.2rem; }
.md-greet {
    font-family: 'Inter', sans-serif;
    font-size: 1.85rem;
    font-weight: 700;
    color: var(--md-text-1);
    letter-spacing: -0.02em;
    line-height: 1.15;
    margin-bottom: 0.25rem;
}
.md-subgreet {
    font-size: 0.95rem;
    color: var(--md-text-2);
    font-weight: 400;
}

/* Quick action chips */
.md-chip {
    background: var(--md-surface);
    border: 1px solid var(--md-border);
    border-radius: 100px;
    padding: 0.55rem 0.95rem;
    font-size: 0.82rem;
    font-weight: 500;
    color: var(--md-text-1);
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    box-shadow: var(--md-shadow-sm);
    transition: all .18s ease;
}
.md-chip:hover { border-color: var(--md-brand-1); transform: translateY(-1px); box-shadow: var(--md-shadow-md); }
/* Style streamlit columns containing chip buttons to match */
div[data-testid="stHorizontalBlock"] .stButton > button[kind="secondary"].md-chip-btn,
.md-chip-row .stButton > button {
    background: var(--md-surface) !important;
    border: 1px solid var(--md-border) !important;
    color: var(--md-text-1) !important;
    border-radius: 100px !important;
    padding: 0.55rem 0.95rem !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
    height: auto !important;
    min-height: 0 !important;
    box-shadow: var(--md-shadow-sm) !important;
}
.md-chip-row .stButton > button:hover {
    border-color: var(--md-brand-1) !important;
    transform: translateY(-1px);
    box-shadow: var(--md-shadow-md) !important;
}
.md-chip-row .stButton > button p,
.md-chip-row .stButton > button div {
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    margin: 0 !important;
}

/* Hero card */
@keyframes mdFloat { 0%, 100% { transform: translateY(0); } 50% { transform: translateY(-4px); } }
.md-hp-green { background: var(--md-soft-green); color: #047857; }
.md-hp-violet { background: var(--md-soft-violet); color: #6d28d9; }
.md-hp-blue { background: var(--md-soft-blue); color: #1d4ed8; }
.md-hp-pink { background: var(--md-soft-pink); color: #be185d; }

/* Composer wrapper */

/* Smart Actions panel */
.md-smart-head {
    display: flex; align-items: center; justify-content: space-between;
    margin: 0.4rem 0 0.6rem 0;
}
.md-smart-title { font-size: 1rem; font-weight: 700; color: var(--md-text-1); }
@media (max-width: 900px) { .md-smart-grid { grid-template-columns: repeat(2, 1fr); } }

/* Right column dashboard cards */
.md-rcard {
    background: var(--md-surface);
    border: 1px solid var(--md-border);
    border-radius: 16px;
    padding: 1rem 1.1rem;
    margin-bottom: 0.85rem;
    box-shadow: var(--md-shadow-sm);
}
.md-rcard-head {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 0.7rem;
}
.md-rcard-title { font-size: 0.88rem; font-weight: 700; color: var(--md-text-1); }
.md-rcard-link { font-size: 0.72rem; color: var(--md-brand-2); font-weight: 600; }
.md-rcard-link-btn {
    text-decoration: none !important;
    display: inline-flex;
    align-items: center;
    border-radius: 8px;
    padding: 0.08rem 0.34rem;
    border: 1px solid transparent;
    transition: all 0.15s ease;
}
.md-rcard-link-btn:hover {
    border-color: #d6e2fb;
    background: #f7fbff;
}

/* Health Overview metric rows */
.md-metric-row {
    display: flex; align-items: center; gap: 0.6rem;
    padding: 0.55rem 0;
    border-top: 1px solid var(--md-border);
}
.md-metric-row:first-of-type { border-top: none; }
.md-metric-icon {
    width: 34px; height: 34px;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 0.95rem;
    flex-shrink: 0;
}
.md-metric-mid { flex: 1; min-width: 0; }
.md-metric-label { font-size: 0.78rem; color: var(--md-text-2); font-weight: 500; }
.md-metric-value { font-size: 0.95rem; color: var(--md-text-1); font-weight: 700; }
.md-metric-status {
    font-size: 0.7rem; font-weight: 700; padding: 0.15rem 0.5rem;
    border-radius: 100px; flex-shrink: 0;
}
.md-status-good { background: var(--md-soft-green); color: #047857; }
.md-status-warn { background: var(--md-soft-amber); color: #92400e; }
.md-status-info { background: var(--md-soft-blue); color: #1d4ed8; }

/* Recent Conversations rows */
.md-conv-row {
    display: flex; align-items: center; gap: 0.55rem;
    padding: 0.5rem 0;
    border-top: 1px solid var(--md-border);
    font-size: 0.78rem;
}
.md-conv-row:first-of-type { border-top: none; }
.md-conv-bubble { width: 26px; height: 26px; border-radius: 8px; background: var(--md-soft-blue); color: #1d4ed8; display: flex; align-items: center; justify-content: center; font-size: 0.85rem; flex-shrink: 0; }
.md-conv-title { flex: 1; color: var(--md-text-1); font-weight: 500; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.md-conv-time { font-size: 0.68rem; color: var(--md-text-3); flex-shrink: 0; }

/* Health Tip card */
.md-tip {
    background: linear-gradient(135deg, #ecfeff, #fff7ed);
    border: 1px solid var(--md-border);
    border-radius: 16px;
    padding: 1rem 1.1rem;
    box-shadow: var(--md-shadow-sm);
    display: flex; align-items: center; gap: 0.8rem;
}
.md-tip-eyebrow { font-size: 0.62rem; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; color: var(--md-brand-2); margin-bottom: 0.2rem; }
.md-tip-title { font-size: 0.95rem; font-weight: 700; color: var(--md-text-1); margin-bottom: 0.2rem; }
.md-tip-desc { font-size: 0.75rem; color: var(--md-text-2); line-height: 1.45; }

/* MediChat new logo */
.md-logo-wrap {
    display: flex; align-items: center; gap: 0.6rem;
    padding: 0.4rem 0.2rem 1rem 0.2rem;
    margin-bottom: 0.4rem;
    border-bottom: 1px solid var(--md-border);
}
.md-logo-mark {
    width: 40px; height: 40px;
    border-radius: 12px;
    background: linear-gradient(135deg, #06b6d4, #155e75);
    display: flex; align-items: center; justify-content: center;
    color: white; font-size: 1.2rem; font-weight: 700;
    box-shadow: 0 4px 12px rgba(6,182,212,0.3);
}
.md-logo-text { font-size: 1.1rem; font-weight: 800; color: var(--md-text-1); letter-spacing: -0.01em; }
.md-logo-sub { font-size: 0.66rem; color: var(--md-text-3); font-weight: 500; }

/* Sidebar nav */
[data-testid="stSidebar"] {
    background: var(--md-surface) !important;
    border-right: 1px solid var(--md-border) !important;
}
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0.15rem !important; }
[data-testid="stSidebar"] [data-testid="element-container"] { margin-bottom: 0 !important; }
[data-testid="stSidebar"] hr { margin: 0.6rem 0 !important; }
[data-testid="stSidebar"] .stButton { margin: 0 !important; }
[data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    border: none !important;
    color: var(--md-text-2) !important;
    text-align: left !important;
    padding: 0.55rem 0.8rem !important;
    border-radius: 10px !important;
    font-weight: 500 !important;
    font-size: 0.86rem !important;
    height: 38px !important;
    min-height: 38px !important;
    margin: 0 !important;
    box-shadow: none !important;
    justify-content: flex-start !important;
    display: flex !important;
    line-height: 1 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: var(--md-bg) !important;
    color: var(--md-text-1) !important;
    transform: none;
}
.md-nav-active .stButton > button {
    background: var(--md-soft-blue) !important;
    color: var(--md-accent-blue) !important;
    font-weight: 600 !important;
}
/* Sign in / Sign out buttons in sidebar should still look like buttons */
[data-testid="stSidebar"] .md-side-action .stButton > button {
    background: var(--md-surface) !important;
    border: 1px solid var(--md-border-strong) !important;
    color: var(--md-text-1) !important;
    justify-content: center !important;
    text-align: center !important;
}
[data-testid="stSidebar"] .md-side-action .stButton > button:hover {
    border-color: var(--md-brand-1) !important;
    background: var(--md-bg) !important;
}

/* Premium card */
.md-premium .stButton > button {
    background: white !important;
    color: #6366f1 !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 0.5rem 0.8rem !important;
    width: 100% !important;
    text-align: center !important;
    justify-content: center !important;
    font-size: 0.78rem !important;
}

/* Sidebar profile chip */
.md-side-profile {
    display: flex; align-items: center; gap: 0.55rem;
    padding: 0.55rem;
    border-radius: 12px;
    background: var(--md-bg);
    margin: 0.5rem 0;
    border: 1px solid var(--md-border);
}
.md-side-avatar {
    width: 32px; height: 32px;
    border-radius: 10px;
    background: linear-gradient(135deg, #06b6d4, #8b5cf6);
    color: white;
    display: flex; align-items: center; justify-content: center;
    font-weight: 700; font-size: 0.85rem; flex-shrink: 0;
}
.md-side-pname { font-size: 0.82rem; font-weight: 600; color: var(--md-text-1); line-height: 1.1; }
.md-side-psub { font-size: 0.67rem; color: #4f6280; font-weight: 500; }

/* Hide old header card and trust strip while we use new ones */
.header-card, .trust-strip { display: none !important; }
.welcome-card, .hero-wrap { display: none !important; }

/* Make the chat input form match the new composer */
.stForm {
    background: transparent !important;
    border: none !important;
    padding: 0 !important;
}

/* ================================================================
   UI POLISH PASS — alignment, overflow, typography, premium feel
   ================================================================ */

/* Container hygiene: prevent cross-element overflow */
.md-rcard, .md-tip, .md-snap-card, .md-wearable-card {
    overflow: hidden;
}
.md-rcard * , .md-tip *, .md-greet, .md-subgreet {
    min-width: 0;
}

/* Section header (md-greet) — guarantee no truncation/overlap */
.md-greet-wrap {
    margin-bottom: 1.4rem;
    padding-right: 0.5rem;
}
.md-greet {
    font-size: 1.9rem;
    font-weight: 700;
    line-height: 1.2;
    word-break: break-word;
    overflow-wrap: anywhere;
    white-space: normal;
    letter-spacing: -0.02em;
}
.md-subgreet {
    font-size: 0.92rem;
    line-height: 1.5;
    word-break: break-word;
    overflow-wrap: anywhere;
    white-space: normal;
    max-width: 60ch;
}

/* Standard icon container — uniform shape, never clip text */
.md-metric-icon, .md-snap-icon, .md-conv-bubble {
    overflow: hidden;
    text-align: center;
    line-height: 1;
}
.md-metric-icon {
    width: 38px !important;
    height: 38px !important;
    border-radius: 11px !important;
    font-size: 1rem !important;
}

/* Status badge — never clipped, always rounded uniform */
.md-metric-status {
    white-space: nowrap;
    overflow: visible;
    line-height: 1.4;
    padding: 0.22rem 0.6rem;
    font-size: 0.68rem;
    flex-shrink: 0;
}

/* Card padding/typography polish */
.md-rcard {
    padding: 1.1rem 1.2rem;
    margin-bottom: 1rem;
    border-radius: 18px;
}
.md-rcard-title {
    font-size: 0.95rem;
    font-weight: 700;
    letter-spacing: -0.005em;
    line-height: 1.3;
}
.md-rcard-link {
    font-size: 0.75rem;
    font-weight: 600;
}
.md-metric-label {
    font-size: 0.76rem;
    line-height: 1.3;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}
.md-metric-value {
    font-size: 1rem;
    font-weight: 700;
    line-height: 1.3;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}
.md-metric-row {
    padding: 0.6rem 0;
    gap: 0.7rem;
}

/* ── Snapshot grid (uniform height + width tiles) ── */
.md-snap-grid {
    display: flex;
    flex-direction: column;
    gap: 0.5rem;
}
.md-snap-tile {
    display: flex;
    align-items: center;
    gap: 0.7rem;
    padding: 0.65rem 0.8rem;
    border-radius: 12px;
    background: var(--md-bg);
    border: 1px solid var(--md-border);
    min-height: 52px;
    overflow: hidden;
}
.md-snap-icon {
    width: 34px;
    height: 34px;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.95rem;
    flex-shrink: 0;
    overflow: hidden;
    line-height: 1;
}
.md-snap-text {
    flex: 1;
    min-width: 0;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 0.5rem;
}
.md-snap-label {
    font-size: 0.8rem;
    color: var(--md-text-2);
    font-weight: 500;
    line-height: 1.2;
    flex: 1;
    min-width: 0;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}
.md-snap-value {
    font-size: 1.05rem;
    font-weight: 700;
    color: var(--md-text-1);
    line-height: 1.2;
    flex-shrink: 0;
}

/* ── Recent Conversations rows ── */
.md-rcard-recent .md-rcard-head { margin-bottom: 0.3rem; }
.md-conv-row {
    padding: 0.6rem 0;
    gap: 0.65rem;
    min-height: 42px;
}
.md-conv-empty {
    color: var(--md-text-3) !important;
    font-style: italic;
    border-top: none !important;
    padding: 0.7rem 0 !important;
    font-size: 0.78rem;
}
.md-view-all-wrap {
    margin: -0.6rem 0 1rem 0;
}
.md-view-all-wrap .stButton > button {
    background: transparent !important;
    color: var(--md-brand-2) !important;
    border: 1px solid var(--md-border) !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
    font-size: 0.78rem !important;
    padding: 0.5rem 0.8rem !important;
    height: auto !important;
}
.md-view-all-wrap .stButton > button:hover {
    background: var(--md-soft-blue) !important;
    border-color: var(--md-brand-1) !important;
}

/* ── Sidebar past chats — clean uniform list ── */
/* Past-chats list styling now lives in the compact-mode @media block —
   that's the single source of truth. */

/* ── Wearable sync card (replaces simulated HR/Steps) ── */
.md-wearable-card {
    display: flex;
    align-items: center;
    gap: 1.1rem;
    padding: 1.4rem 1.5rem;
    background: linear-gradient(135deg, #f0f9ff 0%, #ecfeff 50%, #f5f3ff 100%);
    border: 1px solid var(--md-border);
    border-radius: 18px;
    box-shadow: var(--md-shadow-sm);
    margin-bottom: 1.2rem;
    overflow: hidden;
}
.md-wearable-icon {
    width: 56px;
    height: 56px;
    border-radius: 16px;
    background: linear-gradient(135deg, #06b6d4, #6366f1);
    color: white;
    font-size: 1.6rem;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
    box-shadow: 0 4px 14px rgba(6,182,212,0.3);
    line-height: 1;
}
.md-wearable-body { flex: 1; min-width: 0; }
.md-wearable-title {
    font-size: 1.05rem;
    font-weight: 700;
    color: var(--md-text-1);
    margin-bottom: 0.25rem;
    line-height: 1.3;
}
.md-wearable-desc {
    font-size: 0.82rem;
    color: var(--md-text-2);
    line-height: 1.5;
    margin-bottom: 0.7rem;
    max-width: 65ch;
}
.md-wearable-actions {
    display: flex;
    gap: 0.45rem;
    flex-wrap: wrap;
}
.md-wearable-pill {
    display: inline-flex;
    align-items: center;
    gap: 0.3rem;
    padding: 0.32rem 0.7rem;
    background: white;
    border: 1px solid var(--md-border);
    border-radius: 100px;
    font-size: 0.72rem;
    font-weight: 600;
    color: var(--md-text-2);
    white-space: nowrap;
}
.md-wearable-pill.md-wearable-soon {
    background: #fef3c7;
    border-color: #fde68a;
    color: #92400e;
}

/* ── History list page ── */
.md-history-list {
    display: flex;
    flex-direction: column;
    gap: 0.75rem;
    margin-top: 1rem;
}
.md-history-row {
    display: flex;
    align-items: flex-start;
    gap: 0.8rem;
    padding: 1rem 1.1rem;
    background: var(--md-surface);
    border: 1px solid var(--md-border);
    border-radius: 14px;
    box-shadow: var(--md-shadow-sm);
}
.md-history-bubble {
    width: 40px;
    height: 40px;
    border-radius: 12px;
    background: var(--md-soft-blue);
    color: var(--md-accent-blue);
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1.1rem;
    flex-shrink: 0;
}
.md-history-mid { flex: 1; min-width: 0; }
.md-history-title {
    font-size: 0.95rem;
    font-weight: 700;
    color: var(--md-text-1);
    line-height: 1.3;
    word-break: break-word;
}
.md-history-meta {
    font-size: 0.72rem;
    color: var(--md-text-3);
    margin-top: 0.2rem;
}
.md-history-preview {
    font-size: 0.78rem;
    color: var(--md-text-2);
    margin-top: 0.45rem;
    line-height: 1.4;
    overflow: hidden;
    text-overflow: ellipsis;
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
}

/* Generic: prevent any text from leaking out of badges/buttons */
.md-status-good, .md-status-warn, .md-status-info,
.md-chip, .trust-pill, .md-wearable-pill {
    overflow: visible;
    white-space: nowrap;
    text-overflow: clip;
}

/* Form field spacing inside expanders / appointments to avoid label clipping */
[data-testid="stExpander"] [data-testid="stForm"] label {
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    color: var(--md-text-2) !important;
    margin-bottom: 0.2rem !important;
    white-space: normal !important;
    overflow: visible !important;
    text-overflow: clip !important;
    line-height: 1.3 !important;
}
[data-testid="stExpander"] .stDateInput, [data-testid="stExpander"] .stTimeInput {
    margin-bottom: 0.4rem;
}
[data-testid="stExpander"] .stTextInput input,
[data-testid="stExpander"] .stTextArea textarea {
    font-size: 0.85rem !important;
}

/* Premium negative space: increase gap between sections in main column */
.main .block-container {
    padding-top: 2rem !important;
}
</style>
""", unsafe_allow_html=True)

# ── Final premium cleanup pass (reference-aligned) ─────────────────────
st.markdown("""
<style>
.stApp {
    background: #f7f8fc !important;
}
.main .block-container {
    max-width: 1260px !important;
}

[data-testid="stToolbar"],
[data-testid="stDecoration"],
#MainMenu {
    display: none !important;
}

.md-logo-image {
    width: 58px;
    height: 58px;
    border-radius: 18px;
    object-fit: cover;
    flex-shrink: 0;
    box-shadow: 0 12px 32px rgba(37, 99, 235, 0.18);
}

.md-logo-wrap {
    gap: 0.85rem !important;
    padding-bottom: 1.5rem !important;
}
.md-logo-text {
    font-size: 1.22rem !important;
    letter-spacing: 0 !important;
}
.md-logo-sub {
    font-size: 0.76rem !important;
}

.md-home-greet-wrap {
    text-align: center;
    margin: 0.35rem auto 1.35rem auto;
}
.md-home-greet-wrap .md-subgreet {
    margin: 0 auto;
    max-width: 48ch;
}

.main .stButton > button {
    background: rgba(255, 255, 255, 0.95) !important;
    border: 1px solid #e4eaf3 !important;
    border-radius: 14px !important;
    color: #18233c !important;
    box-shadow: 0 3px 10px rgba(15, 23, 42, 0.04) !important;
    transition: all 0.18s ease !important;
}
.main .stButton > button:hover {
    border-color: #cfd9ea !important;
    background: #ffffff !important;
    transform: translateY(-1px);
}

.stForm [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
.stForm [data-testid="stFormSubmitButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, #3b82f6, #6366f1) !important;
    border: none !important;
    color: #ffffff !important;
    border-radius: 16px !important;
    box-shadow: 0 10px 24px rgba(79, 70, 229, 0.24) !important;
}
.stForm [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"]:hover,
.stForm [data-testid="stFormSubmitButton"] > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #2563eb, #4f46e5) !important;
}

.md-chip-row .stButton > button,
.md-chip-row-compact .stButton > button {
    min-height: 42px !important;
    border-radius: 999px !important;
    font-size: 0.83rem !important;
    font-weight: 560 !important;
    padding: 0.34rem 0.68rem !important;
}


form#home_chat_form [data-testid="stTextArea"] textarea,
[data-testid="stForm"] [data-testid="stTextArea"] textarea {
    border-radius: 20px !important;
    border: 1px solid #e2eaf6 !important;
    min-height: 130px !important;
    font-size: 1.03rem !important;
    line-height: 1.45 !important;
    padding: 0.95rem 1.05rem !important;
    box-shadow: inset 0 1px 2px rgba(15, 23, 42, 0.02) !important;
}

form#home_chat_form [data-testid="stTextArea"] textarea:focus,
[data-testid="stForm"] [data-testid="stTextArea"] textarea:focus {
    border-color: #c6d8f8 !important;
    box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12) !important;
}

form#home_chat_form [data-testid="stFormSubmitButton"] > button,
[data-testid="stForm"] [data-testid="stFormSubmitButton"] > button {
    min-height: 44px !important;
    border-radius: 999px !important;
    font-size: 0.9rem !important;
    font-weight: 600 !important;
}

form#home_chat_form [data-testid="stFormSubmitButton"] button p,
[data-testid="stForm"] [data-testid="stFormSubmitButton"] button p {
    white-space: nowrap !important;
}


@keyframes mdFadeUp {
    from { opacity: 0; transform: translateY(8px); }
    to { opacity: 1; transform: translateY(0); }
}

.md-home-composer-note {
    margin-top: 0.35rem !important;
    margin-bottom: 1.25rem !important;
    color: #64748b !important;
}

.md-accent-purple { background: #f2ebff; color: #7c3aed; }
.md-accent-green { background: #e9fbf3; color: #10b981; }
.md-accent-pink { background: #fff0f6; color: #ef4f85; }
.md-accent-blue { background: #eaf6ff; color: #1d8cf8; }

.md-vision-panel,
.md-voice-panel {
    display: flex;
    align-items: center;
    gap: 0.85rem;
    margin: 0.35rem 0 0.75rem 0;
    padding: 0.9rem 1rem;
    border-radius: 22px;
    background: rgba(255,255,255,0.84);
    border: 1px solid #e6eef8;
    box-shadow: 0 12px 30px rgba(15,23,42,0.045);
    animation: mdFadeUp 0.3s ease both;
}
.md-panel-icon {
    width: 46px;
    height: 46px;
    border-radius: 15px;
    display: flex !important;
    align-items: center;
    justify-content: center;
    background: linear-gradient(135deg, #e0f2fe, #eef2ff);
    color: #2563eb;
    font-size: 1.45rem !important;
    flex-shrink: 0;
}
.md-panel-title {
    font-size: 0.95rem;
    font-weight: 760;
    color: #0f172a;
    line-height: 1.2;
}
.md-panel-subtitle {
    font-size: 0.78rem;
    color: #64748b;
    line-height: 1.45;
    margin-top: 0.18rem;
}
.md-home-file-preview {
    border-radius: 18px !important;
    background: rgba(255,255,255,0.92) !important;
}

.md-snap-card {
    padding: 1.05rem 1.05rem !important;
}
.md-snap-grid {
    gap: 0.38rem !important;
}
.md-snap-tile {
    min-height: 68px !important;
    padding: 0.72rem 0.78rem !important;
    border-radius: 18px !important;
    gap: 0.78rem !important;
}
.md-snap-label {
    font-size: 0.74rem !important;
}
.md-snap-value {
    font-size: 0.92rem !important;
}
.md-spark {
    width: 74px;
    height: 28px;
    flex-shrink: 0;
}
.md-spark path {
    fill: none;
    stroke-width: 2.1;
    stroke-linecap: round;
    opacity: 0.78;
}
.md-line-pink path { stroke: #fb7185; }
.md-line-green path { stroke: #34d399; }
.md-line-purple path { stroke: #a78bfa; }
.md-line-blue path { stroke: #38bdf8; }

.md-tip {
    border-radius: 20px !important;
}

.md-sidebar-bottom {
    position: sticky;
    bottom: 0;
    z-index: 10;
    padding-top: 0.65rem;
    margin-top: 0.85rem;
    background: linear-gradient(to top, rgba(255,255,255,0.98), rgba(255,255,255,0.88) 70%, rgba(255,255,255,0));
}
.md-sidebar-bottom .stButton > button {
    justify-content: center !important;
    text-align: center !important;
    min-height: 42px !important;
    border-radius: 12px !important;
    font-weight: 650 !important;
}

@media (max-width: 980px) {
    .main .block-container {
        padding: 0.95rem 0.8rem 1.8rem 0.8rem !important;
    }
    .md-greet {
        font-size: 1.52rem !important;
    }
    form#home_chat_form,
    [data-testid="stForm"] [data-testid="stTextArea"] {
        border-radius: 20px !important;
        padding: 0.85rem !important;
    }
    .md-smart-grid-buttons .stButton > button {
        min-height: 96px !important;
        font-size: 0.82rem !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ── UX refinement pass (final cascade layer) ─────────────────────────
st.markdown("""
<style>
:root {
    --md-radius-sm: 10px;
    --md-radius-md: 14px;
    --md-radius-lg: 18px;
}

.main .block-container {
    max-width: 1220px !important;
    padding: 1.4rem 1.5rem 2.5rem 1.5rem !important;
}

/* Streamlit uses Material Symbols for expander/menu icons. The older global
   font override made those icons render as text such as "keyboard_double". */
.material-icons,
.material-icons-round,
.material-symbols-outlined,
.material-symbols-rounded,
.material-symbols-sharp,
[class*="material-symbols"],
[class*="material-icons"] {
    font-family: "Material Symbols Rounded", "Material Symbols Outlined", "Material Icons Round", "Material Icons" !important;
    font-weight: normal !important;
    font-style: normal !important;
    line-height: 1 !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-flex !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: "liga" !important;
    -webkit-font-smoothing: antialiased !important;
}
[data-testid="stIconMaterial"],
[data-testid="stIconMaterial"] * {
    font-family: "Material Symbols Rounded", "Material Symbols Outlined", "Material Icons Round", "Material Icons" !important;
    font-weight: normal !important;
    font-style: normal !important;
    line-height: 1 !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    white-space: nowrap !important;
    -webkit-font-feature-settings: "liga" !important;
}
[data-testid="collapsedControl"] [data-testid="stIconMaterial"] {
    font-size: 0 !important;
}
[data-testid="collapsedControl"] [data-testid="stIconMaterial"]::after {
    content: "☰";
    font-size: 1.2rem;
    color: #334155;
}

/* App shell closer to the reference mockup */
.stApp {
    background:
        radial-gradient(circle at 38% 8%, rgba(59,130,246,0.08), transparent 32%),
        radial-gradient(circle at 92% 20%, rgba(139,92,246,0.06), transparent 28%),
        #f7f9fd !important;
}
[data-testid="stSidebar"] {
    width: 292px !important;
    min-width: 292px !important;
    background: rgba(255,255,255,0.94) !important;
    border-right: 1px solid rgba(226,232,240,0.9) !important;
    box-shadow: 12px 0 36px rgba(15,23,42,0.035) !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    padding: 1.65rem 1rem 1.25rem 1rem !important;
}
.md-logo-wrap {
    border-bottom: none !important;
    padding: 0.35rem 0.45rem 1.55rem 0.45rem !important;
    gap: 0.88rem !important;
}
.md-logo-mark {
    width: 58px !important;
    height: 58px !important;
    border-radius: 18px !important;
    background: linear-gradient(135deg, #38bdf8, #2563eb 54%, #8b5cf6) !important;
    box-shadow: 0 14px 32px rgba(37,99,235,0.22) !important;
}
.md-logo-text {
    font-size: 1.24rem !important;
}
.md-logo-sub {
    font-size: 0.76rem !important;
}
[data-testid="stSidebar"] hr {
    margin: 0.75rem 0 !important;
    border-color: #eef2f7 !important;
}
[data-testid="stSidebar"] .stButton > button {
    height: 48px !important;
    min-height: 48px !important;
    padding: 0.65rem 0.85rem !important;
    border-radius: 13px !important;
    font-size: 0.93rem !important;
    color: #334155 !important;
    font-weight: 600 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #f4f7ff !important;
    color: #1d4ed8 !important;
}
.md-nav-active .stButton > button {
    background: linear-gradient(135deg, #eef2ff, #f5f3ff) !important;
    color: #1d4ed8 !important;
    box-shadow: inset 0 0 0 1px rgba(96,165,250,0.15) !important;
}

/* Sidebar nav alignment: fixed icon column, fixed text start, active pill */
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) {
    width: 100% !important;
    height: 52px !important;
    min-height: 52px !important;
    padding: 0 1.05rem !important;
    border-radius: 16px !important;
    justify-content: flex-start !important;
    text-align: left !important;
}
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) > div {
    width: 100% !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
}
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) [data-testid="stIconMaterial"] {
    width: 24px !important;
    min-width: 24px !important;
    height: 24px !important;
    margin-right: 0.86rem !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    color: currentColor !important;
    font-size: 1.18rem !important;
}
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) p,
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) span:not([data-testid="stIconMaterial"]) {
    margin: 0 !important;
    text-align: left !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    line-height: 1.1 !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #eef4ff, #f5f3ff) !important;
    color: #1d4ed8 !important;
    border: 1px solid rgba(147,197,253,0.28) !important;
    box-shadow: inset 0 0 0 1px rgba(96,165,250,0.12), 0 10px 24px rgba(37,99,235,0.08) !important;
}
[data-testid="stSidebar"] .stButton > button[kind="secondary"]:has([data-testid="stIconMaterial"]) {
    background: transparent !important;
    border: 1px solid transparent !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] .stButton > button[kind="secondary"]:has([data-testid="stIconMaterial"]):hover {
    background: #f5f7ff !important;
    border-color: rgba(226,232,240,0.9) !important;
    color: #1d4ed8 !important;
}
.md-side-profile {
    margin-top: 1rem !important;
    padding: 0.75rem !important;
    background: #ffffff !important;
    border: 1px solid #edf2f7 !important;
    box-shadow: var(--md-shadow-sm) !important;
}
.md-side-avatar {
    background: linear-gradient(135deg, #38bdf8, #2563eb) !important;
}
.sb-footer {
    border-top: none !important;
    color: #94a3b8 !important;
    padding-top: 0.7rem !important;
}

/* Keep the app clinical and calm: remove decorative orb treatment. */
.md-hero::before,
.md-hero::after,

.md-name-card,
.md-feedback-panel,
.md-download-card,
.md-file-preview,
.md-inline-note {
    background: var(--md-surface);
    border: 1px solid var(--md-border);
    border-radius: var(--md-radius-lg);
    box-shadow: var(--md-shadow-sm);
.md-name-subtitle {
    color: var(--md-text-2);
    font-size: 0.9rem;
    line-height: 1.55;
}
.md-name-card {
    padding: 1rem 1.1rem;
    margin: 0.7rem 0 0.7rem 0;
}
.md-name-title {
    font-size: 0.98rem;
    font-weight: 750;
    color: var(--md-text-1);
    margin-bottom: 0.15rem;
}
.md-file-preview {
    display: flex;
    align-items: center;
    gap: 0.75rem;
    padding: 0.8rem 1rem;
    margin: 0.7rem 0;
}
.md-file-icon {
    width: 42px;
    height: 42px;
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: var(--md-soft-blue);
    color: var(--md-brand-3);
    font-weight: 800;
    font-size: 0.74rem;
    flex-shrink: 0;
}
.md-file-name {
    font-weight: 700;
    color: var(--md-text-1);
    font-size: 0.88rem;
    overflow-wrap: anywhere;
}
.md-file-help {
    color: var(--md-text-2);
    font-size: 0.76rem;
    line-height: 1.4;
}
.md-inline-note {
    padding: 0.8rem 0.95rem;
    color: var(--md-text-2);
    font-size: 0.82rem;
    line-height: 1.5;
}
.md-home-composer-note {
    font-size: 0.74rem;
    color: #64748b;
    text-align: center;
    margin: -0.3rem 0 1rem 0;
}
.md-home-composer-wrap + div [data-testid="stForm"],
.md-smart-head {
    margin-top: 1rem !important;
}
.md-smart-card + div .stButton > button,
.md-smart-card ~ div .stButton > button {
    border-radius: 14px !important;
}

.bot-bubble,
.user-bubble,
.source-row,
.confidence-row {
    word-break: break-word !important;
    overflow-wrap: break-word !important;
    white-space: pre-wrap !important;
}
.emergency-banner {
    animation: mdEmergencyPulse 2s ease-in-out infinite !important;
}
@keyframes mdEmergencyPulse {
    0%, 100% { box-shadow: 0 6px 20px rgba(220, 38, 38, 0.2); }
    50% { box-shadow: 0 6px 30px rgba(220, 38, 38, 0.42); }
}
.bot-bubble {
    max-width: min(86%, 760px);
}
/* User bubble width is now controlled per-mode in the compact-mode block
   (lets short messages stay one line, only long ones wrap). */
.source-row,
.confidence-row {
    max-width: calc(100% - 50px);
}

.stButton > button p,
.stButton > button div {
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
.main .stButton > button {
    min-width: 0 !important;
}
.main .stButton > button p {
    white-space: normal !important;
    line-height: 1.2 !important;
}

[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea,
[data-testid="stSelectbox"] div[data-baseweb="select"] {
    min-height: 44px;
}

/* Streamlit forms as clean cards, and no more broken expander chrome */
[data-testid="stForm"] {
    background: rgba(255,255,255,0.86) !important;
    border: 1px solid #dbeafe !important;
    border-radius: 18px !important;
    padding: 1rem !important;
    box-shadow: 0 10px 28px rgba(15,23,42,0.045) !important;
}
.md-form-intro {
    margin: 1.1rem 0 0.45rem 0;
    font-size: 0.86rem;
    font-weight: 800;
    color: #0f172a;
}
.md-form-sub {
    color: #64748b;
    font-size: 0.78rem;
    margin-top: -0.25rem;
    margin-bottom: 0.65rem;
}

/* Reference-style right rail */
.md-snap-card,
.md-rcard-recent,
.md-tip {
    border-radius: 20px !important;
    border: 1px solid #e8eef8 !important;
    box-shadow: 0 16px 34px rgba(15,23,42,0.055) !important;
}
.md-snap-tile {
    border: none !important;
    border-radius: 14px !important;
    background: #f7f9fd !important;
    min-height: 58px !important;
}
.md-tip {
    background: linear-gradient(135deg, #ecfeff, #fff7ed) !important;
}

@media (max-width: 980px) {
    [data-testid="stToolbar"],
    [data-testid="stDecoration"] {
        display: none !important;
    }
    [data-testid="collapsedControl"] {
        position: fixed !important;
        top: 0.55rem !important;
        left: 0.65rem !important;
        z-index: 1002 !important;
        background: rgba(255,255,255,0.92);
        border-radius: 10px;
        padding: 0.2rem;
        box-shadow: 0 4px 12px rgba(15,23,42,0.12);
    }
    [data-testid="stSidebar"] {
        width: min(86vw, 340px) !important;
        min-width: min(86vw, 340px) !important;
    }
    [data-testid="stSidebar"][aria-expanded="false"] {
        width: 0 !important;
        min-width: 0 !important;
    }
    .main .block-container {
        padding: 0.95rem 0.8rem 1.9rem 0.8rem !important;
    }
    .md-greet {
        font-size: 1.45rem;
        line-height: 1.25;
    }
    .md-subgreet {
        font-size: 0.86rem;
    }
    .bot-bubble,
    .user-bubble {
        max-width: calc(100% - 46px);
        font-size: 0.86rem;
    }
    .av {
        width: 34px;
        height: 34px;
        border-radius: 10px;
    }
    .bot-label,
    .source-row,
    .confidence-row {
        margin-left: 44px;
    }
}

@media (max-width: 640px) {
    .md-smart-head {
        margin-top: 0.4rem !important;
    }
    .md-file-preview {
        align-items: flex-start;
    }
    .md-tip {
        align-items: flex-start;
    }
    .md-snap-text {
        align-items: flex-start;
        flex-direction: column;
        gap: 0.1rem;
    }
    .confidence-row {
        flex-wrap: wrap;
    }
}
</style>
""", unsafe_allow_html=True)

# ── Client polish pass: sidebar, guest, cards, home controls ───────────
st.markdown("""
<style>
/* Quick option pills */
.md-chip-row-compact .stButton > button {
    min-height: 46px !important;
    border-radius: 999px !important;
    border: 1px solid #d7e8fb !important;
    background: rgba(255,255,255,0.94) !important;
    color: #101827 !important;
    box-shadow: 0 8px 20px rgba(15,23,42,0.045) !important;
    font-size: 0.94rem !important;
    font-weight: 650 !important;
}
.md-chip-row-compact .stButton > button [data-testid="stIconMaterial"] {
    width: 22px !important;
    min-width: 22px !important;
    height: 22px !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    margin-right: 0.45rem !important;
    overflow: visible !important;
    font-size: 1.1rem !important;
}
.md-chip-row-compact .stButton > button:hover {
    border-color: #93c5fd !important;
    box-shadow: 0 12px 28px rgba(59,130,246,0.12) !important;
    transform: translateY(-1px);
}

/* Home composer controls: compact upload/voice pills and round send */
form#home_chat_form [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
form#chat_form {
    background: rgba(255, 255, 255, 0.98) !important;
    border: 1px solid #e7edf6 !important;
    border-radius: 28px !important;
    box-shadow: 0 16px 36px rgba(15, 23, 42, 0.08) !important;
    padding: 1rem 1rem 0.85rem 1rem !important;
}
form#chat_form [data-testid="stTextArea"] textarea {
    border-radius: 20px !important;
    border: 1px solid #e2eaf6 !important;
    min-height: 130px !important;
    font-size: 1.03rem !important;
    line-height: 1.45 !important;
    padding: 0.95rem 1.05rem !important;
}
form#chat_form [data-testid="stTextArea"] textarea:focus {
    border-color: #c6d8f8 !important;
    box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12) !important;
}
form#chat_form [data-testid="stFormSubmitButton"] > button {
    height: 48px !important;
    min-height: 48px !important;
    border-radius: 999px !important;
}
form#chat_form [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
form#chat_form [data-testid="stFormSubmitButton"] > button[kind="primary"] {
    width: 52px !important;
    min-width: 52px !important;
    padding: 0 !important;
    justify-content: center !important;
    box-shadow: 0 14px 30px rgba(79,70,229,0.28) !important;
}
form#chat_form [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"] {
    padding: 0 1rem !important;
    border: 1px solid #d7e8fb !important;
    background: rgba(255,255,255,0.92) !important;
    box-shadow: 0 5px 14px rgba(15,23,42,0.035) !important;
}

/* Smart Actions as real Streamlit buttons, styled like compact glass cards.
   Tightened from 172px tall → 108px (~37% smaller) per user feedback. */
.md-smart-route .stButton > button {
    position: relative !important;
    min-height: 108px !important;
    border-radius: 16px !important;
    padding: 0.7rem 0.8rem 0.65rem 0.8rem !important;
    align-items: flex-start !important;
    justify-content: flex-start !important;
    text-align: left !important;
    background: linear-gradient(180deg, rgba(255,255,255,0.96), rgba(255,255,255,0.84)) !important;
    border: 1px solid rgba(226,232,240,0.96) !important;
    box-shadow: 0 6px 16px rgba(15,23,42,0.04) !important;
    color: #111827 !important;
    transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease !important;
    overflow: visible !important;
}
.md-smart-route .stButton > button:hover {
    transform: translateY(-2px) scale(1.005) !important;
    border-color: rgba(147,197,253,0.95) !important;
    box-shadow: 0 10px 24px rgba(15,23,42,0.07) !important;
}
.md-smart-route .stButton > button::after {
    content: "arrow_forward";
    font-family: "Material Symbols Rounded", "Material Symbols Outlined" !important;
    position: absolute;
    right: 0.6rem;
    bottom: 0.55rem;
    width: 22px;
    height: 22px;
    border-radius: 999px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: rgba(255,255,255,0.94);
    border: 1px solid #e6edf7;
    color: #0f172a;
    font-size: 0.85rem;
    line-height: 1;
}
.md-smart-route .stButton > button [data-testid="stIconMaterial"] {
    width: 32px !important;
    height: 32px !important;
    min-width: 32px !important;
    border-radius: 10px !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    margin: 0 0.55rem 0 0 !important;
    font-size: 1.05rem !important;
    overflow: visible !important;
}
.md-smart-purple .stButton > button [data-testid="stIconMaterial"] { background: #f2ebff; color: #7c3aed !important; }
.md-smart-green .stButton > button [data-testid="stIconMaterial"] { background: #e9fbf3; color: #10b981 !important; }
.md-smart-pink .stButton > button [data-testid="stIconMaterial"] { background: #fff0f6; color: #ef4f85 !important; }
.md-smart-blue .stButton > button [data-testid="stIconMaterial"] { background: #eaf6ff; color: #1d8cf8 !important; }
.md-smart-route .stButton > button p {
    white-space: pre-line !important;
    overflow: visible !important;
    text-overflow: clip !important;
    line-height: 1.3 !important;
    font-size: 0.7rem !important;
    color: #64748b !important;
    font-weight: 520 !important;
    margin: 0 !important;
}
.md-smart-route .stButton > button p::first-line {
    color: #111827;
    font-size: 0.82rem;
    font-weight: 700;
}

/* Guest mode: make the limited dashboard feel intentional */
.md-guest-card .stButton > button {
    min-height: 112px !important;
    border-radius: 22px !important;
    align-items: flex-start !important;
    justify-content: flex-start !important;
    text-align: left !important;
    padding: 0.95rem 1rem !important;
    border: 1px solid #e6eef8 !important;
    background: rgba(255,255,255,0.92) !important;
    box-shadow: 0 14px 32px rgba(15,23,42,0.05) !important;
}
.md-guest-card .stButton > button p {
    white-space: pre-line !important;
    line-height: 1.42 !important;
    font-size: 0.82rem !important;
}
.md-guest-blue .stButton > button [data-testid="stIconMaterial"] { color: #1d8cf8 !important; }
.md-guest-purple .stButton > button [data-testid="stIconMaterial"] { color: #7c3aed !important; }
.md-guest-green .stButton > button [data-testid="stIconMaterial"] { color: #10b981 !important; }

/* Right rail working link button */
.md-rail-link-btn {
    margin: -0.55rem 0 0.85rem 0;
}
.md-rail-link-btn .stButton > button {
    height: 38px !important;
    min-height: 38px !important;
    border-radius: 999px !important;
    border: 1px solid #d7e8fb !important;
    background: rgba(255,255,255,0.88) !important;
    color: #1d4ed8 !important;
    font-size: 0.75rem !important;
    font-weight: 760 !important;
    box-shadow: none !important;
}

/* Past conversations: see compact-mode @media block (single source of truth). */

/* Sidebar icons and language control */
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) {
    overflow: visible !important;
}
[data-testid="stSidebar"] .stButton > button [data-testid="stIconMaterial"] {
    flex: 0 0 24px !important;
    overflow: visible !important;
    line-height: 1 !important;
    text-align: center !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] {
    overflow: visible !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] {
    min-height: 42px !important;
    border-radius: 14px !important;
    background: #ffffff !important;
}

.md-tip-icon.material-symbols-rounded {
    width: 46px;
    height: 46px;
    border-radius: 16px;
    background: linear-gradient(135deg, #dbeafe, #ecfeff);
    color: #1d8cf8;
    display: inline-flex !important;
    align-items: center;
    justify-content: center;
    font-size: 1.55rem !important;
    flex-shrink: 0;
}

/* Smart Actions styling bound to widget keys so Streamlit layout wrappers
   cannot break it. Compacted to 108px tall + smaller icons/text per user. */
.st-key-sa_sym .stButton > button,
.st-key-sa_rec .stButton > button,
.st-key-sa_ins .stButton > button,
.st-key-sa_appt .stButton > button {
    position: relative !important;
    min-height: 108px !important;
    border-radius: 16px !important;
    padding: 0.7rem 0.8rem 0.65rem 0.8rem !important;
    align-items: flex-start !important;
    justify-content: flex-start !important;
    text-align: left !important;
    background: linear-gradient(180deg, rgba(255,255,255,0.96), rgba(255,255,255,0.84)) !important;
    border: 1px solid rgba(226,232,240,0.96) !important;
    box-shadow: 0 6px 16px rgba(15,23,42,0.04) !important;
    color: #111827 !important;
    overflow: visible !important;
    transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease !important;
}
.st-key-sa_sym .stButton > button:hover,
.st-key-sa_rec .stButton > button:hover,
.st-key-sa_ins .stButton > button:hover,
.st-key-sa_appt .stButton > button:hover {
    transform: translateY(-2px) !important;
    border-color: rgba(147,197,253,0.95) !important;
    box-shadow: 0 10px 24px rgba(15,23,42,0.07) !important;
}
.st-key-sa_sym .stButton > button::after,
.st-key-sa_rec .stButton > button::after,
.st-key-sa_ins .stButton > button::after,
.st-key-sa_appt .stButton > button::after {
    content: "arrow_forward";
    font-family: "Material Symbols Rounded", "Material Symbols Outlined" !important;
    position: absolute;
    right: 0.6rem;
    bottom: 0.55rem;
    width: 22px;
    height: 22px;
    border-radius: 999px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: rgba(255,255,255,0.94);
    border: 1px solid #e6edf7;
    color: #0f172a;
    font-size: 0.85rem;
    line-height: 1;
}
.st-key-sa_sym .stButton > button [data-testid="stIconMaterial"],
.st-key-sa_rec .stButton > button [data-testid="stIconMaterial"],
.st-key-sa_ins .stButton > button [data-testid="stIconMaterial"],
.st-key-sa_appt .stButton > button [data-testid="stIconMaterial"] {
    width: 32px !important;
    height: 32px !important;
    min-width: 32px !important;
    border-radius: 10px !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    margin: 0 0.55rem 0 0 !important;
    font-size: 1.05rem !important;
    overflow: visible !important;
}
.st-key-sa_sym .stButton > button [data-testid="stIconMaterial"] { background: #f2ebff; color: #7c3aed !important; }
.st-key-sa_rec .stButton > button [data-testid="stIconMaterial"] { background: #e9fbf3; color: #10b981 !important; }
.st-key-sa_ins .stButton > button [data-testid="stIconMaterial"] { background: #fff0f6; color: #ef4f85 !important; }
.st-key-sa_appt .stButton > button [data-testid="stIconMaterial"] { background: #eaf6ff; color: #1d8cf8 !important; }
.st-key-sa_sym .stButton > button p,
.st-key-sa_rec .stButton > button p,
.st-key-sa_ins .stButton > button p,
.st-key-sa_appt .stButton > button p {
    white-space: pre-line !important;
    overflow: visible !important;
    text-overflow: clip !important;
    line-height: 1.3 !important;
    font-size: 0.7rem !important;
    color: #64748b !important;
    font-weight: 520 !important;
    margin: 0 !important;
}
.st-key-sa_sym .stButton > button p::first-line,
.st-key-sa_rec .stButton > button p::first-line,
.st-key-sa_ins .stButton > button p::first-line,
.st-key-sa_appt .stButton > button p::first-line {
    color: #111827;
    font-size: 0.82rem;
    font-weight: 700;
}

/* Past-chats list styling consolidated in the compact-mode @media block. */

/* Keep icons visible and aligned in left navigation */
[data-testid="stSidebar"] .stButton > button > div {
    overflow: visible !important;
}
[data-testid="stSidebar"] .stButton > button [data-testid="stIconMaterial"] {
    opacity: 1 !important;
    visibility: visible !important;
    width: 22px !important;
    min-width: 22px !important;
    margin-right: 0.72rem !important;
}

/* Sidebar final override: make left navigation visibly different and cleaner */
[data-testid="stSidebar"] {
    width: 272px !important;
    min-width: 272px !important;
    background: #ffffff !important;
    border-right: 1px solid #e8eef9 !important;
    box-shadow: 10px 0 34px rgba(15,23,42,0.04) !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    padding: 1.35rem 0.85rem 1rem 0.85rem !important;
}
.sb-title {
    font-size: 0.68rem !important;
    color: #94a3b8 !important;
    letter-spacing: 0.14em !important;
    margin: 0.62rem 0 0.34rem 0 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button,
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button {
    min-height: 50px !important;
    height: 50px !important;
    border-radius: 14px !important;
    padding: 0 0.8rem !important;
    border: 1px solid transparent !important;
    background: transparent !important;
    box-shadow: none !important;
    color: #334155 !important;
    font-size: 0.92rem !important;
    font-weight: 620 !important;
    justify-content: flex-start !important;
    text-align: left !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button:hover,
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button:hover {
    background: #f6f9ff !important;
    border-color: #e2eafa !important;
    color: #1d4ed8 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button > div,
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button > div {
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    width: 100% !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button [data-testid="stIconMaterial"],
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button [data-testid="stIconMaterial"] {
    width: 22px !important;
    min-width: 22px !important;
    height: 22px !important;
    border-radius: 8px !important;
    background: #eef4ff;
    color: #3666ea !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    margin-right: 0.68rem !important;
    font-size: 1rem !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button p,
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button p {
    margin: 0 !important;
    text-align: left !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button[kind="primary"],
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #eef2ff, #eff6ff) !important;
    border-color: #d7e5fb !important;
    color: #1d4ed8 !important;
    box-shadow: inset 0 0 0 1px rgba(191,219,254,0.42) !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] button[kind="primary"] [data-testid="stIconMaterial"],
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button[kind="primary"] [data-testid="stIconMaterial"] {
    background: #2563eb !important;
    color: #ffffff !important;
}

/* Profile + sign out + new chat button polish */
[data-testid="stSidebar"] div.st-key-profile_logout button,
[data-testid="stSidebar"] div.st-key-profile_logout .stButton > button {
    min-height: 38px !important;
    height: 38px !important;
    border-radius: 11px !important;
    border: 1px solid #e5edf8 !important;
    background: #ffffff !important;
    color: #334155 !important;
    font-weight: 630 !important;
}
[data-testid="stSidebar"] div.st-key-profile_logout button:hover,
[data-testid="stSidebar"] div.st-key-profile_logout .stButton > button:hover {
    border-color: #cfe1fb !important;
    background: #f8fbff !important;
}
/* +New chat button styling moved to compact-mode @media (indigo primary CTA). */

/* Keep sidebar cards cleaner */
.md-side-profile {
    border-radius: 14px !important;
    padding: 0.66rem !important;
}

@media (max-width: 980px) {
    [data-testid="stSidebar"] {
        width: min(84vw, 320px) !important;
        min-width: min(84vw, 320px) !important;
    }
}

/* Header safeguard pills */
.md-home-head-left {
    text-align: left !important;
    margin-bottom: 0.8rem !important;
}
.md-safe-pill .material-symbols-rounded {
    font-size: 0.95rem !important;
    color: #2563eb;
}

/* Quick action cards */
.st-key-qa_headache .stButton > button,
.st-key-qa_tired .stButton > button,
.st-key-qa_symptoms .stButton > button,
.st-key-qa_sleep .stButton > button {
    min-height: 84px !important;
    border-radius: 18px !important;
    padding: 0.86rem 0.92rem !important;
    border: 1px solid #dce9fd !important;
    background: rgba(255,255,255,0.95) !important;
    box-shadow: 0 10px 24px rgba(15,23,42,0.045) !important;
    text-align: left !important;
    justify-content: flex-start !important;
    align-items: flex-start !important;
    color: #0f172a !important;
}
.st-key-qa_headache .stButton > button:hover,
.st-key-qa_tired .stButton > button:hover,
.st-key-qa_symptoms .stButton > button:hover,
.st-key-qa_sleep .stButton > button:hover {
    transform: translateY(-2px) !important;
    border-color: #bfd6fb !important;
    box-shadow: 0 14px 30px rgba(37,99,235,0.11) !important;
}
.st-key-qa_headache .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_tired .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_symptoms .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_sleep .stButton > button [data-testid="stIconMaterial"] {
    width: 34px !important;
    min-width: 34px !important;
    height: 34px !important;
    border-radius: 12px !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    margin-right: 0.55rem !important;
    font-size: 1.05rem !important;
    flex-shrink: 0 !important;
}
.st-key-qa_headache .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_tired .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_symptoms .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_sleep .stButton > button [data-testid="stMarkdownContainer"] {
    flex: 1 1 auto !important;
    min-width: 0 !important;
    overflow: visible !important;
    width: auto !important;
}
.st-key-qa_headache .stButton > button [data-testid="stMarkdownContainer"] p,
.st-key-qa_tired .stButton > button [data-testid="stMarkdownContainer"] p,
.st-key-qa_symptoms .stButton > button [data-testid="stMarkdownContainer"] p,
.st-key-qa_sleep .stButton > button [data-testid="stMarkdownContainer"] p {
    white-space: normal !important;
    text-overflow: clip !important;
    overflow: visible !important;
    max-width: 100% !important;
    width: auto !important;
}
.st-key-qa_headache .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(135deg,#ffe4e6,#fecdd3) !important; color: #e11d48 !important; }
.st-key-qa_tired .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(135deg,#fef3c7,#fde68a) !important; color: #d97706 !important; }
.st-key-qa_symptoms .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(135deg,#ede9fe,#ddd6fe) !important; color: #7c3aed !important; }
.st-key-qa_sleep .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(135deg,#e0e7ff,#c7d2fe) !important; color: #4f46e5 !important; }
.st-key-qa_headache .stButton > button p,
.st-key-qa_tired .stButton > button p,
.st-key-qa_symptoms .stButton > button p,
.st-key-qa_sleep .stButton > button p {
    white-space: pre-line !important;
    line-height: 1.34 !important;
    font-size: 0.8rem !important;
    color: #64748b !important;
}
.st-key-qa_headache .stButton > button p::first-line,
.st-key-qa_tired .stButton > button p::first-line,
.st-key-qa_symptoms .stButton > button p::first-line,
.st-key-qa_sleep .stButton > button p::first-line {
    color: #0f172a;
    font-size: 1rem;
    font-weight: 740;
}

/* Scrollable chat history area in sidebar */
.md-past-chats {
    max-height: 320px;
    overflow-y: auto;
    padding-right: 0.15rem;
}
.md-past-chats::-webkit-scrollbar {
    width: 6px;
}
.md-past-chats::-webkit-scrollbar-thumb {
    background: #d5e3f9;
    border-radius: 999px;
}

/* Center health tip card + learn more CTA */
.st-key-tip_learn_more .stButton > button {
    margin-top: 0.9rem !important;
    min-height: 40px !important;
    border-radius: 12px !important;
    border: 1px solid #cfe1fb !important;
    color: #1d4ed8 !important;
    font-weight: 700 !important;
}

/* Emergency help card */
.st-key-open_emergency_guidance .stButton > button {
    margin-top: 0.6rem !important;
    border-radius: 12px !important;
    min-height: 38px !important;
    font-size: 0.78rem !important;
}

@media (max-width: 980px) {
    .st-key-qa_headache .stButton > button,
    .st-key-qa_tired .stButton > button,
    .st-key-qa_symptoms .stButton > button,
    .st-key-qa_sleep .stButton > button {
        min-height: 74px !important;
    }
    .md-past-chats {
        max-height: 240px;
    }
}
</style>
""", unsafe_allow_html=True)

# ── Annotated QA polish pass (wins over all prior Streamlit/CSS output) ──
st.markdown("""
<style>
/* 5, 13: remove Streamlit chrome/collapse text entirely */
header[data-testid="stHeader"],
[data-testid="stHeader"],
.stAppHeader,
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="collapsedControl"],
#MainMenu,
footer {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
    min-height: 0 !important;
}

/* 11: compact desktop fit */
.main .block-container {
    max-width: 1180px !important;
    padding: 0.65rem 1.15rem 1.55rem 1.15rem !important;
}
.stApp {
    background:
        radial-gradient(circle at 46% 0%, rgba(59, 130, 246, 0.075), transparent 30%),
        radial-gradient(circle at 92% 4%, rgba(139, 92, 246, 0.06), transparent 30%),
        #f7f9fd !important;
}

/* 2, 3, 4: bigger logo, correct title/subtitle spacing */
[data-testid="stSidebar"] {
    width: 270px !important;
    min-width: 270px !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    padding: 1.05rem 0.9rem 1rem !important;
}
.md-logo-wrap {
    display: flex !important;
    align-items: center !important;
    gap: 0.92rem !important;
    padding: 1rem 0.55rem 1.35rem !important;
    margin-bottom: 0.35rem !important;
}
.md-logo-image,
.md-logo-mark {
    width: 64px !important;
    height: 64px !important;
    min-width: 64px !important;
    border-radius: 20px !important;
    object-fit: cover !important;
    box-shadow: 0 16px 36px rgba(37,99,235,0.16) !important;
}
.md-logo-text {
    font-size: 1.24rem !important;
    line-height: 1.05 !important;
    font-weight: 850 !important;
    color: #0f172a !important;
}
.md-logo-sub {
    font-size: 0.78rem !important;
    line-height: 1.25 !important;
    color: #8a9ab3 !important;
    margin-top: 0.25rem !important;
}

/* 1: no filled/highlighted home state, only subtle blue text/icon */
.md-nav-active .stButton > button,
[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: transparent !important;
    border-color: transparent !important;
    box-shadow: none !important;
    color: #2563eb !important;
}
[data-testid="stSidebar"] .stButton > button {
    margin: 0.16rem 0 !important;
    min-height: 44px !important;
    height: 44px !important;
    border-radius: 14px !important;
    font-size: 0.95rem !important;
}
[data-testid="stSidebar"] .stButton > button:has([data-testid="stIconMaterial"]) [data-testid="stIconMaterial"] {
    width: 24px !important;
    min-width: 24px !important;
    margin-right: 0.82rem !important;
    font-size: 1.12rem !important;
    overflow: visible !important;
}

/* 6, 7, 8, 9, 10: sidebar footer stack */
.md-side-safe-pill .material-symbols-rounded {
    font-size: 1rem !important;
    color: #2563eb !important;
}
.md-care-copy,
.md-sidebar-bottom {
    margin-top: 0.6rem !important;
    padding-top: 0 !important;
    position: static !important;
    background: transparent !important;
}
.md-sidebar-bottom .stButton > button,
.st-key-nav_privacy_bottom .stButton > button {
    min-height: 36px !important;
    height: 36px !important;
    justify-content: center !important;
    color: #334155 !important;
    font-weight: 650 !important;
}
.sb-footer {
    text-align: center !important;
    color: #94a3b8 !important;
    font-size: 0.64rem !important;
    line-height: 1.35 !important;
    margin: 0.35rem 0 0 !important;
    padding: 0 !important;
}

/* 12, 14: Apple-like greeting */
.md-home-head-left,
.md-home-greet-wrap {
    text-align: left !important;
    max-width: 1180px !important;
    margin: 0.35rem auto 0.85rem !important;
}
.md-greet {
    font-family: Inter, -apple-system, BlinkMacSystemFont, "SF Pro Display", "Segoe UI", sans-serif !important;
    font-size: clamp(1.9rem, 2.2vw, 2.35rem) !important;
    line-height: 1.08 !important;
    letter-spacing: 0 !important;
    font-weight: 850 !important;
    color: #0f172a !important;
}
.md-subgreet {
    display: block !important;
    margin-top: 0.38rem !important;
    font-size: 1.02rem !important;
    line-height: 1.35 !important;
    color: #64748b !important;
    text-align: left !important;
}

/* 15-22: quick actions should be compact pill cards with only one line */
.st-key-qa_headache .stButton > button,
.st-key-qa_tired .stButton > button,
.st-key-qa_symptoms .stButton > button,
.st-key-qa_sleep .stButton > button {
    min-height: 48px !important;
    height: 48px !important;
    border-radius: 999px !important;
    padding: 0 1rem !important;
    align-items: center !important;
    justify-content: center !important;
    text-align: center !important;
    background: rgba(255,255,255,0.96) !important;
    border: 1px solid #d8eafd !important;
    box-shadow: 0 8px 18px rgba(15,23,42,0.045) !important;
}
.st-key-qa_headache .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_tired .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_symptoms .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_sleep .stButton > button [data-testid="stIconMaterial"] {
    width: 22px !important;
    min-width: 22px !important;
    height: 22px !important;
    margin-right: 0.48rem !important;
    border-radius: 0 !important;
    background: transparent !important;
    font-size: 1.12rem !important;
}
.st-key-qa_headache .stButton > button p,
.st-key-qa_tired .stButton > button p,
.st-key-qa_symptoms .stButton > button p,
.st-key-qa_sleep .stButton > button p {
    white-space: nowrap !important;
    color: #0f172a !important;
    font-size: 0.98rem !important;
    font-weight: 680 !important;
    line-height: 1 !important;
}

/* 23, 24: smaller chat box/buttons */
form#home_chat_form [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
.md-home-composer-note {
    font-size: 0.72rem !important;
    margin: 0.35rem 0 1.05rem !important;
}

/* 25-30: guest feature cards as small pills with exact labels */
.st-key-guest_vision_card .stButton > button,
.st-key-guest_symptom_card .stButton > button,
.st-key-guest_rx_card .stButton > button {
    min-height: 42px !important;
    height: 42px !important;
    border-radius: 999px !important;
    padding: 0 0.8rem !important;
    text-align: center !important;
    justify-content: center !important;
    background: rgba(255,255,255,0.96) !important;
    border: 1px solid #d8eafd !important;
    box-shadow: 0 6px 14px rgba(15,23,42,0.035) !important;
}
.st-key-guest_vision_card .stButton > button p,
.st-key-guest_symptom_card .stButton > button p,
.st-key-guest_rx_card .stButton > button p {
    white-space: nowrap !important;
    font-size: 0.86rem !important;
    font-weight: 650 !important;
    color: #0f172a !important;
}

/* 31, 32: small health tip, no Learn more button rendered */
.st-key-tip_learn_more {
    display: none !important;
}

/* 34: custom compact error */
.md-mini-error {
    margin-top: 0.5rem !important;
    display: inline-flex !important;
    align-items: center !important;
    max-width: 560px !important;
    padding: 0.5rem 0.7rem !important;
    border-radius: 12px !important;
    background: #fff1f2 !important;
    border: 1px solid #fecdd3 !important;
    color: #be123c !important;
    font-size: 0.78rem !important;
    line-height: 1.35 !important;
}

@media (min-width: 1200px) {
    .main .block-container {
        padding-top: 0.55rem !important;
    }
}
@media (max-width: 980px) {
    [data-testid="collapsedControl"] {
        display: none !important;
    }
    .md-greet, .md-subgreet, .md-home-head-left, .md-home-greet-wrap {
        text-align: center !important;
    }
    .md-subgreet {
        margin-left: auto !important;
        margin-right: auto !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar lock: Streamlit's native collapse button can hide navigation ──
st.markdown("""
<style>
/* Keep the navigation visible even if Streamlit previously stored a collapsed state. */
section[data-testid="stSidebar"],
[data-testid="stSidebar"],
[data-testid="stSidebar"][aria-expanded="false"],
[data-testid="stSidebar"][aria-expanded="true"] {
    display: block !important;
    visibility: visible !important;
    opacity: 1 !important;
    width: 270px !important;
    min-width: 270px !important;
    max-width: 270px !important;
    transform: none !important;
    margin-left: 0 !important;
    left: 0 !important;
    z-index: 1000 !important;
}
section[data-testid="stSidebar"] > div,
[data-testid="stSidebar"] > div,
[data-testid="stSidebarContent"],
[data-testid="stSidebarUserContent"] {
    display: block !important;
    visibility: visible !important;
    opacity: 1 !important;
    width: 270px !important;
    min-width: 270px !important;
    transform: none !important;
}

/* Remove every variant of the Streamlit sidebar collapse/expand arrow. */
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="stSidebarCollapseButton"],
[data-testid="stSidebarNavCollapseButton"],
button[aria-label*="sidebar" i],
button[aria-label*="collapse" i],
button[aria-label*="expand" i],
button[title*="sidebar" i],
button[title*="collapse" i],
button[title*="expand" i],
[class*="collapsedControl"],
[class*="sidebarCollapse"],
[class*="SidebarCollapse"],
[class*="collapseButton"],
[class*="CollapseButton"] {
    display: none !important;
    visibility: hidden !important;
    opacity: 0 !important;
    pointer-events: none !important;
    width: 0 !important;
    height: 0 !important;
    min-width: 0 !important;
    min-height: 0 !important;
    overflow: hidden !important;
}

/* The old collapse icon sometimes leaks as raw Material text in Streamlit. */
[data-testid="stSidebar"] *:has(> [data-testid="stMarkdownContainer"] p),
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p:has(span) {
    overflow: visible !important;
}
[data-testid="stSidebar"] button:has([data-testid="stIconMaterial"]) {
    overflow: visible !important;
}

@media (min-width: 981px) {
    .stAppViewContainer,
    [data-testid="stAppViewContainer"] {
        margin-left: 0 !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ── Mockup parity overrides (home dashboard visual match) ─────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"], [data-testid="stAppViewContainer"], .stApp, .stMarkdown, p, span, button, input, textarea {
    font-family: 'Manrope', system-ui, -apple-system, Segoe UI, Roboto, sans-serif !important;
}

.stApp {
    background:
        radial-gradient(circle at 40% -8%, rgba(59,130,246,0.08), transparent 34%),
        #f7f9fe !important;
}
.main .block-container {
    max-width: 1380px !important;
    padding-top: 0.1rem !important;
}
[data-testid="stAppViewContainer"] .main {
    padding-top: 0 !important;
}
[data-testid="stVerticalBlock"] > [data-testid="element-container"]:first-child {
    margin-top: 0 !important;
}

.md-top-icons-fixed .material-symbols-rounded {
    font-size: 1.35rem !important;
    width: 34px;
    height: 34px;
    border-radius: 999px;
    display: inline-flex !important;
    align-items: center;
    justify-content: center;
    background: rgba(255,255,255,0.9);
    border: 1px solid #e7ecf7;
    box-shadow: 0 6px 18px rgba(15,23,42,0.06);
}

.md-home-greet-wrap {
    margin-top: -0.35rem !important;
    margin-bottom: 0.45rem !important;
}
.md-greet {
    font-size: 2.42rem !important;
    letter-spacing: -0.03em !important;
    color: #0f172a !important;
    font-weight: 800 !important;
}
.md-subgreet {
    font-size: 0.92rem !important;
    color: #64748b !important;
    font-weight: 550 !important;
}

.md-logo-mark {
    width: 52px !important;
    height: 52px !important;
    border-radius: 18px !important;
    background: linear-gradient(145deg, #f2f7ff, #eaf2ff) !important;
    border: 1px solid #d9e8ff !important;
    color: #1d8cf8 !important;
    box-shadow: none !important;
}
.md-logo-mark .material-symbols-rounded {
    font-size: 1.6rem !important;
    color: #1d8cf8 !important;
}

/* Sidebar edge + rhythm polish */
[data-testid="stSidebar"] {
    border-right: 1px solid #e9eef8 !important;
    box-shadow: 8px 0 28px rgba(15, 23, 42, 0.04) !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] > div {
    padding-right: 0.12rem !important;
}
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    row-gap: 0.18rem !important;
}
.md-logo-wrap {
    position: relative;
    border-bottom: none !important;
    margin-bottom: 1rem !important;
    padding-bottom: 1.05rem !important;
}
.md-logo-wrap::after {
    content: "";
    position: absolute;
    left: 0;
    right: 0;
    bottom: 0;
    height: 1px;
    background: linear-gradient(
        to right,
        rgba(203, 213, 225, 0.2) 0%,
        rgba(203, 213, 225, 0.95) 25%,
        rgba(203, 213, 225, 0.95) 75%,
        rgba(203, 213, 225, 0.2) 100%
    );
}

.st-key-qa_headache .stButton > button,
.st-key-qa_tired .stButton > button,
.st-key-qa_symptoms .stButton > button,
.st-key-qa_sleep .stButton > button {
    min-height: 48px !important;
    height: 48px !important;
    border-radius: 999px !important;
    padding: 0 1.05rem !important;
    align-items: center !important;
    justify-content: flex-start !important;
    background: rgba(255,255,255,0.96) !important;
    border: 1px solid #dce8fa !important;
    box-shadow: none !important;
}
.st-key-qa_headache .stButton > button p,
.st-key-qa_tired .stButton > button p,
.st-key-qa_symptoms .stButton > button p,
.st-key-qa_sleep .stButton > button p {
    font-size: 0.92rem !important;
    font-weight: 650 !important;
    white-space: nowrap !important;
}

/* Sidebar nav icon/text spacing */
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button {
    padding: 0 0.9rem !important;
    min-height: 48px !important;
    height: 48px !important;
    font-weight: 700 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button [data-testid="stIconMaterial"] {
    margin-right: 1.1rem !important;
    min-width: 24px !important;
    width: 24px !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button p {
    letter-spacing: 0 !important;
    font-weight: 700 !important;
}

/* Put more breathing room between logo/title and first nav item ("Home"). */
[data-testid="stSidebar"] div.st-key-nav_home {
    margin-top: 0.85rem !important;
}

/* Keep privacy + copyright pinned to the true bottom of the sidebar column. */
[data-testid="stSidebarUserContent"] > div > [data-testid="stVerticalBlock"] {
    min-height: calc(100vh - 10px) !important;
    display: flex !important;
    flex-direction: column !important;
}
.md-side-profile {
    margin-top: auto !important;
}
.md-sidebar-bottom {
    margin-top: 0.35rem !important;
}
.sb-footer {
    margin-top: 0.45rem !important;
    margin-bottom: 0.2rem !important;
}

form#home_chat_form [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],

.md-home-composer-note {
    text-align: center !important;
    margin-top: 0.75rem !important;
}

.st-key-sa_sym .stButton > button,
.st-key-sa_rec .stButton > button,
.st-key-sa_ins .stButton > button,
.st-key-sa_appt .stButton > button {
    min-height: 220px !important;
    border-radius: 20px !important;
    box-shadow: 0 16px 34px rgba(15,23,42,0.055) !important;
}

.md-rcard {
    border-radius: 20px !important;
    padding: 1.15rem 1.2rem !important;
}
.md-snap-card {
    margin-top: 0 !important;
}
.md-rcard-title {
    font-size: 1.02rem !important;
}
.md-rcard-link {
    font-size: 0.84rem !important;
    color: #2563eb !important;
}
.md-snap-status {
    font-size: 0.83rem;
    font-weight: 700;
    color: #16a34a;
    margin-left: 0.45rem;
}
.md-spark {
    margin-top: 0.22rem;
}


.sb-title-language,
.md-side-safe-wrap,
.md-care-note,
.md-sidebar-bottom-spacer,
.st-key-home_overview_see_all,
.st-key-view_all_recent {
    display: none !important;
}

.st-key-guest_to_signin .stButton > button {
    border-radius: 999px !important;
    background: linear-gradient(135deg, #2f8cff, #6b4bff) !important;
    color: #fff !important;
    border: none !important;
    font-weight: 700 !important;
}
.st-key-sidebar_premium_cta .stButton > button {
    border-radius: 999px !important;
    background: linear-gradient(135deg, #3b82f6, #6366f1) !important;
    color: #fff !important;
    border: none !important;
    font-weight: 700 !important;
    height: 40px !important;
    min-height: 40px !important;
    margin: 0.25rem 1.6rem 0.4rem 0 !important;
    width: calc(100% - 1.6rem) !important;
    box-sizing: border-box !important;
    box-shadow: 0 8px 18px rgba(99,102,241,0.28) !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease !important;
    letter-spacing: 0.005em !important;
}
.st-key-sidebar_premium_cta .stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 10px 22px rgba(99,102,241,0.36) !important;
}
.st-key-sidebar_premium_cta .stButton > button p {
    color: #fff !important;
    font-weight: 700 !important;
    letter-spacing: 0.005em !important;
}

/* Mockup-styled Premium upsell card */

.md-sidebar-icon-row .material-symbols-rounded {
    width: 32px;
    height: 32px;
    border-radius: 999px;
    border: 1px solid #e3eaf8;
    color: #475569;
    background: #ffffff;
    display: inline-flex !important;
    align-items: center;
    justify-content: center;
    font-size: 1.05rem !important;
}

.md-sidebar-bottom .stButton > button {
    min-height: 36px !important;
    border-radius: 10px !important;
    font-size: 0.82rem !important;
    color: #334155 !important;
}
.sb-footer {
    text-align: center;
    color: #94a3b8 !important;
    font-size: 0.72rem !important;
    margin-top: 0.5rem !important;
}

/* --- Overlap hotfix: cards/chips --- */
.st-key-sa_sym .stButton > button,
.st-key-sa_rec .stButton > button,
.st-key-sa_ins .stButton > button,
.st-key-sa_appt .stButton > button {
    display: flex !important;
    flex-direction: column !important;
    align-items: flex-start !important;
    justify-content: flex-start !important;
    gap: 0.42rem !important;
    text-align: left !important;
    padding: 1rem 1rem 2.25rem 1rem !important;
    min-height: 196px !important;
    overflow: visible !important;
}
.st-key-sa_sym .stButton > button > span:first-child,
.st-key-sa_rec .stButton > button > span:first-child,
.st-key-sa_ins .stButton > button > span:first-child,
.st-key-sa_appt .stButton > button > span:first-child {
    display: block !important;
    width: 48px !important;
    height: 48px !important;
    min-width: 48px !important;
    min-height: 48px !important;
    line-height: 0 !important;
    margin: 0 0 0.72rem 0 !important;
    overflow: visible !important;
}
.st-key-sa_sym .stButton > button [data-testid="stIconMaterial"],
.st-key-sa_rec .stButton > button [data-testid="stIconMaterial"],
.st-key-sa_ins .stButton > button [data-testid="stIconMaterial"],
.st-key-sa_appt .stButton > button [data-testid="stIconMaterial"] {
    margin: 0 !important;
    display: inline-flex !important;
    position: static !important;
}
.st-key-sa_sym .stButton > button [data-testid="stMarkdownContainer"],
.st-key-sa_rec .stButton > button [data-testid="stMarkdownContainer"],
.st-key-sa_ins .stButton > button [data-testid="stMarkdownContainer"],
.st-key-sa_appt .stButton > button [data-testid="stMarkdownContainer"] {
    margin: 0 !important;
    width: 100% !important;
}
.st-key-sa_sym .stButton > button p,
.st-key-sa_rec .stButton > button p,
.st-key-sa_ins .stButton > button p,
.st-key-sa_appt .stButton > button p {
    margin: 0 !important;
    white-space: pre-line !important;
    line-height: 1.35 !important;
    font-size: 0.9rem !important;
}
.st-key-sa_sym .stButton > button p::first-line,
.st-key-sa_rec .stButton > button p::first-line,
.st-key-sa_ins .stButton > button p::first-line,
.st-key-sa_appt .stButton > button p::first-line {
    font-size: 1.03rem !important;
    font-weight: 780 !important;
}

.st-key-qa_headache .stButton > button p,
.st-key-qa_tired .stButton > button p,
.st-key-qa_symptoms .stButton > button p,
.st-key-qa_sleep .stButton > button p {
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}

@media (max-width: 980px) {
    .md-snap-card {
        margin-top: 0 !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ── Final Sidebar Seam Polish ─────────────────────────────────────────
st.markdown("""
<style>
/* Force one continuous canvas. Sidebar uses the same base tone as the app
   (no white panel, no hairline shadow), so there is no visible edge between
   sidebar and main — the mockup's seamless feel. */
[data-testid="stAppViewContainer"] {
    background: #f7f9fe !important;
}
section[data-testid="stSidebar"],
[data-testid="stSidebar"] {
    position: relative !important;
    border-right: none !important;
    background: #f4f5f9 !important;
    box-shadow: none !important;
    overflow: visible !important;
}
section[data-testid="stSidebar"]::after,
[data-testid="stSidebar"]::after,
section[data-testid="stSidebar"]::before,
[data-testid="stSidebar"]::before {
    display: none !important;
    content: none !important;
    background: none !important;
}
/* Streamlit also draws a separator on the sidebar resize handle — kill it. */
[data-testid="stSidebarCollapseButton"],
[data-testid="stSidebarResizeHandle"] {
    display: none !important;
}
[data-testid="stSidebarContent"] {
    border-right: none !important;
    box-shadow: none !important;
    background: transparent !important;
    height: 100% !important;
}
[data-testid="stSidebarUserContent"] {
    background: transparent !important;
}
/* Streamlit's cache-class wrappers default to the sidebar's full 270px width
   even when they sit INSIDE the sidebar's content padding, causing children
   to overflow ~12-14px past the right edge. Constrain them so percentages
   and child widths resolve against the actual padded content area.

   Critical additions (May 2026): stSidebarUserContent + its direct emotion-
   cache wrapper child were missed by the original list. They both default to
   width=270px (the FULL sidebar width) and sit shifted ~11px right inside
   the sidebarContent's 13.6px padding, so the whole sidebar bottom (cards,
   buttons, conv-list rows) renders ~25px past the visible right edge — where
   overflow-x: hidden on stSidebarContent crops the trailing edge of every
   row (e.g. the "go" of "1w ago" timestamps). Force them to inherit the
   actual padded content width. */
[data-testid="stSidebar"] [data-testid="stSidebarContent"] > div,
[data-testid="stSidebar"] [data-testid="stSidebarContent"] > div > div,
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"],
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] > div,
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    max-width: 100% !important;
    width: 100% !important;
    box-sizing: border-box !important;
    padding-left: 0 !important;
    padding-right: 0 !important;
    margin-left: 0 !important;
    margin-right: 0 !important;
    transform: none !important;
}
[data-testid="stMain"],
.main {
    background: transparent !important;
}

/* Keep sidebar internals from drawing harsh dividing lines. */
[data-testid="stSidebar"] hr {
    border: 0 !important;
    height: 1px !important;
    background: linear-gradient(
        to right,
        rgba(203, 213, 225, 0.0),
        rgba(203, 213, 225, 0.75),
        rgba(203, 213, 225, 0.0)
    ) !important;
    margin: 0.72rem 0 !important;
}

/* Match reference sidebar nav styling. */
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    row-gap: 0.02rem !important;
}
[data-testid="stSidebar"] div.st-key-nav_home {
    margin-top: 0.55rem !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button {
    min-height: 46px !important;
    height: 46px !important;
    border-radius: 14px !important;
    padding: 0 0.92rem !important;
    /* Streamlit's inner emotion-cache wrapper renders ~14px wider than the
       sidebar's padded content area, so we offset only on the right side
       to keep the pill visually inset on both edges. */
    margin: 0 1.6rem 0 0 !important;
    width: calc(100% - 1.6rem) !important;
    box-sizing: border-box !important;
    background: transparent !important;
    border: 1px solid transparent !important;
    box-shadow: none !important;
    justify-content: flex-start !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button > div {
    display: flex !important;
    align-items: center !important;
    gap: 2.6rem !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button [data-testid="stIconMaterial"] {
    background: transparent !important;
    width: 24px !important;
    min-width: 24px !important;
    height: 24px !important;
    margin-right: 0 !important;
    color: #7c7a93 !important;
    font-size: 1.06rem !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button p {
    font-weight: 560 !important;
    font-size: 0.84rem !important;
    color: #7c7a93 !important;
    margin-left: 0.62rem !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button:hover {
    background: #eef4ff !important;
    border-color: transparent !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button:hover [data-testid="stIconMaterial"] {
    color: #2563eb !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button:hover p {
    color: #2563eb !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button[kind="primary"] {
    background: #e6f0ff !important;
    border-color: transparent !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button[kind="primary"] [data-testid="stIconMaterial"] {
    color: #2563eb !important;
    background: transparent !important;
    border-radius: 0 !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"] .stButton > button[kind="primary"] p {
    color: #2563eb !important;
    font-weight: 700 !important;
}
.md-nav-active .stButton > button {
    background: #e6f0ff !important;
    border-color: transparent !important;
    box-shadow: none !important;
}
.md-nav-active div[class*="st-key-nav_"] .stButton > button,
[data-testid="stSidebar"] .md-nav-active .stButton > button {
    background: #e6f0ff !important;
    border-color: transparent !important;
    box-shadow: none !important;
}
.md-nav-active .stButton > button [data-testid="stIconMaterial"] {
    color: #2563eb !important;
}
.md-nav-active div[class*="st-key-nav_"] .stButton > button [data-testid="stIconMaterial"],
[data-testid="stSidebar"] .md-nav-active .stButton > button [data-testid="stIconMaterial"] {
    color: #2563eb !important;
}
.md-nav-active .stButton > button p {
    color: #2563eb !important;
    font-weight: 700 !important;
}
[data-testid="stSidebar"] .md-nav-active .stButton > button p {
    color: #2563eb !important;
    font-weight: 700 !important;
}

.md-logo-wrap {
    margin-bottom: 0.78rem !important;
    padding-bottom: 0.95rem !important;
}
.md-logo-mark {
    width: 42px !important;
    height: 42px !important;
    border-radius: 14px !important;
    background: radial-gradient(circle at 30% 30%, #f6f2ff 0%, #e8f0ff 80%) !important;
    border: 1px solid #d6e2fa !important;
    box-shadow: none !important;
}
.md-logo-mark .material-symbols-rounded {
    color: #2f80ed !important;
    font-size: 1.42rem !important;
}

/* Pin lower section to the true bottom of sidebar. */
[data-testid="stSidebarUserContent"] > div > [data-testid="stVerticalBlock"] {
    min-height: calc(100vh - 8px) !important;
    display: flex !important;
    flex-direction: column !important;
}
.md-side-profile {
    margin-top: auto !important;
}
.md-sidebar-bottom {
    margin-top: auto !important;
}
.sb-footer {
    margin-top: 0.35rem !important;
    margin-bottom: 0.2rem !important;
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar Hard-Lock Layout (strict final pass) ─────────────────────
st.markdown("""
<style>
/* Single source of truth for sidebar layout. */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #f9fbff 0%, #f6f8ff 100%) !important;
    border-right: 1px solid #e2e8f0 !important;
    box-shadow: inset -1px 0 0 rgba(255,255,255,0.92) !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    padding: 1.06rem 1rem 2rem 1rem !important;
    overflow-y: clip !important;
    overflow-x: hidden !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] > div > [data-testid="stVerticalBlock"] {
    /* On TALL screens (≥ ~1000px), min-height ensures the column fills the
       viewport so `.md-side-profile { margin-top: auto }` pushes the profile
       chip + everything below it (sign out, recent chats, language, footer)
       down to the bottom edge.

       On SHORTER laptop screens (~800-900px), the content stack is taller
       than the viewport — so we deliberately do NOT cap max-height. The
       column grows to its natural content height and the sidebarContent's
       `overflow-y: auto` (below) gives the user a thin scrollbar to reach
       the footer instead of clipping it. */
    min-height: calc(100vh - 40px) !important;
    display: flex !important;
    flex-direction: column !important;
}

/* Logo divider: short, centred, minimalist line. Replaces the heavy full-width
   border-bottom set on .md-logo-wrap earlier in the stylesheet. */
[data-testid="stSidebar"] .md-logo-wrap {
    margin-bottom: 1.2rem !important;
    padding-bottom: 1.08rem !important;
    position: relative !important;
    z-index: 10 !important;
    min-height: 48px !important;
    border-bottom: none !important;
}
[data-testid="stSidebar"] .md-logo-wrap::after {
    content: "" !important;
    position: absolute !important;
    left: 15% !important;
    right: 15% !important;
    bottom: -0.48rem !important;
    height: 1px !important;
    background: #e2e8f0 !important;
    z-index: 0 !important;
}
[data-testid="stSidebar"] div.st-key-nav_home {
    margin-top: 0.62rem !important;
}

/* Unified nav style for all profiles/pages (guest, signed-in, auth). */
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button {
    min-height: 56px !important;
    height: 56px !important;
    border-radius: 14px !important;
    margin-bottom: 0.52rem !important;
    font-weight: 650 !important;
    font-size: 1.01rem !important;
    background: transparent !important;
    border: 1px solid transparent !important;
    color: #4c5d78 !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    padding-left: 0.64rem !important;
    overflow: visible !important;
    position: relative !important;
    z-index: 2 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button:hover {
    background: #ffffff !important;
    border-color: #e2eaf9 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > span:first-child {
    width: 40px !important;
    min-width: 40px !important;
    height: 40px !important;
    border-radius: 12px !important;
    margin-right: 0.92rem !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    border: 1px solid #e2e8f0 !important;
    background: #f8fbff !important;
    box-shadow: 0 6px 12px rgba(15, 23, 42, 0.04) !important;
}
/* Fallback for Streamlit builds where the icon is rendered as a direct child
   of the button rather than inside the first span wrapper. */
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > [data-testid="stIconMaterial"] {
    width: 40px !important;
    min-width: 40px !important;
    height: 40px !important;
    border-radius: 12px !important;
    margin-right: 0.92rem !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    border: 1px solid #e2e8f0 !important;
    background: #f8fbff !important;
    box-shadow: 0 6px 12px rgba(15, 23, 42, 0.04) !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button [data-testid="stIconMaterial"] {
    font-size: 1.28rem !important;
    color: #5a6f8c !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > span:first-child [data-testid="stIconMaterial"] {
    margin-right: 0 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > div[data-testid="stMarkdownContainer"] {
    margin: 0 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > div[data-testid="stMarkdownContainer"] p {
    margin: 0 !important;
    font-size: 1.01rem !important;
    font-weight: 650 !important;
    color: #4b5d78 !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button[kind="primary"] {
    background: linear-gradient(180deg, #eef4ff, #edf3ff) !important;
    border: 1px solid #d4e3ff !important;
    color: #2563eb !important;
    box-shadow: 0 8px 18px rgba(59, 130, 246, 0.08) !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button[kind="primary"]::before {
    content: "" !important;
    position: absolute !important;
    left: -0.86rem !important;
    top: 10px !important;
    bottom: 10px !important;
    width: 3px !important;
    border-radius: 999px !important;
    background: #2f67ff !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button[kind="primary"] p,
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button[kind="primary"] [data-testid="stIconMaterial"] {
    color: #2563eb !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button[kind="primary"] > span:first-child {
    background: #ffffff !important;
    border-color: #cddcff !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button[kind="primary"] > [data-testid="stIconMaterial"] {
    background: #ffffff !important;
    border-color: #cddcff !important;
}

/* Per-item icon tile colors. */
[data-testid="stSidebar"] div.st-key-nav_home .stButton > button > span:first-child { background: #eef4ff !important; border-color: #d6e3ff !important; }
[data-testid="stSidebar"] div.st-key-nav_home .stButton > button > [data-testid="stIconMaterial"] { background: #eef4ff !important; border-color: #d6e3ff !important; }
[data-testid="stSidebar"] div.st-key-nav_home .stButton > button [data-testid="stIconMaterial"] { color: #2563eb !important; }
[data-testid="stSidebar"] div.st-key-nav_new .stButton > button > span:first-child { background: #eefbff !important; border-color: #cdeef8 !important; }
[data-testid="stSidebar"] div.st-key-nav_new .stButton > button > [data-testid="stIconMaterial"] { background: #eefbff !important; border-color: #cdeef8 !important; }
[data-testid="stSidebar"] div.st-key-nav_new .stButton > button [data-testid="stIconMaterial"] { color: #0e9fbc !important; }
[data-testid="stSidebar"] div.st-key-nav_overview .stButton > button > span:first-child { background: #effcf7 !important; border-color: #d3f2e6 !important; }
[data-testid="stSidebar"] div.st-key-nav_overview .stButton > button > [data-testid="stIconMaterial"] { background: #effcf7 !important; border-color: #d3f2e6 !important; }
[data-testid="stSidebar"] div.st-key-nav_overview .stButton > button [data-testid="stIconMaterial"] { color: #14a26f !important; }
[data-testid="stSidebar"] div.st-key-nav_symptom .stButton > button > span:first-child { background: #f4f0ff !important; border-color: #e5dbff !important; }
[data-testid="stSidebar"] div.st-key-nav_symptom .stButton > button > [data-testid="stIconMaterial"] { background: #f4f0ff !important; border-color: #e5dbff !important; }
[data-testid="stSidebar"] div.st-key-nav_symptom .stButton > button [data-testid="stIconMaterial"] { color: #7c4dff !important; }
[data-testid="stSidebar"] div.st-key-nav_prescription .stButton > button > span:first-child { background: #fff7ed !important; border-color: #fde2bf !important; }
[data-testid="stSidebar"] div.st-key-nav_prescription .stButton > button > [data-testid="stIconMaterial"] { background: #fff7ed !important; border-color: #fde2bf !important; }
[data-testid="stSidebar"] div.st-key-nav_prescription .stButton > button [data-testid="stIconMaterial"] { color: #f59e0b !important; }
[data-testid="stSidebar"] div.st-key-nav_records .stButton > button > span:first-child { background: #eefbfd !important; border-color: #cceef3 !important; }
[data-testid="stSidebar"] div.st-key-nav_records .stButton > button > [data-testid="stIconMaterial"] { background: #eefbfd !important; border-color: #cceef3 !important; }
[data-testid="stSidebar"] div.st-key-nav_records .stButton > button [data-testid="stIconMaterial"] { color: #0f9db3 !important; }
[data-testid="stSidebar"] div.st-key-nav_meds .stButton > button > span:first-child { background: #fff1f2 !important; border-color: #ffd8de !important; }
[data-testid="stSidebar"] div.st-key-nav_meds .stButton > button > [data-testid="stIconMaterial"] { background: #fff1f2 !important; border-color: #ffd8de !important; }
[data-testid="stSidebar"] div.st-key-nav_meds .stButton > button [data-testid="stIconMaterial"] { color: #ef4444 !important; }
[data-testid="stSidebar"] div.st-key-nav_insights .stButton > button > span:first-child { background: #eef2ff !important; border-color: #d8e1ff !important; }
[data-testid="stSidebar"] div.st-key-nav_insights .stButton > button > [data-testid="stIconMaterial"] { background: #eef2ff !important; border-color: #d8e1ff !important; }
[data-testid="stSidebar"] div.st-key-nav_insights .stButton > button [data-testid="stIconMaterial"] { color: #4f46e5 !important; }
[data-testid="stSidebar"] div.st-key-nav_appts .stButton > button > span:first-child { background: #edf4ff !important; border-color: #d8e6ff !important; }
[data-testid="stSidebar"] div.st-key-nav_appts .stButton > button > [data-testid="stIconMaterial"] { background: #edf4ff !important; border-color: #d8e6ff !important; }
[data-testid="stSidebar"] div.st-key-nav_appts .stButton > button [data-testid="stIconMaterial"] { color: #2563eb !important; }
[data-testid="stSidebar"] .md-side-profile-top {
    border-radius: 14px !important;
    border: 1px solid #e6edf9 !important;
    box-shadow: 0 2px 6px rgba(15,23,42,0.03) !important;
    margin-top: 0.3rem !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] {
    margin: 1.48rem 0.25rem 0.9rem 0 !important;
    width: 242px !important;
    max-width: 242px !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] label {
    margin-bottom: 0.34rem !important;
    color: #50617f !important;
    font-weight: 620 !important;
    font-size: 0.86rem !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] > div {
    width: 100% !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] {
    width: 100% !important;
    min-height: 56px !important;
    height: 56px !important;
    border-radius: 18px !important;
    border: 1px solid #dce6fb !important;
    background: #ffffff !important;
    box-shadow: 0 10px 20px rgba(15,23,42,0.05) !important;
    overflow: hidden !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
    min-height: 56px !important;
    height: 56px !important;
    border: none !important;
    border-radius: 18px !important;
    background: transparent !important;
    padding: 0 0.95rem !important;
    display: flex !important;
    align-items: center !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] [role="combobox"] {
    padding: 0 !important;
    font-size: 1rem !important;
    font-weight: 560 !important;
    color: #1f2937 !important;
}

/* Prevent divider/active-state visual overlap. */
[data-testid="stSidebar"] hr {
    margin-top: 0.9rem !important;
    margin-bottom: 0.9rem !important;
}

/* Bottom stack: one flow layout only (no absolute overlap). */
[data-testid="stSidebar"] .md-side-profile {
    margin-top: 0 !important;
    /* Same Streamlit inner-wrapper compensation as the nav pills + premium card. */
    margin-right: 1.6rem !important;
    margin-left: 0 !important;
    width: calc(100% - 1.6rem) !important;
    box-sizing: border-box !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] > div > [data-testid="stVerticalBlock"] > div:has(.md-side-profile-top) {
    margin-top: 0.4rem !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] > div > [data-testid="stVerticalBlock"] > div:has(.md-sidebar-bottom) {
    margin-top: auto !important;
}
[data-testid="stSidebar"] .md-sidebar-bottom {
    position: static !important;
    margin-top: 0.35rem !important;
    padding: 0 0.9rem 0 !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom {
    margin-top: 0 !important;
    margin-bottom: 0.75rem !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button {
    justify-content: center !important;
    text-align: center !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button > span:first-child {
    margin-right: 0 !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button > div[data-testid="stMarkdownContainer"] {
    width: 100% !important;
    margin: 0 !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button > div[data-testid="stMarkdownContainer"] p {
    width: 100% !important;
    text-align: center !important;
}
[data-testid="stSidebar"] .sb-footer {
    position: absolute !important;
    left: 0 !important;
    right: 0 !important;
    bottom: 1.05rem !important;
    margin: 0 !important;
    text-align: center !important;
}

/* Sidebar scroll behavior — locked, no scroll possible:
   - overflow-x stays HIDDEN to crop any horizontal bleed (the earlier
     stSidebarUserContent clipping fix relies on this).
   - overflow-y is HIDDEN so scrolling is completely disabled. Mouse wheel,
     trackpad, keyboard arrow keys — none produce sidebar scroll. Content
     sits truly fixed in place. The Approach-A tightening pass made the
     content fit cleanly inside the viewport at 900px in both auth and
     unauth states, so nothing visible gets cut off.

   The scrollbar-* / ::-webkit-scrollbar rules are redundant now (no scroll
   means no scrollbar) but kept as a safety net in case any nested element
   inside the sidebar tries to spawn its own scroller. */
[data-testid="stSidebar"] {
    overflow-x: clip !important;
    overflow-y: clip !important;
    scrollbar-width: none !important;
    -ms-overflow-style: none !important;
}
[data-testid="stSidebarContent"] {
    overflow-x: clip !important;
    overflow-y: clip !important; /* Disable vertical scrolling */
    scrollbar-width: none !important;
    -ms-overflow-style: none !important;
}
[data-testid="stSidebarUserContent"] {
    overflow-x: clip !important;
    overflow-y: visible !important; /* Prevent vertical clipping of child elements! */
    scrollbar-width: none !important;
    -ms-overflow-style: none !important;
}
[data-testid="stSidebar"]::-webkit-scrollbar,
[data-testid="stSidebarContent"]::-webkit-scrollbar,
[data-testid="stSidebarUserContent"]::-webkit-scrollbar {
    display: none !important;
    width: 0 !important;
    height: 0 !important;
    background: transparent !important;
}

/* Kill phantom scrollHeight: each Streamlit element wrapper inside the
   sidebar can report scrollHeight > clientHeight because of absolutely-
   positioned pseudo-elements and emotion-cache spacing quirks.

   NOTE (May 2026): originally also applied to stMarkdown + stMarkdownContainer,
   but that was clipping the wrapped 2nd line of the footer copyright
   ("All rights reserved." was being cut off vertically). The overflow: clip
   on stSidebar wrappers already locks scroll, so the per-element-container
   clipping below is only needed for the phantom — and it doesn't need to
   extend to inner markdown wrappers that legitimately host wrapped text. */
[data-testid="stSidebar"] [data-testid="stElementContainer"] {
    overflow: hidden !important;
}
[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.md-side-profile),
[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.md-recent-card) {
    overflow: visible !important;
}
/* Explicitly allow the footer wrappers to expand vertically so the
   copyright's wrapped 2nd line shows in full. */
[data-testid="stSidebar"] .md-sidebar-foot,
[data-testid="stSidebar"] .md-sidebar-foot *,
[data-testid="stSidebar"] [data-testid="stMarkdown"]:has(.md-sidebar-foot),
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"]:has(.md-sidebar-foot),
[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.md-sidebar-foot) {
    overflow: visible !important;
    height: auto !important;
    max-height: none !important;
}

/* Remove settings/moon mini row entirely. */

/* Home/chat attachment and voice buttons: home uses icon + label pills. */
[data-testid="stSidebar"] .stButton > button p {
    font-weight: 700 !important;
}
/* Specificity needs to beat .stForm [data-testid="stFormSubmitButton"] > button[kind="secondary"]
   which sets height:44px, border-radius:14px, and a clinical border. */
.st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button,
.st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button,
.st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
.st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
.st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"],
.st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"] {
    min-width: 112px !important;
    width: 112px !important;
    height: 56px !important;
    min-height: 56px !important;
    padding: 0 0.95rem !important;
    border-radius: 999px !important;
    border: 1px solid #dde6f4 !important;
    background: #ffffff !important;
    box-shadow: none !important;
    transition: border-color 0.15s ease, background 0.15s ease, transform 0.15s ease !important;
}
.st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button:hover,
.st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button:hover,
.st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"]:hover,
.st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"]:hover {
    border-color: #cbd8ef !important;
    background: #fbfdff !important;
    transform: translateY(-1px) !important;
}
.st-key-chat_upload_btn button,
.st-key-chat_voice_btn button {
    min-width: 44px !important;
    width: 44px !important;
    height: 42px !important;
    min-height: 42px !important;
    padding: 0 !important;
    border-radius: 999px !important;
}
.st-key-home_upload_btn button > div,
.st-key-home_voice_btn button > div {
    justify-content: center !important;
    gap: 0.42rem !important;
}
.st-key-chat_upload_btn button > div,
.st-key-chat_voice_btn button > div {
    justify-content: center !important;
    gap: 0 !important;
}
.st-key-home_upload_btn button p,
.st-key-home_voice_btn button p {
    display: block !important;
    margin: 0 !important;
    font-size: 1.04rem !important;
    font-weight: 620 !important;
    line-height: 1 !important;
    color: #1f2937 !important;
}
.st-key-chat_upload_btn button p,
.st-key-chat_voice_btn button p {
    display: none !important;
}
.st-key-home_upload_btn button [data-testid="stIconMaterial"],
.st-key-home_voice_btn button [data-testid="stIconMaterial"] {
    margin: 0 !important;
    color: #475569 !important;
    font-size: 18px !important;
}
.st-key-home_upload_btn button:hover [data-testid="stIconMaterial"],
.st-key-home_voice_btn button:hover [data-testid="stIconMaterial"] {
    color: #334155 !important;
}
.st-key-chat_upload_btn button [data-testid="stIconMaterial"],
.st-key-chat_voice_btn button [data-testid="stIconMaterial"] {
    margin: 0 !important;
}
/* Home composer/card layout is controlled by the cross-profile lock block below
   to avoid overlapping legacy selector conflicts. */

/* Home greeting positioning and vertical stacking. */
.md-home-greet-wrap {
    margin: -0.9rem 0 0.85rem 0 !important;
    display: flex !important;
    flex-direction: column !important;
    align-items: flex-start !important;
    gap: 0.18rem !important;
    width: 100% !important;
    max-width: none !important;
    text-align: left !important;
}
.md-home-greet-wrap .md-greet,
.md-home-greet-wrap .md-subgreet {
    margin: 0 !important;
    text-align: left !important;
    align-self: flex-start !important;
}
.md-home-greet-wrap .md-greet {
    font-size: 2.6rem !important;
    font-weight: 700 !important;
    letter-spacing: -0.02em !important;
    line-height: 1.15 !important;
    color: #0b1220 !important;
}
.md-home-greet-wrap .md-subgreet {
    margin-top: 0.4rem !important;
    line-height: 1.25 !important;
    font-size: 0.98rem !important;
    color: #64748b !important;
    font-weight: 500 !important;
}

/* Streamlit's default main block padding-top is 96px (reserves space for the
   deploy header). In headless mode the header is hidden, so we collapse it
   to bring the greeting up to ~24px from the top of the main column. */
.stMainBlockContainer,
[data-testid="stMainBlockContainer"] {
    padding-top: 1.5rem !important;
}
/* Streamlit injects multiple zero-height stElementContainer wrappers (from
   <style> markdowns and the position:fixed top icons div). The default
   stVerticalBlock gap (1rem) compounds these into ~80px of phantom space
   above the first visible element. Tighten — but not to zero, or real
   sibling elements (greeting → chips) collide. */
.stMainBlockContainer > [data-testid="stVerticalBlock"] {
    gap: 0.6rem !important;
    row-gap: 0.6rem !important;
}

/* Fix 3: Chat composer textarea — kill the dark border. Streamlit doesn't
   render id="home_chat_form" on the DOM form, so we scope via the home send
   button key using :has(). The composer wrap keeps its outer border; only
   the inner textarea loses its line. */
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea {
    border: none !important;
    box-shadow: none !important;
    background: #ffffff !important;
    resize: none !important;
    overflow: auto !important;
    min-height: 136px !important;
    height: 136px !important;
    padding: 0.45rem 0.2rem !important;
    font-size: 1.03rem !important;
    color: #334155 !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea::-webkit-resizer {
    display: none !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea::placeholder {
    color: #8b98af !important;
    opacity: 1 !important;
    font-size: 1.03rem !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea:focus {
    outline: none !important;
    border: none !important;
    box-shadow: none !important;
}
/* Trim the inner gap between the textarea and the icon row so the composer
   feels compact rather than airy. */
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stVerticalBlock"] {
    gap: 0.72rem !important;
    row-gap: 0.72rem !important;
}

/* Fix 4a: Soft gradient glow under the composer. Pure decoration. */
.md-composer-glow {
    height: 30px;
    margin: -2px auto 0 auto;
    width: 86%;
    background: linear-gradient(90deg, rgba(59,130,246,0.18) 0%, rgba(139,92,246,0.22) 50%, rgba(59,130,246,0.18) 100%);
    filter: blur(20px);
    opacity: 0.52;
    pointer-events: none;
    border-radius: 999px;
}

/* Fix 4b: Send button — circular gradient with paper-plane icon. Scoped via
   the unique button key. */
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button,
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button[kind="primary"] {
    width: 56px !important;
    min-width: 56px !important;
    height: 56px !important;
    min-height: 56px !important;
    border-radius: 50% !important;
    background: linear-gradient(135deg, #3b82f6 0%, #6366f1 50%, #8b5cf6 100%) !important;
    border: none !important;
    box-shadow: 0 8px 22px rgba(99, 102, 241, 0.34) !important;
    padding: 0 !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease !important;
}
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button:hover {
    transform: scale(1.05) !important;
    box-shadow: 0 12px 28px rgba(99, 102, 241, 0.45) !important;
}
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
    color: #ffffff !important;
    font-size: 1.35rem !important;
    margin: 0 !important;
}
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button p {
    display: none !important;
}

/* Fix 4c: Shield icon before disclaimer text. */
.md-home-composer-note {
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 0.4rem !important;
    width: 100% !important;
}
.md-home-composer-note .md-disclaimer-shield {
    font-size: 14px !important;
    color: var(--md-text-3) !important;
    line-height: 1 !important;
}

/* Tighten the gap between the disclaimer line and the Smart Actions header.
   Streamlit's stElementContainer wrappers prevent margin collapse, so total
   gap = note margin-bottom + head margin-top. */
.md-home-composer-note {
    margin-top: 0.28rem !important;
    margin-bottom: 0.4rem !important;
}
.md-smart-head {
    margin-top: 4.6rem !important;
}

/* More breathing room between greeting and the chips/composer block below. */
.md-home-greet-wrap {
    margin-bottom: 3.55rem !important;
}

/* Brand logo image (replaces icon + text + subtitle). The PNG already includes
   the wordmark and tagline, so we just center it cleanly in the sidebar.
   Streamlit's stMarkdownContainer for sidebar markdown has a ~13px right offset
   vs widget containers (selectbox, profile card), which left this wrap
   misaligned and overflowing the right edge. translateX counter-shifts to
   align with the rest of the sidebar column. */
[data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap {
    /* Transparent container: the PNG already contains the white rounded card.
       Adding a second white background creates a double-card effect.
       Instead we use filter:drop-shadow on the image itself (see below)
       so the shadow wraps the PNG's actual shape, not a rectangular box. */
    display: grid !important;
    place-items: center !important;
    justify-items: center !important;
    justify-content: center !important;
    align-items: center !important;
    width: 100% !important;
    max-width: 242px !important;
    margin: 0 auto 1rem auto !important;
    transform: translateX(-13px) !important;
    padding: 0.5rem 0 0.5rem 0 !important;
    background: transparent !important;
    border: none !important;
    border-radius: 0 !important;
    box-shadow: none !important;
    position: relative !important;
    z-index: 10 !important;
    min-height: 80px !important;
    overflow: visible !important;
    border-bottom: none !important;
}
[data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap > .md-logo-image {
    mix-blend-mode: normal !important;
    max-width: 160px !important;
    width: min(160px, 100%) !important;
    height: auto !important;
    display: block !important;
    margin: 0 auto !important;
    transform: none !important;
    /* Drop shadow hugs the PNG's actual content shape (the internal white card),
       not the rectangular bounding box. This gives definition without a double-card. */
    filter: drop-shadow(0 4px 16px rgba(15, 23, 42, 0.13)) drop-shadow(0 1px 4px rgba(15, 23, 42, 0.07)) !important;
}
/* Removed: the dividing line under the logo wrap (it conflicts with the new
   self-contained app-icon style — the logo tile reads as a single element
   without needing an underline). */
[data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap::after {
    content: none !important;
    display: none !important;
}
[data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap > .md-logo-image {
    max-width: 160px !important;
    width: min(160px, 100%) !important;
    height: auto !important;
    display: block !important;
    margin: 0 auto !important;
    transform: none !important;
    filter: drop-shadow(0 4px 16px rgba(15, 23, 42, 0.13)) drop-shadow(0 1px 4px rgba(15, 23, 42, 0.07)) !important;
}

/* Sidebar selectbox styling — kept subtle, matches surrounding nav. */
[data-testid="stSidebar"] [data-testid="stSelectbox"] {
    margin: 1.48rem 0.25rem 0.9rem 0 !important;
    width: 242px !important;
    max-width: 242px !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] label {
    margin-bottom: 0.34rem !important;
    color: #50617f !important;
    font-weight: 620 !important;
    font-size: 0.86rem !important;
}

/* Daily Health Tip carousel — 4 live-data slides auto-rotating every 3s.
   Pure CSS animation, no JS, no Streamlit re-runs. */
.md-tip-carousel {
    position: relative;
    margin: 1.6rem 0 1rem 0;
    height: 196px;
    border-radius: 22px;
    background: linear-gradient(135deg, #f0f7ff 0%, #f5f1ff 50%, #fef2ff 100%);
    border: 1px solid #dde7fb;
    box-shadow: 0 12px 28px rgba(99,102,241,0.08);
    overflow: hidden;
}
.md-tip-slide {
    position: absolute;
    inset: 0;
    padding: 1.4rem 1.6rem 1.8rem 1.6rem;
    display: block;
    align-items: center;
    opacity: 0;
    animation: tipCycle 20s infinite ease-in-out;
}
/* First child = text column, fills 100% width. */
.md-tip-slide > div:first-child {
    width: 100%;
    min-width: 0;
}
.md-tip-slide:nth-child(1) { animation-delay: 0s; }
.md-tip-slide:nth-child(2) { animation-delay: 5s; }
.md-tip-slide:nth-child(3) { animation-delay: 10s; }
.md-tip-slide:nth-child(4) { animation-delay: 15s; }
@keyframes tipCycle {
    0% { opacity: 0; transform: translateX(14px); }
    3% { opacity: 1; transform: translateX(0); }
    25% { opacity: 1; transform: translateX(0); }
    28% { opacity: 0; transform: translateX(-14px); }
    100% { opacity: 0; }
}
.md-tip-eyebrow {
    font-size: 0.66rem;
    font-weight: 800;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #2563eb;
    margin-bottom: 0.45rem;
}
.md-tip-title {
    font-size: 1.4rem;
    font-weight: 700;
    color: #0f172a;
    line-height: 1.2;
    margin-bottom: 0.4rem;
    letter-spacing: -0.01em;
}
.md-tip-desc {
    font-size: 0.92rem;
    color: #475569;
    line-height: 1.45;
    margin-bottom: 0.7rem;
    /* Use full available width — illustration is absolute-positioned now. */
    max-width: none;
}
.md-tip-metric {
    display: inline-flex;
    align-items: center;
    gap: 0.45rem;
    background: rgba(255,255,255,0.85);
    padding: 0.4rem 0.85rem;
    border-radius: 999px;
    font-size: 0.86rem;
    font-weight: 700;
    color: #1e3a8a;
    border: 1px solid rgba(59,130,246,0.18);
}
.md-tip-metric .material-symbols-rounded {
    font-size: 16px !important;
    color: #3b82f6;
}
/* Illustration is now a soft corner accent — absolute-positioned in the
   top-right of the slide so the text below uses the full tile width.
   Smaller circle + faded opacity keeps it as decoration, not competition. */
.md-tip-illust {
    position: absolute;
    top: 1.2rem;
    right: 1.4rem;
    display: flex;
    align-items: center;
    justify-content: center;
    width: 72px;
    height: 72px;
    border-radius: 50%;
    background: rgba(255,255,255,0.55);
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.8);
    opacity: 0.65;
    pointer-events: none;
}
.md-tip-illust-svg svg {
    width: 52px;
    height: 52px;
    filter: drop-shadow(0 6px 14px rgba(37,99,235,0.18));
}
.md-tip-illust .material-symbols-rounded {
    font-size: 36px !important;
}
.md-tip-water .md-tip-eyebrow { color: #0891b2; }
.md-tip-water .md-tip-metric { color: #0e7490; border-color: rgba(8,145,178,0.22); }
.md-tip-water .md-tip-metric .material-symbols-rounded { color: #06b6d4; }

.md-tip-sleep .md-tip-eyebrow { color: #7c3aed; }
.md-tip-sleep .md-tip-metric { color: #5b21b6; border-color: rgba(124,58,237,0.22); }
.md-tip-sleep .md-tip-metric .material-symbols-rounded { color: #8b5cf6; }
.md-tip-sleep .md-tip-illust .material-symbols-rounded { color: #8b5cf6; }

.md-tip-move .md-tip-eyebrow { color: #16a34a; }
.md-tip-move .md-tip-metric { color: #166534; border-color: rgba(22,163,74,0.22); }
.md-tip-move .md-tip-metric .material-symbols-rounded { color: #16a34a; }
.md-tip-move .md-tip-illust .material-symbols-rounded { color: #16a34a; }

.md-tip-vitals .md-tip-eyebrow { color: #e11d48; }
.md-tip-vitals .md-tip-metric { color: #9f1239; border-color: rgba(225,29,72,0.22); }
.md-tip-vitals .md-tip-metric .material-symbols-rounded { color: #ef4444; }
.md-tip-vitals .md-tip-illust .material-symbols-rounded { color: #ef4444; }

.md-tip-indicators {
    position: absolute;
    bottom: 0.7rem;
    left: 0;
    right: 0;
    display: flex;
    justify-content: center;
    gap: 0.4rem;
    z-index: 5;
}
.md-tip-dot {
    width: 6px;
    height: 6px;
    border-radius: 999px;
    background: rgba(59, 130, 246, 0.22);
    transition: all 0.3s ease;
}
.md-tip-dot.dot-1 { animation: tipDot 20s infinite ease-in-out; animation-delay: 0s; }
.md-tip-dot.dot-2 { animation: tipDot 20s infinite ease-in-out; animation-delay: 5s; }
.md-tip-dot.dot-3 { animation: tipDot 20s infinite ease-in-out; animation-delay: 10s; }
.md-tip-dot.dot-4 { animation: tipDot 20s infinite ease-in-out; animation-delay: 15s; }
@keyframes tipDot {
    0%, 25% { background: #3b82f6; width: 18px; }
    28%, 100% { background: rgba(59, 130, 246, 0.22); width: 6px; }
}

/* Remove left/right borders from the home chat textarea. */
[class*="st-key-home_chat_input_"] [data-testid="stTextAreaRootElement"],
[class*="st-key-home_chat_input_"] [data-baseweb="base-input"] {
    border-left: 0 !important;
    border-right: 0 !important;
    border-top: 0 !important;
    border-bottom: 0 !important;
    box-shadow: none !important;
}
[class*="st-key-home_chat_input_"] textarea {
    border-left: 0 !important;
    border-right: 0 !important;
    border-top: 0 !important;
    border-bottom: 0 !important;
    box-shadow: none !important;
}
[class*="st-key-home_chat_input_"] textarea:focus {
    border: 0 !important;
    box-shadow: none !important;
}

/* ────────────────────────────────────────────────────────────────────────
   Sidebar "posh" polish — final pass. Darker nav labels for stronger
   readability, refined hover lift, gradient divider under the logo, softer
   profile + selectbox cards, polished bottom finish with hairline divider
   above Privacy & Consent.
   ──────────────────────────────────────────────────────────────────────── */

/* Refined sidebar surface — softer vertical gradient, cleaner right edge. */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #fbfcff 0%, #f4f7fd 100%) !important;
    border-right: 1px solid #e6ecf6 !important;
}

/* Nav labels — darker for contrast, subtle hover lift + shadow. */
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button {
    color: #1f2a3d !important;
    transition: background-color 0.18s ease, border-color 0.18s ease,
                box-shadow 0.18s ease, transform 0.18s ease !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > div[data-testid="stMarkdownContainer"] p {
    color: #1f2a3d !important;
    font-weight: 660 !important;
    letter-spacing: -0.005em !important;
    /* Looser line-height so descenders (g, p, y, q, j) render fully and
       don't get clipped by the tight 1.1 default. */
    line-height: 1.45 !important;
    padding-bottom: 1px !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button:hover {
    background: #ffffff !important;
    border-color: #d6e2f6 !important;
    box-shadow: 0 6px 14px rgba(15, 23, 42, 0.06) !important;
    transform: translateY(-1px) !important;
}
[data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button:hover > div[data-testid="stMarkdownContainer"] p {
    color: #0f172a !important;
}

/* Logo divider — gradient hairline that fades at both ends. */
[data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap::after {
    left: 8% !important;
    right: 8% !important;
    height: 1px !important;
    background: linear-gradient(90deg, transparent 0%, #b9c8e6 50%, transparent 100%) !important;
    bottom: -0.6rem !important;
}

/* Profile card: distinct white card that clearly floats above the sidebar
   background. The old border (#e6edf9) and shadow (3% opacity) were too
   subtle on the near-white sidebar, making the chip invisible. */
[data-testid="stSidebar"] .md-side-profile.md-side-profile-top {
    position: relative !important;
    z-index: 5 !important;
    background: #ffffff !important;
    border: 1px solid #c3d3eb !important;
    box-shadow: 0 4px 14px rgba(15, 23, 42, 0.09), 0 1px 4px rgba(15, 23, 42, 0.05) !important;
    padding: 0.78rem 0.86rem !important;
}

/* Language selectbox — tracker-style label, match profile card visual weight. */
[data-testid="stSidebar"] [data-testid="stSelectbox"] {
    margin: 0.95rem 0 0.6rem 0 !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] label {
    color: #94a3b8 !important;
    font-weight: 700 !important;
    font-size: 0.62rem !important;
    letter-spacing: 0.16em !important;
    text-transform: uppercase !important;
    margin-bottom: 0.42rem !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] {
    border-radius: 14px !important;
    border: 1px solid #dbe5f7 !important;
    background: linear-gradient(180deg, #ffffff 0%, #fafcff 100%) !important;
    box-shadow: 0 6px 14px rgba(15, 23, 42, 0.04) !important;
    min-height: 48px !important;
    height: 48px !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
    min-height: 48px !important;
    height: 48px !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] [role="combobox"] {
    color: #1f2a3d !important;
    font-weight: 600 !important;
}

/* Polished bottom finish — hairline divider, refined typography. */
[data-testid="stSidebar"] .md-sidebar-bottom {
    padding-top: 0.85rem !important;
    margin-top: 0.6rem !important;
    border-top: 1px solid #d6e0f0 !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button {
    color: #5a6b86 !important;
    font-size: 0.84rem !important;
    font-weight: 620 !important;
    letter-spacing: 0.02em !important;
    border-radius: 12px !important;
    transition: background-color 0.18s ease, color 0.18s ease !important;
}
[data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button:hover {
    background: #f1f5fc !important;
    color: #1f2a3d !important;
}
[data-testid="stSidebar"] .sb-footer {
    color: #98a4ba !important;
    font-size: 0.66rem !important;
    letter-spacing: 0.04em !important;
    border-top: none !important;
    padding-top: 0.5rem !important;
    bottom: 0.95rem !important;
}

/* ────────────────────────────────────────────────────────────────────────
   Canonical component styling — applies at every viewport width and height.

   HISTORICAL NOTE (May 2026): this block used to be wrapped in
   `@media (max-height: 900px), (max-width: 1500px) { ... }` so it only
   matched on laptop-sized viewports. Over many design iterations the
   "compact" rules became the canonical look, and on wide+tall monitors
   (>1500px AND >900px) the gate stopped matching, falling back to plainer
   baseline rules from earlier iterations — which made the entire app
   "collapse to plain fallback" at low zoom levels on big screens.
   The media wrapper has been removed so the modern look applies always.
   Legitimate responsive media queries elsewhere in the file are kept.
   ──────────────────────────────────────────────────────────────────────── */
    /* Main column: trim outer padding so cards sit closer to the top. */
    [data-testid="stMainBlockContainer"] {
        padding: 0.5rem 2.2rem 0.6rem !important;
    }
    [data-testid="stMain"] [data-testid="stVerticalBlock"] {
        gap: 0.55rem !important;
    }

    /* Greeting + hero headings — pull the eye in faster. */
    .md-home-greet-wrap { margin: -0.6rem 0 0.5rem 0 !important; }
    .md-home-greet-wrap .md-greet { font-size: 1.95rem !important; line-height: 1.1 !important; }
    .md-home-greet-wrap .md-subgreet { font-size: 0.86rem !important; margin-top: 0.25rem !important; }
    .md-hero, .md-hero-card {
        padding: 1rem 1.2rem !important;
    }
    .md-hero-title, .md-hero h1, .md-hero h2 {
        font-size: 1.7rem !important;
        line-height: 1.15 !important;
    }
    .md-hero-subtitle, .md-hero p {
        font-size: 0.86rem !important;
        line-height: 1.35 !important;
    }

    /* Forms (sign-in, composer) — tighter internal padding. */
    [data-testid="stForm"],
    [data-testid="stForm"]:has(.st-key-si_email),
    [data-testid="stForm"]:has(.st-key-su_email) {
        padding: 0.7rem 0.95rem 0.62rem !important;
        border-radius: 18px !important;
    }
    [data-testid="stForm"] [data-testid="stVerticalBlock"] {
        gap: 0.3rem !important;
    }

    /* Auth welcome card — balanced proportions, fills the tile cleanly.
       margin-bottom gives the tabs/columns row below clear separation. */
    .md-auth-welcome-card {
        padding: 1rem 1.6rem 1.1rem 1.3rem !important;
        grid-template-columns: 96px minmax(0, 1fr) !important;
        gap: 1.2rem !important;
        border-radius: 20px !important;
        align-items: center !important;
        margin-bottom: 0.6rem !important;
    }
    .md-auth-shield, .md-auth-shield.md-auth-shield-image {
        width: 88px !important;
        height: 88px !important;
        font-size: 2.4rem !important;
        border-radius: 26px !important;
    }
    .md-auth-welcome-content {
        align-items: center !important;
        text-align: center !important;
        padding: 0 !important;
    }
    .md-auth-welcome-title {
        font-size: clamp(1.45rem, 1.95vw, 1.85rem) !important;
        margin-bottom: 0.28rem !important;
        text-align: center !important;
        width: 100% !important;
    }
    .md-auth-welcome-copy {
        font-size: clamp(0.85rem, 1vw, 0.96rem) !important;
        line-height: 1.4 !important;
        text-align: center !important;
        max-width: none !important;
    }
    .md-auth-chip-row {
        gap: 0.55rem !important;
        margin-top: 0.65rem !important;
        flex-wrap: wrap !important;
        justify-content: center !important;
    }
    .md-auth-chip {
        font-size: 0.78rem !important;
        padding: 0.32rem 0.72rem !important;
    }
    .md-auth-deco-dots { display: none !important; }

    /* "Why sign in?" right-column feature list — aggressively compact.
       margin-top pushes the card down to match the sign-in tile's top edge:
       the left column has a tab-strip (~36px) + its margin (~6px) above the
       form, so we offset by ~3.2rem to visually align both card tops. */
    .md-auth-side-card {
        padding: 0.85rem 0.95rem 0.7rem 0.95rem !important;
        margin-top: 3.2rem !important;
        border-radius: 18px !important;
    }
    .md-auth-side-title { font-size: 1.3rem !important; line-height: 1.1 !important; }
    .md-auth-side-subline { margin: 0.4rem 0 0.7rem 0 !important; }
    .md-auth-benefit {
        grid-template-columns: 34px 1fr !important;
        gap: 0.55rem !important;
        margin-bottom: 0.55rem !important;
    }
    .md-auth-benefit-ic {
        width: 34px !important;
        height: 34px !important;
        border-radius: 11px !important;
        font-size: 1rem !important;
    }
    .md-auth-benefit-title { font-size: 0.82rem !important; line-height: 1.2 !important; }
    .md-auth-benefit-copy { font-size: 0.72rem !important; line-height: 1.3 !important; margin-top: 0.08rem !important; }
    .md-auth-side-bottom {
        margin-top: 0.55rem !important;
        padding-top: 0.5rem !important;
        font-size: 0.72rem !important;
        line-height: 1.3 !important;
    }
    .md-auth-side-bottom .material-symbols-rounded { font-size: 0.9rem !important; }

    /* Sign-in form fields — tighter row heights. */
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInput"],
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInput"] {
        margin-bottom: 0.3rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInput"] label,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInput"] label {
        font-size: 0.82rem !important;
        margin-bottom: 0.18rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) input,
    [data-testid="stForm"]:has(.st-key-su_email) input {
        min-height: 38px !important;
        height: 38px !important;
        font-size: 0.88rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stFormSubmitButton"] > button,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stFormSubmitButton"] > button {
        min-height: 42px !important;
        height: 42px !important;
        font-size: 0.92rem !important;
    }

    /* Main column outer gap — keep cards close together. */
    [data-testid="stMain"] [data-testid="stVerticalBlock"] {
        gap: 0.25rem !important;
    }
    [data-testid="stMainBlockContainer"] {
        padding: 0.3rem 2rem 0.4rem !important;
    }
    /* Tabs pill bar — compact sizing. Keeps the centered pill design but
       trims paddings, font, and gap. The generous margin-top is intentional:
       it gives the welcome card above (with its soft glow ::after) enough
       breathing room so the pill doesn't visually collide. */
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        margin-top: 1.8rem !important;
        margin-bottom: 1rem !important;
        gap: 0.3rem !important;
        padding: 0.26rem !important;
        max-width: 380px !important;
        border-radius: 12px !important;
    }
    [data-testid="stTabs"] button[role="tab"] {
        padding: 0.42rem 0.95rem !important;
        font-size: 0.86rem !important;
        border-radius: 8px !important;
    }
    /* Extra breathing room after the form, before the "or" divider +
       Continue as Guest, so the secondary action doesn't crowd the
       primary Sign in button. */
    .md-auth-signin-actions {
        margin-top: 1.4rem !important;
    }
    /* Push the post-form privacy meta a bit further down for clearer
       grouping with the page footer. */
    .md-auth-meta {
        margin-top: 1.25rem !important;
    }

    /* Sidebar: shrink nav pills + logo to fit 9 buttons + footer in 768px. */
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        padding: 0.55rem 1rem 0.8rem 1rem !important;
    }
    [data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap > .md-logo-image {
        max-width: 130px !important;
        width: min(130px, 100%) !important;
    }
    [data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap {
        padding: 0.5rem 0 0.4rem 0 !important;
        margin-bottom: 0.3rem !important;
        max-width: 242px !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        border-radius: 0 !important;
        position: relative !important;
        z-index: 10 !important;
        min-height: 48px !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button {
        min-height: 32px !important;
        height: 32px !important;
        margin-bottom: 0.1rem !important;
        font-size: 0.84rem !important;
        padding-left: 0.5rem !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > div[data-testid="stMarkdownContainer"] p {
        font-size: 0.84rem !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > span:first-child,
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > [data-testid="stIconMaterial"] {
        width: 22px !important;
        min-width: 22px !important;
        height: 22px !important;
        border-radius: 7px !important;
        margin-right: 0.5rem !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button [data-testid="stIconMaterial"] {
        font-size: 0.9rem !important;
    }
    [data-testid="stSidebar"] div.st-key-nav_home {
        margin-top: 0.5rem !important;
    }
    [data-testid="stSidebar"] .md-side-profile.md-side-profile-top {
        padding: 0.55rem 0.65rem !important;
        position: relative !important;
        z-index: 5 !important;
    }
    [data-testid="stSidebar"] .md-side-avatar {
        width: 32px !important;
        min-width: 32px !important;
        height: 32px !important;
        font-size: 0.82rem !important;
    }
    /* Language picker — minimal text-link style. Sits as a subtle inline
       row at the very bottom of the sidebar (above Privacy & Consent),
       reading as: "🌐  English  ▾" with no chip background by default. On
       hover it picks up the lightest indigo wash so it still reads as
       interactive. Compact 26px, almost invisible until you look for it.
       Label is collapsed in Python. */
    [data-testid="stSidebar"] [data-testid="stSelectbox"] {
        margin-top: 0.5rem !important;
        margin-bottom: 0.5rem !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] > label,
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-testid="stWidgetLabel"] {
        display: none !important;
        height: 0 !important;
        margin: 0 !important;
    }
    /* Outer container - styled as a beautiful white card */
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] {
        min-height: 38px !important;
        height: 38px !important;
        background: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        box-shadow: 0 4px 12px rgba(15, 23, 42, 0.03) !important;
        border-radius: 12px !important;
        cursor: pointer !important;
        position: relative !important;
        padding-left: 2.2rem !important; /* Make room for the globe icon */
        padding-right: 1.8rem !important; /* Make room for the chevron */
        box-sizing: border-box !important;
        transition: background-color 0.2s ease, border-color 0.2s ease, box-shadow 0.2s ease, transform 0.2s ease !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"]:hover {
        background: #f8fafc !important;
        border-color: #cbd5e1 !important;
        box-shadow: 0 6px 16px rgba(15, 23, 42, 0.06) !important;
        transform: translateY(-1px) !important;
    }
    /* Inner wrapper — transparent, resets default padding */
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 38px !important;
        height: 38px !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        border-radius: 0 !important;
        padding: 0 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
    }
    /* Completely remove the combobox input from the visible/interactive layer */
    [data-testid="stSidebar"] [data-testid="stSelectbox"] input[role="combobox"] {
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
        height: 0 !important;
        padding: 0 !important;
        border: 0 !important;
        margin: 0 !important;
        opacity: 0 !important;
        position: absolute !important;
        left: -9999px !important;
        pointer-events: none !important;
        caret-color: transparent !important;
    }
    /* Selected value container styling */
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value] {
        font-size: 0 !important;
        line-height: 1 !important;
        display: flex !important;
        align-items: center !important;
        height: 100% !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value]::before {
        content: attr(value) !important;
        font-size: 0.8rem !important;
        font-weight: 600 !important;
        color: #334155 !important;
        letter-spacing: -0.005em !important;
        line-height: 1 !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value="English"]::before   { content: "English" !important; }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value="Tamil"]::before     { content: "Tamil" !important; }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value="Sinhala"]::before   { content: "Sinhala" !important; }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value="Hindi"]::before     { content: "Hindi" !important; }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div[value="Malayalam"]::before { content: "Malayalam" !important; }

    /* Hairline divider between the language picker and the (removed)
       Privacy & Consent button. Margins zeroed to reclaim vertical space —
       the language pill and the footer are visually distinct enough without
       an explicit hairline between them. */
    [data-testid="stSidebar"] .md-sidebar-bottom-divider {
        height: 0 !important;
        background: transparent !important;
        margin: 0 !important;
    }

    /* ── Profile chip extras (mockup match) ──
       Adds the "● Synced & up to date" status line + chevron at the right. */
    [data-testid="stSidebar"] .md-side-profile.md-side-profile-top {
        position: relative !important;
        z-index: 5 !important;
        display: flex !important;
        align-items: center !important;
        gap: 0.6rem !important;
        padding-right: 1.6rem !important;
    }
    [data-testid="stSidebar"] .md-side-profile-text {
        flex: 1 1 0 !important;
        min-width: 0 !important;
    }
    [data-testid="stSidebar"] .md-side-status {
        display: flex !important;
        align-items: center !important;
        gap: 0.3rem !important;
        font-size: 0.62rem !important;
        font-weight: 600 !important;
        color: #10b981 !important;
        margin-top: 0.2rem !important;
        line-height: 1.2 !important;
    }
    [data-testid="stSidebar"] .md-status-dot {
        width: 6px !important;
        height: 6px !important;
        border-radius: 999px !important;
        background: #10b981 !important;
        box-shadow: 0 0 0 2px rgba(16, 185, 129, 0.18) !important;
        flex-shrink: 0 !important;
    }
    [data-testid="stSidebar"] .md-status-dot.md-status-dot-off {
        background: #94a3b8 !important;
        box-shadow: 0 0 0 2px rgba(148, 163, 184, 0.18) !important;
    }
    [data-testid="stSidebar"] .md-side-chevron {
        position: absolute !important;
        right: 0.55rem !important;
        top: 50% !important;
        transform: translateY(-50%) !important;
        font-size: 1.05rem !important;
        color: #94a3b8 !important;
        -webkit-text-fill-color: #94a3b8 !important;
    }
    /* Sign-out anchor — a 28px circle pinned to the top-right of the profile
       chip. Because it's a TRUE child of the chip (not a Streamlit sibling
       with negative margins), there are no layout side-effects on the
       Recent Chats card below. Triggers ?signout=1 URL handler. */
    [data-testid="stSidebar"] .md-side-profile.md-side-profile-top {
        position: relative !important;
    }
    [data-testid="stSidebar"] .md-side-signout {
        position: relative !important;
        top: auto !important;
        right: auto !important;
        margin: 0 0 0 auto !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 28px !important;
        height: 28px !important;
        border-radius: 50% !important;
        background: transparent !important;
        color: #94a3b8 !important;
        text-decoration: none !important;
        transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease !important;
        border: 1px solid transparent !important;
    }
    [data-testid="stSidebar"] .md-side-signout:hover {
        background: #fef2f2 !important;
        color: #b91c1c !important;
        border-color: #fecaca !important;
    }
    [data-testid="stSidebar"] .md-side-signout .material-symbols-rounded {
        font-size: 1.05rem !important;
        color: inherit !important;
        -webkit-text-fill-color: currentColor !important;
    }

    /* Sign out: keep matched profile-chip dimensions but add the logout icon. */
    [data-testid="stSidebar"] div.st-key-profile_logout button [data-testid="stIconMaterial"] {
        font-size: 1rem !important;
        color: inherit !important;
        -webkit-text-fill-color: currentColor !important;
        margin-right: 0.35rem !important;
    }

    /* ── Recent Chats card (matches mockup) ──
       Single white surface containing header + "+ New chat" pill + conv rows.
       All three are now true DOM children of this card (rendered as one HTML
       block in Python), so the card's border-radius + padding actually
       encloses them. */
    [data-testid="stSidebar"] .md-recent-card {
        background: #ffffff !important;
        border: 1px solid #e6edf9 !important;
        border-radius: 16px !important;
        padding: 0.9rem !important;
        margin-top: 0.6rem !important;
        margin-bottom: 0.5rem !important;
        box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04), 0 2px 6px rgba(15, 23, 42, 0.02) !important;
        box-sizing: border-box !important;
    }
    /* "+ New chat" pill — premium card-style button inside the card. */
    [data-testid="stSidebar"] .md-new-chat-pill {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.35rem !important;
        width: 100% !important;
        box-sizing: border-box !important;
        margin: 0.25rem 0 0.55rem 0 !important;
        padding: 0.45rem 0.6rem !important;
        background: linear-gradient(135deg, #f5f8ff 0%, #eef2ff 100%) !important;
        color: #4f46e5 !important;
        border: 1px solid #dbe4ff !important;
        border-radius: 12px !important;
        font-size: 0.76rem !important;
        font-weight: 600 !important;
        text-decoration: none !important;
        letter-spacing: -0.005em !important;
        box-shadow: 0 2px 5px rgba(99, 102, 241, 0.05) !important;
        transition: all 0.2s ease !important;
    }
    [data-testid="stSidebar"] .md-new-chat-pill:hover {
        background: linear-gradient(135deg, #eef2ff 0%, #e0e7ff 100%) !important;
        border-color: #c7d2fe !important;
        color: #3730a3 !important;
        box-shadow: 0 4px 12px rgba(99, 102, 241, 0.12) !important;
        transform: translateY(-1px) !important;
        text-decoration: none !important;
    }
    [data-testid="stSidebar"] .md-new-chat-pill .material-symbols-rounded {
        font-size: 0.95rem !important;
        color: inherit !important;
        -webkit-text-fill-color: currentColor !important;
        font-variation-settings: 'FILL' 0, 'wght' 600 !important;
    }
    [data-testid="stSidebar"] .md-new-chat-pill:hover .material-symbols-rounded {
        color: inherit !important;
        -webkit-text-fill-color: currentColor !important;
    }
    [data-testid="stSidebar"] .md-recent-head {
        display: flex !important;
        align-items: center !important;
        justify-content: space-between !important;
        margin-bottom: 0.55rem !important;
        padding: 0 0.1rem !important;
    }
    [data-testid="stSidebar"] .md-recent-title {
        font-size: 0.85rem !important;
        font-weight: 700 !important;
        color: #334155 !important;
        letter-spacing: -0.005em !important;
    }
    [data-testid="stSidebar"] .md-recent-seeall {
        font-size: 0.74rem !important;
        font-weight: 600 !important;
        color: #4f46e5 !important;
        text-decoration: none !important;
        transition: color 0.15s ease !important;
    }
    [data-testid="stSidebar"] .md-recent-seeall:hover {
        color: #3730a3 !important;
        text-decoration: underline !important;
    }

    /* + New chat button inside the Recent Chats card — soft indigo fill. */
    [data-testid="stSidebar"] div.st-key-new_chat_btn button,
    [data-testid="stSidebar"] div.st-key-new_chat_btn .stButton > button {
        box-sizing: border-box !important;
        width: 100% !important;
        max-width: 100% !important;
        min-height: 36px !important;
        height: 36px !important;
        padding: 0 0.8rem !important;
        border-radius: 10px !important;
        border: 1px dashed rgba(99, 102, 241, 0.35) !important;
        background: rgba(99, 102, 241, 0.08) !important;
        color: #4f46e5 !important;
        font-weight: 660 !important;
        font-size: 0.82rem !important;
        box-shadow: none !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.35rem !important;
        margin-bottom: 0.5rem !important;
        transition: background 0.15s ease, border-color 0.15s ease !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button:hover {
        background: rgba(99, 102, 241, 0.14) !important;
        border-color: rgba(99, 102, 241, 0.5) !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button p {
        color: #4f46e5 !important;
        margin: 0 !important;
        font-size: 0.82rem !important;
        font-weight: 660 !important;
        line-height: 1 !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button [data-testid="stIconMaterial"] {
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
        font-size: 1rem !important;
    }

    /* Conversation rows — anchor links with flex layout (icon | title | time).
       Title flex-grows + ellipsis on overflow; time stays pinned right. */
    [data-testid="stSidebar"] .md-conv-list {
        display: flex !important;
        flex-direction: column !important;
        gap: 0.12rem !important;
        margin-top: 0.4rem !important;
    }
    /* Each row is now wrapped in .md-conv-row-wrap which contains the main
       click anchor (.md-conv-row) and a tiny sibling × delete anchor
       (.md-conv-del). The wrapper is the flex container; the row anchor
       takes the stretch, the × pins right and reveals on hover. */
    [data-testid="stSidebar"] .md-conv-row-wrap {
        position: relative !important;
        display: flex !important;
        align-items: stretch !important;
        min-width: 0 !important;
    }
    [data-testid="stSidebar"] .md-conv-row-wrap .md-conv-row {
        flex: 1 1 auto !important;
        min-width: 0 !important;
        padding-right: 1.8rem !important; /* reserve room for the × icon */
    }
    [data-testid="stSidebar"] .md-conv-del {
        position: absolute !important;
        right: 0.4rem !important;
        top: 50% !important;
        transform: translateY(-50%) !important;
        width: 20px !important;
        height: 20px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        border-radius: 50% !important;
        color: #94a3b8 !important;
        text-decoration: none !important;
        opacity: 1 !important;
        transition: background 0.15s ease, color 0.15s ease !important;
        z-index: 2 !important;
    }
    [data-testid="stSidebar"] .md-conv-del:hover {
        background: #fef2f2 !important;
        color: #b91c1c !important;
    }
    [data-testid="stSidebar"] .md-conv-del .material-symbols-rounded {
        font-size: 0.85rem !important;
        color: inherit !important;
        -webkit-text-fill-color: currentColor !important;
        font-variation-settings: 'FILL' 0, 'wght' 600 !important;
    }
    [data-testid="stSidebar"] .md-conv-row {
        display: flex !important;
        align-items: center !important;
        gap: 0.4rem !important;
        padding: 0.45rem 0.6rem !important;
        border-radius: 8px !important;
        border-left: 3px solid transparent !important;
        text-decoration: none !important;
        color: #475569 !important;
        background: transparent !important;
        transition: all 0.15s ease !important;
        min-width: 0 !important;
    }
    [data-testid="stSidebar"] .md-conv-row:hover {
        background: #f8fafc !important;
        color: #0f172a !important;
        text-decoration: none !important;
    }
    /* Active conversation row — styled plain just like other rows per user request */
    [data-testid="stSidebar"] .md-conv-row-active {
        background: transparent !important;
        border-left: 3px solid transparent !important;
        color: #475569 !important;
        font-weight: 550 !important;
    }
    [data-testid="stSidebar"] .md-conv-icon {
        font-size: 0.95rem !important;
        color: #94a3b8 !important;
        -webkit-text-fill-color: #94a3b8 !important;
        flex-shrink: 0 !important;
    }
    [data-testid="stSidebar"] .md-conv-row-active .md-conv-icon {
        color: #94a3b8 !important;
        -webkit-text-fill-color: #94a3b8 !important;
    }
    [data-testid="stSidebar"] .md-conv-title {
        flex: 1 1 auto !important;
        min-width: 0 !important;
        font-size: 0.76rem !important;
        font-weight: 550 !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        line-height: 1.45 !important;
        padding-bottom: 2px !important;
    }
    [data-testid="stSidebar"] .md-conv-row-active .md-conv-title {
        font-weight: 550 !important;
    }
    [data-testid="stSidebar"] .md-conv-time {
        flex-shrink: 0 !important;
        font-size: 0.64rem !important;
        font-weight: 600 !important;
        color: #94a3b8 !important;
        margin-left: 0.25rem !important;
        line-height: 1.45 !important;
        padding-bottom: 2px !important;
    }
    [data-testid="stSidebar"] .md-conv-row-active .md-conv-time {
        color: #94a3b8 !important;
    }

    /* Force every sidebar-bottom tile to match the exact same layout offset
       as the navigation buttons (margin-left 0, margin-right 1.6rem, width
       calc(100% - 1.6rem)) so they align perfectly on both edges. */
    [data-testid="stSidebar"] div.st-key-profile_logout,
    [data-testid="stSidebar"] div.st-key-profile_logout button,
    [data-testid="stSidebar"] .md-side-profile.md-side-profile-top,
    [data-testid="stSidebar"] .md-recent-card,
    [data-testid="stSidebar"] [data-testid="stSelectbox"]:has(.st-key-lang_selector),
    [data-testid="stSidebar"] .md-sidebar-foot {
        box-sizing: border-box !important;
        margin-left: 0 !important;
        margin-right: 1.6rem !important;
        width: calc(100% - 1.6rem) !important;
        max-width: none !important;
    }
    /* Profile chip — comfortable padding so 3-line content (name + email +
       sync status) doesn't pinch. The chevron was removed, so the right
       padding is also lighter (1.1 instead of 1.2). */
    [data-testid="stSidebar"] .md-side-profile.md-side-profile-top {
        display: flex !important;
        flex-direction: row !important;
        align-items: center !important;
        justify-content: flex-start !important;
        padding: 0.55rem 0.65rem !important;
        gap: 0.55rem !important;
        position: relative !important;
    }
    [data-testid="stSidebar"] .md-side-profile-text {
        text-align: center !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        flex: 1 1 0 !important;
        min-width: 0 !important;
    }
    [data-testid="stSidebar"] .md-side-avatar {
        width: 32px !important;
        min-width: 32px !important;
        height: 32px !important;
        font-size: 0.8rem !important;
        border-radius: 10px !important;
    }
    [data-testid="stSidebar"] .md-side-pname {
        font-size: 0.76rem !important;
        font-weight: 700 !important;
        line-height: 1.15 !important;
        text-align: center !important;
    }
    [data-testid="stSidebar"] .md-side-psub {
        font-size: 0.64rem !important;
        line-height: 1.2 !important;
        text-align: center !important;
        color: #94a3b8 !important;
    }
    [data-testid="stSidebar"] .md-side-status {
        font-size: 0.58rem !important;
        margin-top: 0.15rem !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.3rem !important;
    }
    [data-testid="stSidebar"] .md-side-chevron {
        display: none !important;
    }
    /* Sign out — icon-only 28px circle overlaid on the profile chip's
       top-right corner. The negative margins pull the button visually up
       onto the chip AND shrink the column contribution from +29px to -20px
       (49px reclaimed) so the column can comfortably fit Recent Chats + 3
       rows + language + footer at 900px viewport without scrolling. */
    /* st-key-profile_logout is no longer rendered (the Streamlit button was
       replaced by an anchor `?signout=1` inside the chip's HTML, removing
       all the negative-margin layout gymnastics that caused the Recent
       Chats card to collide with the chip's bottom). The selectors are
       retained inert so future legacy CSS that references them is harmless. */
    [data-testid="stSidebar"] div.st-key-profile_logout {
        display: none !important;
    }
    [data-testid="stSidebar"] div.st-key-profile_logout button,
    [data-testid="stSidebar"] div.st-key-profile_logout .stButton > button {
        width: 28px !important;
        min-width: 28px !important;
        max-width: 28px !important;
        height: 28px !important;
        min-height: 28px !important;
        padding: 0 !important;
        border-radius: 50% !important;
        background: transparent !important;
        border: 1px solid transparent !important;
        color: #94a3b8 !important;
        -webkit-text-fill-color: #94a3b8 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        margin: 0 !important;
        box-shadow: none !important;
        font-size: 0 !important;
    }
    [data-testid="stSidebar"] div.st-key-profile_logout button:hover,
    [data-testid="stSidebar"] div.st-key-profile_logout .stButton > button:hover {
        background: #fef2f2 !important;
        border-color: #fecaca !important;
        color: #b91c1c !important;
        -webkit-text-fill-color: #b91c1c !important;
        transform: none !important;
    }
    [data-testid="stSidebar"] div.st-key-profile_logout button [data-testid="stIconMaterial"] {
        font-size: 1.05rem !important;
        margin: 0 !important;
        color: inherit !important;
        -webkit-text-fill-color: currentColor !important;
    }
    /* Hide the space-character label so the icon is dead-centered in the
       circle. Also hide any wrapped Streamlit tooltip-trigger inner spans. */
    [data-testid="stSidebar"] div.st-key-profile_logout button [data-testid="stMarkdownContainer"],
    [data-testid="stSidebar"] div.st-key-profile_logout button p {
        display: none !important;
    }
    /* Recent Chats card: ultra-tight to fit at 900px viewport.
       margin-top: 0.7rem (~11px) gives a clean visual gap below the
       profile chip area so the two cards don't collide. */
    /* Recent Chats card: ultra-tight to fit at 900px viewport.
       margin-top: 0.7rem (~11px) gives a clean visual gap below the
       profile chip area so the two cards don't collide. */
    [data-testid="stSidebar"] div.st-key-recent_chats_card {
        background: #ffffff !important;
        border: 1px solid #e6edf9 !important;
        border-radius: 11px !important;
        padding: 0.45rem 0.5rem 0.4rem 0.5rem !important;
        margin-top: 0.7rem !important;
        margin-bottom: 0.25rem !important;
        box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04), 0 2px 6px rgba(15, 23, 42, 0.02) !important;
        box-sizing: border-box !important;
        margin-left: 0 !important;
        margin-right: 1.6rem !important;
        width: calc(100% - 1.6rem) !important;
        max-width: none !important;
    }
    [data-testid="stSidebar"] div.st-key-recent_chats_card [data-testid="stVerticalBlockBorderWrapper"] > div > [data-testid="stVerticalBlock"] {
        gap: 0.2rem !important;
    }
    [data-testid="stSidebar"] div.st-key-recent_chats_card [data-testid="column"] {
        padding: 0 !important;
    }
    [data-testid="stSidebar"] .md-recent-head {
        margin-bottom: 0.25rem !important;
    }
    [data-testid="stSidebar"] .md-recent-title {
        font-size: 0.76rem !important;
        font-weight: 700 !important;
        color: #334155 !important;
        letter-spacing: -0.005em !important;
        line-height: 1.45 !important;
    }
    
    /* See all small link button in header */
    [data-testid="stSidebar"] div.st-key-recent_see_all_btn button,
    [data-testid="stSidebar"] div.st-key-recent_see_all_btn button:active,
    [data-testid="stSidebar"] div.st-key-recent_see_all_btn button:focus {
        background: transparent !important;
        border: none !important;
        color: #4f46e5 !important;
        font-size: 0.64rem !important;
        font-weight: 600 !important;
        padding: 0 !important;
        margin: 0 !important;
        box-shadow: none !important;
        outline: none !important;
        min-height: unset !important;
        height: auto !important;
        float: right !important;
    }
    [data-testid="stSidebar"] div.st-key-recent_see_all_btn button:hover {
        color: #3730a3 !important;
        text-decoration: underline !important;
        background: transparent !important;
    }

    /* New chat — compact 30px, smaller text. */
    [data-testid="stSidebar"] div.st-key-new_chat_btn button {
        box-sizing: border-box !important;
        width: 100% !important;
        min-height: 30px !important;
        height: 30px !important;
        padding: 0 0.6rem !important;
        margin-bottom: 0.35rem !important;
        font-size: 0.72rem !important;
        border-radius: 8px !important;
        border: 1px solid #dbe4ff !important;
        background: linear-gradient(135deg, #f5f8ff 0%, #eef2ff 100%) !important;
        color: #4f46e5 !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 5px rgba(99, 102, 241, 0.05) !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.35rem !important;
        transition: all 0.2s ease !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button:hover {
        background: linear-gradient(135deg, #eef2ff 0%, #e0e7ff 100%) !important;
        border-color: #c7d2fe !important;
        color: #3730a3 !important;
        box-shadow: 0 4px 12px rgba(99, 102, 241, 0.12) !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button p {
        font-size: 0.72rem !important;
        color: #4f46e5 !important;
        margin: 0 !important;
        font-weight: 600 !important;
        line-height: 1 !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button [data-testid="stIconMaterial"] {
        font-size: 0.88rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
    }

    /* Conversation rows — slightly tighter to fit footer at 900px. */
    [data-testid="stSidebar"] div[class*="st-key-conv_select_"] button,
    [data-testid="stSidebar"] div[class*="st-key-conv_active_"] button {
        background: transparent !important;
        border: none !important;
        border-radius: 8px !important;
        color: #475569 !important;
        text-align: left !important;
        font-size: 0.76rem !important;
        font-weight: 550 !important;
        padding: 0.4rem 0.45rem !important;
        min-height: unset !important;
        height: auto !important;
        width: 100% !important;
        box-shadow: none !important;
        display: flex !important;
        align-items: center !important;
        gap: 0.35rem !important;
        transition: all 0.15s ease !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-conv_select_"] button:hover {
        background: #f8fafc !important;
        color: #0f172a !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-conv_active_"] button {
        background: #f1f5f9 !important;
        color: #0f172a !important;
        font-weight: 600 !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-conv_select_"] button p,
    [data-testid="stSidebar"] div[class*="st-key-conv_active_"] button p {
        margin: 0 !important;
        font-size: 0.76rem !important;
        line-height: 1.15 !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }

    /* Delete buttons */
    [data-testid="stSidebar"] div[class*="st-key-conv_del_"] button {
        background: transparent !important;
        border: none !important;
        color: #94a3b8 !important;
        font-size: 0.74rem !important;
        padding: 0 !important;
        margin: 0 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 20px !important;
        height: 20px !important;
        min-width: 20px !important;
        min-height: 20px !important;
        box-shadow: none !important;
        border-radius: 50% !important;
        transition: background 0.15s ease, color 0.15s ease !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-conv_del_"] button:hover {
        background: #fef2f2 !important;
        color: #b91c1c !important;
    }

    /* Two-line sidebar footer — compact. */
    [data-testid="stSidebar"] .md-sidebar-foot {
        margin-top: 0.6rem !important;
        text-align: center !important;
        opacity: 0.95 !important;
    }
    [data-testid="stSidebar"] .md-sidebar-foot-links {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.5rem !important;
        font-size: 0.65rem !important;
        font-weight: 600 !important;
        margin-bottom: 0.22rem !important;
    }
    
    /* Footer buttons */
    [data-testid="stSidebar"] div.st-key-privacy_btn button,
    [data-testid="stSidebar"] div.st-key-help_btn button {
        background: transparent !important;
        border: none !important;
        color: #64748b !important;
        font-size: 0.65rem !important;
        font-weight: 600 !important;
        padding: 0 !important;
        margin: 0 !important;
        min-height: unset !important;
        height: auto !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] div.st-key-privacy_btn button:hover,
    [data-testid="stSidebar"] div.st-key-help_btn button:hover {
        color: #4f46e5 !important;
        text-decoration: underline !important;
        background: transparent !important;
    }
    [data-testid="stSidebar"] div.st-key-privacy_btn button p,
    [data-testid="stSidebar"] div.st-key-help_btn button p {
        font-size: 0.65rem !important;
        margin: 0 !important;
    }
    [data-testid="stSidebar"] div.st-key-privacy_btn button {
        float: right !important;
    }
    [data-testid="stSidebar"] div.st-key-help_btn button {
        float: left !important;
    }
    [data-testid="stSidebar"] .md-sidebar-foot-dot {
        color: #cbd5e1 !important;
        font-weight: 700 !important;
    }
    [data-testid="stSidebar"] .md-sidebar-foot-copy {
        font-size: 0.62rem !important;
        color: #94a3b8 !important; /* Muted copyright text */
        font-weight: 500 !important;
        line-height: 1.3 !important;
        margin-top: 0.15rem !important;
        letter-spacing: 0.01em !important;
    }

    /* (Legacy: text-pill sign-out styling SUPERSEDED. Sign out is now an
       icon-only 28px circle overlaid on the profile chip — see the earlier
       div.st-key-profile_logout rules ~line 6052. This block intentionally
       left empty to avoid a duplicate text-pill cascade win.) */

    /* ── Sidebar chat-history section (Your Chats) ──
       Clean divider above the section, refined "YOUR CHATS" label, primary
       indigo "+ New chat" CTA, and each past chat row as a subtle list item
       with title left + tiny × delete on the right. */
    [data-testid="stSidebar"] [data-testid="stElementContainer"]:has(hr) {
        overflow: visible !important;
        height: auto !important;
    }
    [data-testid="stSidebar"] hr {
        border: none !important;
        height: 1px !important;
        background: #cbd5e1 !important;
        margin-top: 0.85rem !important;
        margin-bottom: 0.85rem !important;
        margin-left: 0 !important;
        margin-right: 1.6rem !important;
        width: calc(100% - 1.6rem) !important;
        display: block !important;
        opacity: 1 !important;
        visibility: visible !important;
    }
    [data-testid="stSidebar"] .sb-title {
        font-size: 0.6rem !important;
        font-weight: 700 !important;
        color: #94a3b8 !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
        margin: 0.4rem 0 0.5rem 0.1rem !important;
    }

    /* + New chat CTA — compact dashed indigo pill inside the Recent card. */
    [data-testid="stSidebar"] div.st-key-new_chat_btn button,
    [data-testid="stSidebar"] div.st-key-new_chat_btn .stButton > button {
        min-height: 30px !important;
        height: 30px !important;
        border-radius: 8px !important;
        border: 1px dashed rgba(99, 102, 241, 0.4) !important;
        background: rgba(99, 102, 241, 0.08) !important;
        color: #4f46e5 !important;
        font-weight: 660 !important;
        font-size: 0.72rem !important;
        box-shadow: none !important;
        margin-bottom: 0.35rem !important;
        transition: background 0.15s ease, border-color 0.15s ease !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button:hover {
        background: rgba(99, 102, 241, 0.15) !important;
        border-color: rgba(99, 102, 241, 0.55) !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button p {
        color: #4f46e5 !important;
        margin: 0 !important;
        font-size: 0.72rem !important;
        font-weight: 660 !important;
        line-height: 1 !important;
    }
    [data-testid="stSidebar"] div.st-key-new_chat_btn button [data-testid="stIconMaterial"] {
        font-size: 0.88rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
    }

    /* Past-chats list (chat history). Selectors use DESCENDANT (.stButton
       button) not direct child (.stButton > button) because the title
       buttons have help= tooltips → Streamlit wraps them in three extra
       divs (stTooltipHoverTarget, stTooltipIcon, etc.). */
    [data-testid="stSidebar"] .md-past-chats {
        display: flex !important;
        flex-direction: column !important;
        gap: 0.18rem !important;
        margin-bottom: 0.4rem !important;
    }
    [data-testid="stSidebar"] .md-past-chats [data-testid="stHorizontalBlock"] {
        gap: 0.2rem !important;
        align-items: center !important;
        flex-wrap: nowrap !important;
    }
    [data-testid="stSidebar"] .md-past-chats [data-testid="stColumn"] {
        min-width: 0 !important;
        padding: 0 !important;
    }
    /* Pin the delete column to a fixed 32px so it always stays visible. */
    [data-testid="stSidebar"] .md-past-chats [data-testid="stColumn"]:last-child {
        flex: 0 0 32px !important;
        width: 32px !important;
        max-width: 32px !important;
    }
    /* Chat title button — full row width minus the × button. */
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_open_"] button {
        min-height: 34px !important;
        height: 34px !important;
        max-height: 34px !important;
        padding: 0 0.65rem !important;
        border-radius: 9px !important;
        border: 1px solid transparent !important;
        background: transparent !important;
        color: #475569 !important;
        font-weight: 550 !important;
        font-size: 0.76rem !important;
        text-align: left !important;
        justify-content: flex-start !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        box-shadow: none !important;
        display: flex !important;
        align-items: center !important;
        width: 100% !important;
        transition: background 0.15s ease, color 0.15s ease !important;
    }
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_open_"] button:hover {
        background: #f1f5fc !important;
        color: #1f2a3d !important;
        border-color: transparent !important;
    }
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_open_"] button p,
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_open_"] button [data-testid="stMarkdownContainer"] {
        margin: 0 !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        text-align: left !important;
        font-size: 0.76rem !important;
        line-height: 1.2 !important;
        max-width: 100% !important;
        -webkit-line-clamp: 1 !important;
        display: block !important;
        width: 100% !important;
    }
    /* Active chat row — indigo wash + indigo text. */
    [data-testid="stSidebar"] .md-past-active [class*="st-key-conv_open_"] button {
        background: linear-gradient(180deg, #eef0ff 0%, #e0e7ff 100%) !important;
        color: #3730a3 !important;
        font-weight: 660 !important;
    }
    [data-testid="stSidebar"] .md-past-active [class*="st-key-conv_open_"] button p {
        color: #3730a3 !important;
        font-weight: 660 !important;
    }
    /* Delete × button: tiny ghost square pinned to the right column. */
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_del_"] button {
        min-width: 28px !important;
        width: 28px !important;
        max-width: 28px !important;
        min-height: 28px !important;
        height: 28px !important;
        padding: 0 !important;
        border-radius: 7px !important;
        border: none !important;
        background: transparent !important;
        color: #cbd5e1 !important;
        font-size: 0.92rem !important;
        font-weight: 600 !important;
        line-height: 1 !important;
        box-shadow: none !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        transition: background 0.15s ease, color 0.15s ease !important;
    }
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_del_"] button:hover {
        background: #fef2f2 !important;
        color: #b91c1c !important;
    }
    [data-testid="stSidebar"] .md-past-chats [class*="st-key-conv_del_"] button p {
        margin: 0 !important;
        line-height: 1 !important;
        font-size: 1rem !important;
    }
    /* Globe glyph at the start — sits tight against the text. */
    /* Globe glyph and chevron styling now fully integrated into the global selectbox card declaration */
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"]::before {
        content: "🌐";
        position: absolute !important;
        left: 0.8rem !important;
        top: 50% !important;
        transform: translateY(-50%) !important;
        font-size: 0.9rem !important;
        line-height: 1 !important;
        pointer-events: none !important;
        opacity: 0.9 !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] svg {
        position: absolute !important;
        right: 0.8rem !important;
        top: 50% !important;
        transform: translateY(-50%) !important;
        width: 12px !important;
        height: 12px !important;
        fill: #64748b !important;
        opacity: 0.9 !important;
    }
    [data-testid="stSidebar"] .md-sidebar-bottom {
        padding-top: 0.55rem !important;
        margin-top: 0.4rem !important;
    }
    [data-testid="stSidebar"] div.st-key-nav_privacy_bottom {
        margin-bottom: 1rem !important;
    }
    [data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button {
        min-height: 36px !important;
        height: 36px !important;
    }
    [data-testid="stSidebar"] .sb-footer {
        bottom: 0.5rem !important;
    }

    /* ────────────────────────────────────────────────────────────────────
       Apple-style polish — home dashboard at 1366×768.
       Goals: nothing wraps mid-word, vertical rhythm reads cleanly, every
       tile fits its content on one line where possible, icons sit in
       softly tinted square containers (consistent across chips + cards),
       send button matches the brand indigo without dominating.
       ──────────────────────────────────────────────────────────────────── */

    /* Greeting — slightly tighter, brand black. */
    .md-home-greet-wrap .md-greet {
        font-size: 1.7rem !important;
        line-height: 1.08 !important;
        font-weight: 720 !important;
        letter-spacing: -0.018em !important;
    }
    .md-home-greet-wrap .md-subgreet {
        font-size: 0.82rem !important;
        margin-top: 0.18rem !important;
        color: #64748b !important;
    }

    /* ── Quick-action chips ──
       Pill row sits clearly below greeting (1.1rem breathing room). Each
       chip: short single-line label, soft indigo-tinted icon square,
       generous internal padding. Hover gives a small lift + indigo border
       tint, no shape change. Icons are forced to monochrome indigo so
       emoji-style material symbols don't break the unified palette. */
    [data-testid="stHorizontalBlock"]:has(.st-key-qa_headache) {
        margin-top: 1.1rem !important;
        margin-bottom: 0.4rem !important;
    }
    .st-key-qa_headache .stButton > button,
    .st-key-qa_tired .stButton > button,
    .st-key-qa_symptoms .stButton > button,
    .st-key-qa_sleep .stButton > button {
        min-height: 48px !important;
        height: 48px !important;
        padding: 0 0.85rem !important;
        border-radius: 14px !important;
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%) !important;
        border: 1px solid #e6ecf6 !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.85) inset,
            0 2px 5px rgba(15, 23, 42, 0.03) !important;
        gap: 0.6rem !important;
        transition:
            transform 0.18s ease,
            border-color 0.18s ease,
            box-shadow 0.2s ease !important;
    }
    .st-key-qa_headache .stButton > button p,
    .st-key-qa_tired .stButton > button p,
    .st-key-qa_symptoms .stButton > button p,
    .st-key-qa_sleep .stButton > button p {
        font-size: 0.83rem !important;
        font-weight: 620 !important;
        color: #1f2a3d !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        -webkit-line-clamp: 1 !important;
        line-height: 1.1 !important;
        letter-spacing: -0.005em !important;
        margin: 0 !important;
    }
    /* Force a single, on-theme indigo for every quick-chip icon — even
       emoji-style material symbols. */
    .st-key-qa_headache .stButton > button [data-testid="stIconMaterial"],
    .st-key-qa_tired .stButton > button [data-testid="stIconMaterial"],
    .st-key-qa_symptoms .stButton > button [data-testid="stIconMaterial"],
    .st-key-qa_sleep .stButton > button [data-testid="stIconMaterial"] {
        width: 26px !important;
        height: 26px !important;
        min-width: 26px !important;
        font-size: 1.05rem !important;
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.14) 0%, rgba(139, 92, 246, 0.12) 100%) !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
        border-radius: 8px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        font-variation-settings: 'FILL' 0, 'wght' 500, 'GRAD' 0, 'opsz' 24 !important;
    }
    .st-key-qa_headache .stButton > button:hover,
    .st-key-qa_tired .stButton > button:hover,
    .st-key-qa_symptoms .stButton > button:hover,
    .st-key-qa_sleep .stButton > button:hover {
        background: #ffffff !important;
        border-color: rgba(99, 102, 241, 0.35) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 8px 18px rgba(79, 70, 229, 0.1) !important;
        transform: translateY(-1px) !important;
    }
    .st-key-qa_headache .stButton > button:hover [data-testid="stIconMaterial"],
    .st-key-qa_tired .stButton > button:hover [data-testid="stIconMaterial"],
    .st-key-qa_symptoms .stButton > button:hover [data-testid="stIconMaterial"],
    .st-key-qa_sleep .stButton > button:hover [data-testid="stIconMaterial"] {
        background: linear-gradient(135deg, #4f46e5 0%, #6366f1 100%) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }

    /* ── Composer (Apple-style) ──
       The composer reads as one clean white card: textarea sits flush at
       top with no inner box, action row below it has icon-only Upload /
       Voice ghost circles on the left and a refined indigo Send pill on
       the right. No double-borders, no grey defaults from baseweb. */
    [data-testid="stForm"]:has(.st-key-home_send_btn) {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 22px !important;
        padding: 0.95rem 1rem 0.85rem 1rem !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 8px 24px rgba(15, 23, 42, 0.05) !important;
    }
    /* Textarea + baseweb wrappers — strip all background + borders so the
       form's white surface flows through. */
    [data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextAreaRootElement"],
    [data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextAreaRootElement"] [data-baseweb="base-input"] {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        padding: 0 !important;
    }
    [data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea {
        min-height: 72px !important;
        height: 72px !important;
        font-size: 0.92rem !important;
        padding: 0.2rem 0.2rem !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        color: #1f2a3d !important;
    }
    [data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea::placeholder {
        color: #94a3b8 !important;
        font-size: 0.92rem !important;
    }
    [data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea:focus {
        outline: none !important;
        border: none !important;
        box-shadow: none !important;
    }

    /* Upload + Voice — small clean pills with paperclip/mic icon + text
       label. Both fill 100% of their column so they have identical width
       and align perfectly along their left edges; content is centered
       inside each pill. Override the original 112×56 pill rules so they
       read as compact secondary actions instead of dominant chips. */
    .st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
    .st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
    .st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"],
    .st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"] {
        min-width: 0 !important;
        width: 100% !important;
        min-height: 36px !important;
        height: 36px !important;
        padding: 0 0.85rem !important;
        border-radius: 999px !important;
        border: 1px solid #e6ecf6 !important;
        background: #ffffff !important;
        box-shadow: none !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.45rem !important;
        transition:
            background 0.18s ease,
            border-color 0.18s ease,
            transform 0.15s ease !important;
    }
    /* Kill the baseweb 8px margin-right on the inner icon wrap so the
       icon and label sit side-by-side without a stray gap. */
    .st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button > span:first-child,
    .st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button > span:first-child {
        margin: 0 !important;
    }
    .st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button:hover,
    .st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button:hover {
        background: #f7f9ff !important;
        border-color: rgba(99, 102, 241, 0.32) !important;
        transform: translateY(-1px) !important;
    }
    .st-key-home_upload_btn button p,
    .st-key-home_voice_btn button p {
        display: block !important;
        margin: 0 !important;
        font-size: 0.83rem !important;
        font-weight: 600 !important;
        line-height: 1 !important;
        color: #1f2a3d !important;
    }
    .st-key-home_upload_btn button [data-testid="stIconMaterial"],
    .st-key-home_voice_btn button [data-testid="stIconMaterial"] {
        margin: 0 !important;
        color: #1f2a3d !important;
        -webkit-text-fill-color: #1f2a3d !important;
        font-size: 1rem !important;
    }

    /* Send button — round indigo gradient ball, the primary action. */
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button[kind="primary"] {
        min-width: 42px !important;
        width: 42px !important;
        min-height: 42px !important;
        height: 42px !important;
        border-radius: 50% !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.2) 0%, rgba(255,255,255,0) 55%),
            linear-gradient(135deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        border: none !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 8px 18px rgba(79, 70, 229, 0.34) !important;
        padding: 0 !important;
        transition: transform 0.15s ease, box-shadow 0.2s ease !important;
    }
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button:hover {
        transform: translateY(-1px) scale(1.04) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 12px 24px rgba(79, 70, 229, 0.42) !important;
    }
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        font-size: 1.15rem !important;
        margin: 0 !important;
    }
    /* The submit button has a 2-column grid (icon | label). With a blank
       label (" "), the empty stMarkdownContainer still occupies a grid
       track, pushing the icon left. Collapse the whole label container so
       the icon sits dead-center. Switch the button to flex centering too
       so any residual grid-gap doesn't bias the layout. */
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0 !important;
    }
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stMarkdownContainer"] {
        display: none !important;
    }
    /* Baseweb wraps the icon in an outer <span> with margin-right: 8px (the
       gap it would use between icon + label). With the label hidden, that
       margin shoves the icon left. Zero it out so the icon sits dead-center. */
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button > span:first-child {
        margin: 0 !important;
    }
    /* Glow under composer — softer in compact mode. */
    .md-composer-glow {
        height: 18px !important;
        opacity: 0.35 !important;
    }
    .md-home-composer-note {
        font-size: 0.72rem !important;
        margin-top: 0.6rem !important;
        margin-bottom: 0.4rem !important;
        color: #94a3b8 !important;
    }

    /* ── Chat composer (active-chat page) ──
       Mirror every Apple-style rule from the home composer but scoped to
       the chat_* keys. Identical visual: clean white shell, Upload + Voice
       text pills, round indigo Send ball with a perfectly centered arrow. */
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 22px !important;
        padding: 0.95rem 1rem 0.85rem 1rem !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 8px 24px rgba(15, 23, 42, 0.05) !important;
    }
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextAreaRootElement"],
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextAreaRootElement"] [data-baseweb="base-input"] {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        padding: 0 !important;
    }
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] textarea {
        min-height: 72px !important;
        height: 72px !important;
        font-size: 0.92rem !important;
        padding: 0.2rem 0.2rem !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        color: #1f2a3d !important;
    }
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] textarea::placeholder {
        color: #94a3b8 !important;
        font-size: 0.92rem !important;
    }
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] textarea:focus {
        outline: none !important;
        border: none !important;
        box-shadow: none !important;
    }
    /* Upload + Voice + Clear — same pill style as home. */
    .st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
    .st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button[kind="secondary"],
    .st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"],
    .st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"],
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"] {
        min-width: 0 !important;
        width: 100% !important;
        min-height: 36px !important;
        height: 36px !important;
        padding: 0 0.85rem !important;
        border-radius: 999px !important;
        border: 1px solid #e6ecf6 !important;
        background: #ffffff !important;
        box-shadow: none !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.45rem !important;
        transition:
            background 0.18s ease,
            border-color 0.18s ease,
            color 0.18s ease,
            transform 0.15s ease !important;
    }
    .st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button > span:first-child,
    .st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button > span:first-child,
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button > span:first-child {
        margin: 0 !important;
    }
    /* Upload + Voice hover: indigo tint. */
    .st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button:hover,
    .st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button:hover {
        background: #f7f9ff !important;
        border-color: rgba(99, 102, 241, 0.32) !important;
        transform: translateY(-1px) !important;
    }
    /* Clear hover: soft red danger tint to signal destructive action. */
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button:hover {
        background: #fef2f2 !important;
        border-color: #fecaca !important;
        color: #b91c1c !important;
        transform: translateY(-1px) !important;
    }
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button:hover [data-testid="stIconMaterial"],
    .st-key-chat_clear_btn [data-testid="stFormSubmitButton"] > button:hover p {
        color: #b91c1c !important;
        -webkit-text-fill-color: #b91c1c !important;
    }
    .st-key-chat_upload_btn button p,
    .st-key-chat_voice_btn button p,
    .st-key-chat_clear_btn button p {
        display: block !important;
        margin: 0 !important;
        font-size: 0.83rem !important;
        font-weight: 600 !important;
        line-height: 1 !important;
        color: #1f2a3d !important;
    }
    .st-key-chat_upload_btn button [data-testid="stIconMaterial"],
    .st-key-chat_voice_btn button [data-testid="stIconMaterial"],
    .st-key-chat_clear_btn button [data-testid="stIconMaterial"] {
        margin: 0 !important;
        color: #1f2a3d !important;
        -webkit-text-fill-color: #1f2a3d !important;
        font-size: 1rem !important;
    }
    /* Chat Send — same round indigo ball as home Send. Selector chain
       matches the older :has() override at line ~7087 so this wins. */
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button,
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button[kind="primary"],
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button,
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button[kind="primary"] {
        min-width: 42px !important;
        width: 42px !important;
        min-height: 42px !important;
        height: 42px !important;
        border-radius: 50% !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.2) 0%, rgba(255,255,255,0) 55%),
            linear-gradient(135deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        border: none !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 8px 18px rgba(79, 70, 229, 0.34) !important;
        padding: 0 !important;
        transition: transform 0.15s ease, box-shadow 0.2s ease !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0 !important;
    }
    /* Optical centering: the send paper-plane glyph tilts up-right so its
       geometric center isn't its visual center. Nudge it 1px left + 1px
       down so it reads centered in the round button. */
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
        transform: translate(-1px, 1px) !important;
    }
    .st-key-home_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
        transform: translate(-1px, 1px) !important;
    }
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button:hover {
        transform: translateY(-1px) scale(1.04) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 12px 24px rgba(79, 70, 229, 0.42) !important;
    }
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        font-size: 1.15rem !important;
        margin: 0 !important;
    }
    /* Strip the empty-label container + baseweb icon margin so the arrow
       sits dead-center in the round button. */
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stMarkdownContainer"] {
        display: none !important;
    }
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] > button > span:first-child {
        margin: 0 !important;
    }
    /* Push the round send button flush to the right edge of the composer.
       The send lives in column 5 of the chat form's action row; force that
       column + its inner wrappers to right-align so the 42px ball sits in
       the corner instead of pinned to the left of its column. */
    [data-testid="stForm"]:has(.st-key-chat_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_send_btn) [data-testid="stColumn"]:last-child,
    [data-testid="stForm"]:has(.st-key-chat_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_send_btn) [data-testid="stColumn"]:last-child > div,
    [data-testid="stForm"]:has(.st-key-chat_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_send_btn) [data-testid="stColumn"]:last-child .st-key-chat_send_btn {
        display: flex !important;
        justify-content: flex-end !important;
        align-items: center !important;
        margin-left: auto !important;
        padding-right: 0 !important;
    }
    .st-key-chat_send_btn [data-testid="stFormSubmitButton"] {
        margin-left: auto !important;
        display: flex !important;
        justify-content: flex-end !important;
        width: 100% !important;
    }

    /* ── Chat conversation polish (Apple-style) ───────────────────────
       Fixes badge overlap, refines the bubbles, makes Yes/No + Download
       + Clear all share the brand indigo language. */

    /* MEDICHAT label above each bot bubble: clear top margin + overflow
       visible so the label cap doesn't get clipped by parent. */
    .bot-label {
        font-size: 0.62rem !important;
        margin-top: 1.6rem !important;
        margin-bottom: 0.35rem !important;
        margin-left: 44px !important;
        color: #94a3b8 !important;
        letter-spacing: 0.12em !important;
        line-height: 1.4 !important;
        overflow: visible !important;
        padding-top: 4px !important;
    }
    /* Make sure parent containers don't clip the label. */
    [data-testid="stMain"] .stMarkdown:has(.bot-label),
    [data-testid="stMain"] .stMarkdown:has(.bot-bubble) {
        overflow: visible !important;
    }

    /* Bot bubble: chat-bubble feel matching the user side — clean white
       surface, soft border, mirrored corner truncation (top-left is the
       "tail" closest to the avatar). No left accent stripe (the avatar +
       header inside the bubble already signal the speaker). */
    .bot-bubble {
        background: #ffffff !important;
        border-radius: 4px 18px 18px 18px !important;
        border: 1px solid #e6ecf6 !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.95) inset,
            0 4px 14px rgba(15, 23, 42, 0.04) !important;
        padding: 0.95rem 1.15rem !important;
        font-size: 0.92rem !important;
        line-height: 1.6 !important;
        max-width: 100% !important;
    }
    /* Kill the legacy purple accent stripe — it wasn't aligning cleanly
       with the new in-bubble header + caused a visual seam at the top. */
    .bot-bubble::before {
        display: none !important;
        content: none !important;
    }

    /* User bubble: soft lavender wash (matches mockup), dark indigo text.
       Bubble width hugs its content so short messages stay on one line and
       only long ones wrap. */
    .user-bubble {
        background: linear-gradient(135deg, #e0e7ff 0%, #e0f2fe 100%) !important;
        color: #1e293b !important;
        border-radius: 18px 18px 4px 18px !important;
        padding: 0.75rem 1.1rem !important;
        font-size: 0.92rem !important;
        font-weight: 600 !important;
        line-height: 1.45 !important;
        max-width: 100% !important;
        width: fit-content !important;
        margin-left: auto !important;
        border: 1px solid #c7d2fe !important;
        box-shadow: 0 4px 14px rgba(99, 102, 241, 0.05) !important;
    }
    /* User stack should NOT cap width — let the bubble grow naturally up
       to its container so short messages stay on one line. */
    .user-stack {
        max-width: 70% !important;
        width: auto !important;
        display: inline-flex !important;
        flex-direction: column !important;
        align-items: flex-end !important;
        margin-left: auto !important;
    }
    .user-wrap {
        margin-bottom: 1.6rem !important;
        gap: 0.75rem !important;
    }

    /* Avatars: smaller + indigo for bot, soft grey for user. */
    .av {
        width: 32px !important;
        height: 32px !important;
        border-radius: 10px !important;
        font-size: 0.78rem !important;
        box-shadow: 0 2px 6px rgba(79, 70, 229, 0.18) !important;
    }
    .av-bot {
        background: linear-gradient(135deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        color: #ffffff !important;
    }
    .av-user {
        background: linear-gradient(135deg, #e0e7ff 0%, #c7d2fe 100%) !important;
        color: #4338ca !important;
    }

    /* ── Chat page hero (matches mockup) ──
       Big bold title at the top-left, privacy indicator floats top-right. */
    .md-chat-hero {
        display: flex !important;
        align-items: flex-start !important;
        justify-content: space-between !important;
        gap: 1rem !important;
        margin: 0.2rem 0 2.8rem 0 !important;
        flex-wrap: wrap !important;
    }
    .md-chat-hero-text { min-width: 0 !important; }
    .md-chat-hero-title {
        font-size: 1.9rem !important;
        font-weight: 740 !important;
        color: #0f172a !important;
        letter-spacing: -0.022em !important;
        line-height: 1.15 !important;
        margin-bottom: 0.25rem !important;
        display: flex !important;
        align-items: center !important;
        gap: 0.5rem !important;
    }
    /* Shield is now an inline SVG (not a Material Symbols glyph) — renders
       as a proper filled shield silhouette with a soft check inside,
       identical across all browsers regardless of icon font loading. */
    .md-chat-hero-shield {
        display: inline-block !important;
        width: 1.6rem !important;
        height: 1.6rem !important;
        color: #4f46e5 !important;
        fill: currentColor !important;
        flex-shrink: 0 !important;
        vertical-align: -0.25em !important;
        overflow: visible !important;
    }
    .md-chat-hero-sub {
        font-size: 0.88rem !important;
        color: #64748b !important;
        line-height: 1.4 !important;
    }
    .md-chat-hero-privacy {
        display: inline-flex !important;
        align-items: center !important;
        gap: 0.4rem !important;
        font-size: 0.78rem !important;
        font-weight: 600 !important;
        color: #475569 !important;
        background: transparent !important;
        flex-shrink: 0 !important;
        margin-top: 0.45rem !important;
    }
    .md-chat-hero-privacy .material-symbols-rounded {
        font-size: 0.95rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
    }

    /* ── Bot bubble redesigned to match mockup ──
       Logo as avatar (circular), "MediChat AI" + sparkle header INSIDE the
       bubble, timestamp right-aligned below the bubble. */
    .av-bot.av-bot-image {
        background: transparent !important;
        border: 1px solid #e6ecf6 !important;
        box-shadow: 0 2px 6px rgba(15, 23, 42, 0.06) !important;
        padding: 2px !important;
        overflow: hidden !important;
    }
    .av-bot.av-bot-image img {
        width: 100% !important;
        height: 100% !important;
        object-fit: contain !important;
        display: block !important;
        border-radius: 8px !important;
    }
    .bot-stack {
        flex: 1 1 auto;
        min-width: 0;
        max-width: 88%;
    }
    .user-stack {
        display: flex;
        flex-direction: column;
        align-items: flex-end;
        max-width: 78%;
    }
    /* Header line inside the bot bubble: "MediChat AI" name + sparkle. */
    .bot-bubble-head {
        display: flex;
        align-items: center;
        gap: 0.35rem;
        margin-bottom: 0.4rem;
    }
    .bot-bubble-name {
        font-size: 0.84rem !important;
        font-weight: 700 !important;
        color: #4f46e5 !important;
        letter-spacing: -0.005em !important;
    }
    .bot-bubble-spark {
        font-size: 0.85rem !important;
        color: #6366f1 !important;
        -webkit-text-fill-color: #6366f1 !important;
    }
    /* Hide the OLD external bot-label since we now have the header inside. */
    .bot-label {
        display: none !important;
    }
    /* Timestamps below user + bot messages. */
    .bubble-ts {
        font-size: 0.66rem !important;
        color: #94a3b8 !important;
        font-weight: 500 !important;
        margin-top: 0.3rem !important;
        text-align: right !important;
    }
    .bot-ts {
        font-size: 0.66rem !important;
        color: #94a3b8 !important;
        font-weight: 500 !important;
        margin: 0.3rem 0 1.2rem 44px !important;
        text-align: right !important;
        max-width: calc(88% - 0px) !important;
    }

    /* Memory card sparkle next to title (matches mockup). */
    .memory-title { display: inline-flex !important; align-items: center !important; gap: 0.35rem !important; }
    .memory-sparkle {
        font-size: 0.9rem !important;
        color: #6366f1 !important;
        -webkit-text-fill-color: #6366f1 !important;
    }
    /* Memory body: bullet lines of comma-joined values. */
    .memory-body {
        margin-top: 0.15rem !important;
        padding-left: 0.4rem !important;
    }
    .memory-line {
        font-size: 0.85rem !important;
        line-height: 1.5 !important;
        color: #475569 !important;
        margin: 0.25rem 0 !important;
    }
    .memory-bullet {
        color: #94a3b8 !important;
        font-weight: 700 !important;
        margin-right: 0.45rem !important;
    }
    .memory-line-label {
        font-weight: 700 !important;
        color: #0f172a !important;
    }
    .memory-line-text {
        color: #475569 !important;
    }

    /* Meta label pills (Sources / Confidence) — highlighted indigo chips
       so the section headers pop visually instead of fading into the row. */
    .meta-label {
        display: inline-flex !important;
        align-items: center !important;
        gap: 0.32rem !important;
        font-size: 0.7rem !important;
        font-weight: 800 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.08em !important;
        color: #4338ca !important;
        -webkit-text-fill-color: #4338ca !important;
        padding: 0.28rem 0.65rem 0.28rem 0.6rem !important;
        background: linear-gradient(135deg, #eef0ff 0%, #f5f3ff 100%) !important;
        border: 1px solid #c7d2fe !important;
        border-radius: 999px !important;
        white-space: nowrap !important;
        height: 24px !important;
        box-shadow: 0 1px 2px rgba(79, 70, 229, 0.08) !important;
    }
    .meta-label-sources .material-symbols-rounded {
        font-size: 0.88rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
        font-variation-settings: 'FILL' 1, 'wght' 600, 'GRAD' 0, 'opsz' 20 !important;
    }
    .meta-label-conf {
        margin-left: 0.6rem !important;
        background: linear-gradient(135deg, #ecfdf5 0%, #eef2ff 100%) !important;
        border-color: #bbf7d0 !important;
        color: #047857 !important;
        -webkit-text-fill-color: #047857 !important;
    }

    /* Unified meta row — source tags + confidence pill + bar + RAG text
       all on ONE line, aligned to the bot bubble's left edge. Scrolls
       horizontally if the screen is too narrow so nothing wraps awkwardly. */
    .meta-row {
        margin-left: 44px !important;
        margin-top: 0.7rem !important;
        margin-bottom: 1.1rem !important;
        display: flex !important;
        align-items: center !important;
        flex-wrap: nowrap !important;
        gap: 0.4rem !important;
        line-height: 1.4 !important;
        overflow-x: auto !important;
        scrollbar-width: none !important;
    }
    .meta-row::-webkit-scrollbar { display: none !important; }
    .meta-row .rag-text {
        font-size: 0.68rem !important;
        color: #64748b !important;
        white-space: nowrap !important;
        margin-left: 0.15rem !important;
    }
    /* Legacy single-row classes still work if anything else uses them. */
    .source-row,
    .confidence-row {
        margin-left: 44px !important;
        display: flex !important;
        align-items: center !important;
        flex-wrap: nowrap !important;
        gap: 0.4rem !important;
        line-height: 1.4 !important;
        overflow-x: auto !important;
        scrollbar-width: none !important;
    }
    .source-row::-webkit-scrollbar,
    .confidence-row::-webkit-scrollbar { display: none !important; }
    .source-row { margin-top: 0.6rem !important; margin-bottom: 0 !important; }
    .confidence-row { margin-top: 0.7rem !important; margin-bottom: 1rem !important; }

    /* Engine + source + confidence pills: consistent height, no overflow. */
    .engine-badge,
    .source-tag,
    .confidence-pill {
        font-size: 0.66rem !important;
        padding: 0.22rem 0.6rem !important;
        border-radius: 999px !important;
        white-space: nowrap !important;
        line-height: 1 !important;
        margin: 0 !important;
        height: 22px !important;
        display: inline-flex !important;
        align-items: center !important;
    }
    .source-tag { background: #eef2ff !important; border-color: #c7d2fe !important; color: #4338ca !important; }
    .engine-claude { background: #fef3e8 !important; border-color: #fed7aa !important; color: #9a3412 !important; }
    .engine-groq { background: #ecfdf5 !important; border-color: #a7f3d0 !important; color: #047857 !important; }
    .engine-vision { background: #f3e8ff !important; border-color: #d8b4fe !important; color: #6b21a8 !important; }
    .conf-high   { background: #ecfdf5 !important; border-color: #a7f3d0 !important; color: #047857 !important; }
    .conf-medium { background: #fef3c7 !important; border-color: #fde68a !important; color: #92400e !important; }
    .conf-low    { background: #fee2e2 !important; border-color: #fecaca !important; color: #991b1b !important; }

    .confidence-bar {
        width: 90px !important;
        height: 5px !important;
        background: #e2e8f0 !important;
        border-radius: 999px !important;
        margin-left: 0.1rem !important;
    }
    .confidence-row > span:last-child {
        font-size: 0.66rem !important;
        color: #64748b !important;
        margin-left: 0.1rem !important;
    }

    /* Feedback row: compact "Was this helpful?" + tiny Yes/No chips. */
    .md-feedback-wrap {
        margin: 1.2rem 0 1rem 0 !important;
    }
    .md-feedback-q {
        text-align: center !important;
        font-size: 0.78rem !important;
        color: #94a3b8 !important;
        margin-bottom: 0.45rem !important;
        font-weight: 600 !important;
        letter-spacing: -0.005em !important;
    }
    .st-key-chat_helpful .stButton > button,
    .st-key-chat_not_helpful .stButton > button {
        min-height: 30px !important;
        height: 30px !important;
        min-width: 0 !important;
        width: 100% !important;
        padding: 0 0.75rem !important;
        border-radius: 999px !important;
        border: 1px solid #e6ecf6 !important;
        background: #ffffff !important;
        font-size: 0.78rem !important;
        font-weight: 600 !important;
        color: #475569 !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
        transition: background 0.15s ease, border-color 0.15s ease, color 0.15s ease, transform 0.15s ease !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 0.3rem !important;
    }
    /* Kill the baseweb 8px stray margin on the icon wrap. */
    .st-key-chat_helpful .stButton > button > span:first-child,
    .st-key-chat_not_helpful .stButton > button > span:first-child {
        margin: 0 !important;
    }
    .st-key-chat_helpful .stButton > button [data-testid="stIconMaterial"],
    .st-key-chat_not_helpful .stButton > button [data-testid="stIconMaterial"] {
        font-size: 0.9rem !important;
    }
    .st-key-chat_helpful .stButton > button:hover {
        background: #ecfdf5 !important;
        border-color: #a7f3d0 !important;
        color: #047857 !important;
        transform: translateY(-1px) !important;
    }
    .st-key-chat_helpful .stButton > button:hover [data-testid="stIconMaterial"] {
        color: #047857 !important;
        -webkit-text-fill-color: #047857 !important;
    }
    .st-key-chat_not_helpful .stButton > button:hover {
        background: #fef2f2 !important;
        border-color: #fecaca !important;
        color: #b91c1c !important;
        transform: translateY(-1px) !important;
    }
    .st-key-chat_not_helpful .stButton > button:hover [data-testid="stIconMaterial"] {
        color: #b91c1c !important;
        -webkit-text-fill-color: #b91c1c !important;
    }

    /* Download Chat (primary indigo gradient) + Doctor Visit (ghost). */
    .st-key-dl_chat_btn .stButton > button {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: none !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        color: #ffffff !important;
        font-weight: 660 !important;
        font-size: 0.9rem !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 8px 18px rgba(79, 70, 229, 0.28) !important;
        transition: transform 0.15s ease, box-shadow 0.2s ease !important;
    }
    .st-key-dl_chat_btn .stButton > button:hover {
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 12px 24px rgba(79, 70, 229, 0.36) !important;
    }
    .st-key-dl_summary_btn .stButton > button {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(99, 102, 241, 0.22) !important;
        background: linear-gradient(180deg, #ffffff 0%, #f7f9ff 100%) !important;
        color: #4338ca !important;
        font-weight: 650 !important;
        font-size: 0.9rem !important;
        box-shadow: none !important;
    }
    .st-key-dl_summary_btn .stButton > button:hover {
        border-color: rgba(99, 102, 241, 0.42) !important;
        background: #eef0ff !important;
    }

    /* Clear chat: subtle ghost with muted danger tint on hover. */
    .st-key-main_clear_btn .stButton > button,
    .st-key-main_clear_btn_voice .stButton > button,
    .st-key-main_clear_btn_plain .stButton > button {
        min-height: 38px !important;
        height: 38px !important;
        border-radius: 10px !important;
        border: 1px solid #e6ecf6 !important;
        background: #ffffff !important;
        color: #64748b !important;
        font-weight: 600 !important;
        font-size: 0.84rem !important;
        box-shadow: none !important;
        transition: background 0.15s ease, border-color 0.15s ease, color 0.15s ease !important;
    }
    .st-key-main_clear_btn .stButton > button:hover,
    .st-key-main_clear_btn_voice .stButton > button:hover,
    .st-key-main_clear_btn_plain .stButton > button:hover {
        background: #fef2f2 !important;
        border-color: #fecaca !important;
        color: #b91c1c !important;
    }
    .st-key-main_clear_btn .stButton > button [data-testid="stIconMaterial"],
    .st-key-main_clear_btn_voice .stButton > button [data-testid="stIconMaterial"],
    .st-key-main_clear_btn_plain .stButton > button [data-testid="stIconMaterial"] {
        font-size: 1rem !important;
    }

    /* "Download your conversation:" header + helper line: cleaner type. */
    [data-testid="stMain"] .stMarkdown p strong:only-child {
        font-size: 0.9rem !important;
    }

    /* ── Chat conversation vertical rhythm ──
       Generous, consistent gaps between every element so the page reads
       like a real conversation, not a wall of stacked text. */

    /* Memory card → space below it before the first user message. */
    .memory-card {
        margin-bottom: 1.6rem !important;
        padding: 1rem 1.2rem !important;
        border-radius: 14px !important;
    }
    /* Each message row (bot or user): clear vertical breathing room. */
    .bot-wrap,
    .user-wrap {
        margin-bottom: 1.4rem !important;
        margin-top: 0.4rem !important;
    }
    /* Confidence row sits between two messages → push the next user row
       down so it never crashes into the bar. */
    .confidence-row + div,
    .confidence-row ~ .stElementContainer .user-wrap {
        margin-top: 1.2rem !important;
    }
    /* Feedback row spacing (Yes/No chips columns container). */
    [data-testid="stMain"] [data-testid="stHorizontalBlock"]:has(.st-key-chat_helpful) {
        gap: 0.5rem !important;
        margin-bottom: 0.6rem !important;
    }
    /* Divider above "Download your conversation:" gets margin top/bottom. */
    [data-testid="stMain"] hr {
        margin-top: 1.4rem !important;
        margin-bottom: 1rem !important;
        border-color: #e2e8f0 !important;
    }
    /* "Download your conversation:" heading → space below before the
       Download Chat / Doctor Visit Summary buttons. */
    [data-testid="stMain"] .stMarkdown:has(p > strong:only-child) {
        margin-bottom: 0.6rem !important;
    }
    /* Download / Doctor Visit columns → clear gap from the helper line. */
    [data-testid="stMain"] [data-testid="stHorizontalBlock"]:has(.st-key-dl_chat_btn) {
        margin-bottom: 0.4rem !important;
    }
    /* Composer form (chat) sits below the download section → clear top
       gap so it reads as its own zone. */
    [data-testid="stForm"]:has(.st-key-chat_upload_btn) {
        margin-top: 1.2rem !important;
        margin-bottom: 0.9rem !important;
    }
    /* Clear chat button: space above so it doesn't hug the composer. */
    div[class*="st-key-main_clear_btn"] {
        margin-top: 0.8rem !important;
    }
    /* Final disclaimer text: space above. */
    .md-home-composer-note {
        margin-top: 1rem !important;
    }

    /* ── Smart Actions header ── */
    .md-smart-head {
        margin-top: 5rem !important;
        margin-bottom: 0.95rem !important;
    }
    .md-smart-title {
        font-size: 0.86rem !important;
        font-weight: 700 !important;
        color: #1f2a3d !important;
        letter-spacing: -0.005em !important;
    }

    /* ── Smart Actions cards (Apple-style premium polish) ──
       Crisp white card with a subtle vertical gradient and a glossy top-
       edge inset highlight. Each card is named with its own --accent CSS
       variable so the icon container, hover border, hover shadow, and
       arrow all pick up the matching color in one place. */
    .st-key-sa_sym  .stButton > button { --accent: 124, 58, 237; --accent-soft: 237, 233, 254; --accent-soft2: 221, 214, 254; }
    .st-key-sa_rec  .stButton > button { --accent:  16, 185, 129; --accent-soft: 209, 250, 229; --accent-soft2: 167, 243, 208; }
    .st-key-sa_ins  .stButton > button { --accent: 219,  39, 119; --accent-soft: 252, 231, 243; --accent-soft2: 251, 207, 232; }
    .st-key-sa_appt .stButton > button { --accent:  37,  99, 235; --accent-soft: 219, 234, 254; --accent-soft2: 191, 219, 254; }
    .st-key-sa_sym .stButton > button,
    .st-key-sa_rec .stButton > button,
    .st-key-sa_ins .stButton > button,
    .st-key-sa_appt .stButton > button {
        position: relative !important;
        min-height: 152px !important;
        height: 152px !important;
        padding: 1.05rem 0.7rem 1.55rem 0.7rem !important;
        border-radius: 18px !important;
        background: linear-gradient(180deg, #ffffff 0%, #f8faff 100%) !important;
        border: 1px solid rgba(15, 23, 42, 0.06) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.95) inset,
            0 1px 2px rgba(15, 23, 42, 0.04),
            0 8px 20px rgba(15, 23, 42, 0.04) !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: flex-start !important;
        text-align: center !important;
        gap: 0.6rem !important;
        transition:
            transform 0.22s cubic-bezier(0.16, 1, 0.3, 1),
            box-shadow 0.22s ease,
            border-color 0.22s ease !important;
    }
    /* Hover: lift, deepen shadow with the card's accent tint, border picks
       up the accent. Smooth easing for the iOS spring feel. */
    .st-key-sa_sym .stButton > button:hover,
    .st-key-sa_rec .stButton > button:hover,
    .st-key-sa_ins .stButton > button:hover,
    .st-key-sa_appt .stButton > button:hover {
        transform: translateY(-2px) !important;
        border-color: rgba(var(--accent), 0.35) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.95) inset,
            0 2px 4px rgba(15, 23, 42, 0.04),
            0 14px 32px rgba(var(--accent), 0.18) !important;
    }
    .st-key-sa_sym .stButton > button:active,
    .st-key-sa_rec .stButton > button:active,
    .st-key-sa_ins .stButton > button:active,
    .st-key-sa_appt .stButton > button:active {
        transform: translateY(0) !important;
    }
    /* Markdown container — centered, full width so wrap is symmetric. */
    .st-key-sa_sym .stButton > button > div[data-testid="stMarkdownContainer"],
    .st-key-sa_rec .stButton > button > div[data-testid="stMarkdownContainer"],
    .st-key-sa_ins .stButton > button > div[data-testid="stMarkdownContainer"],
    .st-key-sa_appt .stButton > button > div[data-testid="stMarkdownContainer"] {
        flex: 0 0 auto !important;
        width: 100% !important;
        max-width: 100% !important;
        text-align: center !important;
    }
    /* Icon span — centered at top. */
    .st-key-sa_sym .stButton > button > span:first-child,
    .st-key-sa_rec .stButton > button > span:first-child,
    .st-key-sa_ins .stButton > button > span:first-child,
    .st-key-sa_appt .stButton > button > span:first-child {
        flex: 0 0 auto !important;
        margin: 0 0 0.15rem 0 !important;
        display: inline-flex !important;
        justify-content: center !important;
    }
    .st-key-sa_sym .stButton > button:hover,
    .st-key-sa_rec .stButton > button:hover,
    .st-key-sa_ins .stButton > button:hover,
    .st-key-sa_appt .stButton > button:hover {
        transform: translateY(-1px) !important;
        border-color: #d6e2f6 !important;
        box-shadow: 0 8px 18px rgba(15, 23, 42, 0.06) !important;
    }
    /* Icon tile — premium gradient background derived from the card's
       --accent. Inset highlight on top + soft drop shadow tinted with the
       accent give the iOS app-icon depth. */
    .st-key-sa_sym .stButton > button [data-testid="stIconMaterial"],
    .st-key-sa_rec .stButton > button [data-testid="stIconMaterial"],
    .st-key-sa_ins .stButton > button [data-testid="stIconMaterial"],
    .st-key-sa_appt .stButton > button [data-testid="stIconMaterial"] {
        width: 34px !important;
        height: 34px !important;
        min-width: 34px !important;
        border-radius: 10px !important;
        font-size: 1.12rem !important;
        font-weight: 500 !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        margin: 0 !important;
        color: rgb(var(--accent)) !important;
        -webkit-text-fill-color: rgb(var(--accent)) !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.55) 0%, rgba(255,255,255,0) 60%),
            linear-gradient(135deg, rgb(var(--accent-soft)) 0%, rgb(var(--accent-soft2)) 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.6) inset,
            0 0 0 1px rgba(var(--accent), 0.08),
            0 4px 10px rgba(var(--accent), 0.16) !important;
        transition: transform 0.22s cubic-bezier(0.16, 1, 0.3, 1), box-shadow 0.22s ease !important;
    }
    /* Hover: icon lifts subtly + shadow deepens. */
    .st-key-sa_sym .stButton > button:hover [data-testid="stIconMaterial"],
    .st-key-sa_rec .stButton > button:hover [data-testid="stIconMaterial"],
    .st-key-sa_ins .stButton > button:hover [data-testid="stIconMaterial"],
    .st-key-sa_appt .stButton > button:hover [data-testid="stIconMaterial"] {
        transform: translateY(-1px) scale(1.04) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.7) inset,
            0 0 0 1px rgba(var(--accent), 0.14),
            0 8px 18px rgba(var(--accent), 0.28) !important;
    }
    .st-key-sa_sym .stButton > button p,
    .st-key-sa_rec .stButton > button p,
    .st-key-sa_ins .stButton > button p,
    .st-key-sa_appt .stButton > button p {
        font-size: 0.64rem !important;
        line-height: 1.3 !important;
        color: #6b7280 !important;
        font-weight: 530 !important;
        margin: 0 !important;
        /* Break only between words, not characters. Override the inherited
           `overflow-wrap: anywhere` that was breaking text per-letter. */
        overflow-wrap: break-word !important;
        word-break: normal !important;
        white-space: normal !important;
    }
    /* Inner markdown container — full width minus icon. */
    .st-key-sa_sym .stButton > button > div[data-testid="stMarkdownContainer"],
    .st-key-sa_rec .stButton > button > div[data-testid="stMarkdownContainer"],
    .st-key-sa_ins .stButton > button > div[data-testid="stMarkdownContainer"],
    .st-key-sa_appt .stButton > button > div[data-testid="stMarkdownContainer"] {
        flex: 1 1 auto !important;
        min-width: 0 !important;
        width: 100% !important;
        max-width: 100% !important;
        overflow: visible !important;
    }
    /* Tighten the outer icon span — kill any baseweb margin/padding so it
       sits flush at the icon's content size (~32px) instead of inflating to
       ~48px and starving the text column. */
    .st-key-sa_sym .stButton > button > span:first-child,
    .st-key-sa_rec .stButton > button > span:first-child,
    .st-key-sa_ins .stButton > button > span:first-child,
    .st-key-sa_appt .stButton > button > span:first-child {
        margin: 0 !important;
        padding: 0 !important;
        width: auto !important;
        min-width: 0 !important;
    }
    /* Bold title — refined SF Pro-style weight + tight tracking. */
    .st-key-sa_sym .stButton > button p strong,
    .st-key-sa_rec .stButton > button p strong,
    .st-key-sa_ins .stButton > button p strong,
    .st-key-sa_appt .stButton > button p strong {
        font-size: 0.82rem !important;
        font-weight: 700 !important;
        color: #0f172a !important;
        line-height: 1.2 !important;
        letter-spacing: -0.015em !important;
        display: inline !important;
    }
    /* Arrow chip — ghost circle bottom-right, picks up the card's accent
       on hover and slides forward like an iOS disclosure indicator. */
    .st-key-sa_sym .stButton > button::after,
    .st-key-sa_rec .stButton > button::after,
    .st-key-sa_ins .stButton > button::after,
    .st-key-sa_appt .stButton > button::after {
        content: "arrow_forward" !important;
        font-family: "Material Symbols Rounded", "Material Symbols Outlined" !important;
        position: absolute !important;
        bottom: 0.5rem !important;
        right: 0.5rem !important;
        left: auto !important;
        transform: none !important;
        width: 20px !important;
        height: 20px !important;
        font-size: 0.78rem !important;
        font-weight: 500 !important;
        line-height: 1 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        border-radius: 999px !important;
        background: rgba(var(--accent), 0.1) !important;
        border: none !important;
        color: rgb(var(--accent)) !important;
        -webkit-text-fill-color: rgb(var(--accent)) !important;
        transition:
            background 0.22s ease,
            color 0.22s ease,
            transform 0.22s cubic-bezier(0.16, 1, 0.3, 1),
            box-shadow 0.22s ease !important;
    }
    .st-key-sa_sym .stButton > button:hover::after,
    .st-key-sa_rec .stButton > button:hover::after,
    .st-key-sa_ins .stButton > button:hover::after,
    .st-key-sa_appt .stButton > button:hover::after {
        background: rgb(var(--accent)) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        transform: translateX(3px) !important;
        box-shadow: 0 4px 10px rgba(var(--accent), 0.35) !important;
    }

    /* ────────────────────────────────────────────────────────────────────
       Apple-style polish — page hero (Prescription Reader, etc.)
       Replaces the heavy bold dark-navy heading with a refined two-column
       layout: tinted icon square + title + muted subtitle. Sits as a soft
       card at the top of the page so the form below feels intentional. */
    .md-page-hero {
        display: grid;
        grid-template-columns: 56px minmax(0, 1fr);
        gap: 1rem;
        align-items: center;
        margin: 0.4rem 0 1.3rem 0;
        padding: 1.1rem 1.25rem;
        border-radius: 18px;
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%);
        border: 1px solid #e6ecf6;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 6px 18px rgba(15, 23, 42, 0.04);
    }
    .md-page-hero-ic {
        width: 56px;
        height: 56px;
        border-radius: 16px;
        display: flex;
        align-items: center;
        justify-content: center;
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.12) 0%, rgba(139, 92, 246, 0.12) 100%);
        color: #4f46e5;
    }
    .md-page-hero-ic .material-symbols-rounded {
        font-size: 1.55rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
    }
    .md-page-hero-text { min-width: 0; }
    .md-page-hero-title {
        font-size: 1.45rem;
        font-weight: 720;
        color: #0f172a;
        line-height: 1.15;
        letter-spacing: -0.018em;
        margin-bottom: 0.2rem;
    }
    .md-page-hero-sub {
        font-size: 0.86rem;
        line-height: 1.45;
        color: #64748b;
    }

    /* ── Prescription Reader form polish ──
       Style the form shell as a single clean card. The file uploader
       dropzone becomes a soft tinted area with a refined dashed border;
       the "Read prescription" submit picks up our brand indigo gradient. */
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 18px !important;
        padding: 1.1rem !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 8px 22px rgba(15, 23, 42, 0.04) !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFileUploaderDropzone"],
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFileUploader"] section {
        background: linear-gradient(180deg, #f7f9ff 0%, #eef1ff 100%) !important;
        border: 1.5px dashed rgba(99, 102, 241, 0.32) !important;
        border-radius: 14px !important;
        padding: 1.4rem 1.1rem !important;
        transition: background 0.18s ease, border-color 0.18s ease !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFileUploaderDropzone"]:hover,
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFileUploader"] section:hover {
        background: linear-gradient(180deg, #eef1ff 0%, #e7eaff 100%) !important;
        border-color: rgba(99, 102, 241, 0.55) !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFileUploaderDropzoneInstructions"] {
        color: #475569 !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFileUploaderDropzone"] button {
        background: #ffffff !important;
        border: 1px solid #d8e4fa !important;
        color: #4f46e5 !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
    }
    /* Text input ("Optional context") inside the rx form. */
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stTextInput"] label {
        font-size: 0.82rem !important;
        color: #475569 !important;
        font-weight: 600 !important;
        margin-bottom: 0.3rem !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stTextInputRootElement"] {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 12px !important;
        box-shadow: none !important;
        min-height: 42px !important;
        height: 42px !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stTextInputRootElement"]:focus-within {
        border-color: rgba(99, 102, 241, 0.45) !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12) !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stTextInput"] input {
        font-size: 0.9rem !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }
    /* Read prescription / Clear buttons. */
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"] {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: none !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 8px 18px rgba(79, 70, 229, 0.28) !important;
        color: #ffffff !important;
        font-weight: 660 !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"]:hover {
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 12px 24px rgba(79, 70, 229, 0.36) !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"] {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(99, 102, 241, 0.22) !important;
        background: linear-gradient(180deg, #ffffff 0%, #f7f9ff 100%) !important;
        color: #4338ca !important;
        font-weight: 650 !important;
        box-shadow: none !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"]:hover {
        border-color: rgba(99, 102, 241, 0.42) !important;
        background: #eef0ff !important;
    }
    /* Checkbox inside the records/rx form — refined indigo accent. */
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stCheckbox"] {
        margin: 0.2rem 0 0.6rem 0 !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stCheckbox"] label {
        font-size: 0.85rem !important;
        color: #475569 !important;
        gap: 0.55rem !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stCheckbox"] label > div:first-child {
        width: 18px !important;
        height: 18px !important;
        border-radius: 5px !important;
        border-color: rgba(99, 102, 241, 0.4) !important;
    }
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stCheckbox"] label > div:first-child[data-checked="true"],
    [data-testid="stForm"]:has([data-testid="stFileUploader"]):not(:has(.st-key-home_send_btn)) [data-testid="stCheckbox"] input:checked ~ div {
        background: linear-gradient(135deg, #4f46e5 0%, #6366f1 100%) !important;
        border-color: #4f46e5 !important;
    }
    /* "No records uploaded yet." empty state — softer, less italic-heavy. */
    [data-testid="stMain"] .md-rcard[style*="No records"],
    [data-testid="stMain"] .md-rcard[style*="text-align:center"]:has-text("No records") {
        background: linear-gradient(180deg, #f7f9ff 0%, #eef1ff 100%) !important;
        border: 1px dashed rgba(99, 102, 241, 0.25) !important;
        color: #64748b !important;
        font-style: normal !important;
        font-weight: 500 !important;
    }

    /* ────────────────────────────────────────────────────────────────────
       Apple-style polish — Symptom Assessment page.
       Refines the bespoke .assessment-card / .question-bubble / progress
       bar to use the brand indigo palette and Inter (not DM Serif). Quick
       select buttons become clean indigo-tinted chips, and the Next/Cancel
       row gets a primary-vs-ghost hierarchy that aligns cleanly.
       ──────────────────────────────────────────────────────────────────── */
    .assessment-card {
        border-radius: 18px !important;
        padding: 1.2rem 1.4rem 1.3rem 1.4rem !important;
        margin-bottom: 1.1rem !important;
        border: 1px solid #e6ecf6 !important;
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 6px 18px rgba(15, 23, 42, 0.04) !important;
    }
    /* Assessment card head — icon + title/subtitle column, matching the
       page-hero pattern used on Health Overview, Medications, etc. */
    .assessment-card-head {
        display: grid !important;
        grid-template-columns: 56px minmax(0, 1fr) !important;
        gap: 1rem !important;
        align-items: center !important;
        margin-bottom: 1.05rem !important;
    }
    .assessment-head-text { min-width: 0 !important; }
    .assessment-title {
        font-family: 'Inter', sans-serif !important;
        font-size: 1.35rem !important;
        font-weight: 720 !important;
        color: #0f172a !important;
        letter-spacing: -0.018em !important;
        line-height: 1.15 !important;
        margin-bottom: 0.2rem !important;
    }
    .assessment-subtitle {
        color: #64748b !important;
        font-size: 0.86rem !important;
        line-height: 1.45 !important;
        margin-bottom: 0 !important;
    }
    .progress-label {
        font-size: 0.74rem !important;
        font-weight: 650 !important;
        margin-bottom: 0.4rem !important;
    }
    .progress-label span:first-child {
        color: #1f2a3d !important;
        letter-spacing: 0.03em !important;
    }
    .progress-label span:last-child {
        color: #4f46e5 !important;
        font-weight: 700 !important;
    }
    .progress-bar-wrap {
        height: 6px !important;
        background: #eef1ff !important;
        border: none !important;
        border-radius: 999px !important;
    }
    .progress-bar-fill {
        background: linear-gradient(90deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        border-radius: 999px !important;
    }

    /* Question bubble — indigo accent border on the left, soft tinted bg. */
    .question-bubble {
        background: linear-gradient(180deg, #f7f9ff 0%, #eef1ff 100%) !important;
        border: 1px solid #e3e8fb !important;
        border-left: 3px solid #4f46e5 !important;
        border-radius: 14px !important;
        padding: 0.95rem 1.15rem !important;
        margin: 0.85rem 0 0.9rem 0 !important;
        font-size: 0.96rem !important;
        font-weight: 600 !important;
        color: #0f172a !important;
        line-height: 1.45 !important;
    }

    /* Quick-select option buttons — clean white chips with indigo on hover/active. */
    [data-testid="stMain"] .stButton > button[data-testid="stBaseButton-secondary"][class*="opt_"],
    [data-testid="stMain"] div[class*="st-key-opt_"] .stButton > button {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: 1px solid #e6ecf6 !important;
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%) !important;
        color: #1f2a3d !important;
        font-size: 0.88rem !important;
        font-weight: 600 !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.85) inset,
            0 1px 2px rgba(15, 23, 42, 0.02) !important;
        transition:
            background 0.18s ease,
            border-color 0.18s ease,
            box-shadow 0.2s ease,
            transform 0.15s ease !important;
    }
    [data-testid="stMain"] div[class*="st-key-opt_"] .stButton > button:hover {
        background: linear-gradient(180deg, #f7f9ff 0%, #eef1ff 100%) !important;
        border-color: rgba(99, 102, 241, 0.4) !important;
        color: #4338ca !important;
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 6px 14px rgba(79, 70, 229, 0.1) !important;
    }

    /* Assessment form (text input + Next / Cancel row). */
    [data-testid="stForm"][class*="assessment_form_"] {
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
        box-shadow: none !important;
        margin-top: 0.5rem !important;
    }
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stTextInputRootElement"] {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 12px !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
        min-height: 44px !important;
        height: 44px !important;
        margin-bottom: 0.7rem !important;
    }
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stTextInputRootElement"]:focus-within {
        border-color: rgba(99, 102, 241, 0.45) !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12) !important;
    }
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stTextInput"] input {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        font-size: 0.9rem !important;
    }
    /* Next button — primary indigo gradient. */
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"] {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: none !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 8px 18px rgba(79, 70, 229, 0.28) !important;
        color: #ffffff !important;
        font-weight: 660 !important;
    }
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"]:hover {
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 12px 24px rgba(79, 70, 229, 0.36) !important;
    }
    /* Cancel button — ghost. */
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"] {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(99, 102, 241, 0.22) !important;
        background: linear-gradient(180deg, #ffffff 0%, #f7f9ff 100%) !important;
        color: #4338ca !important;
        font-weight: 650 !important;
        box-shadow: none !important;
    }
    [data-testid="stForm"][class*="assessment_form_"] [data-testid="stFormSubmitButton"] > button[kind="secondaryFormSubmit"]:hover {
        border-color: rgba(99, 102, 241, 0.42) !important;
        background: #eef0ff !important;
    }

    /* ────────────────────────────────────────────────────────────────────
       Apple-style polish — Health Overview page.
       Wearable card uses our soft white shell; metric tiles get cleaner
       padding; Save buttons (sleep/steps/heart rate) and the +1 glass /
       Reset buttons all pick up the indigo primary / ghost hierarchy.
       ──────────────────────────────────────────────────────────────────── */
    .md-wearable-card {
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%) !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 18px !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 6px 18px rgba(15, 23, 42, 0.04) !important;
        padding: 1.05rem 1.15rem !important;
    }
    .md-wearable-icon {
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.14) 0%, rgba(139, 92, 246, 0.12) 100%) !important;
        color: #4f46e5 !important;
        border-radius: 14px !important;
    }
    .md-wearable-title {
        font-size: 0.98rem !important;
        font-weight: 700 !important;
        color: #0f172a !important;
        letter-spacing: -0.005em !important;
    }
    .md-wearable-desc {
        color: #64748b !important;
        font-size: 0.84rem !important;
        line-height: 1.45 !important;
    }
    .md-wearable-pill {
        background: linear-gradient(180deg, #ffffff 0%, #f7f9ff 100%) !important;
        border: 1px solid #e6ecf6 !important;
        color: #4f46e5 !important;
        font-weight: 600 !important;
        font-size: 0.76rem !important;
        padding: 0.32rem 0.7rem !important;
        border-radius: 999px !important;
    }
    .md-wearable-pill.md-wearable-soon {
        background: #f1f5f9 !important;
        color: #94a3b8 !important;
        border-color: #e2e8f0 !important;
    }

    /* Metric tile cards — softer surface, refined typography. */
    [data-testid="stMain"] .md-rcard:has(.md-metric-row) {
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%) !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 16px !important;
        box-shadow: 0 2px 6px rgba(15, 23, 42, 0.03) !important;
        padding: 0.8rem 1rem !important;
    }
    [data-testid="stMain"] .md-metric-label {
        font-size: 0.78rem !important;
        color: #64748b !important;
        font-weight: 600 !important;
        letter-spacing: 0.005em !important;
    }
    [data-testid="stMain"] .md-metric-value {
        font-size: 1.05rem !important;
        font-weight: 720 !important;
        color: #0f172a !important;
    }

    /* Save sleep / Save steps / Save heart rate buttons — indigo primary. */
    [data-testid="stForm"][class*="sleep_form_today"] [data-testid="stFormSubmitButton"] > button,
    [data-testid="stForm"][class*="steps_form_today"] [data-testid="stFormSubmitButton"] > button,
    [data-testid="stForm"][class*="hr_form_today"] [data-testid="stFormSubmitButton"] > button {
        min-height: 42px !important;
        height: 42px !important;
        border-radius: 12px !important;
        border: none !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 6px 14px rgba(79, 70, 229, 0.24) !important;
        color: #ffffff !important;
        font-weight: 660 !important;
        font-size: 0.88rem !important;
    }
    [data-testid="stForm"][class*="sleep_form_today"] [data-testid="stFormSubmitButton"] > button:hover,
    [data-testid="stForm"][class*="steps_form_today"] [data-testid="stFormSubmitButton"] > button:hover,
    [data-testid="stForm"][class*="hr_form_today"] [data-testid="stFormSubmitButton"] > button:hover {
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 10px 20px rgba(79, 70, 229, 0.32) !important;
    }
    /* Number-input arrows inside the save forms — keep them tidy. */
    [data-testid="stForm"][class*="sleep_form_today"] [data-testid="stNumberInput"],
    [data-testid="stForm"][class*="steps_form_today"] [data-testid="stNumberInput"],
    [data-testid="stForm"][class*="hr_form_today"] [data-testid="stNumberInput"] {
        margin-bottom: 0.55rem !important;
    }
    [data-testid="stForm"][class*="sleep_form_today"] [data-testid="stNumberInput"] label,
    [data-testid="stForm"][class*="steps_form_today"] [data-testid="stNumberInput"] label,
    [data-testid="stForm"][class*="hr_form_today"] [data-testid="stNumberInput"] label {
        font-size: 0.78rem !important;
        color: #64748b !important;
        font-weight: 600 !important;
        margin-bottom: 0.28rem !important;
    }
    [data-testid="stForm"][class*="sleep_form_today"] [data-testid="stNumberInput"] > div > div,
    [data-testid="stForm"][class*="steps_form_today"] [data-testid="stNumberInput"] > div > div,
    [data-testid="stForm"][class*="hr_form_today"] [data-testid="stNumberInput"] > div > div {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 12px !important;
        box-shadow: none !important;
        min-height: 42px !important;
        height: 42px !important;
    }
    [data-testid="stForm"][class*="sleep_form_today"] [data-testid="stNumberInput"] input,
    [data-testid="stForm"][class*="steps_form_today"] [data-testid="stNumberInput"] input,
    [data-testid="stForm"][class*="hr_form_today"] [data-testid="stNumberInput"] input {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        font-size: 0.92rem !important;
        font-weight: 600 !important;
        color: #1f2a3d !important;
    }

    /* "+1 glass" / "Reset" buttons — ghost pills, paired symmetrically. */
    div[class*="st-key-water_inc"] .stButton > button,
    div[class*="st-key-water_reset"] .stButton > button {
        min-height: 42px !important;
        height: 42px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(99, 102, 241, 0.22) !important;
        background: linear-gradient(180deg, #ffffff 0%, #f7f9ff 100%) !important;
        color: #4338ca !important;
        font-weight: 650 !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
        font-size: 0.88rem !important;
    }
    div[class*="st-key-water_inc"] .stButton > button:hover,
    div[class*="st-key-water_reset"] .stButton > button:hover {
        border-color: rgba(99, 102, 241, 0.42) !important;
        background: #eef0ff !important;
        transform: translateY(-1px) !important;
    }
    /* "+1 glass" gets a hint of primary tint to signal the more common action. */
    div[class*="st-key-water_inc"] .stButton > button {
        background: linear-gradient(180deg, #eef0ff 0%, #e3e7ff 100%) !important;
        border-color: rgba(99, 102, 241, 0.32) !important;
    }

    /* ────────────────────────────────────────────────────────────────────
       Apple-style polish — Medications + Appointments forms.
       Both forms share the same structure: intro header + 2-col text inputs
       + select/textarea + full-width primary save button. The intro is now
       moved inside the form (in Python) so the whole thing reads as one
       cohesive card. Same hierarchy as other redesigned forms.
       ──────────────────────────────────────────────────────────────────── */
    [data-testid="stForm"]:has(.md-form-intro) {
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%) !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 18px !important;
        padding: 1.2rem 1.25rem 1.1rem 1.25rem !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 8px 22px rgba(15, 23, 42, 0.04) !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) .md-form-intro {
        font-size: 1.05rem !important;
        font-weight: 720 !important;
        color: #0f172a !important;
        letter-spacing: -0.005em !important;
        margin-bottom: 0.25rem !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) .md-form-sub {
        font-size: 0.84rem !important;
        color: #64748b !important;
        line-height: 1.45 !important;
        margin-bottom: 0.95rem !important;
    }
    /* Text inputs + select + textarea labels — small, muted tracker style. */
    [data-testid="stForm"]:has(.md-form-intro) label[data-testid="stWidgetLabel"] {
        font-size: 0.78rem !important;
        color: #64748b !important;
        font-weight: 600 !important;
        margin-bottom: 0.3rem !important;
    }
    /* Text input fields — clean white, indigo focus. */
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextInputRootElement"] {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 12px !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
        min-height: 42px !important;
        height: 42px !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextInputRootElement"]:focus-within {
        border-color: rgba(99, 102, 241, 0.45) !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12) !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextInputRootElement"] [data-baseweb="base-input"] {
        background: transparent !important;
        border: none !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextInput"] input {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        font-size: 0.9rem !important;
    }
    /* Selectbox (Frequency) — match the text-input shell. */
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stSelectbox"] div[data-baseweb="select"] {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 12px !important;
        min-height: 42px !important;
        height: 42px !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stSelectbox"] div[data-baseweb="select"]:focus-within,
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stSelectbox"] div[data-baseweb="select"]:hover {
        border-color: rgba(99, 102, 241, 0.35) !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 40px !important;
        height: 40px !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stSelectbox"] [role="combobox"] {
        color: #1f2a3d !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
    }
    /* Textarea (Notes) — same shell as text inputs but taller. */
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextAreaRootElement"] {
        background: #ffffff !important;
        border: 1px solid #e6ecf6 !important;
        border-radius: 12px !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextAreaRootElement"]:focus-within {
        border-color: rgba(99, 102, 241, 0.45) !important;
        box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12) !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextAreaRootElement"] [data-baseweb="base-input"] {
        background: transparent !important;
        border: none !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stTextArea"] textarea {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        font-size: 0.9rem !important;
        padding: 0.55rem 0.7rem !important;
    }
    /* Save medication / Save appointment — primary indigo gradient. */
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"] {
        min-height: 44px !important;
        height: 44px !important;
        border-radius: 12px !important;
        border: none !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 55%, #8b5cf6 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 8px 18px rgba(79, 70, 229, 0.28) !important;
        color: #ffffff !important;
        font-weight: 660 !important;
        margin-top: 0.5rem !important;
    }
    [data-testid="stForm"]:has(.md-form-intro) [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"]:hover {
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 12px 24px rgba(79, 70, 229, 0.36) !important;
    }

    /* ────────────────────────────────────────────────────────────────────
       Reusable Apple-style empty-state card (Ai Insights, Recent
       Conversations, no-records, no-meds, etc.). Tinted indigo bulb icon
       on a soft white card with friendly copy explaining the next step. */
    .md-empty-card {
        text-align: center;
        padding: 2.4rem 1.6rem;
        background: linear-gradient(180deg, #ffffff 0%, #fafbff 100%);
        border: 1px solid #e6ecf6;
        border-radius: 18px;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 6px 18px rgba(15, 23, 42, 0.04);
        margin-top: 0.6rem;
    }
    .md-empty-icon {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 56px;
        height: 56px;
        border-radius: 16px;
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.14) 0%, rgba(139, 92, 246, 0.12) 100%);
        margin: 0 auto 0.8rem auto;
    }
    .md-empty-icon .material-symbols-rounded {
        font-size: 1.55rem !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
    }
    .md-empty-title {
        font-size: 1.05rem;
        font-weight: 720;
        color: #0f172a;
        letter-spacing: -0.005em;
        margin-bottom: 0.4rem;
    }
    .md-empty-copy {
        color: #64748b;
        font-size: 0.88rem;
        line-height: 1.5;
        max-width: 420px;
        margin: 0 auto;
    }

    /* ────────────────────────────────────────────────────────────────────
       Classic ℞ (Rx) prescription symbol — used on the Prescription Reader
       page hero, the sidebar nav, and the Smart Action card. Material
       Symbols has no Rx glyph, so we render the Unicode U+211E
       "PRESCRIPTION TAKE" character in a serif face for the authentic
       pharmacy-pad look. For the nav + smart card we override Streamlit's
       rendered material icon via a ::before swap. */
    .md-rx-glyph {
        font-family: 'DM Serif Display', 'Times New Roman', 'Cambria', serif !important;
        font-weight: 700 !important;
        font-size: 1.95rem !important;
        line-height: 1 !important;
        color: #4f46e5 !important;
        -webkit-text-fill-color: #4f46e5 !important;
        letter-spacing: -0.02em !important;
        display: inline-block !important;
        transform: translateY(1px) !important;
    }
    /* Nav menu + Smart Action card Prescription Reader use Material
       Symbols `prescriptions` (paper with Rx). Streamlit's icon param
       rejects non-emoji Unicode (so we can't pass ℞ directly there). The
       page hero — which we render via raw HTML — still uses the real ℞
       serif glyph for the authentic pharmacy-pad look. */
/* (End of canonical-styling block — was previously the @media closing brace.) */
</style>
""", unsafe_allow_html=True)

# ── Cross-Profile Main UI Lock (guest + signed-in, not page-specific) ─────────
st.markdown("""
<style>
/* Keep composer/card styling consistent for every profile state. */
[data-testid="stForm"]:has(.st-key-home_send_btn),
[data-testid="stForm"]:has(.st-key-chat_upload_btn) {
    background: #ffffff !important;
    border: 1px solid #e4ecf8 !important;
    border-radius: 24px !important;
    box-shadow: 0 14px 32px rgba(15, 23, 42, 0.07) !important;
    padding: 0.9rem 0.9rem 0.82rem 0.9rem !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"],
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] {
    background: #ffffff !important;
    border: none !important;
    box-shadow: none !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] > div,
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] > div,
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextAreaRootElement"],
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextAreaRootElement"],
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] [data-baseweb="base-input"],
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] [data-baseweb="base-input"] {
    background: #ffffff !important;
    border: none !important;
    box-shadow: none !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea,
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] textarea {
    border: none !important;
    box-shadow: none !important;
    background: #ffffff !important;
    resize: none !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea::-webkit-resizer,
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stTextArea"] textarea::-webkit-resizer {
    display: none !important;
}

/* Force one-line action row alignment (upload + voice + send) on both forms. */
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn),
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) {
    display: flex !important;
    flex-wrap: nowrap !important;
    align-items: center !important;
    min-height: 56px !important;
    gap: 0.72rem !important;
    column-gap: 0.72rem !important;
}

/* Home: upload/voice pills. */
.st-key-home_upload_btn [data-testid="stFormSubmitButton"] > button,
.st-key-home_voice_btn [data-testid="stFormSubmitButton"] > button {
    min-height: 56px !important;
    height: 56px !important;
    min-width: 112px !important;
    width: 112px !important;
    border-radius: 999px !important;
}

/* Active-chat mode: icon pills for upload/voice stay aligned with send. */
.st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button,
.st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button {
    width: 44px !important;
    min-width: 44px !important;
    height: 44px !important;
    min-height: 44px !important;
    border-radius: 999px !important;
    border: 1px solid #dce6f7 !important;
    background: #ffffff !important;
    box-shadow: 0 4px 12px rgba(15, 23, 42, 0.06) !important;
    padding: 0 !important;
}
.st-key-chat_upload_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"],
.st-key-chat_voice_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
    color: #475569 !important;
    font-size: 1.08rem !important;
}

/* Active-chat send button matches home send button shape/gradient. */
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button,
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"],
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button[kind="primary"] {
    width: 56px !important;
    min-width: 56px !important;
    height: 56px !important;
    min-height: 56px !important;
    border-radius: 50% !important;
    border: none !important;
    background: linear-gradient(135deg, #3b82f6 0%, #6366f1 50%, #8b5cf6 100%) !important;
    box-shadow: 0 8px 22px rgba(99, 102, 241, 0.34) !important;
    color: #ffffff !important;
    padding: 0 !important;
}
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button:hover {
    transform: scale(1.05) !important;
    box-shadow: 0 12px 28px rgba(99, 102, 241, 0.45) !important;
}
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button p {
    display: none !important;
}
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
    color: #ffffff !important;
    font-size: 1.35rem !important;
}

/* Right rail: live Health Passport card to fill empty area with useful data. */
.md-passport-card {
    background: linear-gradient(180deg, rgba(255,255,255,0.99), rgba(246,250,255,0.98));
    border: 1px solid #d9e6fb !important;
}
.md-passport-head {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 0.55rem;
    margin-bottom: 0.42rem;
}
.md-passport-title-wrap {
    display: flex;
    align-items: center;
    gap: 0.45rem;
}
.md-passport-title-wrap .material-symbols-rounded {
    color: #2563eb;
    font-size: 1.15rem !important;
}
.md-passport-title {
    font-size: 0.96rem;
    font-weight: 740;
    color: #0f172a;
}
.md-passport-pct {
    font-size: 0.8rem;
    font-weight: 760;
    color: #2563eb;
}
.md-passport-sub {
    color: #5b6b84;
    font-size: 0.76rem;
    margin-bottom: 0.6rem;
}
.md-passport-progress {
    height: 8px;
    border-radius: 999px;
    background: #eaf0fb;
    overflow: hidden;
    margin-bottom: 0.68rem;
}
.md-passport-fill {
    height: 100%;
    border-radius: 999px;
    background: linear-gradient(90deg, #3b82f6 0%, #6366f1 56%, #8b5cf6 100%);
}
.md-passport-check {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 0.6rem;
    padding: 0.3rem 0;
    border-top: 1px solid #e7edf8;
}
.md-passport-check:first-of-type {
    border-top: none;
}
.md-passport-check-left {
    display: inline-flex;
    align-items: center;
    gap: 0.45rem;
    min-width: 0;
}
.md-passport-check-left .material-symbols-rounded {
    font-size: 1.02rem !important;
}
.md-passport-check.ok .md-passport-check-left .material-symbols-rounded {
    color: #10b981;
}
.md-passport-check.todo .md-passport-check-left .material-symbols-rounded {
    color: #94a3b8;
}
.md-passport-check-label {
    color: #334155;
    font-size: 0.8rem;
    font-weight: 560;
}
.md-passport-status {
    font-size: 0.68rem;
    font-weight: 700;
    border-radius: 999px;
    padding: 0.16rem 0.44rem;
    white-space: nowrap;
}
.md-passport-status.ok {
    background: #e8fbf3;
    color: #047857;
}
.md-passport-status.todo {
    background: #eef2ff;
    color: #475569;
}

.st-key-home_passport_records .stButton > button,
.st-key-home_passport_overview .stButton > button,
.st-key-home_passport_meds .stButton > button {
    min-height: 40px !important;
    height: 40px !important;
    border-radius: 12px !important;
    font-size: 0.8rem !important;
    font-weight: 650 !important;
}

/* Final quick-action tile remodel: clean icons, clean spacing, no crumbled text. */
.st-key-qa_headache .stButton > button,
.st-key-qa_tired .stButton > button,
.st-key-qa_symptoms .stButton > button,
.st-key-qa_sleep .stButton > button {
    min-height: 56px !important;
    height: 56px !important;
    border-radius: 999px !important;
    padding: 0 0.92rem !important;
    border: 1px solid #d9e6fb !important;
    background: #ffffff !important;
    box-shadow: 0 8px 20px rgba(15,23,42,0.045) !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    text-align: left !important;
    gap: 0 !important;
}
.st-key-qa_headache .stButton > button:hover,
.st-key-qa_tired .stButton > button:hover,
.st-key-qa_symptoms .stButton > button:hover,
.st-key-qa_sleep .stButton > button:hover {
    border-color: #bfd7fb !important;
    box-shadow: 0 12px 26px rgba(59,130,246,0.10) !important;
    transform: translateY(-1px) !important;
}
.st-key-qa_headache .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_tired .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_symptoms .stButton > button [data-testid="stIconMaterial"],
.st-key-qa_sleep .stButton > button [data-testid="stIconMaterial"] {
    width: 28px !important;
    min-width: 28px !important;
    height: 28px !important;
    border-radius: 999px !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    margin-right: 0.55rem !important;
    font-size: 1rem !important;
    border: 1px solid rgba(255,255,255,0.65) !important;
}
.st-key-qa_headache .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(145deg,#ffe7ee,#ffd7e2) !important; color: #e11d48 !important; }
.st-key-qa_tired .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(145deg,#fff5d9,#ffeeb8) !important; color: #c97a00 !important; }
.st-key-qa_symptoms .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(145deg,#efeaff,#e1d7ff) !important; color: #7c3aed !important; }
.st-key-qa_sleep .stButton > button [data-testid="stIconMaterial"] { background: linear-gradient(145deg,#e7edff,#d7e1ff) !important; color: #4f46e5 !important; }

.st-key-qa_headache .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_tired .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_symptoms .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_sleep .stButton > button [data-testid="stMarkdownContainer"] {
    flex: 1 1 auto !important;
    min-width: 0 !important;
    overflow: hidden !important;
}
.st-key-qa_headache .stButton > button p,
.st-key-qa_tired .stButton > button p,
.st-key-qa_symptoms .stButton > button p,
.st-key-qa_sleep .stButton > button p {
    margin: 0 !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    font-size: 0.97rem !important;
    font-weight: 690 !important;
    line-height: 1.15 !important;
    color: #1f2f46 !important;
}
.st-key-qa_headache .stButton > button p::first-line,
.st-key-qa_tired .stButton > button p::first-line,
.st-key-qa_symptoms .stButton > button p::first-line,
.st-key-qa_sleep .stButton > button p::first-line {
    font-size: 0.97rem !important;
    font-weight: 690 !important;
    color: #1f2f46 !important;
}

/* Final send-button alignment lock: center arrow/icon perfectly. */
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button,
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button {
    display: grid !important;
    place-items: center !important;
    text-align: center !important;
    padding: 0 !important;
    line-height: 1 !important;
}
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button > div,
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button > div {
    display: grid !important;
    place-items: center !important;
    margin: 0 !important;
    width: auto !important;
}
.st-key-home_send_btn [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"],
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
    margin: 0 !important;
    line-height: 1 !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 1em !important;
    height: 1em !important;
    transform: translate(0, 0) !important;
}

/* Hard-lock composer action row at 100% browser zoom (desktop + laptop). */
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn),
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) {
    display: flex !important;
    flex-wrap: nowrap !important;
    align-items: center !important;
    gap: 0.72rem !important;
    column-gap: 0.72rem !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn) [data-testid="stColumn"],
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"] {
    display: flex !important;
    align-items: center !important;
    min-height: 56px !important;
    padding-top: 0 !important;
    padding-bottom: 0 !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn) [data-testid="stColumn"]:nth-child(1),
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn) [data-testid="stColumn"]:nth-child(2) {
    flex: 0 0 116px !important;
    min-width: 116px !important;
    max-width: 116px !important;
    width: 116px !important;
    margin: 0 !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn) [data-testid="stColumn"]:nth-child(3) {
    flex: 1 1 auto !important;
    min-width: 0 !important;
}
[data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stHorizontalBlock"]:has(.st-key-home_send_btn) [data-testid="stColumn"]:nth-child(4),
[data-testid="stForm"]:has(.st-key-chat_upload_btn) [data-testid="stHorizontalBlock"]:has(.st-key-chat_upload_btn) [data-testid="stColumn"]:nth-child(4) {
    flex: 0 0 56px !important;
    min-width: 56px !important;
    max-width: 56px !important;
    width: 56px !important;
    margin: 0 !important;
    justify-content: flex-end !important;
}
.st-key-home_upload_btn [data-testid="stFormSubmitButton"],
.st-key-home_voice_btn [data-testid="stFormSubmitButton"],
.st-key-home_send_btn [data-testid="stFormSubmitButton"],
.st-key-chat_upload_btn [data-testid="stFormSubmitButton"],
.st-key-chat_voice_btn [data-testid="stFormSubmitButton"] {
    margin: 0 !important;
    padding: 0 !important;
    width: 100% !important;
    display: flex !important;
    align-items: center !important;
}

/* Compact first-screen desktop pass so home fits cleanly without scroll. */
@media (min-width: 1200px) and (min-height: 760px) {
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        padding: 0.74rem 0.8rem 0.7rem 0.8rem !important;
    }
    [data-testid="stSidebar"] .md-logo-wrap {
        margin-bottom: 0.62rem !important;
        padding-bottom: 0.62rem !important;
        position: relative !important;
        z-index: 10 !important;
        min-height: 48px !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button {
        min-height: 34px !important;
        height: 34px !important;
        margin-bottom: 0.1rem !important;
        font-size: 0.85rem !important;
        padding-left: 0.45rem !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > span:first-child,
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > [data-testid="stIconMaterial"] {
        width: 24px !important;
        min-width: 24px !important;
        height: 24px !important;
        border-radius: 7px !important;
        margin-right: 0.5rem !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button [data-testid="stIconMaterial"] {
        font-size: 0.9rem !important;
    }
    [data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap > .md-logo-image {
        max-width: 140px !important;
        width: min(140px, 100%) !important;
    }
    [data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap {
        padding: 0.4rem 0 0.4rem 0 !important;
        margin-bottom: 0.75rem !important;
        max-width: 242px !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        border-radius: 0 !important;
        position: relative !important;
        z-index: 10 !important;
        min-height: 70px !important;
    }
    [data-testid="stSidebar"] .md-logo-wrap.md-logo-image-wrap > .md-logo-image {
        max-width: 155px !important;
        width: min(155px, 100%) !important;
        filter: drop-shadow(0 4px 16px rgba(15, 23, 42, 0.13)) drop-shadow(0 1px 4px rgba(15, 23, 42, 0.07)) !important;
    }
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        padding: 0.5rem 0.85rem 0.55rem 0.85rem !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] {
        margin: 0.45rem 0 0.3rem 0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"],
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 36px !important;
        height: 36px !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] label {
        margin-bottom: 0.22rem !important;
    }
    [data-testid="stSidebar"] .md-side-profile.md-side-profile-top {
        padding: 0.5rem 1.1rem 0.5rem 0.55rem !important;
        position: relative !important;
        z-index: 5 !important;
    }
    [data-testid="stSidebar"] .md-side-avatar {
        width: 28px !important;
        min-width: 28px !important;
        height: 28px !important;
        font-size: 0.76rem !important;
    }
    [data-testid="stSidebar"] .md-side-pname { font-size: 0.78rem !important; }
    [data-testid="stSidebar"] .md-side-psub { font-size: 0.62rem !important; }
    [data-testid="stSidebar"] div.st-key-nav_privacy_bottom .stButton > button {
        min-height: 32px !important;
        height: 32px !important;
        font-size: 0.78rem !important;
    }
    [data-testid="stSidebar"] div.st-key-nav_privacy_bottom {
        margin-bottom: 0.5rem !important;
    }
    [data-testid="stSidebar"] .md-sidebar-bottom {
        padding-top: 0.4rem !important;
        margin-top: 0.25rem !important;
    }
    [data-testid="stSidebar"] .sb-footer {
        font-size: 0.6rem !important;
        bottom: 0.35rem !important;
    }
    [data-testid="stSidebar"] div[class*="st-key-nav_"]:not(.st-key-nav_privacy_bottom) .stButton > button > div[data-testid="stMarkdownContainer"] p {
        font-size: 0.88rem !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] {
        margin-top: 0.68rem !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"],
    [data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 38px !important;
        height: 38px !important;
        border-radius: 12px !important;
    }
    [data-testid="stSidebar"] .md-side-profile-top {
        margin-top: 0.1rem !important;
    }

    .stMainBlockContainer,
    [data-testid="stMainBlockContainer"] {
        padding-top: 0.64rem !important;
        padding-bottom: 0.3rem !important;
    }
    .stMainBlockContainer > [data-testid="stVerticalBlock"] {
        gap: 0.3rem !important;
        row-gap: 0.3rem !important;
    }
    .md-home-greet-wrap {
        margin: 0.12rem 0 1.02rem 0 !important;
    }
    .md-home-greet-wrap .md-greet {
        font-size: 1.92rem !important;
        line-height: 1.08 !important;
    }
    .md-home-greet-wrap .md-subgreet {
        margin-top: 0.1rem !important;
        font-size: 0.9rem !important;
    }
    [data-testid="stHorizontalBlock"]:has(.st-key-qa_headache),
    [data-testid="stHorizontalBlock"]:has(.st-key-qa_tired),
    [data-testid="stHorizontalBlock"]:has(.st-key-qa_symptoms),
    [data-testid="stHorizontalBlock"]:has(.st-key-qa_sleep) {
        margin-top: 0.28rem !important;
    }
    .st-key-qa_headache .stButton > button,
    .st-key-qa_tired .stButton > button,
    .st-key-qa_symptoms .stButton > button,
    .st-key-qa_sleep .stButton > button {
        min-height: 54px !important;
        height: 54px !important;
        padding: 0 0.82rem !important;
    }
    .st-key-qa_headache .stButton > button p,
    .st-key-qa_tired .stButton > button p,
    .st-key-qa_symptoms .stButton > button p,
    .st-key-qa_sleep .stButton > button p {
        font-size: 0.9rem !important;
    }
    [data-testid="stForm"]:has(.st-key-home_send_btn) {
        padding: 0.62rem 0.66rem 0.6rem 0.66rem !important;
        border-radius: 20px !important;
    }
    [data-testid="stForm"]:has(.st-key-home_send_btn) [data-testid="stTextArea"] textarea {
        min-height: 94px !important;
        height: 94px !important;
        font-size: 0.95rem !important;
    }
    .md-home-composer-note {
        margin-top: 0.06rem !important;
        margin-bottom: 0.12rem !important;
        font-size: 0.72rem !important;
    }
    .md-smart-head {
        margin-top: 0.54rem !important;
    }
    .st-key-sa_sym .stButton > button,
    .st-key-sa_rec .stButton > button,
    .st-key-sa_ins .stButton > button,
    .st-key-sa_appt .stButton > button {
        min-height: 108px !important;
        border-radius: 16px !important;
        padding: 0.7rem 0.8rem 0.65rem 0.8rem !important;
    }
    .st-key-sa_sym .stButton > button p,
    .st-key-sa_rec .stButton > button p,
    .st-key-sa_ins .stButton > button p,
    .st-key-sa_appt .stButton > button p {
        font-size: 0.7rem !important;
        line-height: 1.3 !important;
    }
    .st-key-sa_sym .stButton > button p::first-line,
    .st-key-sa_rec .stButton > button p::first-line,
    .st-key-sa_ins .stButton > button p::first-line,
    .st-key-sa_appt .stButton > button p::first-line {
        font-size: 0.82rem !important;
        font-weight: 700 !important;
    }
    .st-key-sa_sym .stButton > button [data-testid="stIconMaterial"],
    .st-key-sa_rec .stButton > button [data-testid="stIconMaterial"],
    .st-key-sa_ins .stButton > button [data-testid="stIconMaterial"],
    .st-key-sa_appt .stButton > button [data-testid="stIconMaterial"] {
        width: 32px !important;
        height: 32px !important;
        min-width: 32px !important;
        font-size: 1.05rem !important;
        border-radius: 10px !important;
        margin: 0 0.55rem 0 0 !important;
    }
    .st-key-sa_sym .stButton > button::after,
    .st-key-sa_rec .stButton > button::after,
    .st-key-sa_ins .stButton > button::after,
    .st-key-sa_appt .stButton > button::after {
        right: 0.6rem !important;
        bottom: 0.55rem !important;
        width: 22px !important;
        height: 22px !important;
        font-size: 0.85rem !important;
    }
    .md-tip-carousel {
        margin: 0.58rem 0 0.38rem 0 !important;
        height: 136px !important;
        border-radius: 16px !important;
    }
    /* Tip carousel in compact mode — text now spans the full tile width.
       The illustration is absolute-positioned in the top-right corner as
       a soft decorative accent; not part of any grid track. */
    .md-tip-slide {
        padding: 0.85rem 1rem 0.95rem 1rem !important;
        display: block !important;
        gap: 0 !important;
        grid-template-columns: none !important;
    }
    .md-tip-slide > div:first-child {
        width: 100% !important;
        min-width: 0 !important;
    }
    .md-tip-title {
        font-size: 1rem !important;
        margin-bottom: 0.18rem !important;
    }
    .md-tip-desc {
        font-size: 0.78rem !important;
        margin-bottom: 0.4rem !important;
        line-height: 1.4 !important;
        max-width: none !important;
        padding-right: 64px !important;
    }
    .md-tip-metric {
        font-size: 0.74rem !important;
        padding: 0.24rem 0.6rem !important;
    }
    .md-tip-illust {
        position: absolute !important;
        top: 0.85rem !important;
        right: 0.95rem !important;
        width: 50px !important;
        height: 50px !important;
        opacity: 0.6 !important;
    }
    .md-tip-illust .material-symbols-rounded {
        font-size: 26px !important;
    }
    .md-rcard {
        padding: 0.76rem 0.84rem !important;
        border-radius: 14px !important;
    }
    .md-snap-grid {
        gap: 0.42rem !important;
    }
    .md-passport-sub {
        margin-bottom: 0.22rem !important;
        font-size: 0.72rem !important;
    }
    .md-passport-progress {
        margin-bottom: 0.36rem !important;
    }
    .md-passport-check {
        padding: 0.08rem 0 !important;
    }
    .md-passport-check:nth-of-type(n+4) {
        display: none !important;
    }
    .st-key-home_passport_records .stButton > button,
    .st-key-home_passport_overview .stButton > button,
    .st-key-home_passport_meds .stButton > button {
        min-height: 32px !important;
        height: 32px !important;
        font-size: 0.74rem !important;
    }
    div.st-key-home_passport_records,
    div.st-key-home_passport_overview,
    div.st-key-home_passport_meds {
        display: none !important;
    }
}

/* Final desktop text-overlap safety pass (home cards + quick actions + overview rows). */
.st-key-qa_headache .stButton > button,
.st-key-qa_tired .stButton > button,
.st-key-qa_symptoms .stButton > button,
.st-key-qa_sleep .stButton > button {
    min-width: 0 !important;
    gap: 0 !important;
    align-items: center !important;
}
.st-key-qa_headache .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_tired .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_symptoms .stButton > button [data-testid="stMarkdownContainer"],
.st-key-qa_sleep .stButton > button [data-testid="stMarkdownContainer"] {
    min-width: 0 !important;
    width: 100% !important;
    overflow: hidden !important;
}
.st-key-qa_headache .stButton > button p,
.st-key-qa_tired .stButton > button p,
.st-key-qa_symptoms .stButton > button p,
.st-key-qa_sleep .stButton > button p {
    margin: 0 !important;
    white-space: normal !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    display: -webkit-box !important;
    -webkit-line-clamp: 2 !important;
    -webkit-box-orient: vertical !important;
    line-height: 1.14 !important;
}

.st-key-sa_sym .stButton > button,
.st-key-sa_rec .stButton > button,
.st-key-sa_ins .stButton > button,
.st-key-sa_appt .stButton > button {
    min-width: 0 !important;
    overflow: hidden !important;
    /* padding-right intentionally NOT set here; the compact-mode block
       hides the arrow chip so no right-side buffer is needed. */
}
.st-key-sa_sym .stButton > button [data-testid="stMarkdownContainer"],
.st-key-sa_rec .stButton > button [data-testid="stMarkdownContainer"],
.st-key-sa_ins .stButton > button [data-testid="stMarkdownContainer"],
.st-key-sa_appt .stButton > button [data-testid="stMarkdownContainer"] {
    width: 100% !important;
    min-width: 0 !important;
}
.st-key-sa_sym .stButton > button p,
.st-key-sa_rec .stButton > button p,
.st-key-sa_ins .stButton > button p,
.st-key-sa_appt .stButton > button p {
    margin: 0 !important;
    white-space: pre-line !important;
    overflow-wrap: anywhere !important;
    word-break: normal !important;
    line-height: 1.3 !important;
}
.st-key-sa_sym .stButton > button p::first-line,
.st-key-sa_rec .stButton > button p::first-line,
.st-key-sa_ins .stButton > button p::first-line,
.st-key-sa_appt .stButton > button p::first-line {
    line-height: 1.22 !important;
}

.md-snap-tile {
    min-height: 64px !important;
}
.md-snap-text {
    min-width: 0 !important;
}
.md-snap-text > div {
    min-width: 0 !important;
}
.md-snap-label {
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
.md-snap-value {
    white-space: nowrap !important;
    font-size: 0.9rem !important;
    line-height: 1.18 !important;
}
.md-snap-status {
    white-space: nowrap !important;
    flex-shrink: 0 !important;
}

/* ════════════════════════════════════════════════════════════════════════
   FINAL NUKE — Past-chats sidebar list (chat history).
   Always-applied (no media query), maximum specificity. Streamlit renders
   each chat title button TWICE (one wrapped in stTooltipHoverTarget, one
   inline) — we target ALL of them by widget-key wildcard, so it doesn't
   matter how Streamlit nests the actual <button>.
   ──────────────────────────────────────────────────────────────────────── */

/* ── Container + row layout ──────────────────────────────────────────── */
[data-testid="stSidebar"] .md-past-chats {
    display: flex !important;
    flex-direction: column !important;
    gap: 0.18rem !important;
    margin-bottom: 0.45rem !important;
    max-height: 320px !important;
    overflow-y: auto !important;
    padding-right: 0.15rem !important;
}
[data-testid="stSidebar"] .md-past-chats [data-testid="stHorizontalBlock"] {
    gap: 0.18rem !important;
    align-items: center !important;
    flex-wrap: nowrap !important;
}
[data-testid="stSidebar"] .md-past-chats [data-testid="stColumn"] {
    min-width: 0 !important;
    padding: 0 !important;
}
/* Pin the × delete column to a hard 32px so it can never get pushed off. */
[data-testid="stSidebar"] .md-past-chats [data-testid="stColumn"]:last-child {
    flex: 0 0 32px !important;
    width: 32px !important;
    max-width: 32px !important;
    min-width: 32px !important;
}

/* ── Chat title button (any button inside st-key-conv_open_*) ─────── */
[data-testid="stSidebar"] [class*="st-key-conv_open_"] button,
[data-testid="stSidebar"] [class*="st-key-conv_open_"] button[kind="secondary"],
[data-testid="stSidebar"] [class*="st-key-conv_open_"] [data-testid="stBaseButton-secondary"] {
    min-height: 34px !important;
    height: 34px !important;
    max-height: 34px !important;
    padding: 0 0.6rem !important;
    border-radius: 9px !important;
    border: 1px solid transparent !important;
    background: transparent !important;
    color: #475569 !important;
    font-weight: 550 !important;
    font-size: 0.76rem !important;
    text-align: left !important;
    justify-content: flex-start !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    box-shadow: none !important;
    display: flex !important;
    align-items: center !important;
    width: 100% !important;
    transition: background 0.15s ease, color 0.15s ease !important;
}
[data-testid="stSidebar"] [class*="st-key-conv_open_"] button:hover {
    background: #f1f5fc !important;
    color: #1f2a3d !important;
    border-color: transparent !important;
}
[data-testid="stSidebar"] [class*="st-key-conv_open_"] button p,
[data-testid="stSidebar"] [class*="st-key-conv_open_"] button [data-testid="stMarkdownContainer"],
[data-testid="stSidebar"] [class*="st-key-conv_open_"] button [data-testid="stMarkdownContainer"] p {
    margin: 0 !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    text-align: left !important;
    font-size: 0.76rem !important;
    font-weight: 550 !important;
    line-height: 1.2 !important;
    max-width: 100% !important;
    -webkit-line-clamp: 1 !important;
    display: block !important;
    width: 100% !important;
    color: inherit !important;
}
/* Streamlit renders title button TWICE. The first child of .stButton is
   the visible tooltip-wrapped button; the second child is an inline
   duplicate (class e8vg11g19) that doubles row height. Hide it via
   nth-child (broader browser support than :has). */
[data-testid="stSidebar"] [class*="st-key-conv_open_"] .stButton > div + div {
    display: none !important;
}
[data-testid="stSidebar"] [class*="st-key-conv_open_"] .stButton > div:nth-child(n + 2) {
    display: none !important;
}
/* Keep the sidebar row strictly inside the sidebar's right edge. */
[data-testid="stSidebar"] .md-past-chats {
    padding-right: 0.4rem !important;
    max-width: 100% !important;
    box-sizing: border-box !important;
}
[data-testid="stSidebar"] .md-past-chats [data-testid="stHorizontalBlock"] {
    max-width: 100% !important;
    overflow: hidden !important;
}

/* Active chat row — indigo wash + bolder indigo text. */
[data-testid="stSidebar"] .md-past-active [class*="st-key-conv_open_"] button {
    background: linear-gradient(180deg, #eef0ff 0%, #e0e7ff 100%) !important;
    color: #3730a3 !important;
    font-weight: 660 !important;
}
[data-testid="stSidebar"] .md-past-active [class*="st-key-conv_open_"] button p,
[data-testid="stSidebar"] .md-past-active [class*="st-key-conv_open_"] button [data-testid="stMarkdownContainer"] p {
    color: #3730a3 !important;
    font-weight: 660 !important;
}

/* ── Delete × button (any button inside st-key-conv_del_*) ────────── */
[data-testid="stSidebar"] [class*="st-key-conv_del_"] button,
[data-testid="stSidebar"] [class*="st-key-conv_del_"] button[kind="secondary"],
[data-testid="stSidebar"] [class*="st-key-conv_del_"] [data-testid="stBaseButton-secondary"] {
    min-width: 28px !important;
    width: 28px !important;
    max-width: 28px !important;
    min-height: 28px !important;
    height: 28px !important;
    max-height: 28px !important;
    padding: 0 !important;
    border-radius: 7px !important;
    border: none !important;
    background: transparent !important;
    color: #cbd5e1 !important;
    font-size: 0.92rem !important;
    font-weight: 600 !important;
    line-height: 1 !important;
    box-shadow: none !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    transition: background 0.15s ease, color 0.15s ease !important;
}
[data-testid="stSidebar"] [class*="st-key-conv_del_"] button:hover {
    background: #fef2f2 !important;
    color: #b91c1c !important;
}
[data-testid="stSidebar"] [class*="st-key-conv_del_"] button p,
[data-testid="stSidebar"] [class*="st-key-conv_del_"] button [data-testid="stMarkdownContainer"] p {
    margin: 0 !important;
    line-height: 1 !important;
    font-size: 1rem !important;
    color: inherit !important;
}

</style>
""", unsafe_allow_html=True)

GROQ_API_KEY = _safe_secret("GROQ_API_KEY", os.environ.get("GROQ_API_KEY", ""))
GROQ_ACTIVE = bool(GROQ_API_KEY)
groq_client = Groq(api_key=GROQ_API_KEY) if GROQ_ACTIVE else None

ANTHROPIC_API_KEY = _safe_secret("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY", ""))
CLAUDE_MODEL = _safe_secret("CLAUDE_MODEL", "claude-haiku-4-5")
anthropic_client = None
if ANTHROPIC_AVAILABLE and ANTHROPIC_API_KEY:
    try:
        anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY)
    except Exception as e:
        print("Anthropic init failed:", e)
        anthropic_client = None

CLAUDE_ACTIVE = anthropic_client is not None

def transcribe_voice_note(audio_file):
    """Transcribe a browser-recorded or uploaded audio note with Groq Whisper."""
    if not GROQ_ACTIVE or groq_client is None:
        return ""
    if not audio_file:
        return ""
    try:
        audio_file.seek(0)
    except Exception:
        pass
    try:
        audio_bytes = audio_file.read()
    except Exception:
        try:
            audio_bytes = audio_file.getvalue()
        except Exception:
            audio_bytes = b""
    if not audio_bytes:
        return ""
    file_name = getattr(audio_file, "name", "voice-note.wav") or "voice-note.wav"
    mime_type = getattr(audio_file, "type", "audio/wav") or "audio/wav"
    try:
        transcription = groq_client.audio.transcriptions.create(
            file=(file_name, audio_bytes, mime_type),
            model=os.environ.get("GROQ_TRANSCRIPTION_MODEL", "whisper-large-v3-turbo"),
            response_format="json",
            temperature=0,
        )
        return (getattr(transcription, "text", "") or "").strip()
    except Exception as e:
        print("Voice transcription failed:", e)
        return ""

# ── Language Config ───────────────────────────────────────────────────
LANGUAGES = {
    "English": {
        "flag": "🇦🇺",
        "greeting": "Hello! How can I help you today?",
        "welcome_text": "I am MediChat, your friendly Ai health assistant.<br>I remember everything you tell me during our conversation.<br><br>Or switch to Symptom Check for a guided assessment!",
        "placeholder": "Type your health question here...",
        "send_btn": "Send to MediChat",
        "clear_btn": "Clear",
        "upload_label": "Upload a medical image (optional)",
        "question_label": "Your question",
        "helpful": "Was this conversation helpful?",
        "yes": "Yes",
        "no": "No",
        "thanks_helpful": "Thank you! Glad MediChat was helpful.",
        "thanks_not": "Thank you for your feedback. We will keep improving!",
        "download_chat": "Download your conversation:",
        "download_chat_btn": "Download Chat as PDF",
        "download_assess_btn": "Download Assessment Report as PDF",
        "symptom_title": "Symptom Assessment",
        "symptom_subtitle": "Answer a few quick questions and MediChat will generate a personalised health assessment.",
        "quick_select": "Quick select:",
        "next": "Next",
        "cancel": "Cancel",
        "answers_so_far": "Your answers so far",
        "new_assessment": "Start New Assessment",
        "switch_chat": "Switch to Free Chat",
        "report_title": "MediChat Assessment Report",
        "symptoms_reported": "Symptoms Reported",
        "possible_conditions": "Possible Conditions",
        "what_to_do": "What To Do Next",
        "summary": "MediChat Summary",
        "disclaimer_short": "This assessment is for information only and is NOT a medical diagnosis. Please consult a qualified healthcare professional.",
        "free_chat": "Free Chat",
        "symptom_check": "Symptom Check",
        "lang_instruction": "IMPORTANT: You MUST respond entirely in English.",
    },
    "Tamil": {
        "flag": "🇱🇰",
        "greeting": "வணக்கம்! இன்று நான் உங்களுக்கு எப்படி உதவலாம்?",
        "welcome_text": "நான் MediChat, உங்கள் நட்பான Ai சுகாதார உதவியாளர்.<br>உரையாடலில் நீங்கள் சொல்வதை நான் நினைவில் வைத்திருப்பேன்.<br><br>வழிகாட்டப்பட்ட மதிப்பீட்டிற்கு அறிகுறி சரிபார்ப்புக்கு மாறலாம்!",
        "placeholder": "உங்கள் உடல்நல கேள்வியை இங்கே தட்டச்சு செய்யுங்கள்...",
        "send_btn": "MediChat க்கு அனுப்பவும்",
        "clear_btn": "அழிக்கவும்",
        "upload_label": "மருத்துவ படத்தை பதிவேற்றவும் (விரும்பினால்)",
        "question_label": "உங்கள் கேள்வி",
        "helpful": "இந்த உரையாடல் உதவியாக இருந்ததா?",
        "yes": "ஆம்",
        "no": "இல்லை",
        "thanks_helpful": "நன்றி! MediChat உதவியாக இருந்தது மகிழ்ச்சி.",
        "thanks_not": "உங்கள் கருத்துக்கு நன்றி. நாங்கள் தொடர்ந்து மேம்படுவோம்!",
        "download_chat": "உரையாடலை பதிவிறக்கவும்:",
        "download_chat_btn": "அரட்டையை PDF ஆக பதிவிறக்கவும்",
        "download_assess_btn": "மதிப்பீட்டு அறிக்கையை PDF ஆக பதிவிறக்கவும்",
        "symptom_title": "அறிகுறி மதிப்பீடு",
        "symptom_subtitle": "சில விரைவான கேள்விகளுக்கு பதிலளிக்கவும், MediChat உங்களுக்கு தனிப்பயன் சுகாதார மதிப்பீட்டை உருவாக்கும்.",
        "quick_select": "விரைவு தேர்வு:",
        "next": "அடுத்து",
        "cancel": "ரத்து செய்",
        "answers_so_far": "இதுவரை உங்கள் பதில்கள்",
        "new_assessment": "புதிய மதிப்பீட்டை தொடங்கவும்",
        "switch_chat": "இலவச அரட்டைக்கு மாறவும்",
        "report_title": "MediChat மதிப்பீட்டு அறிக்கை",
        "symptoms_reported": "தெரிவிக்கப்பட்ட அறிகுறிகள்",
        "possible_conditions": "சாத்தியமான நிலைமைகள்",
        "what_to_do": "அடுத்து என்ன செய்வது",
        "summary": "MediChat சுருக்கம்",
        "disclaimer_short": "இந்த மதிப்பீடு தகவல் நோக்கங்களுக்காக மட்டுமே. தயவுசெய்து தகுதிவாய்ந்த மருத்துவரை அணுகவும்.",
        "free_chat": "இலவச அரட்டை",
        "symptom_check": "அறிகுறி சரிபார்ப்பு",
        "lang_instruction": "IMPORTANT: You MUST respond entirely in Tamil (தமிழ்). All your responses must be in Tamil language.",
    },
    "Sinhala": {
        "flag": "🇱🇰",
        "greeting": "ආයුබෝවන්! අද මට ඔබට කෙසේ උදව් කළ හැකිද?",
        "welcome_text": "මම MediChat, ඔබේ මිත්‍රශීලී Ai සෞඛ්‍ය සහායකයා.<br>ඔබ කියන සෑම දෙයක්ම මම මතක තබා ගනිමි.<br><br>මඟ පෙන්වූ තක්සේරු කිරීම සඳහා රෝග ලක්ෂණ පරීක්ෂාවට මාරු වන්න!",
        "placeholder": "ඔබේ සෞඛ්‍ය ප්‍රශ්නය මෙහි ටයිප් කරන්න...",
        "send_btn": "MediChat වෙත යවන්න",
        "clear_btn": "හිස් කරන්න",
        "upload_label": "වෛද්‍ය රූපයක් උඩුගත කරන්න (විකල්ප)",
        "question_label": "ඔබේ ප්‍රශ්නය",
        "helpful": "මෙම සංවාදය ප්‍රයෝජනවත් වූවාද?",
        "yes": "ඔව්",
        "no": "නැහැ",
        "thanks_helpful": "ස්තූතියි! MediChat ප්‍රයෝජනවත් වූ බව සතුටක්.",
        "thanks_not": "ඔබේ ප්‍රතිචාරයට ස්තූතියි. අපි දිගටම වැඩිදියුණු කරන්නෙමු!",
        "download_chat": "ඔබේ සංවාදය බාගත කරන්න:",
        "download_chat_btn": "Chat PDF ලෙස බාගත කරන්න",
        "download_assess_btn": "තක්සේරු වාර්තාව PDF ලෙස බාගත කරන්න",
        "symptom_title": "රෝග ලක්ෂණ තක්සේරු කිරීම",
        "symptom_subtitle": "ප්‍රශ්න කිහිපයකට පිළිතුරු දෙන්න, MediChat ඔබට පෞද්ගලික සෞඛ්‍ය තක්සේරුවක් ජනනය කරනු ඇත.",
        "quick_select": "ඉක්මන් තේරීම:",
        "next": "ඊළඟ",
        "cancel": "අවලංගු කරන්න",
        "answers_so_far": "මෙතෙක් ඔබේ පිළිතුරු",
        "new_assessment": "නව තක්සේරු කිරීමක් ආරම්භ කරන්න",
        "switch_chat": "නිදහස් Chat වෙත මාරු වන්න",
        "report_title": "MediChat තක්සේරු වාර්තාව",
        "symptoms_reported": "වාර්තා කළ රෝග ලක්ෂණ",
        "possible_conditions": "හැකි තත්ත්වයන්",
        "what_to_do": "ඊළඟට කළ යුතු දේ",
        "summary": "MediChat සාරාංශය",
        "disclaimer_short": "මෙම තක්සේරු කිරීම තොරතුරු පමණක් වේ. සුදුසුකම් ලත් වෛද්‍යවරයෙකු හමුවෙන්න.",
        "free_chat": "නිදහස් Chat",
        "symptom_check": "රෝග ලක්ෂණ පරීක්ෂාව",
        "lang_instruction": "IMPORTANT: You MUST respond entirely in Sinhala (සිංහල). All your responses must be in Sinhala language.",
    },
    "Hindi": {
        "flag": "🇮🇳",
        "greeting": "नमस्ते! आज मैं आपकी कैसे मदद कर सकता हूं?",
        "welcome_text": "मैं MediChat हूं, आपका मित्रवत Ai स्वास्थ्य सहायक.<br>आप जो कुछ भी बताते हैं मैं याद रखता हूं.<br><br>निर्देशित मूल्यांकन के लिए लक्षण जांच पर स्विच करें!",
        "placeholder": "अपना स्वास्थ्य प्रश्न यहाँ टाइप करें...",
        "send_btn": "MediChat को भेजें",
        "clear_btn": "साफ करें",
        "upload_label": "चिकित्सा छवि अपलोड करें (वैकल्पिक)",
        "question_label": "आपका प्रश्न",
        "helpful": "क्या यह बातचीत मददगार थी?",
        "yes": "हाँ",
        "no": "नहीं",
        "thanks_helpful": "धन्यवाद! खुशी है कि MediChat मददगार रहा।",
        "thanks_not": "आपकी प्रतिक्रिया के लिए धन्यवाद। हम लगातार सुधार करते रहेंगे!",
        "download_chat": "अपनी बातचीत डाउनलोड करें:",
        "download_chat_btn": "चैट PDF के रूप में डाउनलोड करें",
        "download_assess_btn": "मूल्यांकन रिपोर्ट PDF डाउनलोड करें",
        "symptom_title": "लक्षण मूल्यांकन",
        "symptom_subtitle": "कुछ त्वरित प्रश्नों का उत्तर दें और MediChat आपका व्यक्तिगत स्वास्थ्य मूल्यांकन तैयार करेगा।",
        "quick_select": "त्वरित चयन:",
        "next": "अगला",
        "cancel": "रद्द करें",
        "answers_so_far": "अब तक के आपके उत्तर",
        "new_assessment": "नया मूल्यांकन शुरू करें",
        "switch_chat": "फ्री चैट पर स्विच करें",
        "report_title": "MediChat मूल्यांकन रिपोर्ट",
        "symptoms_reported": "रिपोर्ट किए गए लक्षण",
        "possible_conditions": "संभावित स्थितियां",
        "what_to_do": "आगे क्या करें",
        "summary": "MediChat सारांश",
        "disclaimer_short": "यह मूल्यांकन केवल सूचना के लिए है। कृपया योग्य चिकित्सक से परामर्श लें।",
        "free_chat": "फ्री चैट",
        "symptom_check": "लक्षण जांच",
        "lang_instruction": "IMPORTANT: You MUST respond entirely in Hindi (हिन्दी). All your responses must be in Hindi language.",
    },
    "Malayalam": {
        "flag": "🇮🇳",
        "greeting": "നമസ്കാരം! ഇന്ന് ഞാൻ നിങ്ങളെ എങ്ങനെ സഹായിക്കാം?",
        "welcome_text": "ഞാൻ MediChat, നിങ്ങളുടെ സൗഹൃദ Ai ആരോഗ്യ സഹായി.<br>നിങ്ങൾ പറയുന്നതെല്ലാം ഞാൻ ഓർത്തിരിക്കും.<br><br>ഗൈഡഡ് അസസ്മെൻ്റിനായി സിംപ്റ്റം ചെക്കിലേക്ക് മാറുക!",
        "placeholder": "നിങ്ങളുടെ ആരോഗ്യ ചോദ്യം ഇവിടെ ടൈപ്പ് ചെയ്യുക...",
        "send_btn": "MediChat-ലേക്ക് അയയ്ക്കുക",
        "clear_btn": "മായ്ക്കുക",
        "upload_label": "മെഡിക്കൽ ചിത്രം അപ്ലോഡ് ചെയ്യുക (ഐച്ഛികം)",
        "question_label": "നിങ്ങളുടെ ചോദ്യം",
        "helpful": "ഈ സംഭാഷണം സഹായകരമായിരുന്നോ?",
        "yes": "അതെ",
        "no": "ഇല്ല",
        "thanks_helpful": "നന്ദി! MediChat സഹായകരമായതിൽ സന്തോഷം.",
        "thanks_not": "നിങ്ങളുടെ ഫീഡ്‌ബാക്കിന് നന്ദി. ഞങ്ങൾ മെച്ചപ്പെടുത്തുന്നത് തുടരും!",
        "download_chat": "നിങ്ങളുടെ സംഭാഷണം ഡൗൺലോഡ് ചെയ്യുക:",
        "download_chat_btn": "ചാറ്റ് PDF ആയി ഡൗൺലോഡ് ചെയ്യുക",
        "download_assess_btn": "അസസ്മെൻ്റ് റിപ്പോർട്ട് PDF ഡൗൺലോഡ് ചെയ്യുക",
        "symptom_title": "ലക്ഷണ വിലയിരുത്തൽ",
        "symptom_subtitle": "ചില ചോദ്യങ്ങൾക്ക് ഉത്തരം നൽകൂ, MediChat നിങ്ങൾക്ക് വ്യക്തിഗത ആരോഗ്യ വിലയിരുത്തൽ തയ്യാറാക്കും.",
        "quick_select": "വേഗ തിരഞ്ഞെടുക്കൽ:",
        "next": "അടുത്തത്",
        "cancel": "റദ്ദാക്കുക",
        "answers_so_far": "ഇതുവരെ നിങ്ങളുടെ ഉത്തരങ്ങൾ",
        "new_assessment": "പുതിയ വിലയിരുത്തൽ ആരംഭിക്കുക",
        "switch_chat": "ഫ്രീ ചാറ്റിലേക്ക് മാറുക",
        "report_title": "MediChat വിലയിരുത്തൽ റിപ്പോർട്ട്",
        "symptoms_reported": "റിപ്പോർട്ട് ചെയ്ത ലക്ഷണങ്ങൾ",
        "possible_conditions": "സാധ്യമായ അവസ്ഥകൾ",
        "what_to_do": "അടുത്തതായി എന്ത് ചെയ്യണം",
        "summary": "MediChat സംഗ്രഹം",
        "disclaimer_short": "ഈ വിലയിരുത്തൽ വിവരങ്ങൾക്ക് മാത്രമുള്ളതാണ്. യോഗ്യതയുള്ള ഡോക്ടറെ സമീപിക്കുക.",
        "free_chat": "ഫ്രീ ചാറ്റ്",
        "symptom_check": "ലക്ഷണ പരിശോധന",
        "lang_instruction": "IMPORTANT: You MUST respond entirely in Malayalam (മലയാളം). All your responses must be in Malayalam language.",
    },
}

@st.cache_resource(show_spinner=False)
def load_rag_system():
    """Initializes embeddings and mounts the clinical database flat index with strict offline fallbacks."""
    try:
        embedder = SentenceTransformer("all-MiniLM-L6-v2")
    except Exception as e:
        print("Embedding transformer initialization failed, using local model mirror:", e)
        return None, None, []

    pubmed_docs = []
    dialog_docs = []
    pubmed_target = max(500, MEDICAL_REFERENCE_TARGET // 2)
    dialog_target = max(500, MEDICAL_REFERENCE_TARGET - pubmed_target)
    
    # Attempt PubMedQA pipeline loading
    try:
        pubmed = load_dataset("qiaojin/PubMedQA", "pqa_labeled", split="train[:" + str(pubmed_target) + "]", timeout=10)
        pubmed_docs = ["[PubMed Research]\nQuestion: " + i["question"] + "\nAnswer: " + i["long_answer"] for i in pubmed]
    except Exception as e:
        print("PubMedQA repository offline. Switched to secure backup arrays.", e)

    # Attempt MedDialog pipeline loading
    dialog_sources = [
        ("BinKhoaLe1812/MedDialog-EN-100k", "train[:" + str(dialog_target) + "]", "input", "output"),
        ("shibing624/medical", "train[:" + str(dialog_target) + "]", "instruction", "output")
    ]
    
    for ds_name, ds_split, in_key, out_key in dialog_sources:
        try:
            meddialog = load_dataset(ds_name, split=ds_split, timeout=10)
            dialog_docs = [
                "[Doctor-Patient Conversation]\nPatient: " + str(i.get(in_key, "")) + "\nDoctor: " + str(i.get(out_key, "")) 
                for i in meddialog if i.get(in_key) and i.get(out_key)
            ]
            if dialog_docs:
                break
        except Exception:
            continue

    docs = pubmed_docs + dialog_docs
    
    # Production Safeguard: If all external datasets fail, mount the complete core clinical fallback matrix
    if not docs:
        docs = [
            "[PubMed Research]\nQuestion: What causes severe headaches?\nAnswer: Headaches stem from muscle tension, severe dehydration, stress, vascular migraines, or blood pressure issues. Sudden, thunderclap headaches require urgent screening to rule out subarachnoid haemorrhage or clinical meningitis.",
            "[PubMed Research]\nQuestion: What are the cardinal markers of acute myocardial infarction?\nAnswer: Clinical indicators include crushing central chest tightness, radiating jaw or left arm discomfort, acute dyspnea, diaphoresis, and sudden nausea. This layout requires immediate 000 activation.",
            "[PubMed Research]\nQuestion: How is chronic primary hypertension audited?\nAnswer: Audiological management incorporates radical dietary reductions, structural aerobic exercise, and pharmacological management via ACE inhibitors, calcium channel blockers, or beta-blockers."
        ]

    try:
        embeddings = embedder.encode(docs, show_progress_bar=False)
        idx = faiss.IndexFlatL2(embeddings.shape[1])
        idx.add(embeddings.astype("float32"))
        return embedder, idx, docs
    except Exception as initialization_error:
        print("FAISS structural assembly failure:", initialization_error)
        return None, None, docs

with st.spinner("Loading MediChat knowledge base..."):
    try:
        embedder, index, documents = load_rag_system()
    except Exception as e:
        st.error("MediChat knowledge base failed to load. Please refresh the page in a moment. Error: " + str(e))
        st.stop()

MEDICAL_REFERENCE_COUNT = len(documents)

AU_TOP_DRUGS_FALLBACK = {
    "amoxicillin", "amoxicillin clavulanate", "azithromycin", "cephalexin", "doxycycline", "trimethoprim",
    "metformin", "gliclazide", "empagliflozin", "sitagliptin", "insulin glargine",
    "atorvastatin", "rosuvastatin", "simvastatin", "ezetimibe",
    "amlodipine", "perindopril", "lisinopril", "ramipril", "losartan", "valsartan", "metoprolol",
    "aspirin", "clopidogrel", "apixaban", "rivaroxaban", "warfarin",
    "salbutamol", "budesonide", "fluticasone", "tiotropium", "montelukast",
    "levothyroxine", "prednisone", "prednisolone", "omeprazole", "esomeprazole", "pantoprazole",
    "ibuprofen", "naproxen", "diclofenac", "paracetamol", "codeine", "tramadol",
    "sertraline", "escitalopram", "fluoxetine", "mirtazapine", "venlafaxine",
    "quetiapine", "olanzapine", "risperidone", "lithium",
    "ondansetron", "metoclopramide", "loperamide", "dimenhydrinate",
    "nitrofurantoin", "ciprofloxacin", "valaciclovir", "acyclovir",
}

def load_known_drugs():
    csv_path = _safe_secret("AU_DRUGS_CSV_PATH", os.environ.get("AU_DRUGS_CSV_PATH", ""))
    known = set(AU_TOP_DRUGS_FALLBACK)
    if not csv_path:
        return known
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            for line in f:
                token = (line or "").strip().lower()
                if token and token not in {"drug", "name", "medication"}:
                    known.add(token)
    except Exception as e:
        print("Known-drug list load failed:", e)
    return known

KNOWN_DRUGS = load_known_drugs()

PRESCRIPTION_PROMPT = """You are a clinical transcription assistant. Your only job is to read the prescription image and transcribe what is written.
You are NOT giving medical advice, NOT diagnosing, and NOT prescribing treatment.

Read this order:
1. Prescriber/clinic header
2. Patient identifiers
3. Medication line(s)
4. Dose, frequency, route, quantity, repeats
5. Date and signature

Output this exact format:

**MEDICATION**
Reading: <best transcription>
Matches known drug: <yes/no>
Confidence: high | medium | low

**STRENGTH / DOSE**
Reading: <exact text or not specified>
Confidence: high | medium | low

**FREQUENCY / DIRECTIONS**
As written: <exact text>
Plain English: <short rewrite>
Confidence: high | medium | low

**ROUTE**
Reading: <oral/topical/inhaled/injection/not specified>

**QUANTITY**
Reading: <exact text or not specified>

**REFILLS**
Reading: <exact text or not specified>

**PRESCRIBER**
Name: <as written or unclear>
Date: <as written or unclear>

**OVERALL CONFIDENCE**: high | medium | low
**ILLEGIBLE SECTIONS**: <list regions that are genuinely unreadable>

**IMPORTANT**: Transcription only. Do not use this output as dosing or diagnosis advice. Confirm with a pharmacist or prescriber.
"""

def preprocess_prescription(image_bytes):
    img = Image.open(io.BytesIO(image_bytes))
    img = ImageOps.exif_transpose(img)
    if img.mode != "RGB":
        img = img.convert("RGB")
    max_dim = 1568
    if max(img.size) > max_dim:
        img.thumbnail((max_dim, max_dim), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.35)
    img = ImageEnhance.Sharpness(img).enhance(1.4)
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=92)
    return out.getvalue()

def validate_drug_name(reading):
    candidate = (reading or "").lower().strip()
    if not candidate:
        return {"match": "none"}
    if candidate in KNOWN_DRUGS:
        return {"match": "exact", "drug": candidate}
    close = difflib.get_close_matches(candidate, KNOWN_DRUGS, n=3, cutoff=0.76)
    if close:
        return {"match": "close", "candidates": close}
    return {"match": "none"}

def extract_medication_reading(transcribed_text):
    m = re.search(r"\*\*MEDICATION\*\*.*?Reading:\s*(.+)", transcribed_text or "", flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return ""
    line = (m.group(1) or "").splitlines()[0]
    return line.strip()[:120]

def read_prescription(image_bytes, user_note="", lang_instruction=""):
    processed = preprocess_prescription(image_bytes)
    b64 = base64.standard_b64encode(processed).decode("utf-8")
    prompt = PRESCRIPTION_PROMPT + ("\n\n" + lang_instruction if lang_instruction else "")
    if user_note and user_note.strip():
        prompt += "\n\nContext note from user: " + user_note.strip()[:200]

    reading = ""
    model_used = "unavailable"
    if CLAUDE_ACTIVE:
        try:
            first = anthropic_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=1600,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                        {"type": "text", "text": prompt},
                    ],
                }],
                temperature=0.2,
            )
            reading = first.content[0].text
            model_used = CLAUDE_MODEL
        except Exception as e:
            print("Prescription reader Claude pass failed:", e)

    if not reading:
        try:
            r = groq_client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": "data:image/jpeg;base64," + b64}},
                    ],
                }],
                temperature=0.2,
                max_tokens=1800,
            )
            reading = r.choices[0].message.content
            model_used = "groq-vision"
        except Exception as e:
            print("Prescription reader Groq pass failed:", e)
            return {"reading": "I could not process this prescription image right now. Please try again with a clearer photo.", "model_used": "error", "overall_confidence": "low"}

    low_conf = "overall confidence**: low" in reading.lower() or "overall confidence: low" in reading.lower()
    if low_conf and CLAUDE_ACTIVE:
        try:
            second = anthropic_client.messages.create(
                model=os.environ.get("CLAUDE_RX_ESCALATION_MODEL", "claude-opus-4-7"),
                max_tokens=2200,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                        {"type": "text", "text": prompt + "\n\nA prior pass was low confidence. Re-check carefully using common Australian prescriptions."},
                    ],
                }],
                temperature=0.2,
            )
            reading = second.content[0].text
            model_used = os.environ.get("CLAUDE_RX_ESCALATION_MODEL", "claude-opus-4-7")
        except Exception as e:
            print("Prescription reader escalation failed:", e)

    med_reading = extract_medication_reading(reading)
    validation = validate_drug_name(med_reading)
    if validation.get("match") == "close":
        reading += "\n\n**Drug name check:** Closest known AU medications: " + ", ".join(validation.get("candidates", []))
    elif validation.get("match") == "exact":
        reading += "\n\n**Drug name check:** Exact match to known AU medication: " + validation.get("drug", "")

    overall = "medium"
    overall_match = re.search(r"\*\*OVERALL CONFIDENCE\*\*\s*:\s*(high|medium|low)", reading, flags=re.IGNORECASE)
    if overall_match:
        overall = overall_match.group(1).lower()

    return {"reading": reading, "model_used": model_used, "overall_confidence": overall, "drug_validation": validation}

def looks_like_prescription_request(text):
    t = (text or "").lower()
    if not t:
        return False
    keys = [
        "prescription", "script", "handwriting", "doctor wrote", "doctor writing",
        "medicine name", "medication name", "can you read this", "what does this say",
        "dose on script", "rx", "chemist",
    ]
    return any(k in t for k in keys)

def encode_image(f):
    img = Image.open(f)
    if img.mode != "RGB":
        img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")

def extract_pdf_text(uploaded_file):
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return ""
    try:
        uploaded_file.seek(0)
        reader = PdfReader(uploaded_file)
        text_parts = []
        for page in reader.pages[:20]:
            try:
                text_parts.append(page.extract_text() or "")
            except Exception:
                continue
        full_text = "\n".join(text_parts).strip()
        if len(full_text) > 8000:
            full_text = full_text[:8000] + "\n\n[Document truncated for length]"
        return full_text
    except Exception as e:
        print("PDF extraction failed:", e)
        return ""

def medichat_pdf_analysis(question, pdf_text, all_messages, lang_instruction=""):
    memory = extract_patient_memory(all_messages)
    memory_context = build_memory_context(memory)

    system = (
        "You are MediChat, a warm and clinically competent AI health companion. "
        "A patient has uploaded a medical PDF report (likely a blood test, lab result, or clinical letter). "
        "Your job is to read the report, identify any abnormal values, explain what they mean in plain language, "
        "and highlight what the patient should pay attention to.\n\n"
        "RULES:\n"
        "- Never use em-dashes or en-dashes. Use commas, semicolons, or periods.\n"
        "- Be warm, conversational, and clear. Avoid jargon. Explain medical terms in plain words.\n"
        "- Highlight ABNORMAL values clearly first, then briefly summarise normal results.\n"
        "- If everything is normal, say so positively.\n"
        "- Suggest specific, actionable next steps based on what's abnormal.\n"
        "- Do not provide a final diagnosis. Use probability language and recommend clinician confirmation.\n"
        "- Do not prescribe medicines, initiate treatment plans, or provide exact dosing changes.\n"
        "- Mention patient by name if known, sparingly.\n"
        "- One brief disclaimer at most. The app shows a permanent disclaimer.\n\n"
        "FORMATTING RULES (CRITICAL for readability):\n"
        "- DO NOT use markdown headings (no #, ##, ###). They render badly in chat bubbles.\n"
        "- For section labels, write a short bold line like **What stands out:** on its own line, then continue on a new line.\n"
        "- Always put a blank line between sections.\n"
        "- Use **bold** sparingly for key values, drug names, or label headers.\n"
        "- Use bullet points (- item) for lists, each on its own line.\n"
        "- Keep paragraphs short, 2-3 sentences max.\n\n"
    )
    if lang_instruction:
        system += lang_instruction + "\n\n"
    if memory_context:
        system += "What this patient has told you in conversation:\n" + memory_context + "\n\n"

    system += "PATIENT'S UPLOADED REPORT:\n" + pdf_text + "\n\n"
    system += "PATIENT'S QUESTION ABOUT THE REPORT:\n" + (question or "Please review this report and tell me what stands out.")

    if CLAUDE_ACTIVE:
        try:
            resp = anthropic_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=1500,
                system=system,
                messages=[{"role": "user", "content": question or "Please review this report and tell me what stands out."}],
                temperature=0.4,
            )
            return resp.content[0].text, "claude"
        except Exception as e:
            print("Claude PDF analysis failed, falling back to Groq:", e)

    msgs = [{"role": "system", "content": system}, {"role": "user", "content": question or "Please review this report and tell me what stands out."}]
    r = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=msgs,
        temperature=0.4,
        max_tokens=1500,
    )
    return r.choices[0].message.content, "groq"

def extract_patient_memory(messages):
    memory = {"symptoms": [], "conditions": [], "medications": []}
    symptom_phrases = ["i have","i feel","i am feeling","i've been feeling","i'm experiencing","i suffer from","my back hurts","my chest","my stomach","i feel pain","feeling dizzy","feeling nauseous","i have a fever","i have a cough","i have a headache","shortness of breath","i've been tired","i feel tired","i have pain","i feel sick"]
    condition_phrases = ["i have diabetes","i have hypertension","i have asthma","i have cancer","i am diabetic","i am hypertensive","i was diagnosed with","i have high blood pressure","i have low blood pressure","i have depression","i have anxiety","i have heart","i have kidney"]
    medication_phrases = ["i am taking","i take","i was prescribed","i'm on","taking medication","prescribed me","i have an inhaler","i take tablets","i take pills"]
    for msg in messages:
        if msg.get("role") == "user" and msg.get("type") == "text":
            content = msg["content"].lower()
            for phrase in symptom_phrases:
                if phrase in content:
                    pos = content.find(phrase)
                    snippet = content[pos:pos+60].strip()
                    if snippet and snippet not in memory["symptoms"]:
                        memory["symptoms"].append(snippet)
                    break
            for phrase in condition_phrases:
                if phrase in content:
                    pos = content.find(phrase)
                    snippet = content[pos:pos+60].strip()
                    if snippet and snippet not in memory["conditions"]:
                        memory["conditions"].append(snippet)
                    break
            for phrase in medication_phrases:
                if phrase in content:
                    pos = content.find(phrase)
                    snippet = content[pos:pos+60].strip()
                    if snippet and snippet not in memory["medications"]:
                        memory["medications"].append(snippet)
                    break
    return memory

def build_memory_context(memory):
    parts = []
    if memory["symptoms"]:
        parts.append("Reported symptoms: " + ", ".join(memory["symptoms"]))
    if memory["conditions"]:
        parts.append("Mentioned conditions: " + ", ".join(memory["conditions"]))
    if memory["medications"]:
        parts.append("Referenced medications: " + ", ".join(memory["medications"]))
    return "\n".join(parts)

EMERGENCY_KEYWORDS = [
    "suicide", "suicidal", "kill myself", "end my life", "want to die", "self harm",
    "can't breathe", "cant breathe", "unable to breathe", "not breathing", "struggling to breathe",
    "chest pain", "crushing chest", "heart attack", "heart racing dangerously",
    "stroke", "having a stroke", "face drooping", "slurred speech",
    "unconscious", "passed out", "fainted", "blacking out", "lose consciousness",
    "severe bleeding", "heavy bleeding", "bleeding a lot", "cannot stop bleeding",
    "overdose", "poisoned", "took too many pills",
    "choking",
    "severe allergic reaction", "anaphylaxis", "throat closing",
    "seizure", "having a seizure", "convulsing",
    "cannot move", "can't move", "paralysed", "paralyzed",
]

EMERGENCY_SYMPTOM_CLUSTERS = [
    {
        "name": "Possible cardiac event",
        "required": [["chest", "chest pressure", "chest tight"]],
        "any_of": [["breath", "breathless", "out of breath"], ["arm pain", "jaw pain", "left arm"], ["sweat", "sweating"], ["nausea", "vomit"]],
        "min_any": 2,
    },
    {
        "name": "Possible stroke",
        "required": [],
        "any_of": [["face droop", "drooping"], ["slurred", "speech"], ["weakness one side", "numb one side"], ["sudden confusion"], ["sudden severe headache"]],
        "min_any": 2,
    },
    {
        "name": "Possible severe allergic reaction",
        "required": [],
        "any_of": [["swelling"], ["throat", "tongue"], ["hives", "rash all over"], ["difficulty breathing", "trouble breathing"]],
        "min_any": 3,
    },
    {
        "name": "Possible syncope/pre-syncope episode",
        "required": [],
        "any_of": [["faint", "faintish", "feel faint"], ["dizzy", "lightheaded"], ["blank", "going blank"], ["nausea"], ["heart racing", "heartbeat raises", "palpitations"], ["breath", "breathless"]],
        "min_any": 3,
    },
    {
        "name": "Possible severe asthma exacerbation",
        "required": [["asthma", "asthmatic"]],
        "any_of": [["can't breathe", "cant breathe", "struggling to breathe"], ["inhaler not working", "inhaler isn't helping"], ["blue lips", "blue fingers"], ["cannot speak", "can't talk"]],
        "min_any": 1,
    },
    {
        "name": "Possible hyperglycemic crisis",
        "required": [["diabetes", "diabetic"]],
        "any_of": [["excessive thirst"], ["frequent urination"], ["confusion"], ["fruity breath"], ["vomiting"], ["extreme fatigue"]],
        "min_any": 3,
    },
]

# Patterns that indicate the message is meta/docs/test rather than a real
# symptom report. Used to suppress emergency + triage detection on pasted
# instructions, code blocks, or app feedback.
META_INDICATORS = [
    "streamlit", "redeploy", "redeploying", "deployment", "deploy",
    "tier 1", "tier 2", "tier 3", "tier 4", "tier 5",
    "test path", "test plan", "test scenario", "test case",
    "feature", "merging", "merged", "commit", "git push", "branch",
    "github", "claude.ai/code", "session_01", "session_state",
    "pull request", " pr #", "rebase",
    "what you'll see", "what you’ll see",
    "shipped to main", "redeploy in",
]

def is_meta_text(text):
    if not text:
        return False
    t = text.lower()
    score = sum(1 for ind in META_INDICATORS if ind in t)
    has_md = ("**" in t) or ("###" in t) or ("```" in t)
    if has_md:
        score += 1
    if len(t) > 1500:
        score += 1
    return score >= 2

def detect_emergency(text, conversation_text=""):
    if is_meta_text(text):
        return False, None
    if not text and not conversation_text:
        return False, None
    combined = (text + " " + conversation_text).lower()
    for kw in EMERGENCY_KEYWORDS:
        if kw in combined:
            return True, "Emergency keyword detected"
    for cluster in EMERGENCY_SYMPTOM_CLUSTERS:
        required_met = True
        for req_group in cluster["required"]:
            if not any(term in combined for term in req_group):
                required_met = False
                break
        if not required_met:
            continue
        matches = 0
        for any_group in cluster["any_of"]:
            if any(term in combined for term in any_group):
                matches += 1
        if matches >= cluster["min_any"]:
            return True, cluster["name"]
    return False, None

# ── Triage tier scoring (Manchester-style 5-tier) ────────────────────
URGENT_KEYWORDS = [
    "severe pain", "worst pain", "intense pain", "unbearable pain",
    "high fever", "fever over 39", "fever 40", "very high temperature",
    "vomiting blood", "blood in stool", "blood in urine", "coughing up blood",
    "cannot keep anything down", "can't keep food down",
    "severe dehydration", "haven't urinated in", "no urine for",
    "severe headache that won't go", "thunderclap headache",
    "vision changes", "double vision", "lost vision",
    "new confusion", "disoriented", "very confused",
    "severe abdominal pain", "rigid abdomen",
    "rapid heartbeat", "racing heart for hours",
    "severe dizziness", "cannot stand",
    "broken bone", "deformed limb", "bone visible",
    "deep cut", "deep wound", "wound won't stop bleeding",
    "burn larger than", "blistering burn",
    "severe asthma attack", "wheezing badly",
    "severe panic", "cannot stop shaking",
    "pregnancy bleeding", "severe pregnancy pain",
]
CONCERN_KEYWORDS = [
    "fever for", "persistent fever", "fever won't go", "fever lasting",
    "headache for days", "persistent headache", "daily headaches",
    "cough for weeks", "persistent cough", "cough lasting",
    "rash spreading", "rash worse", "growing rash",
    "weight loss", "losing weight", "lost weight",
    "tired all the time", "exhausted constantly", "no energy for weeks",
    "trouble sleeping for", "insomnia for",
    "persistent pain", "pain for days", "pain getting worse",
    "frequent urination", "burning urination", "painful urination",
    "diarrhoea for", "diarrhea for", "constipation for",
    "swelling that won't", "lump that's growing", "new lump",
    "bleeding gums often", "easy bruising",
    "new mole", "mole changing", "changing mole",
    "irregular periods", "missed period",
    "anxious all the time", "feeling depressed", "low mood for weeks",
    "joint pain for", "stiff joints",
    "shortness of breath on stairs", "out of breath easily",
]

def assess_triage_tier(text, conversation_text="", memory=None):
    """Return a 5-tier triage assessment with reasons.
    Tier 1: emergency (call 000). Tier 5: self-care.
    Returns Tier 5 (no banner) for clearly meta / non-symptom text.
    """
    text_l = (text or "").lower()
    conv_l = (conversation_text or "").lower()
    combined = (text_l + " " + conv_l).strip()
    memory = memory or {}

    # Meta detection: skip triage when the user's text is clearly app/docs/test
    # rather than a first-person symptom report. Avoids false positives on
    # pasted release notes, code blocks, or meta questions about MediChat itself.
    if is_meta_text(text):
        return {
            "tier": 5, "label": "Self-care appropriate",
            "icon": "⚪", "color": "#64748b",
            "bg": "linear-gradient(135deg,#64748b,#475569)",
            "next_step": "",
            "reasons": [],
        }

    is_emerg, emerg_reason = detect_emergency(text, conversation_text)
    if is_emerg:
        return {
            "tier": 1,
            "label": "Emergency",
            "icon": "🔴",
            "color": "#dc2626",
            "bg": "linear-gradient(135deg,#dc2626,#991b1b)",
            "next_step": "Call 000 (Australia) or your local emergency number now. Do not wait. Do not drive yourself.",
            "reasons": [emerg_reason or "Emergency pattern detected"],
        }

    matched_urgent = [kw for kw in URGENT_KEYWORDS if kw in combined]
    if matched_urgent:
        return {
            "tier": 2,
            "label": "Urgent care today",
            "icon": "🟠",
            "color": "#ea580c",
            "bg": "linear-gradient(135deg,#ea580c,#c2410c)",
            "next_step": "See a GP, urgent care clinic, or ED today, ideally within 2-4 hours.",
            "reasons": matched_urgent[:3],
        }

    matched_concern = [kw for kw in CONCERN_KEYWORDS if kw in combined]
    has_chronic = bool(memory.get("conditions"))
    if matched_concern or (has_chronic and any(s in combined for s in ["worse", "new symptom", "different"])):
        reasons = matched_concern[:3] if matched_concern else ["Chronic condition + new or worsening symptom"]
        return {
            "tier": 3,
            "label": "GP within 24-48 hrs",
            "icon": "🟡",
            "color": "#ca8a04",
            "bg": "linear-gradient(135deg,#ca8a04,#a16207)",
            "next_step": "Book a GP appointment for today or tomorrow. Watch for worsening signs.",
            "reasons": reasons,
        }

    has_symptoms = bool(memory.get("symptoms")) or any(w in combined for w in ["pain", "ache", "sore", "tired", "dizzy", "nausea"])
    if has_symptoms:
        return {
            "tier": 4,
            "label": "GP within a week",
            "icon": "🟢",
            "color": "#16a34a",
            "bg": "linear-gradient(135deg,#16a34a,#15803d)",
            "next_step": "Book a routine GP appointment in the next few days to discuss your symptoms.",
            "reasons": (memory.get("symptoms") or [])[:3] or ["Active symptoms reported"],
        }

    return {
        "tier": 5,
        "label": "Self-care appropriate",
        "icon": "⚪",
        "color": "#64748b",
        "bg": "linear-gradient(135deg,#64748b,#475569)",
        "next_step": "Self-care and monitoring is reasonable. Reach out if anything new or worsening.",
        "reasons": [],
    }

DRUG_INTERACTIONS = {
    "antihistamine": {
        "drugs": ["dramamine", "dimenhydrinate", "bonine", "meclizine", "benadryl", "diphenhydramine", "chlorpheniramine", "promethazine"],
        "conditions": ["asthma", "copd", "glaucoma", "bph", "enlarged prostate"],
        "warning": "Antihistamines can dry respiratory secretions (worsening asthma/COPD), raise intraocular pressure (glaucoma), or cause urinary retention (BPH)."
    },
    "nsaid": {
        "drugs": ["ibuprofen", "advil", "nurofen", "naproxen", "aleve", "aspirin", "diclofenac", "voltaren"],
        "conditions": ["hypertension", "high blood pressure", "kidney", "renal", "ulcer", "gastritis", "asthma", "heart failure", "anticoagulant", "warfarin", "blood thinner"],
        "warning": "NSAIDs can raise blood pressure, worsen kidney function, cause GI bleeding (especially with ulcers or blood thinners), and trigger bronchospasm in some asthmatics."
    },
    "decongestant": {
        "drugs": ["pseudoephedrine", "sudafed", "phenylephrine", "oxymetazoline"],
        "conditions": ["hypertension", "high blood pressure", "heart disease", "arrhythmia", "thyroid", "hyperthyroid", "diabetes", "glaucoma", "bph"],
        "warning": "Decongestants raise blood pressure and heart rate, destabilise cardiac rhythm, worsen hyperthyroidism, and can affect glaucoma or urinary retention."
    },
    "paracetamol_high_dose": {
        "drugs": ["paracetamol", "acetaminophen", "panadol", "tylenol"],
        "conditions": ["liver", "hepatitis", "cirrhosis", "heavy alcohol", "alcoholic"],
        "warning": "High-dose or chronic paracetamol use can cause severe liver damage, especially with pre-existing liver disease or heavy alcohol use."
    },
    "bismuth": {
        "drugs": ["pepto-bismol", "bismuth subsalicylate"],
        "conditions": ["aspirin allergy", "kidney", "renal", "bleeding disorder", "anticoagulant", "warfarin"],
        "warning": "Bismuth subsalicylate contains salicylate (aspirin-like), unsafe with aspirin allergy, kidney disease, or blood thinners."
    },
    "ppi_h2": {
        "drugs": ["omeprazole", "esomeprazole", "pantoprazole", "ranitidine", "famotidine", "zantac"],
        "conditions": ["osteoporosis", "kidney", "c. diff", "magnesium deficiency"],
        "warning": "Long-term PPI use can reduce calcium/magnesium absorption (bone risk), affect kidney function, and increase C. diff infection risk."
    },
}

def check_drug_interactions(response_text, memory):
    if not response_text:
        return []
    text_lower = response_text.lower()
    conditions_text = " ".join(memory.get("conditions", [])).lower() + " " + " ".join(memory.get("medications", [])).lower()
    if not conditions_text.strip():
        return []
    alerts = []
    for class_name, info in DRUG_INTERACTIONS.items():
        drug_hit = next((d for d in info["drugs"] if d in text_lower), None)
        if not drug_hit:
            continue
        condition_hits = [c for c in info["conditions"] if c in conditions_text]
        if condition_hits:
            alerts.append({
                "drug": drug_hit.title(),
                "conditions": condition_hits,
                "warning": info["warning"]
            })
    return alerts

def calculate_confidence(distances):
    if not distances or len(distances) == 0:
        return "low", 0
    avg_dist = sum(distances) / len(distances)
    if avg_dist < 0.8:
        return "high", round((1 - avg_dist / 2) * 100)
    elif avg_dist < 1.3:
        return "medium", round((1 - avg_dist / 2) * 100)
    else:
        return "low", max(20, round((1 - avg_dist / 2.5) * 100))

def get_sources_used(idxs):
    pubmed_split = max(1, len(documents) // 2)
    pubmed_count = sum(1 for i in idxs if i < pubmed_split)
    dialog_count = sum(1 for i in idxs if i >= pubmed_split)
    sources = []
    if pubmed_count > 0:
        sources.append("PubMed Research (" + str(pubmed_count) + ")")
    if dialog_count > 0:
        sources.append("Doctor-Patient Data (" + str(dialog_count) + ")")
    return sources

def sanitize_rag_context(raw_context):
    if not raw_context:
        return ""
    leaked_meds = [
        "subutex", "neurontin", "gabapentin", "remeron", "mirtazapine",
        "zoloft", "sertraline", "klonopin", "clonazepam", "synthroid",
        "levothyroxine", "xanax", "prozac", "lexapro", "wellbutrin",
        "lisinopril", "metformin", "atorvastatin", "amlodipine",
    ]
    text = raw_context
    patterns_to_strip = [
        r"(?i)i['']?m on [^.]*?\.",
        r"(?i)i take [^.]*?\.",
        r"(?i)i am taking [^.]*?\.",
        r"(?i)i have been on [^.]*?\.",
        r"(?i)i was prescribed [^.]*?\.",
        r"(?i)my medications? (are|include)[^.]*?\.",
        r"(?i)currently on [^.]*?\.",
    ]
    for pat in patterns_to_strip:
        text = re.sub(pat, "", text)
    sentences = re.split(r'(?<=[.!?])\s+', text)
    clean = []
    for s in sentences:
        s_lower = s.lower()
        if any(m in s_lower for m in leaked_meds):
            continue
        clean.append(s)
    result = " ".join(clean).strip()
    return result if len(result) > 100 else raw_context

def markdown_to_html(text):
    if not text:
        return ""

    raw = text
    raw = re.sub(r"^#{1,6}\s+(.+?)$", lambda m: "\n\n**" + m.group(1).strip().rstrip(":") + ":**\n", raw, flags=re.MULTILINE)
    raw = re.sub(r"\s*#{2,6}\s+", " ", raw)
    raw = re.sub(r"(?<=[.!?:\)])\s+\*\s+(?=[A-Z])", "\n- ", raw)
    raw = re.sub(r"\*\*\s*\*\s+(?=[A-Z])", "**\n- ", raw)
    raw = re.sub(r"(?<=[a-zA-Z\.])\s+\*\s+(?=[A-Z])", "\n- ", raw)
    raw = re.sub(r"(?<=[.!?])\s+(\*\*[^\*\n]{2,40}:\*\*)", r"\n\n\1\n\n", raw)
    raw = re.sub(r"(?<=[a-zA-Z\.\)])\s+(?=-\s+[A-Z])", r"\n", raw)
    raw = re.sub(r"\n{3,}", "\n\n", raw)

    try:
        import markdown as _md
        html_text = _md.markdown(html.escape(raw, quote=False))
    except ImportError:
        safe = html.escape(raw)
        html_text = "<p>" + safe.replace("\n\n", "</p><p>").replace("\n", "<br>") + "</p>"
        html_text = re.sub(r"\*\*([^\*\n]+)\*\*", r"<strong>\1</strong>", html_text)

    html_text = html_text.replace("<p>", "<p class='md-p'>")
    html_text = html_text.replace("<ul>", "<ul class='md-ul'>")
    html_text = html_text.replace("<ol>", "<ol class='md-ol'>")
    html_text = html_text.replace("<hr>", "<hr class='md-hr'/>")
    html_text = html_text.replace("<hr />", "<hr class='md-hr'/>")
    for level in range(1, 7):
        html_text = re.sub(
            r"<h" + str(level) + r"[^>]*>(.*?)</h" + str(level) + r">",
            r"<p class='md-p'><strong>\1</strong></p>",
            html_text,
            flags=re.DOTALL
        )
    return html_text

def strip_excessive_disclaimers(text):
    if not text:
        return text

    patterns = [
        r"\*?\*?Disclaimer:[^\n]*\*?\*?\s*",
        r"\*?\*?Please note:[^\n]*not a doctor[^\n]*\*?\*?\s*",
        r"Please note that I'?m not a doctor[^.]*\.\s*",
        r"I'?m not a doctor[,.]?[^.]*\.\s*",
        r"(I want to|I'd like to) emphasi[sz]e that I'?m not a doctor[^.]*\.\s*",
        r"(My|This) response is not a substitute for professional medical (advice|diagnosis)[^.]*\.\s*",
        r"This (conversation|response) is (not a substitute|for general information)[^.]*\.\s*",
        r"Always consult (with )?a (qualified )?(healthcare|medical) professional[^.]*\.\s*",
        r"Please consult (with )?a (qualified )?(healthcare|medical) professional[^.]*\.\s*",
        r"It'?s (important|essential|crucial) to (consult|speak with) a (doctor|healthcare professional|medical professional)[^.]*\.\s*",
    ]
    cleaned = text
    for pat in patterns:
        cleaned = re.sub(pat, "", cleaned, flags=re.IGNORECASE)

    cleaned = re.sub(r"\s+[\u2014\u2013]\s+", ", ", cleaned)
    cleaned = cleaned.replace("\u2014", ", ")
    cleaned = cleaned.replace("\u2013", ", ")
    cleaned = re.sub(r",\s*,", ",", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    cleaned = cleaned.strip()
    return cleaned

def medichat_rag_stream(question, all_messages, lang_instruction="", patient_name="", pdf_context="", image_context="", past_chats_summary=""):
    emb = embedder.encode([question]).astype("float32")
    distances, idxs = index.search(emb, k=6)
    raw_context = "\n\n---\n\n".join([documents[i] for i in idxs[0]])
    clean_context = sanitize_rag_context(raw_context)
    sources = get_sources_used(idxs[0])
    confidence_level, confidence_pct = calculate_confidence(distances[0].tolist())
    memory = extract_patient_memory(all_messages)
    memory_context = build_memory_context(memory)

    history = []
    for m in all_messages[-12:]:
        if m.get("type") == "text":
            history.append({"role": m["role"], "content": m["content"]})

    last_user_message = question.lower() if question else ""
    dissatisfaction_signals = [
        "not helping", "isn't helping", "not useful", "doesn't help",
        "you are not helping", "keep repeating", "already said", "same thing",
        "useless", "unhelpful", "rude", "cold",
    ]
    escalation_needed = any(sig in last_user_message for sig in dissatisfaction_signals)
    tone_complaint = any(w in last_user_message for w in ["rude", "cold", "robotic", "unfriendly"])

    system = (
        "You are MediChat — a warm, thoughtful AI health companion. "
        "You care about the person in front of you. You speak like a caring GP who happens to also be a good friend: "
        "kind, genuinely interested, never robotic, never preachy.\n\n"

        "HARD RULES (NEVER BREAK THESE)\n\n"

        "RULE 1 — NEVER FABRICATE PATIENT HISTORY:\n"
        "The <patient_history> block below shows EXACTLY what this patient has told you. "
        "The <reference_knowledge> block is generic medical information — it does NOT describe this patient. "
        "You must NEVER say 'you mentioned', 'you said', 'you told me', or 'you're taking' unless "
        "the thing you're referencing appears LITERALLY in <patient_history> or in the conversation turns below.\n\n"

        "RULE 2 — RED FLAG SCREENING FIRST:\n"
        "For these presentations, screen for danger signs BEFORE general advice:\n"
        "Sudden/severe headache, chest pain, sudden SOB, severe abdominal pain, neuro symptoms.\n"
        "If red flags present: advise emergency care clearly and without hedging.\n\n"

        "RULE 3 — NO REPETITION:\n"
        "Look at your own previous messages in this conversation. Never repeat the same advice twice. "
        "Each response must add something new.\n\n"

        "RULE 4 — ONE DISCLAIMER MAX:\n"
        "The app already shows a disclaimer. Only add 'consult a doctor' phrasing when it's genuinely the most important thing to say. "
        "Do NOT end every response with a disclaimer.\n\n"

        "RULE 5 — NO EM-DASHES OR EN-DASHES:\n"
        "Never use em-dashes (\u2014) or en-dashes (\u2013) in your responses. Use commas, semicolons, colons, periods.\n\n"

        "RULE 6 — NO MARKDOWN HEADINGS:\n"
        "Never use # ## ### markdown headings. If you need a section label, write it as a short bold line: **Section label:** then continue on a new line.\n\n"

        "RULE 7 — NO REPEATED GREETINGS OR RESTATEMENTS:\n"
        "ONLY greet the patient in your VERY FIRST message. On every subsequent turn, jump straight into your answer.\n"
        "NEVER start with 'Hi there', 'Hi [name]', 'Hello', or any greeting after the first turn.\n"
        "NEVER start responses with 'I can see you've uploaded...', 'Looking at your blood work...', 'Based on your report...', 'Thanks for sharing...'\n\n"

        "RULE 8 — MATCH RESPONSE LENGTH TO QUESTION:\n"
        "Short casual questions get short conversational answers (1-3 sentences). "
        "Long detailed questions get structured answers with bold labels.\n\n"

        "RULE 9 — REMEMBER THE CONVERSATION:\n"
        "If the patient already shared something, DO NOT ask them to repeat it.\n\n"

        "RULE 10 — CLINICAL GUARDRAILS:\n"
        "Do not give a final diagnosis. Do not prescribe medicine. Do not provide exact dosage calculations or dose changes. "
        "Use risk-aware language, safety-net advice, and escalation triggers.\n\n"
    )

    if tone_complaint:
        system += (
            "TONE FEEDBACK DETECTED:\n"
            "The patient has told you your tone felt off. "
            "Apologize briefly and genuinely in ONE short sentence, then show warmth.\n\n"
        )

    if escalation_needed:
        system += (
            "ESCALATION TRIGGER:\n"
            "The patient said your previous responses weren't helpful. This response must be noticeably different: "
            "more specific, more committed, with a concrete intervention.\n\n"
        )

    if patient_name:
        system += "Patient's first name: " + patient_name + ". Use naturally and sparingly (max once per response).\n\n"
    if lang_instruction:
        system += lang_instruction + "\n\n"

    system += "<patient_history>\n"
    if memory_context:
        system += "What this patient has told you across all your conversations with them:\n" + memory_context + "\n"
    else:
        system += "This patient has NOT stated any conditions, medications, or chronic illnesses yet. Do not assume any exist.\n"
    system += "</patient_history>\n\n"

    if past_chats_summary:
        system += (
            "<past_conversations>\n"
            "This patient has had earlier separate conversations with you. Use these to recognise references "
            "to 'last time' or 'remember when'. Each entry shows when the chat happened, its main topic, and a short summary.\n"
            + past_chats_summary +
            "\n</past_conversations>\n\n"
            "When the patient asks if you remember a past chat, refer to <past_conversations> by topic naturally. "
            "Do not say 'I have no memory'. You DO have access to the topics and summaries listed above.\n\n"
        )

    if pdf_context:
        system += (
            "<recent_uploaded_report>\n"
            "IMPORTANT: The patient JUST uploaded this medical report and you already analyzed it. "
            "ANY question they ask now is likely a follow-up to this report. "
            "REPORT CONTENT:\n"
            + pdf_context[:6000]
            + "\n</recent_uploaded_report>\n\n"
        )

    if image_context:
        system += (
            "<recent_uploaded_image>\n"
            "IMPORTANT: The patient JUST uploaded a medical image and you already analyzed it visually. "
            "ANY question they ask now is likely a follow-up about that image. "
            "YOUR EARLIER VISUAL ANALYSIS:\n"
            + image_context[:3000]
            + "\n</recent_uploaded_image>\n\n"
        )

    system += (
        "<reference_knowledge>\n"
        "The following is generic medical information retrieved to help you reason about this query. "
        "It is NOT about the current patient.\n\n"
        + clean_context + "\n"
        "</reference_knowledge>\n"
    )

    full_response = ""

    if CLAUDE_ACTIVE:
        try:
            anthropic_messages = history + [{"role": "user", "content": question}]
            normalized = []
            for m in anthropic_messages:
                role = m["role"]
                if role not in ("user", "assistant"):
                    continue
                content = m.get("content", "")
                if normalized and normalized[-1]["role"] == role:
                    normalized[-1]["content"] += "\n\n" + content
                else:
                    normalized.append({"role": role, "content": content})
            if not normalized or normalized[0]["role"] != "user":
                normalized = [{"role": "user", "content": question}]

            with anthropic_client.messages.stream(
                model=CLAUDE_MODEL,
                max_tokens=1024,
                system=system,
                messages=normalized,
                temperature=0.55,
            ) as claude_stream:
                for text in claude_stream.text_stream:
                    if text:
                        full_response += text
                        yield ("chunk", text, full_response)
            yield ("done", full_response, {"memory": memory, "sources": sources, "confidence": confidence_level, "confidence_pct": confidence_pct, "engine": "claude"})
            return
        except Exception as e:
            print("Claude stream failed, falling back to Groq:", e)
            full_response = ""

    if groq_client is None:
        raise RuntimeError(
            "No AI backend is available. Check that GROQ_API_KEY (and optionally ANTHROPIC_API_KEY) "
            "are set in your Streamlit Cloud secrets."
        )

    msgs = [{"role": "system", "content": system}] + history + [{"role": "user", "content": question}]
    try:
        stream = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=msgs,
            temperature=0.55,
            max_tokens=1024,
            stream=True,
            timeout=30.0,
        )
    except Exception as stream_exception:
        import traceback as _tb
        print(f"Groq streaming error: {str(stream_exception)}")
        _tb.print_exc()
        raise RuntimeError(f"Groq API error: {str(stream_exception)}") from stream_exception
    for chunk in stream:
        delta = chunk.choices[0].delta.content or ""
        if delta:
            full_response += delta
            yield ("chunk", delta, full_response)
    yield ("done", full_response, {"memory": memory, "sources": sources, "confidence": confidence_level, "confidence_pct": confidence_pct, "engine": "groq"})

def medichat_vision(question, b64, all_messages, lang_instruction=""):
    memory = extract_patient_memory(all_messages)
    memory_context = build_memory_context(memory)
    prompt = question.strip() if question.strip() else "Please analyse this medical image."
    memory_note = ("\n\nPatient context: " + memory_context) if memory_context else ""
    lang_note = ("\n\n" + lang_instruction) if lang_instruction else ""

    system_text = (
        "You are MediChat, a warm clinical AI companion. The patient has shared a medical image (skin condition, "
        "X-ray, rash, mole, wound, scan, etc.). Analyse it carefully and respond in plain language.\n\n"
        "STRUCTURE YOUR RESPONSE WITH THESE BOLD SECTIONS:\n"
        "**What I see:**\n"
        "(2-3 sentences describing visual findings in plain language)\n\n"
        "**What this could suggest:**\n"
        "(Possible conditions, hedged appropriately.)\n\n"
        "**What you should do:**\n"
        "(Clear next steps.)\n\n"
        "FORMATTING RULES:\n"
        "- Never use em-dashes or en-dashes. Use commas, semicolons, colons.\n"
        "- Never use # ## ### markdown headings.\n"
        "- Use bullet points (- item) on their own lines for lists.\n"
        "- Do not provide a final diagnosis or medication dosage instructions.\n"
        "- One brief disclaimer at most."
        + memory_note + lang_note
    )

    if CLAUDE_ACTIVE:
        try:
            resp = anthropic_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=1500,
                system=system_text,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                        {"type": "text", "text": prompt},
                    ]
                }],
                temperature=0.4,
            )
            return resp.content[0].text, "claude"
        except Exception as e:
            print("Claude vision failed, falling back to Groq:", e)

    r = groq_client.chat.completions.create(
        model="meta-llama/llama-4-scout-17b-16e-instruct",
        messages=[{"role": "user", "content": [
            {"type": "text", "text": system_text + "\n\nQuestion: " + prompt},
            {"type": "image_url", "image_url": {"url": "data:image/jpeg;base64," + b64}}
        ]}],
        temperature=0.5, max_tokens=1024
    )
    return r.choices[0].message.content, "groq-vision"

def clean_text(text):
    replacements = {"\u2019": "'", "\u2018": "'", "\u201c": '"', "\u201d": '"', "\u2013": "-", "\u2014": "-", "\u2022": "-"}
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text.encode("latin-1", errors="replace").decode("latin-1")

def generate_doctor_visit_summary(messages, patient_name=""):
    if not messages:
        return None, None

    transcript_parts = []
    for msg in messages:
        role = msg.get("role", "")
        content = msg.get("content", "")
        if not content:
            continue
        if role == "user":
            transcript_parts.append("Patient said: " + content)
        elif role == "assistant":
            transcript_parts.append("MediChat replied: " + content)
    transcript = "\n\n".join(transcript_parts)

    name_line = ("Patient name: " + patient_name + "\n") if patient_name and patient_name != "Guest" else ""

    system_prompt = (
        "You are a medical scribe assistant. The patient had a conversation with an AI health companion. "
        "Your job is to produce a CONCISE, STRUCTURED summary the patient can hand to their actual GP at their next appointment. "
        "Use ONLY information from the conversation. Do NOT invent symptoms, dates, or details.\n\n"
        "OUTPUT FORMAT (use these exact section labels with bold markdown):\n\n"
        "**Patient overview:**\n"
        "(One sentence describing who this is and what brought them to the chat)\n\n"
        "**Symptoms reported:**\n"
        "- (each symptom on its own bullet)\n\n"
        "**Duration and pattern:**\n"
        "(When symptoms started, how often, what triggers them. If unknown, say 'not discussed'.)\n\n"
        "**Relevant medical history mentioned:**\n"
        "- (each condition or medication on its own bullet, or 'None mentioned' if not discussed)\n\n"
        "**Concerns raised by patient:**\n"
        "(What the patient is worried about, in their own framing)\n\n"
        "**MediChat's preliminary assessment:**\n"
        "(Brief summary of what the AI suggested, hedged appropriately)\n\n"
        "**Suggested questions for the GP:**\n"
        "- (3-5 specific questions the patient should ask their doctor)\n\n"
        "RULES:\n"
        "- Never use em-dashes or en-dashes. Use commas, semicolons, colons.\n"
        "- Never use # ## ### markdown headings.\n"
        "- Be factual. No fluff. No greetings. No sign-offs.\n"
    )

    user_prompt = (
        name_line +
        "Generate a doctor visit prep summary from this conversation:\n\n" +
        transcript[:8000]
    )

    summary_text = ""
    if CLAUDE_ACTIVE:
        try:
            resp = anthropic_client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=1500,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}],
                temperature=0.3,
            )
            summary_text = resp.content[0].text
        except Exception as e:
            print("Claude summary failed, falling back to Groq:", e)

    if not summary_text:
        try:
            r = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                temperature=0.3,
                max_tokens=1500,
            )
            summary_text = r.choices[0].message.content
        except Exception as e:
            print("Groq summary failed:", e)
            return None, None

    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.set_fill_color(33, 118, 174)
    pdf.rect(0, 0, 210, 38, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_y(8)
    pdf.cell(0, 10, "Doctor Visit Summary", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7, "Prepared by MediChat for: " + (patient_name if patient_name and patient_name != "Guest" else "Patient"), ln=True, align="C")
    pdf.set_y(28)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, "Generated on " + datetime.now().strftime("%B %d, %Y at %I:%M %p"), ln=True, align="C")

    pdf.set_y(46)

    pdf.set_fill_color(237, 246, 252)
    pdf.set_text_color(20, 66, 114)

    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5, "Bring this summary to your GP appointment. It captures what you discussed with MediChat, including symptoms, history, and questions to ask. This is NOT a diagnosis. Your GP will make the clinical judgement.", fill=True, border=0)
    pdf.ln(4)

    pdf.set_text_color(40, 40, 40)
    cleaned_summary = clean_text(summary_text)

    lines = cleaned_summary.split("\n")
    for line in lines:
        stripped = line.strip()
        if not stripped:
            pdf.ln(2)
            continue
        bold_match = re.match(r"^\*\*(.+?)\*\*:?$", stripped)
        if bold_match:
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(30, 58, 54)
            pdf.cell(0, 6, bold_match.group(1).rstrip(":") + ":", ln=True)
            pdf.set_text_color(40, 40, 40)
            continue
        bullet_match = re.match(r"^[\-\*]\s+(.+)$", stripped)
        if bullet_match:
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(25)
            pdf.cell(5, 5, "-", ln=False)
            pdf.multi_cell(0, 5, bullet_match.group(1))
            continue
        pdf.set_font("Helvetica", "", 10)
        plain = re.sub(r"\*\*([^\*]+)\*\*", r"\1", stripped)
        pdf.multi_cell(0, 5, plain)

    pdf.ln(8)
    pdf.set_draw_color(168, 197, 189)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(3)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 4, "MediChat is a research prototype, not a medical device. This summary is generated by Ai from a chat conversation and may contain errors or omissions. Always rely on your qualified healthcare professional for diagnosis and treatment decisions.")
    pdf.ln(2)
    pdf.cell(0, 4, APP_VERSION_LABEL, ln=True, align="C")

    pdf_bytes = bytes(pdf.output(dest="S"))
    return pdf_bytes, summary_text

def generate_chat_pdf(messages):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)
    pdf.set_fill_color(33, 118, 174)
    pdf.rect(0, 0, 210, 35, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_y(8)
    pdf.cell(0, 10, "MediChat - Conversation Export", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7, "Generated: " + datetime.now().strftime("%B %d, %Y at %I:%M %p"), ln=True, align="C")
    pdf.set_y(43)
    pdf.set_text_color(146, 64, 14)
    pdf.set_fill_color(255, 251, 235)
    pdf.set_font("Helvetica", "", 8)
    pdf.multi_cell(0, 5, "DISCLAIMER: This conversation is for informational purposes only. Always consult a qualified healthcare professional.", fill=True)
    pdf.ln(4)
    for msg in messages:
        role = msg.get("role", "")
        content = msg.get("content", "")
        msg_type = msg.get("type", "text")
        if not content:
            continue
        if role == "user":
            pdf.set_fill_color(42, 143, 197)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 7, "  You", ln=True, fill=True)
            pdf.set_fill_color(237, 246, 252)
            pdf.set_text_color(12, 45, 72)

            pdf.set_font("Helvetica", "", 9)
            display = "[Medical image uploaded]" + (" - " + clean_text(content) if content else "") if msg_type == "image" else clean_text(content)
            pdf.multi_cell(0, 6, "  " + display, fill=True)
        else:
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 7, "  MediChat", ln=True, fill=True)
            pdf.set_fill_color(248, 250, 252)
            pdf.set_text_color(51, 65, 85)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 6, "  " + clean_text(content), fill=True)
        pdf.ln(3)
    pdf.set_y(-18)
    pdf.set_text_color(148, 163, 184)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(0, 5, APP_VERSION_LABEL, align="C")
    return bytes(pdf.output())

def generate_assessment_pdf(parsed, data, report_date):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)
    pdf.set_fill_color(33, 118, 174)
    pdf.rect(0, 0, 210, 40, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_y(7)
    pdf.cell(0, 10, "MediChat Assessment Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, "Pre-Consultation Health Summary", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7, "Generated: " + report_date, ln=True, align="C")
    pdf.set_y(48)
    urgency = parsed.get("urgency", "See a doctor soon")
    urgency_lower = urgency.lower()
    if "emergency" in urgency_lower or "now" in urgency_lower:
        pdf.set_fill_color(254, 226, 226)
        pdf.set_text_color(153, 27, 27)
    elif "urgent" in urgency_lower or "today" in urgency_lower:
        pdf.set_fill_color(255, 251, 235)
        pdf.set_text_color(146, 64, 14)
    else:
        pdf.set_fill_color(240, 253, 244)
        pdf.set_text_color(22, 101, 52)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 10, "URGENCY: " + clean_text(urgency), ln=True, fill=True, align="C")
    pdf.ln(5)

    def section_header(title):
        pdf.set_fill_color(33, 118, 174)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, "  " + title, ln=True, fill=True)
        pdf.ln(2)


    def info_row(label, value):
        pdf.set_fill_color(248, 250, 252)
        pdf.set_text_color(51, 65, 85)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(50, 7, "  " + label + ":", fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 7, " " + clean_text(str(value)), ln=True, fill=True)
        pdf.ln(1)

    section_header("SYMPTOMS REPORTED")
    info_row("Main Symptom", data.get("main_symptom", ""))
    info_row("Duration", data.get("duration", ""))
    info_row("Severity", data.get("severity", ""))
    info_row("Pattern", data.get("pattern", ""))
    other = data.get("other_symptoms", "")
    if other and other.lower() not in ["no", "none", "n/a", "no other symptoms"]:
        info_row("Other Symptoms", other)
    info_row("Known Conditions", data.get("known_conditions", "Not specified"))
    info_row("Current Medications", data.get("current_medications", "Not specified"))
    info_row("Red-flag Symptoms", data.get("red_flags", "Not specified"))
    info_row("Age Group", data.get("age", ""))
    info_row("Biological Sex", data.get("gender", ""))
    pdf.ln(4)
    section_header("POSSIBLE CONDITIONS")
    pdf.set_text_color(51, 65, 85)
    pdf.set_font("Helvetica", "", 9)
    for c in parsed.get("conditions", []):
        if c.strip():
            pdf.set_fill_color(240, 253, 250)
            pdf.cell(8, 7, "", fill=True)
            pdf.cell(0, 7, "- " + clean_text(c.strip()), ln=True)
            pdf.ln(1)
    pdf.ln(3)
    section_header("WHAT TO DO NEXT")
    pdf.set_text_color(51, 65, 85)
    pdf.set_font("Helvetica", "", 9)
    for i, s in enumerate(parsed.get("next_steps", []), 1):
        if s.strip():
            pdf.set_fill_color(248, 250, 252)
            pdf.multi_cell(0, 7, str(i) + ". " + clean_text(s.strip()), fill=True)
            pdf.ln(1)
    pdf.ln(3)
    section_header("MEDICHAT SUMMARY")
    summary = parsed.get("summary", "")
    if summary:
        pdf.set_fill_color(240, 253, 250)
        pdf.set_text_color(19, 78, 74)
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(0, 7, clean_text(summary), fill=True)
    pdf.ln(5)
    pdf.set_fill_color(255, 251, 235)
    pdf.set_text_color(146, 64, 14)
    pdf.set_font("Helvetica", "B", 8)
    pdf.multi_cell(0, 6, "IMPORTANT DISCLAIMER: This report was generated by MediChat Ai and does NOT constitute a medical diagnosis. Please share this with your doctor for professional evaluation and treatment.", fill=True)
    pdf.set_y(-18)
    pdf.set_text_color(148, 163, 184)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(0, 5, APP_VERSION_LABEL, align="C")
    return bytes(pdf.output())

ASSESSMENT_STAGES = [
    {"key": "main_symptom", "question": "What is your main symptom or health concern today?", "hint": "Please describe clearly e.g. chest pain, headache, fever, shortness of breath, dizziness...", "options": ["Chest pain", "Headache", "Fever", "Cough", "Dizziness", "Stomach pain", "Fatigue", "Other (type below)"]},
    {"key": "duration", "question": "How long have you been experiencing this?", "hint": "", "options": ["Just started today", "A few days", "About a week", "More than 2 weeks", "Over a month"]},
    {"key": "severity", "question": "How severe is it on a scale of 1 to 10? (1 = mild, 10 = unbearable)", "hint": "", "options": ["1-2 (Mild)", "3-4 (Moderate)", "5-6 (Significant)", "7-8 (Severe)", "9-10 (Unbearable)"]},
    {"key": "pattern", "question": "Is it constant or does it come and go?", "hint": "", "options": ["Constant, always there", "Comes and goes", "Getting worse over time", "Getting better", "Only happens sometimes"]},
    {"key": "other_symptoms", "question": "Are you experiencing any other symptoms alongside this?", "hint": "e.g. nausea, dizziness, fever, fatigue... or select None", "options": ["No other symptoms", "Nausea or vomiting", "Fever or chills", "Dizziness", "Fatigue or weakness", "Other (type below)"]},
    {"key": "known_conditions", "question": "Do you have any known medical conditions relevant to this symptom?", "hint": "This helps make triage safer and more accurate.", "options": ["No known conditions", "Diabetes", "Asthma", "High blood pressure", "Heart condition", "Pregnancy", "Other (type below)"]},
    {"key": "current_medications", "question": "Are you currently taking any medications?", "hint": "List known meds if relevant. You can type specific names below.", "options": ["No current medications", "Yes, regular medications", "Yes, as-needed medications", "Not sure", "Other (type below)"]},
    {"key": "red_flags", "question": "Any red-flag symptoms right now?", "hint": "Select any urgent danger signs if present.", "options": ["No red flags", "Chest pain or pressure", "Trouble breathing", "Fainting or confusion", "Severe bleeding", "One-sided weakness or slurred speech", "Other (type below)"]},
    {"key": "age", "question": "How old are you?", "hint": "", "options": ["Under 18", "18-30", "31-45", "46-60", "61-75", "Over 75"]},
    {"key": "gender", "question": "What is your biological sex? (helps with medical accuracy)", "hint": "", "options": ["Male", "Female", "Prefer not to say"]},
]

def generate_assessment_report(assessment_data, lang_instruction=""):
    emb = embedder.encode([assessment_data.get("main_symptom", "")]).astype("float32")
    _, idxs = index.search(emb, k=5)
    context = "\n\n---\n\n".join([documents[i] for i in idxs[0]])
    lang_note = ("\n" + lang_instruction) if lang_instruction else ""
    prompt = (
        "You are an experienced clinical AI assistant.\n\n"
        "Patient symptom assessment data:\n"
        "- Main symptom: " + assessment_data.get("main_symptom", "Not specified") + "\n"
        "- Duration: " + assessment_data.get("duration", "Not specified") + "\n"
        "- Severity: " + assessment_data.get("severity", "Not specified") + "\n"
        "- Pattern: " + assessment_data.get("pattern", "Not specified") + "\n"
        "- Other symptoms: " + assessment_data.get("other_symptoms", "None") + "\n"
        "- Known conditions: " + assessment_data.get("known_conditions", "Not specified") + "\n"
        "- Current medications: " + assessment_data.get("current_medications", "Not specified") + "\n"
        "- Red-flag symptoms: " + assessment_data.get("red_flags", "Not specified") + "\n"
        "- Age group: " + assessment_data.get("age", "Not specified") + "\n"
        "- Biological sex: " + assessment_data.get("gender", "Not specified") + "\n\n"
        "INSTRUCTIONS:\n"
        "1. If the main symptom looks like a typo, interpret the most likely intended symptom.\n"
        "2. Consider all symptoms holistically.\n"
        "3. Provide realistic, evidence-based possible conditions.\n"
        "4. Urgency must reflect severity, red-flag symptoms, age, and comorbidity risk.\n"
        "5. If any red flags are present, escalate urgency and explicitly recommend urgent or emergency care.\n"
        "6. Do not overdiagnose, use cautious probability language.\n"
        + lang_note + "\n\n"
        "Medical research context:\n" + context + "\n\n"
        "Respond in EXACTLY this format:\n"
        "URGENCY: [one of: Self-care at home / See a doctor soon / Seek urgent care today / Go to emergency NOW]\n"
        "CONDITIONS: [condition 1] | [condition 2] | [condition 3]\n"
        "NEXT STEPS: [step 1] | [step 2] | [step 3]\n"
        "SUMMARY: [2-3 warm, clear sentences summarising the assessment]\n"
        "SAFETY: [one line with emergency fallback if symptoms worsen]\n"
    )
    r = groq_client.chat.completions.create(model="llama-3.3-70b-versatile", messages=[{"role": "user", "content": prompt}], temperature=0.3, max_tokens=1024)
    return r.choices[0].message.content

def parse_report(report_text):
    parsed = {"urgency": "", "conditions": [], "next_steps": [], "summary": "", "safety": ""}
    for line in report_text.strip().split("\n"):
        line = line.strip()
        if line.startswith("URGENCY:"):
            parsed["urgency"] = line.replace("URGENCY:", "").strip()
        elif line.startswith("CONDITIONS:"):
            parsed["conditions"] = [c.strip() for c in line.replace("CONDITIONS:", "").split("|") if c.strip()]
        elif line.startswith("NEXT STEPS:"):
            parsed["next_steps"] = [s.strip() for s in line.replace("NEXT STEPS:", "").split("|") if s.strip()]
        elif line.startswith("SUMMARY:"):
            parsed["summary"] = line.replace("SUMMARY:", "").strip()
        elif line.startswith("SAFETY:"):
            parsed["safety"] = line.replace("SAFETY:", "").strip()
if "session_started" not in st.session_state:
    st.session_state.session_started = True
    st.session_state.messages = []
    st.session_state.qcount = 0
    st.session_state.feedback = {}
    st.session_state.nav_clicked = False
    st.session_state.patient_memory = {"symptoms": [], "conditions": [], "medications": []}
    st.session_state.uploader_key = 0
    st.session_state.last_pdf_context = ""
    st.session_state.last_pdf_name = ""
    st.session_state.last_image_context = ""
    st.session_state.mode = "chat"
    st.session_state.assessment_stage = 0
    st.session_state.assessment_data = {}
    st.session_state.assessment_complete = False
    st.session_state.assessment_report = None
    st.session_state.assessment_parsed = None
    st.session_state.selected_language = "English"
    st.session_state.patient_name = ""
    st.session_state.emergency_detected = False
    st.session_state.emergency_reason = ""
    st.session_state.triage_assessment = None
    st.session_state.last_sources = []
    st.session_state.eval_log = []
    st.session_state.response_times = []
    st.session_state.admin_authenticated = False
    st.session_state.admin_attempt_failed = False
    st.session_state.chat_input_key = 0
    st.session_state.is_authenticated = False
    st.session_state.is_guest = False
    st.session_state.user_email_hash = ""
    st.session_state.user_email_display = ""
    st.session_state.auth_error = ""
    st.session_state.auth_view = "choose"
    st.session_state.current_conversation_id = ""
    st.session_state.pending_user_input = ""
    st.session_state.rx_reader_result = None
    st.session_state.rx_uploader_key = 0
    st.session_state.home_show_vision_upload = False
    st.session_state.home_show_voice = False
    st.session_state.voice_audio_key = 0


with st.sidebar:
    _brand_logo_uri = get_brand_logo_data_uri()
    if _brand_logo_uri:
        st.markdown(
            '<div class="md-logo-wrap md-logo-image-wrap">'
            '<img class="md-logo-image" src="' + _brand_logo_uri + '" alt="' + ui_escape(APP_TITLE) + '"/>'
            '</div>',
            unsafe_allow_html=True
        )
    else:
        # Fallback if the asset is missing — keep the Material Symbol mark + text.
        _logo_block = '<div class="md-logo-mark"><span class="material-symbols-rounded">medical_services</span></div>'
        st.markdown(
            '<div class="md-logo-wrap">'
            + _logo_block +
            '<div>'
            '<div class="md-logo-text">' + APP_TITLE + '</div>'
            '<div class="md-logo-sub">' + APP_SUBTITLE + '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True
        )

    # ── Primary nav ─────────────────────────────────────────────
    _mode = st.session_state.mode
    nav_items = [
        ("home", "Home", "chat", ":material/home:"),
        ("overview", "Health Overview", "overview", ":material/monitoring:"),
        ("symptom", "Symptoms Checker", "assessment", ":material/stethoscope:"),
        ("prescription", "Prescription Reader", "rx_reader", ":material/prescriptions:"),
        ("records", "Health Records", "records", ":material/lab_profile:"),
        ("meds", "Medications", "medications", ":material/pill:"),
        ("insights", "Ai Insights", "insights", ":material/auto_awesome:"),
        ("appts", "Appointments", "appointments", ":material/calendar_month:"),
    ]
    _nav_clicked = st.session_state.get("nav_clicked", False)
    for nav_key, nav_label, target_mode, nav_icon in nav_items:
        # Home and New Chat both target chat mode; differentiate by message presence
        # so exactly one is active at a time. Both also gated on nav_clicked so that
        # nothing is highlighted on a fresh load — the active state appears only
        # after the user explicitly navigates.
        if nav_key == "home":
            is_active = _nav_clicked and (_mode == "chat") and not st.session_state.messages
        elif nav_key == "new":
            is_active = _nav_clicked and (_mode == "chat") and bool(st.session_state.messages)
        else:
            is_active = (target_mode == _mode)
        _btn_type = "primary" if is_active else "secondary"
        active_cls = "md-nav-active" if is_active else ""
        st.markdown('<div class="' + active_cls + '">', unsafe_allow_html=True)
        if st.button(nav_label, key="nav_" + nav_key, use_container_width=True, icon=nav_icon, type=_btn_type):
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.session_state.nav_clicked = True
            if nav_key == "new":
                start_new_chat_session()
            else:
                st.session_state.mode = target_mode
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # Divider sits between Appointments (last nav item) and the profile chip.
    st.markdown("---")

    # ── Profile chip — sits where the Premium card used to. Live-updates with auth state. ──
    if st.session_state.is_authenticated:
        _profile_em = st.session_state.user_email_display or ""
        _saved_name = (st.session_state.patient_name or "").strip()
        if _saved_name and _saved_name.lower() != "guest":
            _profile_nm = _saved_name
        elif _profile_em and "@" in _profile_em:
            _profile_nm = _profile_em.split("@", 1)[0].replace(".", " ").replace("_", " ").title() or "Patient"
        else:
            _profile_nm = "Patient"
        _profile_in = (_profile_nm[0] if _profile_nm and _profile_nm != "Patient" else (_profile_em[0] if _profile_em else "P")).upper()
        _profile_sub = _profile_em if _profile_em else "View profile"
    else:
        _profile_nm = "Guest"
        _profile_in = "G"
        _profile_sub = "Sign in to save your data"
    # Inject the custom CSS overrides specifically for styling the native Streamlit buttons in the sidebar
    st.markdown("""
<style>
/* Hide empty anchor containers to reclaim space */
div.element-container:has(.md-recent-card-inside),
div.element-container:has(.md-recent-seeall-anchor),
div.element-container:has(.md-new-chat-anchor),
div.element-container:has(.md-conv-select-anchor),
div.element-container:has(.md-conv-del-anchor),
div.element-container:has(.md-side-signout-anchor),
div.element-container:has(.md-privacy-btn-anchor),
div.element-container:has(.md-help-btn-anchor) {
    display: none !important;
}

/* Custom sidebar native button overrides */
div[data-testid="stVerticalBlockBorderWrapper"]:has(.md-recent-card-inside) {
    background: #ffffff !important;
    border: 1px solid #e6edf9 !important;
    border-radius: 11px !important;
    padding: 0.45rem 0.5rem 0.4rem 0.5rem !important;
    margin-top: 0.7rem !important;
    margin-bottom: 0.25rem !important;
    box-shadow: 0 8px 24px rgba(15, 23, 42, 0.04), 0 2px 6px rgba(15, 23, 42, 0.02) !important;
    box-sizing: border-box !important;
    position: relative !important;
    overflow: visible !important;
}
div[data-testid="stVerticalBlockBorderWrapper"]:has(.md-recent-card-inside) [data-testid="stVerticalBlock"] {
    gap: 0.2rem !important;
}

/* Sign out button */
.element-container:has(.md-side-signout-anchor) + .element-container button {
    position: absolute !important;
    top: -55px !important; /* Pull up to align with the top profile card */
    right: 12px !important;
    width: 28px !important;
    height: 28px !important;
    min-width: 28px !important;
    min-height: 28px !important;
    padding: 0 !important;
    border-radius: 50% !important;
    background: transparent !important;
    border: 1px solid transparent !important;
    color: #94a3b8 !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    z-index: 10 !important;
    box-shadow: none !important;
}
.element-container:has(.md-side-signout-anchor) + .element-container button:hover {
    background: #fef2f2 !important;
    border-color: #fecaca !important;
    color: #b91c1c !important;
}

/* See all button in card header */
.element-container:has(.md-recent-seeall-anchor) + .element-container button {
    position: absolute !important;
    top: 8px !important; /* Align with Recent Chats header title */
    right: 8px !important;
    background: transparent !important;
    border: none !important;
    color: #4f46e5 !important;
    font-size: 0.74rem !important;
    font-weight: 600 !important;
    padding: 0 !important;
    margin: 0 !important;
    box-shadow: none !important;
    min-height: unset !important;
    height: auto !important;
    width: auto !important;
    z-index: 5 !important;
    display: inline-block !important;
}
.element-container:has(.md-recent-seeall-anchor) + .element-container button:hover {
    color: #3730a3 !important;
    text-decoration: underline !important;
    background: transparent !important;
}

/* + New chat button inside the card */
.element-container:has(.md-new-chat-anchor) + .element-container button {
    box-sizing: border-box !important;
    width: 100% !important;
    min-height: 30px !important;
    height: 30px !important;
    padding: 0 0.6rem !important;
    margin-bottom: 0.35rem !important;
    font-size: 0.72rem !important;
    border-radius: 8px !important;
    border: 1px solid #dbe4ff !important;
    background: linear-gradient(135deg, #f5f8ff 0%, #eef2ff 100%) !important;
    color: #4f46e5 !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 5px rgba(99, 102, 241, 0.05) !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 0.35rem !important;
    transition: all 0.2s ease !important;
}
.element-container:has(.md-new-chat-anchor) + .element-container button:hover {
    background: linear-gradient(135deg, #eef2ff 0%, #e0e7ff 100%) !important;
    border-color: #c7d2fe !important;
    color: #3730a3 !important;
    box-shadow: 0 4px 12px rgba(99, 102, 241, 0.12) !important;
}
.element-container:has(.md-new-chat-anchor) + .element-container button p {
    font-size: 0.72rem !important;
    color: #4f46e5 !important;
    margin: 0 !important;
    font-weight: 600 !important;
    line-height: 1 !important;
}
.element-container:has(.md-new-chat-anchor) + .element-container button [data-testid="stIconMaterial"] {
    font-size: 0.88rem !important;
    color: #4f46e5 !important;
    -webkit-text-fill-color: #4f46e5 !important;
}

/* Conversation row buttons */
.element-container:has(.md-conv-select-anchor) + .element-container button {
    background: transparent !important;
    border: none !important;
    border-radius: 8px !important;
    color: #475569 !important;
    text-align: left !important;
    font-size: 0.76rem !important;
    font-weight: 550 !important;
    padding: 0.4rem 0.45rem !important;
    min-height: unset !important;
    height: auto !important;
    width: 100% !important;
    box-shadow: none !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    gap: 0.35rem !important;
    transition: all 0.15s ease !important;
}
.element-container:has(.md-conv-select-anchor) + .element-container button:hover {
    background: #f8fafc !important;
    color: #0f172a !important;
}
.element-container:has(.md-conv-select-anchor.md-active) + .element-container button {
    background: #f1f5f9 !important;
    color: #0f172a !important;
    font-weight: 600 !important;
}
.element-container:has(.md-conv-select-anchor) + .element-container button p {
    margin: 0 !important;
    font-size: 0.76rem !important;
    line-height: 1.15 !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
.element-container:has(.md-conv-select-anchor) + .element-container button [data-testid="stIconMaterial"] {
    font-size: 0.95rem !important;
    color: #94a3b8 !important;
    -webkit-text-fill-color: #94a3b8 !important;
}

/* Flex layout columns for conversation rows */
div[data-testid="stHorizontalBlock"]:has(.md-conv-select-anchor) {
    display: flex !important;
    flex-direction: row !important;
    align-items: center !important;
    justify-content: space-between !important;
    width: 100% !important;
    gap: 0.1rem !important;
    margin-bottom: 0.1rem !important;
}
div[data-testid="stHorizontalBlock"]:has(.md-conv-select-anchor) div[data-testid="column"] {
    padding: 0 !important;
    margin: 0 !important;
    min-width: unset !important;
}

/* Style delete button */
.element-container:has(.md-conv-del-anchor) + .element-container button {
    background: transparent !important;
    border: none !important;
    color: #94a3b8 !important;
    font-size: 0.74rem !important;
    padding: 0 !important;
    margin: 0 !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 20px !important;
    height: 20px !important;
    min-width: 20px !important;
    min-height: 20px !important;
    box-shadow: none !important;
    border-radius: 50% !important;
    transition: background 0.15s ease, color 0.15s ease !important;
}
.element-container:has(.md-conv-del-anchor) + .element-container button:hover {
    background: #fef2f2 !important;
    color: #b91c1c !important;
}

/* Privacy & Terms and Help Center footer buttons */
.element-container:has(.md-privacy-btn-anchor) + .element-container button,
.element-container:has(.md-help-btn-anchor) + .element-container button {
    background: transparent !important;
    border: none !important;
    color: #64748b !important;
    font-size: 0.65rem !important;
    font-weight: 600 !important;
    padding: 0 !important;
    margin: 0 !important;
    min-height: unset !important;
    height: auto !important;
    box-shadow: none !important;
}
.element-container:has(.md-privacy-btn-anchor) + .element-container button:hover,
.element-container:has(.md-help-btn-anchor) + .element-container button:hover {
    color: #4f46e5 !important;
    text-decoration: underline !important;
    background: transparent !important;
}
.element-container:has(.md-privacy-btn-anchor) + .element-container button p,
.element-container:has(.md-help-btn-anchor) + .element-container button p {
    font-size: 0.65rem !important;
    margin: 0 !important;
}
.element-container:has(.md-privacy-btn-anchor) + .element-container button {
    float: right !important;
}
.element-container:has(.md-help-btn-anchor) + .element-container button {
    float: left !important;
}
</style>
""", unsafe_allow_html=True)

    # Profile chip: avatar + name/email + "Synced & up to date" status.
    _sync_dot = '<span class="md-status-dot"></span>Synced &amp; up to date' if st.session_state.is_authenticated else '<span class="md-status-dot md-status-dot-off"></span>Guest mode'
    st.markdown(
        '<div class="md-side-profile md-side-profile-top">'
        '<div class="md-side-avatar">' + ui_escape(_profile_in) + '</div>'
        '<div class="md-side-profile-text">'
        '<div class="md-side-pname">' + ui_text(_profile_nm, 30) + '</div>'
        '<div class="md-side-psub">' + ui_text(_profile_sub, 40) + '</div>'
        '<div class="md-side-status">' + _sync_dot + '</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # Sign-out icon button (overlaid absolutely on top of profile card top-right)
    if st.session_state.is_authenticated:
        st.markdown('<div class="md-side-signout-anchor"></div>', unsafe_allow_html=True)
        if st.button(" ", key="profile_logout", icon=":material/logout:"):
            for k in ["is_authenticated", "is_guest", "user_email_hash", "user_email_display", "patient_name", "patient_memory", "messages", "qcount", "feedback", "last_sources", "last_pdf_context", "last_image_context", "rx_reader_result", "rx_uploader_key"]:
                if k in st.session_state:
                    if k in ("is_authenticated", "is_guest"):
                        st.session_state[k] = False
                    elif k == "patient_memory":
                        st.session_state[k] = {"symptoms": [], "conditions": [], "medications": []}
                    elif k == "messages":
                        st.session_state[k] = []
                    elif k == "rx_reader_result":
                        st.session_state[k] = None
                    elif k == "rx_uploader_key":
                        st.session_state[k] = 0
                    elif k == "qcount":
                        st.session_state[k] = 0
                    elif k == "feedback":
                        st.session_state[k] = {}
                    else:
                        st.session_state[k] = "" if isinstance(st.session_state[k], str) else st.session_state[k]
            st.session_state.current_conversation_id = ""
            st.rerun()

    # Spacing div to prevent clipping/overlapping of Recent Chats card header under profile chip
    st.markdown('<div style="height: 14px;"></div>', unsafe_allow_html=True)

    if st.session_state.is_authenticated:
        # == Recent Chats card (wrapped in container so grandparent :has targets it) ==
        _convs = list_conversations(st.session_state.user_email_hash, limit=3)
        _active_id = st.session_state.current_conversation_id

        with st.container(border=True):
            # Class markup to identify this st.container wrapper block
            st.markdown('<div class="md-recent-card-inside"></div>', unsafe_allow_html=True)

            # Card header title
            st.markdown('<div class="md-recent-head"><div class="md-recent-title">Recent Chats</div></div>', unsafe_allow_html=True)
            
            # Card header "See all" button
            st.markdown('<div class="md-recent-seeall-anchor"></div>', unsafe_allow_html=True)
            if st.button("See all", key="recent_see_all_btn"):
                st.session_state.mode = "history"
                st.rerun()

            # "+ New chat" button inside card
            st.markdown('<div class="md-new-chat-anchor"></div>', unsafe_allow_html=True)
            if st.button("＋ New chat", key="new_chat_btn", use_container_width=True):
                start_new_chat_session()
                st.rerun()

            # Conversation rows
            if _convs:
                for _c in _convs:
                    _is_active = _c["id"] == _active_id
                    _title = (_c.get("title") or "Chat")[:40]
                    _lu = _c.get("last_updated")
                    _ago = ""
                    try:
                        if _lu and hasattr(_lu, "strftime"):
                            _delta = datetime.utcnow() - (_lu.replace(tzinfo=None) if _lu.tzinfo else _lu)
                            _h = int(_delta.total_seconds() // 3600)
                            if _h < 1:
                                _ago = str(max(1, int(_delta.total_seconds() // 60))) + "m ago"
                            elif _h < 24:
                                _ago = str(_h) + "h ago"
                            elif _h < 168:
                                _ago = str(_h // 24) + "d ago"
                            else:
                                _ago = str(_h // 168) + "w ago"
                    except Exception:
                        _ago = ""
                    
                    _btn_label = _title + ("  " + _ago if _ago else "")
                    _conv_key_prefix = "conv_active_" if _is_active else "conv_select_"
                    
                    # Conversation row select/delete column layout
                    _row_col1, _row_col2 = st.columns([0.85, 0.15])
                    with _row_col1:
                        st.markdown(f'<div class="md-conv-select-anchor{" md-active" if _is_active else ""}"></div>', unsafe_allow_html=True)
                        if st.button(_btn_label, icon=":material/description:", key=_conv_key_prefix + _c["id"], use_container_width=True):
                            _conv_obj = load_conversation(st.session_state.user_email_hash, _c["id"])
                            if _conv_obj is not None:
                                st.session_state.current_conversation_id = _c["id"]
                                st.session_state.messages = _conv_obj.get("messages", []) or []
                                st.session_state.qcount = sum(1 for m in st.session_state.messages if m.get("role") == "user")
                                st.session_state.feedback = {}
                                st.session_state.last_sources = []
                                st.session_state.emergency_detected = False
                                st.session_state.mode = "chat"
                            st.rerun()
                    with _row_col2:
                        st.markdown('<div class="md-conv-del-anchor"></div>', unsafe_allow_html=True)
                        if st.button("✕", key="conv_del_" + _c["id"], use_container_width=True):
                            delete_conversation(st.session_state.user_email_hash, _c["id"])
                            if st.session_state.get("current_conversation_id") == _c["id"]:
                                st.session_state.current_conversation_id = ""
                                st.session_state.messages = []
                            st.rerun()

    elif st.session_state.is_guest:
        pass

    L = LANGUAGES[st.session_state.selected_language]

    # ── Sidebar bottom: language picker → Privacy & Consent → footer ──
    st.markdown('<div class="md-sidebar-bottom">', unsafe_allow_html=True)

    # Language selector sits immediately above the Privacy & Consent button.
    _lang_keys_top = list(LANGUAGES.keys())
    _current_lang_top = st.session_state.get("selected_language", "English")
    _lang_idx_top = _lang_keys_top.index(_current_lang_top) if _current_lang_top in _lang_keys_top else 0
    st.selectbox(
        "Language",
        options=_lang_keys_top,
        index=_lang_idx_top,
        key="lang_selector",
        label_visibility="collapsed",
    )
    # Lock the combobox input to read-only so the picker behaves as a
    # tap-to-open button.
    import streamlit.components.v1 as _components_lang
    _components_lang.html(
        """
        <script>
        (function(){
            function lock(){
                try {
                    var doc = window.parent.document;
                    var inputs = doc.querySelectorAll('[data-testid="stSidebar"] [data-testid="stSelectbox"] input[role="combobox"]');
                    inputs.forEach(function(input){
                        input.setAttribute('readonly', '');
                        input.setAttribute('tabindex', '-1');
                        input.setAttribute('inputmode', 'none');
                    });
                } catch (e) {}
            }
            lock();
            setTimeout(lock, 200);
            setTimeout(lock, 800);
            setInterval(lock, 1500);
        })();
        </script>
        """,
        height=0,
    )
    if st.session_state.lang_selector != st.session_state.selected_language:
        st.session_state.selected_language = st.session_state.lang_selector
        st.rerun()

    # Footer: Privacy & Terms · Help Center + copyright
    _foot_col1, _foot_col2, _foot_col3 = st.columns([0.45, 0.1, 0.45])
    with _foot_col1:
        st.markdown('<div class="md-privacy-btn-anchor"></div>', unsafe_allow_html=True)
        if st.button("Privacy & Terms", key="privacy_btn"):
            st.session_state.mode = "privacy"
            st.rerun()
    with _foot_col2:
        st.markdown('<div class="md-sidebar-foot-dot">·</div>', unsafe_allow_html=True)
    with _foot_col3:
        st.markdown('<div class="md-help-btn-anchor"></div>', unsafe_allow_html=True)
        if st.button("Help Center", key="help_btn"):
            st.session_state.mode = "privacy"
            st.rerun()
    st.markdown(
        '<div class="md-sidebar-foot-copy">© 2026 ' + APP_TITLE + '. All rights reserved.</div>',
        unsafe_allow_html=True
    )

L = LANGUAGES[st.session_state.selected_language]

ADMIN_PASSWORD = _safe_secret("ADMIN_PASSWORD", os.environ.get("ADMIN_PASSWORD", "MediChatAdmin@2026"))
_query_params = st.query_params
_admin_requested = _query_params.get("admin", "") != ""
_force_auth_requested = str(_query_params.get("force_auth", "") or "").strip().lower() in {"1", "true", "yes", "on"}
_mode_from_url = str(_query_params.get("mode", "") or "").strip()
_url_modes = {"chat", "overview", "assessment", "records", "rx_reader", "medications", "appointments", "insights", "history", "privacy"}

if _force_auth_requested:
    st.session_state.is_authenticated = False
    st.session_state.is_guest = False
    st.session_state.user_email_hash = ""
    st.session_state.user_email_display = ""
    st.session_state.current_conversation_id = ""
    st.session_state.messages = []
    st.session_state.mode = "chat"
    try:
        del st.query_params["force_auth"]
    except Exception:
        pass

if _mode_from_url in _url_modes:
    if st.session_state.mode != _mode_from_url:
        st.session_state.mode = _mode_from_url
    try:
        del st.query_params["mode"]
    except Exception:
        pass

# ?conv=<id> → load that conversation into the chat view. Lets the Recent
# Chats sidebar render as anchor links (giving us proper title-left +
# time-right flex layout that st.button can't).
_conv_id_from_url = str(_query_params.get("conv", "") or "").strip()
if _conv_id_from_url and st.session_state.is_authenticated and st.session_state.user_email_hash:
    _conv_obj = load_conversation(st.session_state.user_email_hash, _conv_id_from_url)
    if _conv_obj is not None:
        st.session_state.current_conversation_id = _conv_id_from_url
        st.session_state.messages = _conv_obj.get("messages", []) or []
        st.session_state.qcount = sum(1 for m in st.session_state.messages if m.get("role") == "user")
        st.session_state.feedback = {}
        st.session_state.last_sources = []
        st.session_state.emergency_detected = False
        st.session_state.mode = "chat"
    try:
        del st.query_params["conv"]
    except Exception:
        pass

# ?new_chat=1 → start a fresh chat session. Mirrors the ?conv handler so the
# "+ New chat" pill inside the Recent Chats card can be a true anchor link
# (nested inside the card's HTML) instead of a Streamlit-wrapped st.button
# rendered as a sibling.
_new_chat_from_url = str(_query_params.get("new_chat", "") or "").strip()
if _new_chat_from_url:
    start_new_chat_session()
    try:
        del st.query_params["new_chat"]
    except Exception:
        pass

# ?signout=1 → sign the user out. Lets the small logout icon inside the
# profile chip (a true HTML child of the chip, not a Streamlit button) clear
# session state without the negative-margin layout tricks that caused the
# Recent Chats card to collide with the chip's bottom.
_signout_from_url = str(_query_params.get("signout", "") or "").strip()
if _signout_from_url:
    for k in ["is_authenticated", "is_guest", "user_email_hash", "user_email_display", "patient_name", "patient_memory", "messages", "qcount", "feedback", "last_sources", "last_pdf_context", "last_image_context", "rx_reader_result", "rx_uploader_key"]:
        if k in st.session_state:
            if k in ("is_authenticated", "is_guest"):
                st.session_state[k] = False
            elif k == "patient_memory":
                st.session_state[k] = {"symptoms": [], "conditions": [], "medications": []}
            elif k == "messages":
                st.session_state[k] = []
            elif k == "rx_reader_result":
                st.session_state[k] = None
            elif k == "rx_uploader_key":
                st.session_state[k] = 0
            elif k == "qcount":
                st.session_state[k] = 0
            elif k == "feedback":
                st.session_state[k] = {}
            else:
                st.session_state[k] = "" if isinstance(st.session_state[k], str) else st.session_state[k]
    st.session_state.current_conversation_id = ""
    try:
        del st.query_params["signout"]
    except Exception:
        pass
    st.rerun()

# ?del_conv=<id> → delete a saved conversation. Powers the small × icon on
# each Recent Chats row (a true HTML child of the row's anchor structure,
# not a separate Streamlit widget). Uses the existing delete_conversation()
# function — no data-logic change, just a new UI invocation path.
_del_conv_from_url = str(_query_params.get("del_conv", "") or "").strip()
if _del_conv_from_url and st.session_state.is_authenticated and st.session_state.user_email_hash:
    delete_conversation(st.session_state.user_email_hash, _del_conv_from_url)
    # If the deleted conversation was the active one, clear it so a stale
    # active highlight doesn't persist in the sidebar.
    if st.session_state.get("current_conversation_id") == _del_conv_from_url:
        st.session_state.current_conversation_id = ""
        st.session_state.messages = []
    try:
        del st.query_params["del_conv"]
    except Exception:
        pass
    st.rerun()

if st.session_state.is_guest and not st.session_state.is_authenticated:
    _guest_allowed_modes = {
        "chat", "overview", "assessment", "records", "rx_reader",
        "medications", "appointments", "insights", "history", "privacy"
    }
    if st.session_state.mode not in _guest_allowed_modes:
        st.session_state.mode = "chat"

if _admin_requested and not st.session_state.admin_authenticated:
    st.markdown(
        '<div style="background:linear-gradient(135deg,#1f2937,#111827);color:white;padding:1.5rem 2rem;border-radius:16px;margin:2rem auto;max-width:500px;box-shadow:0 8px 30px rgba(0,0,0,0.2);">'
        '<div style="text-align:center;margin-bottom:1.2rem;">'
        '<div style="font-size:2.5rem;margin-bottom:0.5rem;">🔒</div>'
        '<div style="font-size:1.3rem;font-weight:700;margin-bottom:0.3rem;">Admin Access Required</div>'
        '<div style="font-size:0.85rem;opacity:0.8;">Research &amp; Evaluation Dashboard</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    pg_c1, pg_c2, pg_c3 = st.columns([1, 2, 1])
    with pg_c2:
        with st.form(key="admin_login_form", clear_on_submit=True):
            admin_pw_input = st.text_input("Password", type="password", placeholder="Enter admin password", label_visibility="collapsed")
            login_btn = st.form_submit_button("Unlock Analytics", use_container_width=True)
        if login_btn:
            if admin_pw_input == ADMIN_PASSWORD:
                st.session_state.admin_authenticated = True
                st.session_state.admin_attempt_failed = False
                st.rerun()
            else:
                st.session_state.admin_attempt_failed = True
                st.rerun()
        if st.session_state.admin_attempt_failed:
            st.error("Incorrect password. Access denied.")
        st.caption("Authorised team members only. This dashboard contains anonymised research data.")
    st.stop()

_is_admin = st.session_state.admin_authenticated

# ── Patient Profile Auth Gate ────────────────────────────────────────
# Only enforced when Firebase is connected and user is not admin.
# Users can also continue as Guest (no persistence).
if (not _is_admin) and (not st.session_state.is_authenticated) and (not st.session_state.is_guest) and st.session_state.mode != "privacy":
    st.markdown("""
    <style>
    /* Premium auth page polish, scoped to auth-only render path. */
    [data-testid="stForm"] {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        padding: 0 !important;
    }
    /* Sidebar/nav styles are intentionally NOT overridden here so auth, guest,
       and signed-in profiles share the same global sidebar design system. */

    .md-auth-welcome-card {
        position: relative;
        display: grid;
        grid-template-columns: 124px minmax(0, 1fr);
        align-items: center;
        gap: 1.25rem;
        padding: 1.9rem 2.2rem 1.85rem 1.6rem;
        border-radius: 28px;
        background:
            radial-gradient(circle at 10% 12%, rgba(191, 219, 254, 0.42), transparent 46%),
            radial-gradient(circle at 93% 90%, rgba(167, 139, 250, 0.2), transparent 48%),
            linear-gradient(132deg, rgba(255,255,255,0.995) 0%, rgba(249,252,255,0.995) 50%, rgba(240,246,255,0.99) 100%);
        border: 1px solid #d4e3ff;
        box-shadow: 0 22px 48px rgba(59, 130, 246, 0.11), 0 5px 16px rgba(15, 23, 42, 0.06);
        overflow: hidden;
        isolation: isolate;
    }
    .md-auth-welcome-card::before {
        content: "";
        position: absolute;
        left: -15%;
        bottom: -58%;
        width: 82%;
        height: 94%;
        border-radius: 999px;
        background: radial-gradient(ellipse at center, rgba(129, 140, 248, 0.24) 0%, rgba(147, 197, 253, 0.1) 42%, transparent 78%);
        transform: rotate(-8deg);
        pointer-events: none;
        z-index: -1;
    }
    .md-auth-welcome-card::after {
        content: "";
        position: absolute;
        left: -16%;
        right: -16%;
        bottom: -64px;
        height: 128px;
        border-radius: 999px;
        background: radial-gradient(120% 100% at 50% 0%, rgba(167, 139, 250, 0.18), rgba(147, 197, 253, 0.14) 42%, rgba(255, 255, 255, 0) 78%);
        pointer-events: none;
        opacity: 0.56;
    }
    .md-auth-deco-dots {
        position: absolute;
        right: 2.2rem;
        top: 1.25rem;
        width: 90px;
        height: 48px;
        background-image: radial-gradient(rgba(99, 102, 241, 0.28) 1.4px, transparent 1.4px);
        background-size: 10px 10px;
        opacity: 0.55;
        pointer-events: none;
    }
    .md-auth-shield {
        width: 112px;
        height: 112px;
        border-radius: 33px;
        display: flex !important;
        align-items: center;
        justify-content: center;
        color: #ffffff;
        flex-shrink: 0;
        font-size: 3rem !important;
        background:
            radial-gradient(circle at 26% 20%, rgba(255,255,255,0.56), transparent 44%),
            linear-gradient(145deg, #3b82f6 0%, #6366f1 58%, #8b5cf6 100%);
        box-shadow: 0 20px 38px rgba(99, 102, 241, 0.34);
        position: relative;
        left: auto;
        top: auto;
        transform: none;
        margin-left: 0.2rem;
    }
    .md-auth-shield.md-auth-shield-image {
        width: 118px;
        height: 118px;
        border-radius: 0;
        background: transparent !important;
        box-shadow: none !important;
        margin-left: 0;
    }
    .md-auth-shield.md-auth-shield-image img {
        width: 100%;
        height: 100%;
        object-fit: contain;
        display: block;
        transform: none;
        transform-origin: center center;
        filter: drop-shadow(0 14px 24px rgba(59, 130, 246, 0.26));
    }
    .md-auth-welcome-content {
        flex: 1 1 auto;
        min-width: 0;
        text-align: center;
        padding: 0 0.3rem 0 0;
        display: flex;
        flex-direction: column;
        align-items: center;
    }
    .md-auth-welcome-title {
        font-size: clamp(1.9rem, 2.7vw, 2.65rem);
        font-weight: 820;
        letter-spacing: -0.02em;
        color: #0f172a;
        line-height: 1.1;
        margin: 0 0 0.45rem 0;
        text-align: center;
        width: 100%;
    }
    .md-auth-welcome-copy {
        font-size: clamp(1.02rem, 1.38vw, 1.22rem);
        color: #334155;
        line-height: 1.5;
        max-width: 62ch;
        white-space: normal;
        text-align: center;
    }
    .md-auth-chip-row {
        display: flex;
        gap: 0.86rem;
        flex-wrap: wrap;
        row-gap: 0.72rem;
        margin-top: 1.28rem;
        position: relative;
        z-index: 1;
        justify-content: center;
        width: 100%;
    }
    .md-auth-chip {
        display: inline-flex;
        align-items: center;
        gap: 0.5rem;
        padding: 0.62rem 1.12rem;
        border-radius: 15px;
        background: rgba(255,255,255,0.92);
        border: 1px solid #dbe6fd;
        box-shadow: 0 10px 18px rgba(15, 23, 42, 0.06);
        color: #334155;
        font-size: 0.98rem;
        font-weight: 630;
    }
    .md-auth-chip .material-symbols-rounded {
        color: #2563eb;
        font-size: 1rem !important;
    }

    /* Tabs — centered glass pill bar (Apple-style frosted rail). The rail
       is a translucent white wash with a subtle backdrop blur so it picks
       up the page tone behind it; a hairline border + soft inner highlight
       sell the "glass" depth. The active tab fills with the brand indigo
       gradient and a colored drop shadow that matches the Sign in button. */
    [data-baseweb="tab-list"] {
        gap: 0.4rem !important;
        border-bottom: none !important;
        margin-top: 1.5rem !important;
        margin-bottom: 1rem !important;
        justify-content: center !important;
        background: rgba(255, 255, 255, 0.42) !important;
        backdrop-filter: blur(18px) saturate(160%) !important;
        -webkit-backdrop-filter: blur(18px) saturate(160%) !important;
        border: 1px solid rgba(255, 255, 255, 0.6) !important;
        border-radius: 14px !important;
        padding: 0.32rem !important;
        max-width: 440px !important;
        margin-left: auto !important;
        margin-right: auto !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.6) inset,
            0 0 0 1px rgba(15, 23, 42, 0.04),
            0 8px 24px rgba(15, 23, 42, 0.05) !important;
    }
    button[role="tab"] {
        padding: 0.55rem 1.25rem !important;
        font-size: 0.96rem !important;
        font-weight: 600 !important;
        color: #5a6b86 !important;
        border: none !important;
        border-bottom: none !important;
        border-radius: 10px !important;
        background: transparent !important;
        transition:
            color 0.18s ease,
            background 0.2s ease,
            box-shadow 0.22s ease,
            transform 0.18s ease !important;
    }
    button[role="tab"]:hover {
        color: #1f2a3d !important;
        background: rgba(99, 102, 241, 0.06) !important;
    }
    button[role="tab"][aria-selected="true"] {
        color: #ffffff !important;
        font-weight: 700 !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.16) 0%, rgba(255,255,255,0) 55%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 55%, #7c6ff2 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 6px 14px rgba(79, 70, 229, 0.28),
            0 2px 5px rgba(79, 70, 229, 0.15) !important;
        text-shadow: 0 1px 0 rgba(15, 23, 42, 0.14) !important;
    }
    button[role="tab"][aria-selected="true"]:hover {
        color: #ffffff !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.2) 0%, rgba(255,255,255,0) 55%),
            linear-gradient(108deg, #4338ca 0%, #5b54e8 55%, #7466f3 100%) !important;
    }
    /* Streamlit's underline indicator — superseded by the pill fill. */
    [data-baseweb="tab-highlight"] {
        display: none !important;
    }
    [data-baseweb="tab-panel"] {
        padding-top: 0.85rem !important;
        border-top: none !important;
    }
    [data-baseweb="tab-list"] + div,
    [data-baseweb="tab-list"] + div > div {
        border-top: none !important;
        box-shadow: none !important;
    }

    [data-testid="stForm"]:has(.st-key-si_email),
    [data-testid="stForm"]:has(.st-key-su_email) {
        background: rgba(255,255,255,0.98) !important;
        border: 1px solid #dbe6fb !important;
        border-radius: 28px !important;
        padding: 1.82rem 1.72rem 1.38rem 1.72rem !important;
        box-shadow: 0 22px 44px rgba(15, 23, 42, 0.08), 0 3px 10px rgba(15, 23, 42, 0.04) !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInput"] label,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInput"] label {
        font-size: 0.98rem !important;
        color: #1e293b !important;
        font-weight: 630 !important;
        margin-bottom: 0.35rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInput"],
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInput"] {
        margin-bottom: 0.42rem !important;
    }
    /* The white surface + blue border live on the ROOT wrapper, not the
       <input>, so the icon sits inside the same field instead of next to a
       separate grey baseweb container. The inner input is transparent. */
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInput"] input,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInput"] input {
        min-height: 54px !important;
        border-radius: 0 !important;
        border: none !important;
        background: transparent !important;
        box-shadow: none !important;
        font-size: 1.03rem !important;
        line-height: 1.3 !important;
        padding: 0.78rem 0.95rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputRootElement"],
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputRootElement"] {
        width: 100% !important;
        max-width: 100% !important;
        min-height: 56px !important;
        height: 56px !important;
        margin-left: 0 !important;
        margin-right: 0 !important;
        display: flex !important;
        align-items: center !important;
        overflow: hidden !important;
        background: #ffffff !important;
        border: 1px solid #d8e4fa !important;
        border-radius: 15px !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.02) !important;
        transition: border-color 0.18s ease, box-shadow 0.18s ease !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputRootElement"]:focus-within,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputRootElement"]:focus-within {
        border-color: #93c5fd !important;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.14) !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputRootElement"] [data-baseweb="base-input"],
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputRootElement"] [data-baseweb="base-input"] {
        min-height: 56px !important;
        height: 56px !important;
        display: flex !important;
        align-items: center !important;
        flex: 1 1 auto !important;
        min-width: 0 !important;
        width: auto !important;
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }
    /* Icon column — transparent, but with a subtle hairline divider on its
       right edge so the user can read at a glance where the click target
       (typing area) begins. */
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputRootElement"] > div:has([data-testid="stTextInputIcon"]),
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputRootElement"] > div:has([data-testid="stTextInputIcon"]) {
        background: transparent !important;
        border: none !important;
        border-right: 1px solid #e6ecf6 !important;
        box-shadow: none !important;
        padding-right: 0.55rem !important;
        margin-right: 0.55rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputIcon"],
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputIcon"] {
        width: 40px !important;
        min-width: 40px !important;
        height: 40px !important;
        margin-left: 0.56rem !important;
        margin-right: 0.44rem !important;
        border-radius: 11px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        background: transparent !important;
        border: none !important;
    }
    /* Focus state is now handled on the root wrapper via :focus-within above. */
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInput"] [data-testid="stIconMaterial"],
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInput"] [data-testid="stIconMaterial"] {
        color: #6366f1 !important;
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
        font-size: 1.16rem !important;
    }
    /* Show/hide password "eye" button on PIN field — strip baseweb's grey. */
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputRootElement"] button,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputRootElement"] button {
        background: transparent !important;
        border: none !important;
        color: #6366f1 !important;
        margin-right: 0.4rem !important;
    }
    [data-testid="stForm"]:has(.st-key-si_email) [data-testid="stTextInputRootElement"] button:hover,
    [data-testid="stForm"]:has(.st-key-su_email) [data-testid="stTextInputRootElement"] button:hover {
        background: rgba(99, 102, 241, 0.08) !important;
    }
    /* Primary submit button — premium glossy indigo gradient with a top
       highlight (inner 1px), a colored drop shadow, and an arrow that
       glides to the right on hover. The hover state lifts the button
       slightly and deepens the shadow for an alive, tactile feel. */
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button,
    .st-key-auth_signup_submit [data-testid="stFormSubmitButton"] > button {
        height: 60px !important;
        min-height: 60px !important;
        border-radius: 16px !important;
        border: none !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4f46e5 0%, #6366f1 42%, #7c6ff2 72%, #8b5cf6 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.22) inset,
            0 -1px 0 rgba(15, 23, 42, 0.08) inset,
            0 14px 30px rgba(79, 70, 229, 0.32),
            0 4px 10px rgba(79, 70, 229, 0.18) !important;
        color: #ffffff !important;
        font-size: 1.05rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.005em !important;
        text-shadow: 0 1px 0 rgba(15, 23, 42, 0.18) !important;
        transition:
            transform 0.18s ease,
            box-shadow 0.22s ease,
            background 0.25s ease !important;
    }
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button:hover,
    .st-key-auth_signup_submit [data-testid="stFormSubmitButton"] > button:hover {
        transform: translateY(-1px) !important;
        background:
            linear-gradient(180deg, rgba(255,255,255,0.22) 0%, rgba(255,255,255,0) 50%),
            linear-gradient(108deg, #4338ca 0%, #5b54e8 42%, #7466f3 72%, #8b5cf6 100%) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.26) inset,
            0 -1px 0 rgba(15, 23, 42, 0.1) inset,
            0 18px 38px rgba(79, 70, 229, 0.38),
            0 6px 14px rgba(79, 70, 229, 0.22) !important;
    }
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button:active,
    .st-key-auth_signup_submit [data-testid="stFormSubmitButton"] > button:active {
        transform: translateY(0) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.18) inset,
            0 8px 18px rgba(79, 70, 229, 0.28) !important;
    }
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button {
        position: relative !important;
        padding-right: 3.25rem !important;
    }
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button [data-testid="stIconMaterial"] {
        display: none !important;
    }
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button::after {
        content: "→";
        position: absolute;
        right: 1.25rem;
        top: 50%;
        transform: translateY(-54%);
        font-size: 1.55rem;
        font-weight: 400;
        opacity: 0.96;
        line-height: 1;
        transition: transform 0.22s ease, opacity 0.18s ease;
    }
    .st-key-auth_signin_submit [data-testid="stFormSubmitButton"] > button:hover::after {
        transform: translateY(-54%) translateX(4px);
        opacity: 1;
    }

    .md-auth-signin-actions {
        margin-top: 0.95rem;
        padding: 0;
    }
    .md-auth-or-divider {
        position: relative;
        text-align: center;
        margin: 0.55rem 0 0.9rem 0;
    }
    .md-auth-or-divider::before {
        content: "";
        position: absolute;
        left: 0;
        right: 0;
        top: 50%;
        border-top: 1px solid #e2e8f5;
        transform: translateY(-50%);
    }
    .md-auth-or-divider span {
        position: relative;
        background: transparent;
        color: #64748b;
        padding: 0 0.72rem;
        font-size: 0.95rem;
        font-weight: 560;
    }
    /* "Continue as Guest" — ghost button companion to the primary Sign in.
       Indigo-tinted surface and border so it visually belongs to the same
       family without competing for the eye. Hover fills with a slightly
       deeper indigo wash and lifts subtly to match the primary's motion. */
    .md-auth-signin-actions .stButton > button {
        height: 54px !important;
        min-height: 54px !important;
        border-radius: 16px !important;
        border: 1px solid rgba(99, 102, 241, 0.22) !important;
        background: linear-gradient(180deg, #ffffff 0%, #f6f8ff 100%) !important;
        color: #4338ca !important;
        font-size: 1.02rem !important;
        font-weight: 650 !important;
        letter-spacing: 0.005em !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 2px 6px rgba(79, 70, 229, 0.06) !important;
        transition:
            transform 0.18s ease,
            box-shadow 0.2s ease,
            border-color 0.18s ease,
            background 0.22s ease !important;
    }
    .md-auth-signin-actions .stButton > button [data-testid="stIconMaterial"] {
        color: #6366f1 !important;
        font-size: 1.18rem !important;
        margin-right: 0.18rem !important;
    }
    .md-auth-signin-actions .stButton > button:hover {
        border-color: rgba(99, 102, 241, 0.42) !important;
        background: linear-gradient(180deg, #f5f6ff 0%, #edefff 100%) !important;
        color: #3730a3 !important;
        transform: translateY(-1px) !important;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.95) inset,
            0 8px 18px rgba(79, 70, 229, 0.12) !important;
    }
    .md-auth-signin-actions .stButton > button:hover [data-testid="stIconMaterial"] {
        color: #4f46e5 !important;
    }
    .md-auth-signin-actions .stButton > button:active {
        transform: translateY(0) !important;
        background: #ebedff !important;
        box-shadow: 0 2px 6px rgba(79, 70, 229, 0.1) !important;
    }
    .md-auth-meta {
        margin-top: 0.95rem;
        padding-top: 0.95rem;
        border-top: 1px solid #e4ebf9;
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        gap: 1rem;
        flex-wrap: wrap;
    }
    .md-auth-security-note {
        display: flex;
        align-items: flex-start;
        gap: 0.55rem;
        color: #475569;
        font-size: 0.92rem;
        line-height: 1.5;
    }
    .md-auth-security-note .material-symbols-rounded {
        color: #2563eb;
        font-size: 1.18rem !important;
        margin-top: 0.08rem;
    }
    a.md-auth-forgot-link,
    .md-auth-forgot-row a.md-auth-forgot-link {
        color: #4f46e5 !important;
        text-decoration: none !important;
        font-size: 0.88rem;
        font-weight: 600;
        white-space: nowrap;
        transition: color 0.15s ease;
    }
    a.md-auth-forgot-link:hover,
    .md-auth-forgot-row a.md-auth-forgot-link:hover {
        color: #4338ca !important;
        text-decoration: none !important;
    }
    /* Inline row directly below the PIN field — right-aligned link, the
       conventional place auth-aware users look for "Forgot PIN?". */
    .md-auth-forgot-row {
        display: flex;
        justify-content: flex-end;
        margin: -0.18rem 0 0.62rem 0;
    }
    /* When the meta row is the only thing under the form (the Forgot link
       has been moved inline), center the security note across the full width
       so it doesn't look orphaned on the left. */
    .md-auth-meta.md-auth-meta-solo {
        justify-content: center;
        text-align: center;
    }
    .md-auth-meta.md-auth-meta-solo .md-auth-security-note {
        justify-content: center;
    }

    /* Guest tab — soft card explaining the no-account experience, with the
       Continue as Guest button spaced clearly below. The card uses a
       subtle indigo wash (matching the welcome chips) so the tab content
       feels framed without competing with the white form area on the
       neighbouring tabs. */
    .md-auth-guest-intro {
        margin: 0.8rem 0 1.25rem 0;
        padding: 1.1rem 1.25rem;
        border-radius: 16px;
        background: linear-gradient(180deg, #f6f7ff 0%, #eef1ff 100%);
        border: 1px solid #e3e8fb;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.8) inset,
            0 6px 18px rgba(15, 23, 42, 0.04);
    }
    .md-auth-guest-title {
        display: flex;
        align-items: center;
        gap: 0.55rem;
        font-size: 1.05rem;
        font-weight: 730;
        color: #1f2a3d;
        letter-spacing: -0.005em;
        margin-bottom: 0.45rem;
    }
    .md-auth-guest-title .material-symbols-rounded {
        color: #4f46e5;
        font-size: 1.18rem !important;
    }
    .md-auth-guest-copy {
        color: #475569;
        font-size: 0.92rem;
        line-height: 1.55;
    }
    /* When the actions wrapper sits inside the Guest tab there is no
       preceding "or" divider, so reset the margin-top to align cleanly
       with the intro card above. */
    .md-auth-signin-actions.md-auth-signin-actions-guest {
        margin-top: 0 !important;
    }

    .md-auth-privacy-foot {
        margin-top: 1.2rem;
        color: #64748b;
        font-size: 0.95rem;
        line-height: 1.6;
    }
    .md-auth-privacy-foot a {
        color: #1d4ed8 !important;
        font-weight: 620;
    }

    .md-auth-side-card {
        border-radius: 24px;
        background: linear-gradient(180deg, rgba(255,255,255,0.99), rgba(249,251,255,0.99));
        border: 1px solid #dbe6fb;
        box-shadow: 0 18px 38px rgba(15, 23, 42, 0.07);
        padding: 1.55rem 1.45rem 1.3rem 1.45rem;
        margin-top: 1.42rem;
        max-width: 372px;
        margin-left: auto;
    }
    .md-auth-side-title {
        font-size: 2rem;
        font-weight: 820;
        color: #0f172a;
        margin: 0;
        line-height: 1.12;
    }
    .md-auth-side-subline {
        width: 46px;
        height: 2.5px;
        border-radius: 999px;
        background: #5b6cf9;
        margin: 0.68rem 0 1.2rem 0;
    }
    .md-auth-benefit {
        display: grid;
        grid-template-columns: 56px 1fr;
        gap: 0.9rem;
        align-items: start;
        margin-bottom: 1.35rem;
    }
    .md-auth-benefit-ic {
        width: 56px;
        height: 56px;
        border-radius: 20px;
        display: flex !important;
        align-items: center;
        justify-content: center;
        color: #2563eb;
        font-size: 1.5rem !important;
        background: linear-gradient(145deg, #eef4ff, #e8f0ff);
    }
    /* Tonal variation within a single blue→indigo family — keeps the card
       cohesive while still giving each row a subtle distinct accent. */
    .md-auth-benefit:nth-of-type(2) .md-auth-benefit-ic { color: #4f46e5; background: linear-gradient(145deg, #eef2ff, #e7ecff); }
    .md-auth-benefit:nth-of-type(3) .md-auth-benefit-ic { color: #6366f1; background: linear-gradient(145deg, #f0f1ff, #e8eaff); }
    .md-auth-benefit-title {
        margin: 0;
        font-size: 1.01rem;
        font-weight: 730;
        color: #0f172a;
        line-height: 1.35;
    }
    .md-auth-benefit-copy {
        margin-top: 0.18rem;
        color: #475569;
        font-size: 0.91rem;
        line-height: 1.5;
    }
    .md-auth-side-bottom {
        margin-top: 1.28rem;
        padding-top: 1rem;
        border-top: 1px dashed #dbe5f1;
        display: flex;
        gap: 0.6rem;
        color: #334155;
        font-size: 0.93rem;
        line-height: 1.5;
    }
    .md-auth-side-bottom .material-symbols-rounded {
        color: #4f46e5;
        font-size: 1.1rem !important;
        margin-top: 0.05rem;
    }

    @media (max-width: 1280px) {
        .md-auth-welcome-title {
            font-size: 2rem;
        }
    }
    @media (max-width: 980px) {
        .md-auth-side-card {
            max-width: 100%;
        }
        .md-auth-welcome-card {
            padding: 1.35rem 1.15rem;
            border-radius: 22px;
            gap: 1rem;
        }
        .md-auth-shield {
            width: 80px;
            height: 80px;
            border-radius: 24px;
            font-size: 2.2rem !important;
            position: static;
            transform: none;
        }
        .md-auth-chip-row {
            gap: 0.5rem;
        }
        .md-auth-chip {
            font-size: 0.8rem;
            padding: 0.42rem 0.68rem;
        }
        .md-auth-welcome-copy {
            white-space: normal;
        }
        .md-auth-welcome-content {
            padding-left: 0;
            padding-right: 0;
        }
        [data-testid="stForm"]:has(.st-key-si_email),
        [data-testid="stForm"]:has(.st-key-su_email) {
            padding: 1.15rem 1rem 1.05rem 1rem !important;
            border-radius: 22px !important;
        }
        .md-auth-signin-actions {
            padding: 0;
        }
    }
    @media (max-width: 720px) {
        [data-baseweb="tab-list"] {
            gap: 1.05rem !important;
        }
        button[role="tab"] {
            font-size: 0.93rem !important;
            padding-bottom: 0.7rem !important;
        }
        .md-auth-meta {
            flex-direction: column;
            align-items: flex-start;
        }
    }
    </style>
    """, unsafe_allow_html=True)

    _auth_shield_html = '<div class="md-auth-shield material-symbols-rounded">shield_person</div>'
    _auth_shield_uri = load_asset_data_uri("auth_welcome_shield.png")
    if _auth_shield_uri:
        _auth_shield_html = (
            '<div class="md-auth-shield md-auth-shield-image">'
            '<img src="' + _auth_shield_uri + '" alt="MediChat welcome shield icon">'
            '</div>'
        )

    st.markdown(
        '<div class="md-auth-welcome-card">'
        '<div class="md-auth-deco-dots"></div>'
        + _auth_shield_html +
        '<div class="md-auth-welcome-content">'
        '<div class="md-auth-welcome-title">Welcome to MediChat</div>'
        '<div class="md-auth-welcome-copy">Sign in to save your health profile across visits, or continue as a guest for a one-time chat.</div>'
        '<div class="md-auth-chip-row">'
        '<span class="md-auth-chip"><span class="material-symbols-rounded">lock</span>Private</span>'
        '<span class="md-auth-chip"><span class="material-symbols-rounded">person_check</span>Secure guest mode</span>'
        '<span class="md-auth-chip"><span class="material-symbols-rounded">verified</span>APP + HIPAA standard</span>'
        '<span class="md-auth-chip"><span class="material-symbols-rounded">verified_user</span>Health data protected</span>'
        '</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    auth_main_col, auth_side_col = st.columns([2.35, 0.85], gap="large")

    with auth_main_col:
        view = st.session_state.auth_view
        if view == "choose":
            tab_signin, tab_signup, tab_guest = st.tabs(["Sign in", "Create profile", "Guest"])
            with tab_signin:
                with st.form("signin_form", clear_on_submit=False):
                    si_email = st.text_input("Email", placeholder="you@example.com", key="si_email", icon=":material/mail:")
                    si_pin = st.text_input("4-8 digit PIN", type="password", max_chars=8, key="si_pin", icon=":material/lock:")
                    st.markdown(
                        '<div class="md-auth-forgot-row">'
                        '<a class="md-auth-forgot-link" href="' + ui_escape(PRIVACY_POLICY_URL) + '" target="_blank">Forgot PIN?</a>'
                        '</div>',
                        unsafe_allow_html=True
                    )
                    si_btn = st.form_submit_button(
                        "Sign in",
                        key="auth_signin_submit",
                        use_container_width=True,
                        type="primary",
                        icon=":material/arrow_forward:"
                    )

                st.markdown('<div class="md-auth-signin-actions">', unsafe_allow_html=True)
                st.markdown('<div class="md-auth-or-divider"><span>or</span></div>', unsafe_allow_html=True)
                if st.button("Continue as Guest", key="go_guest_btn_signin", use_container_width=True, icon=":material/person:"):
                    st.session_state.is_guest = True
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

                st.markdown(
                    '<div class="md-auth-meta md-auth-meta-solo">'
                    '<div class="md-auth-security-note">'
                    '<span class="material-symbols-rounded">shield_lock</span>'
                    '<span>We use secure, privacy-first authentication for your health information.</span>'
                    '</div>'
                    '</div>',
                    unsafe_allow_html=True
                )

                if si_btn:
                    if not si_email or "@" not in si_email or not si_pin or not si_pin.isdigit() or len(si_pin) < 4 or len(si_pin) > 8:
                        st.error("Please enter a valid email and a 4-8 digit numeric PIN.")
                    else:
                        profile, status = authenticate_profile(si_email, si_pin)
                        if status == "ok":
                            st.session_state.is_authenticated = True
                            st.session_state.user_email_hash = profile["email_hash"]
                            st.session_state.user_email_display = si_email.strip()
                            st.session_state.patient_name = profile.get("name", "") or ""
                            st.session_state.patient_memory = profile.get("patient_memory", {"symptoms": [], "conditions": [], "medications": []})
                            st.session_state.selected_language = profile.get("language", "English")
                            recent = list_conversations(profile["email_hash"], limit=1)
                            if recent:
                                conv = load_conversation(st.session_state.user_email_hash, recent[0]["id"])
                                if conv:
                                    st.session_state.current_conversation_id = recent[0]["id"]
                                    st.session_state.messages = conv.get("messages", []) or []
                                    st.session_state.qcount = sum(1 for m in st.session_state.messages if m.get("role") == "user")
                            else:
                                st.session_state.current_conversation_id = ""
                                st.session_state.messages = []
                                st.session_state.qcount = 0
                            st.success("Welcome back" + ((", " + profile.get("name", "")) if profile.get("name") else "") + ". Loading your profile…")
                            st.rerun()
                        elif status == "not_found":
                            st.error("No profile found for that email. Switch to 'Create profile' to start one.")
                        elif status == "wrong_pin":
                            st.error("Incorrect PIN. Try again or recover your account by creating a new profile with a different email.")

            with tab_signup:
                with st.form("signup_form", clear_on_submit=False):
                    su_email = st.text_input("Email", placeholder="you@example.com", key="su_email", icon=":material/mail:")
                    su_name = st.text_input("First name (optional)", key="su_name", max_chars=30, icon=":material/person:")
                    su_pin = st.text_input("Choose a 4-8 digit PIN", type="password", max_chars=8, key="su_pin", icon=":material/lock:")
                    su_pin2 = st.text_input("Confirm PIN", type="password", max_chars=8, key="su_pin2", icon=":material/password:")
                    su_btn = st.form_submit_button(
                        "Create profile",
                        key="auth_signup_submit",
                        use_container_width=True,
                        type="primary",
                        icon=":material/person_add:"
                    )
                if su_btn:
                    if not su_email or "@" not in su_email:
                        st.error("Please enter a valid email.")
                    elif not su_pin or not su_pin.isdigit() or len(su_pin) < 4 or len(su_pin) > 8:
                        st.error("PIN must be 4-8 digits, numbers only.")
                    elif su_pin != su_pin2:
                        st.error("PINs do not match.")
                    elif get_profile(hash_email(su_email)) is not None:
                        st.error("A profile already exists for that email. Sign in instead.")
                    else:
                        profile = create_profile(su_email, su_pin, su_name)
                        if profile is None:
                            st.error("Could not create profile. Please try again in a moment.")
                        else:
                            st.session_state.is_authenticated = True
                            st.session_state.user_email_hash = profile["email_hash"]
                            st.session_state.user_email_display = su_email.strip()
                            st.session_state.patient_name = profile.get("name", "") or ""
                            st.success("Profile created. Welcome to MediChat.")
                            st.rerun()

            with tab_guest:
                st.markdown(
                    '<div class="md-auth-guest-intro">'
                    '<div class="md-auth-guest-title">'
                    '<span class="material-symbols-rounded">visibility_off</span>'
                    'No account, no trace'
                    '</div>'
                    '<div class="md-auth-guest-copy">'
                    'Try MediChat without signing up. Your conversation lives only in this tab. Close it and everything is forgotten. '
                    'Switch to a profile any time for continuity across visits.'
                    '</div>'
                    '</div>',
                    unsafe_allow_html=True
                )
                st.markdown('<div class="md-auth-signin-actions md-auth-signin-actions-guest">', unsafe_allow_html=True)
                if st.button("Continue as Guest", use_container_width=True, key="go_guest_btn", icon=":material/person:"):
                    st.session_state.is_guest = True
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        st.markdown(
            '<div class="md-auth-privacy-foot">'
            'We store only irreversible email hashes and salted PIN hashes, designed to align with Australian Privacy Principles (APP), HIPAA-style safeguards, and Notifiable Data Breach commitments. '
            '<a href="' + ui_escape(PRIVACY_POLICY_URL) + '" target="_blank">Read the full Privacy Policy →</a>'
            '</div>',
            unsafe_allow_html=True
        )

    with auth_side_col:
        st.markdown(
            '<div class="md-auth-side-card">'
            '<h3 class="md-auth-side-title">Why sign in?</h3>'
            '<div class="md-auth-side-subline"></div>'
            '<div class="md-auth-benefit">'
            '<div class="md-auth-benefit-ic material-symbols-rounded">folder_managed</div>'
            '<div><p class="md-auth-benefit-title">Save your health history</p><div class="md-auth-benefit-copy">Keep your conversations, symptoms, and insights in one place.</div></div>'
            '</div>'
            '<div class="md-auth-benefit">'
            '<div class="md-auth-benefit-ic material-symbols-rounded">badge</div>'
            '<div><p class="md-auth-benefit-title">Access your Health Passport</p><div class="md-auth-benefit-copy">View and manage your verified health information anytime.</div></div>'
            '</div>'
            '<div class="md-auth-benefit">'
            '<div class="md-auth-benefit-ic material-symbols-rounded">sync</div>'
            '<div><p class="md-auth-benefit-title">Sync reports & medications</p><div class="md-auth-benefit-copy">Automatically sync your reports and medications across devices.</div></div>'
            '</div>'
            '<div class="md-auth-side-bottom">'
            '<span class="material-symbols-rounded">shield_locked</span>'
            '<span>Your health data is encrypted and protected at all times.</span>'
            '</div>'
            '</div>',
            unsafe_allow_html=True
        )
    st.stop()

if _is_admin:
    admin_c1, admin_c2 = st.columns([5, 1])
    with admin_c2:
        if st.button("🔒 Logout", key="admin_logout"):
            st.session_state.admin_authenticated = False
            st.session_state.mode = "chat"
            st.rerun()
    # Admin gets a compact mode switcher to access Analytics
    if st.session_state.mode != "eval":
        if st.button("📊 Open Admin Analytics", key="open_eval_btn"):
            st.session_state.mode = "eval"
            st.rerun()
else:
    if st.session_state.mode == "eval":
        st.session_state.mode = "chat"
    # Mode is now driven by the sidebar nav (Home / Symptoms Checker).

st.markdown('<div class="disclaimer">MediChat offers general health information, not personal medical advice. Please consult a qualified doctor for any health concerns that need diagnosis or treatment.</div>', unsafe_allow_html=True)

if st.session_state.emergency_detected:
    reason = st.session_state.get("emergency_reason", "Emergency indicators detected")
    st.markdown(
        '<div class="emergency-banner">'
        '<div class="emergency-title">🚨 This May Be a Medical Emergency</div>'
        '<div class="emergency-text"><strong>Detected pattern:</strong> ' + ui_text(reason, 120) + '. Based on what you described, you may need immediate medical attention. Please stop and call emergency services now. Do not wait.</div>'
        '<div class="emergency-number">📞 Call 000 (Australia)</div>'
        '<div style="font-size:0.75rem;margin-top:0.5rem;opacity:0.9;">Other countries: 911 (USA) | 999 (UK) | 112 (EU) | 119 (Sri Lanka) | 102 (India)</div>'
        '</div>',
        unsafe_allow_html=True
    )
    cols = st.columns([3, 1])
    with cols[1]:
        if st.button("Dismiss alert", key="dismiss_emergency"):
            st.session_state.emergency_detected = False
            st.rerun()

# Triage tier banner removed per UI/UX refinement pass to clear screen real estate.

home_user_input = ""
home_submit = False
home_upload_clicked = False
home_voice_clicked = False
home_uploaded_image = None
home_vision_analyze = False
home_empty_chat = st.session_state.mode == "chat" and not st.session_state.messages

if st.session_state.mode == "chat":
    # ── New Dashboard Home (only on empty chat) ─────────────────────
    if not st.session_state.messages:
        _local_now = get_user_local_now()
        _hour = _local_now.hour
        if 5 <= _hour < 12:
            _tod = "morning"
        elif 12 <= _hour < 17:
            _tod = "afternoon"
        elif 17 <= _hour < 21:
            _tod = "evening"
        else:
            _tod = "night"
        _disp_name = (st.session_state.patient_name if st.session_state.patient_name and st.session_state.patient_name != "Guest" else "")
        _greet = "Good " + _tod + (", " + _disp_name if _disp_name else "") + " 👋"
        st.markdown(
            '<div class="md-greet-wrap md-home-greet-wrap md-home-head-left">'
            '<div class="md-greet">' + ui_text(_greet, 80) + '</div>'
            '<div class="md-subgreet">How can I help you today?</div>'
            '</div>',
            unsafe_allow_html=True
        )

        home_main, home_side = st.columns([2.38, 0.95], gap="medium")

        with home_main:
            # Quick action cards
            qa_cols = st.columns(4, gap="small")
            qa_specs = [
                ("qa_headache", "Headache", "I have a headache and would like to understand what might be causing it.", ":material/psychology:"),
                ("qa_tired", "Low energy", "I have been feeling unusually tired lately. What could be the reason?", ":material/battery_low:"),
                ("qa_symptoms", "Check symptoms", "_route_assessment", ":material/monitor_heart:"),
                ("qa_sleep", "Better sleep", "Can you suggest ways to improve my sleep quality?", ":material/bedtime:"),
            ]
            for i, (qa_key, qa_label, qa_query, qa_icon) in enumerate(qa_specs):
                with qa_cols[i]:
                    if st.button(qa_label, key=qa_key, use_container_width=True, icon=qa_icon):
                        if qa_query == "_route_assessment":
                            st.session_state.mode = "assessment"
                            st.rerun()
                        else:
                            st.session_state.pending_user_input = qa_query
                            st.rerun()

            with st.form("home_chat_form", clear_on_submit=True):
                home_user_input = st.text_area(
                    "Start a chat",
                    placeholder="Ask anything about your health...",
                    label_visibility="collapsed",
                    height=100,
                    key="home_chat_input_" + str(st.session_state.chat_input_key),
                )
                ac1, ac2, ac_spacer, ac3 = st.columns([0.56, 0.56, 4.0, 0.56], vertical_alignment="center")
                with ac1:
                    home_upload_clicked = st.form_submit_button("Upload", key="home_upload_btn", icon=":material/attach_file:", use_container_width=True)
                with ac2:
                    home_voice_clicked = st.form_submit_button("Voice", key="home_voice_btn", icon=":material/mic:", use_container_width=True)
                with ac3:
                    home_submit = st.form_submit_button(" ", key="home_send_btn", icon=":material/send:", use_container_width=True, type="primary")
            st.markdown('<div class="md-composer-glow"></div>', unsafe_allow_html=True)
            if home_upload_clicked:
                st.session_state.home_show_vision_upload = True
                st.session_state.home_show_voice = False
                st.rerun()
            if home_voice_clicked:
                st.session_state.home_show_voice = True
                st.session_state.home_show_vision_upload = False
                st.rerun()

            if st.session_state.get("home_show_vision_upload", False):
                st.markdown(
                    '<div class="md-vision-panel">'
                    '<div class="md-panel-icon material-symbols-rounded">image_search</div>'
                    '<div><div class="md-panel-title">Vision Ai upload</div>'
                    '<div class="md-panel-subtitle">Upload an X-ray, scan photo, skin image, or blood report PDF for cautious visual/report analysis.</div></div>'
                    '</div>',
                    unsafe_allow_html=True
                )
                home_uploaded_image = st.file_uploader(
                    "Upload for Vision Ai",
                    type=["jpg", "jpeg", "png", "pdf"],
                    label_visibility="collapsed",
                    key="home_vision_uploader_" + str(st.session_state.uploader_key),
                    help="Upload a medical image or PDF report for Vision Ai analysis.",
                )
                if home_uploaded_image:
                    is_home_pdf = home_uploaded_image.name.lower().endswith(".pdf")
                    if is_home_pdf:
                        st.markdown(
                            '<div class="md-file-preview md-home-file-preview">'
                            '<div class="md-file-icon">PDF</div>'
                            '<div style="flex:1;min-width:0;">'
                            '<div class="md-file-name">' + ui_text(home_uploaded_image.name, 90) + '</div>'
                            '<div class="md-file-help">Ready for Vision Ai report review.</div>'
                            '</div>'
                            '</div>',
                            unsafe_allow_html=True
                        )
                    else:
                        st.image(home_uploaded_image, caption="Ready for Vision Ai analysis", use_column_width=True)
                vu1, vu2 = st.columns([1, 1])
                with vu1:
                    home_vision_analyze = st.button("Analyze with Vision Ai", key="home_vision_analyze", type="primary", use_container_width=True, icon=":material/image_search:")
                    if home_vision_analyze and not home_uploaded_image:
                        st.error("Please upload an image or PDF first.")
                        home_vision_analyze = False
                with vu2:
                    if st.button("Cancel upload", key="home_vision_cancel", use_container_width=True, icon=":material/close:"):
                        st.session_state.home_show_vision_upload = False
                        st.session_state.uploader_key = st.session_state.get("uploader_key", 0) + 1
                        st.rerun()

            if st.session_state.get("home_show_voice", False):
                st.markdown(
                    '<div class="md-voice-panel">'
                    '<div class="md-panel-icon material-symbols-rounded">mic</div>'
                    '<div><div class="md-panel-title">Voice note</div>'
                    '<div class="md-panel-subtitle">Record or upload a short voice question. MediChat will transcribe it before answering.</div></div>'
                    '</div>',
                    unsafe_allow_html=True
                )
                voice_audio = None
                if hasattr(st, "audio_input"):
                    voice_audio = st.audio_input("Record your question", key="voice_audio_" + str(st.session_state.voice_audio_key))
                else:
                    voice_audio = st.file_uploader(
                        "Upload a voice note",
                        type=["wav", "mp3", "m4a", "ogg", "webm"],
                        key="voice_upload_" + str(st.session_state.voice_audio_key),
                        help="Your Streamlit version does not expose microphone recording, so upload an audio note instead.",
                    )
                vc1, vc2 = st.columns([1, 1])
                with vc1:
                    if st.button("Transcribe voice", key="voice_transcribe", type="primary", use_container_width=True, icon=":material/graphic_eq:"):
                        transcript = transcribe_voice_note(voice_audio)
                        if transcript:
                            st.session_state.pending_user_input = transcript
                            st.session_state.home_show_voice = False
                            st.session_state.voice_audio_key = st.session_state.get("voice_audio_key", 0) + 1
                            st.rerun()
                        else:
                            st.error("I could not transcribe that audio. Please try a clearer recording or type the question.")
                with vc2:
                    if st.button("Cancel voice", key="voice_cancel", use_container_width=True, icon=":material/close:"):
                        st.session_state.home_show_voice = False
                        st.session_state.voice_audio_key = st.session_state.get("voice_audio_key", 0) + 1
                        st.rerun()

            st.markdown(
                '<div class="md-home-composer-note"><span class="material-symbols-rounded md-disclaimer-shield">verified_user</span>MediChat Ai can make mistakes. Please consult a healthcare professional for medical advice.</div>',
                unsafe_allow_html=True
            )

            # ── Smart Actions (available for all users) ───────────────
            st.markdown('<div class="md-smart-head"><div class="md-smart-title">Smart Actions</div></div>', unsafe_allow_html=True)
            # Markdown bold (**...**) on the title makes it bold even when
            # it wraps to 2 lines (unlike CSS ::first-line which only styles
            # the first VISUAL line). The two trailing spaces before \n
            # render a line break in markdown, separating title + subtitle.
            sa_specs = [
                ("sa_sym", "**Vision Ai**  \nAnalyze images and X-rays", "vision", ":material/image_search:", "md-smart-purple"),
                ("sa_rec", "**Health Records**  \nUpload medical reports", "records", ":material/medical_information:", "md-smart-green"),
                ("sa_ins", "**Prescription Reader**  \nScan prescriptions", "rx_reader", ":material/prescriptions:", "md-smart-pink"),
                ("sa_appt", "**Appointments**  \nManage appointments", "appointments", ":material/calendar_month:", "md-smart-blue"),
            ]
            sa_cols = st.columns(4, gap="small")
            for i, (sk, label, action, icon_name, accent_cls) in enumerate(sa_specs):
                with sa_cols[i]:
                    if st.button(label, key=sk, use_container_width=True, icon=icon_name):
                        if action == "vision":
                            st.session_state.mode = "chat"
                            st.session_state.home_show_voice = False
                            st.session_state.home_show_vision_upload = True
                        else:
                            st.session_state.mode = action
                        st.rerun()

            # ── Daily Health Tip carousel (4 live-data slides, 3s rotation) ──
            _tip_daily = get_daily_metrics()
            _tip_water_raw = _tip_daily.get("water_glasses")
            _tip_sleep_raw = _tip_daily.get("sleep_hours")
            _tip_steps_raw = _tip_daily.get("steps")
            _tip_hr_raw = _tip_daily.get("heart_rate_resting")
            _tip_water_metric = (str(int(_tip_water_raw)) + " / 8 glasses today") if _tip_water_raw is not None else "Not logged today. Tap Health Overview to log"
            _tip_sleep_metric = ("Last night: " + ("%.1f" % float(_tip_sleep_raw)) + "h") if _tip_sleep_raw else "Sleep not logged. Log it on Health Overview"
            _tip_steps_metric = (format(int(_tip_steps_raw), ",") + " / 10,000 steps") if _tip_steps_raw else "Steps not logged. Log them on Health Overview"
            _tip_hr_metric = ("Resting HR " + str(int(_tip_hr_raw)) + " BPM") if _tip_hr_raw else "Heart rate not logged. Log it on Health Overview"
            _tip_glass_svg = (
                '<svg viewBox="0 0 86 100" xmlns="http://www.w3.org/2000/svg" aria-hidden="true">'
                '<defs><linearGradient id="mdGlassC" x1="0" y1="0" x2="0" y2="1">'
                '<stop offset="0%" stop-color="#bfdbfe" stop-opacity="0.55"/>'
                '<stop offset="100%" stop-color="#60a5fa" stop-opacity="0.85"/>'
                '</linearGradient></defs>'
                '<ellipse cx="43" cy="92" rx="30" ry="4" fill="#1e3a8a" opacity="0.08"/>'
                '<path d="M22 30 L64 30 L60 88 Q60 92 56 92 L30 92 Q26 92 26 88 Z" fill="url(#mdGlassC)" stroke="#3b82f6" stroke-width="1.6" stroke-linejoin="round"/>'
                '<ellipse cx="43" cy="30" rx="21" ry="3.4" fill="#dbeafe" stroke="#3b82f6" stroke-width="1.4"/>'
                '<path d="M30 42 Q34 46 32 52 Q30 56 33 60" stroke="#ffffff" stroke-width="1.8" stroke-linecap="round" fill="none" opacity="0.85"/>'
                '<circle cx="48" cy="58" r="2" fill="#ffffff" opacity="0.7"/>'
                '<circle cx="38" cy="72" r="1.4" fill="#ffffff" opacity="0.6"/>'
                '</svg>'
            )
            _tip_html = (
                '<div class="md-tip-carousel">'
                '<div class="md-tip-slide md-tip-water">'
                '<div>'
                '<div class="md-tip-eyebrow">Hydration tip</div>'
                '<div class="md-tip-title">Stay hydrated</div>'
                '<div class="md-tip-desc">Drinking enough water helps maintain energy and supports overall health.</div>'
                '<div class="md-tip-metric"><span class="material-symbols-rounded">water_drop</span>' + _tip_water_metric + '</div>'
                '</div>'
                '<div class="md-tip-illust md-tip-illust-svg">' + _tip_glass_svg + '</div>'
                '</div>'
                '<div class="md-tip-slide md-tip-sleep">'
                '<div>'
                '<div class="md-tip-eyebrow">Sleep insight</div>'
                '<div class="md-tip-title">Wind down earlier</div>'
                '<div class="md-tip-desc">Quality sleep boosts recovery, mood, and immunity. Aim for 7–9 hours every night.</div>'
                '<div class="md-tip-metric"><span class="material-symbols-rounded">bedtime</span>' + _tip_sleep_metric + '</div>'
                '</div>'
                '<div class="md-tip-illust"><span class="material-symbols-rounded">bedtime</span></div>'
                '</div>'
                '<div class="md-tip-slide md-tip-move">'
                '<div>'
                '<div class="md-tip-eyebrow">Movement</div>'
                '<div class="md-tip-title">Keep moving</div>'
                '<div class="md-tip-desc">A short walk after meals helps regulate blood sugar and energy levels through the day.</div>'
                '<div class="md-tip-metric"><span class="material-symbols-rounded">directions_walk</span>' + _tip_steps_metric + '</div>'
                '</div>'
                '<div class="md-tip-illust"><span class="material-symbols-rounded">directions_walk</span></div>'
                '</div>'
                '<div class="md-tip-slide md-tip-vitals">'
                '<div>'
                '<div class="md-tip-eyebrow">Vitals check</div>'
                '<div class="md-tip-title">Heart in good rhythm</div>'
                '<div class="md-tip-desc">A resting heart rate of 60–100 BPM is typical for healthy adults. Keep moving and resting well.</div>'
                '<div class="md-tip-metric"><span class="material-symbols-rounded">favorite</span>' + _tip_hr_metric + '</div>'
                '</div>'
                '<div class="md-tip-illust"><span class="material-symbols-rounded">favorite</span></div>'
                '</div>'
                '<div class="md-tip-indicators">'
                '<span class="md-tip-dot dot-1"></span>'
                '<span class="md-tip-dot dot-2"></span>'
                '<span class="md-tip-dot dot-3"></span>'
                '<span class="md-tip-dot dot-4"></span>'
                '</div>'
                '</div>'
            )
            st.markdown(_tip_html, unsafe_allow_html=True)

        with home_side:
            # ── Profile Snapshot ────────────────────────────────────────
                _mem_now = st.session_state.patient_memory or {}
                _cond_n = len(_mem_now.get("conditions") or [])
                _med_n = len(_mem_now.get("medications") or [])
                _sym_n = len(_mem_now.get("symptoms") or [])
                _convs_total = 0
                _last_visit_str = "-"
                if st.session_state.is_authenticated and st.session_state.user_email_hash:
                    _all_convs = list_conversations(st.session_state.user_email_hash, limit=100)
                    _convs_total = len(_all_convs)
                    if _all_convs and _all_convs[0].get("last_updated"):
                        try:
                            _lu = _all_convs[0]["last_updated"]
                            _delta = datetime.utcnow() - (_lu.replace(tzinfo=None) if _lu.tzinfo else _lu)
                            _h = int(_delta.total_seconds() // 3600)
                            _last_visit_str = (str(int(_delta.total_seconds() // 60)) + "m ago") if _h < 1 else (str(_h) + "h ago" if _h < 24 else str(_h // 24) + "d ago")
                        except Exception:
                            _last_visit_str = "recently"

                _snap_title = "Health Overview"
                _daily = get_daily_metrics()
                # Real values only — no fake fallbacks. Missing → "-" with no status pill.
                _hr = _daily.get("heart_rate_resting")
                _steps_val = _daily.get("steps")
                _sleep_hours = _daily.get("sleep_hours")
                _water_count = _daily.get("water_glasses")

                _heart_rate_display = (str(int(_hr)) + " BPM") if _hr else "-"
                _hr_status, _hr_cls = heart_rate_status(_hr)
                _heart_rate_status_text = _hr_status or ""

                _steps_display = (f"{int(_steps_val):,} / 10,000") if _steps_val else "-"
                _steps_status_text = (str(min(100, int(round((int(_steps_val) / 10000) * 100)))) + "%") if _steps_val else ""

                _sleep_display = (f"{float(_sleep_hours):.1f}h") if _sleep_hours else "-"
                _sl_status, _sl_cls = sleep_status(_sleep_hours)
                _sleep_status_text = _sl_status or ""

                _wc_int = int(_water_count) if _water_count is not None else None
                _water_display = (f"{_wc_int} / 8 glasses") if _wc_int is not None else "-"
                _water_status_text = (str(min(100, int(round((_wc_int / 8) * 100)))) + "%") if _wc_int else ""

                _tiles = [
                    ("md-accent-pink", "favorite", "Heart Rate", _heart_rate_display, _heart_rate_status_text, "md-line-pink"),
                    ("md-accent-green", "directions_walk", "Steps", _steps_display, _steps_status_text, "md-line-green"),
                    ("md-accent-purple", "bedtime", "Sleep", _sleep_display, _sleep_status_text, "md-line-purple"),
                    ("md-accent-blue", "water_drop", "Water Intake", _water_display, _water_status_text, "md-line-blue"),
                ]
                snap_html = (
                    '<div class="md-rcard md-snap-card">'
                    '<div class="md-rcard-head"><div class="md-rcard-title">' + _snap_title + '</div><span class="md-rcard-link md-rcard-link-btn" style="cursor: default;">See all</span></div>'
                    '<div class="md-snap-grid">'
                )
                for _cls, _icon, _lbl, _val, _status, _line_cls in _tiles:
                    snap_html += (
                        '<div class="md-snap-tile">'
                        '<div class="md-snap-icon ' + _cls + ' material-symbols-rounded">' + ui_escape(_icon) + '</div>'
                        '<div class="md-snap-text">'
                        '<div><div class="md-snap-label">' + ui_text(_lbl, 40) + '</div>'
                        '<div class="md-snap-value">' + ui_text(_val, 40) + '</div></div>'
                        '<div class="md-snap-status">' + ui_text(_status, 20) + '</div>'
                        '</div>'
                        '<svg class="md-spark ' + _line_cls + '" viewBox="0 0 96 28" aria-hidden="true"><path d="M2 18 C12 18 14 11 24 14 S38 24 48 15 S62 2 72 10 S84 19 94 13" /></svg>'
                        '</div>'
                    )
                snap_html += '</div></div>'
                st.markdown(snap_html, unsafe_allow_html=True)
                st.markdown('<div class="md-rail-link-btn">', unsafe_allow_html=True)
                if st.button("See all health data →", key="home_overview_see_all", use_container_width=True):
                    st.session_state.mode = "overview"
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

                # Recent Conversations (real, from Firestore)
                recent_html = '<div class="md-rcard md-rcard-recent"><div class="md-rcard-head"><div class="md-rcard-title">Recent Conversations</div><span class="md-rcard-link md-rcard-link-btn" style="cursor: default;">See all</span></div>'
                if st.session_state.is_authenticated and st.session_state.user_email_hash:
                    _recent = list_conversations(st.session_state.user_email_hash, limit=4)
                else:
                    _recent = []  # Real conversations only: no mock entries.
                if _recent:
                    for _r in _recent:
                        _rt = (_r.get("title") or "Chat")[:32]
                        _ru = _r.get("last_updated")
                        try:
                            if _ru and hasattr(_ru, "strftime"):
                                _delta = datetime.utcnow() - _ru.replace(tzinfo=None) if _ru.tzinfo else datetime.utcnow() - _ru
                                _hours = int(_delta.total_seconds() // 3600)
                                if _hours < 1:
                                    _ago = str(int(_delta.total_seconds() // 60)) + "m ago"
                                elif _hours < 24:
                                    _ago = str(_hours) + "h ago"
                                else:
                                    _ago = str(_hours // 24) + "d ago"
                            else:
                                _ago = ""
                        except Exception:
                            _ago = ""
                        recent_html += '<div class="md-conv-row"><div class="md-conv-bubble material-symbols-rounded">chat_bubble</div><div class="md-conv-title">' + ui_text(_rt, 40) + '</div><div class="md-conv-time">' + ui_text(_ago, 20) + '</div></div>'
                else:
                    if st.session_state.is_authenticated:
                        recent_html += '<div class="md-conv-row md-conv-empty">No recent conversations yet. Start a chat to see it here.</div>'
                    else:
                        recent_html += '<div class="md-conv-row md-conv-empty">Sign in to save and revisit your conversations.</div>'
                recent_html += '</div>'
                st.markdown(recent_html, unsafe_allow_html=True)
                st.markdown('<div class="md-view-all-wrap">', unsafe_allow_html=True)
                if st.button("View all chats →", key="view_all_recent", use_container_width=True):
                    st.session_state.mode = "history"
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

                # Live Health Passport (fills empty right-rail area with a real, usable feature).
                _records_total = len(list_health_records())
                _meds_total = len(list_medications())
                _appts_total = len(list_appointments())
                _vitals_logged = bool(_hr or _steps_val or _sleep_hours or (_water_count is not None and int(_water_count) > 0))
                _has_history = bool(_convs_total) if st.session_state.is_authenticated else bool(st.session_state.messages)
                _profile_ready = bool(st.session_state.is_authenticated and st.session_state.user_email_hash)
                _passport_checks = [
                    ("Profile linked", _profile_ready),
                    ("Chat history", _has_history),
                    ("Vitals logged", _vitals_logged),
                    ("Records uploaded", _records_total > 0),
                    ("Medications synced", _meds_total > 0),
                    ("Appointments tracked", _appts_total > 0),
                ]
                _passport_done = sum(1 for _, _ok in _passport_checks if _ok)
                _passport_total = len(_passport_checks)
                _passport_pct = int(round((_passport_done / _passport_total) * 100)) if _passport_total else 0

                _passport_html = (
                    '<div class="md-rcard md-passport-card">'
                    '<div class="md-passport-head">'
                    '<div class="md-passport-title-wrap"><span class="material-symbols-rounded">badge</span><div class="md-passport-title">Health Passport</div></div>'
                    '<div class="md-passport-pct">' + str(_passport_pct) + '% ready</div>'
                    '</div>'
                    '<div class="md-passport-sub">Live profile completeness across vitals, records, medications, and appointments.</div>'
                    '<div class="md-passport-progress"><div class="md-passport-fill" style="width:' + str(_passport_pct) + '%;"></div></div>'
                )
                for _label, _ok in _passport_checks:
                    _row_cls = "ok" if _ok else "todo"
                    _icon = "check_circle" if _ok else "radio_button_unchecked"
                    _status = "Complete" if _ok else "Pending"
                    _passport_html += (
                        '<div class="md-passport-check ' + _row_cls + '">'
                        '<div class="md-passport-check-left">'
                        '<span class="material-symbols-rounded">' + _icon + '</span>'
                        '<span class="md-passport-check-label">' + ui_text(_label, 40) + '</span>'
                        '</div>'
                        '<span class="md-passport-status ' + _row_cls + '">' + _status + '</span>'
                        '</div>'
                    )
                _passport_html += '</div>'
                st.markdown(_passport_html, unsafe_allow_html=True)

                _pp1, _pp2 = st.columns(2, gap="small")
                with _pp1:
                    if st.button("Open Records", key="home_passport_records", use_container_width=True, icon=":material/folder_managed:"):
                        st.session_state.mode = "records"
                        st.rerun()
                with _pp2:
                    if st.button("Update Vitals", key="home_passport_overview", use_container_width=True, icon=":material/monitoring:"):
                        st.session_state.mode = "overview"
                        st.rerun()
                if st.button("Sync Medications", key="home_passport_meds", use_container_width=True, icon=":material/pill:"):
                    st.session_state.mode = "medications"
                    st.rerun()
                # Daily Health Tip carousel is rendered below Smart Actions in home_main.



    show_hero = False

    if show_hero:
        st.markdown(
            '<div class="md-name-card">'
            '<div class="md-name-title">Personalise the conversation</div>'
            '<div class="md-name-subtitle">Add a first name if you want MediChat to greet you more naturally. You can skip this and start chatting right away.</div>'
            '</div>',
            unsafe_allow_html=True
        )

        with st.form(key="name_form", clear_on_submit=True):
            name_cols = st.columns([3, 1, 1])
            with name_cols[0]:
                name_typed = st.text_input("", placeholder="Type your first name...", label_visibility="collapsed")
            with name_cols[1]:
                name_submit = st.form_submit_button("Save")
            with name_cols[2]:
                skip_name = st.form_submit_button("Skip")
        if name_submit and name_typed.strip():
            st.session_state.patient_name = name_typed.strip()[:20]
            if st.session_state.is_authenticated and st.session_state.user_email_hash:
                persist_profile_state(st.session_state.user_email_hash, name=st.session_state.patient_name)
            st.rerun()
        if skip_name:
            st.session_state.patient_name = "Guest"
            st.rerun()
    elif not st.session_state.messages and st.session_state.patient_name:
        pass

    else:
        user_initial = "U"
        if st.session_state.patient_name and st.session_state.patient_name != "Guest":
            user_initial = st.session_state.patient_name[0].upper()
        user_name_label = st.session_state.patient_name if st.session_state.patient_name and st.session_state.patient_name != "Guest" else "You"

        # Page hero at top of every chat conversation (matches the design
        # mockup: title + subtitle + privacy indicator).
        st.markdown(
            '<div class="md-chat-hero">'
            '<div class="md-chat-hero-text">'
            '<div class="md-chat-hero-title">AI Health Conversation '
            # Inline SVG shield — zero font dependency, renders identically
            # across all browsers as a proper filled shield silhouette with
            # a soft check mark inside. Replaces the Material Symbols
            # `shield` glyph which was rendering ambiguously at small sizes.
            '<svg class="md-chat-hero-shield" xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor" aria-hidden="true">'
            '<path d="M12 2 3.5 5.5v6c0 5.2 3.6 9.9 8.5 11 4.9-1.1 8.5-5.8 8.5-11v-6L12 2z"/>'
            '<path d="M10.6 14.4 8 11.8l-1.1 1.1 3.7 3.7 7-7-1.1-1.1-5.9 5.9z" fill="#ffffff"/>'
            '</svg></div>'
            '<div class="md-chat-hero-sub">Private, supportive guidance for your symptoms</div>'
            '</div>'
            '<div class="md-chat-hero-privacy">'
            '<span class="material-symbols-rounded">lock</span>'
            'Your conversation is private'
            '</div>'
            '</div>',
            unsafe_allow_html=True
        )

        # MediChat brand logo loaded once for bot avatars across the conversation.
        _bot_avatar_uri = get_brand_logo_data_uri()
        _bot_avatar_html = (
            '<div class="av av-bot av-bot-image"><img src="' + _bot_avatar_uri + '" alt="MediChat AI"></div>'
            if _bot_avatar_uri
            else '<div class="av av-bot">M</div>'
        )

        for msg in st.session_state.messages:
            role = msg.get("role", "")
            content = msg.get("content", "")
            msg_type = msg.get("type", "text")
            msg_ts = msg.get("ts", "")
            ts_html = ('<div class="bubble-ts">' + ui_escape(msg_ts) + '</div>') if msg_ts else ""
            if role == "user":
                safe_content = ui_lines(content)
                safe_initial = ui_text(user_initial, 2)
                if msg_type == "image":
                    st.markdown('<span class="image-tag">Medical image uploaded for analysis</span>', unsafe_allow_html=True)
                if content or msg_type != "image":
                    st.markdown(
                        '<div class="user-wrap">'
                        '<div class="user-stack">'
                        '<div class="user-bubble">' + safe_content + '</div>'
                        + ts_html +
                        '</div>'
                        '<div class="av av-user"><span class="material-symbols-rounded" style="font-size: 1.25rem;">person</span></div>'
                        '</div>',
                        unsafe_allow_html=True
                    )
            else:
                # Bot message — logo avatar + bubble with inline "MediChat AI"
                # header (with sparkle), the response text, and a right-aligned
                # timestamp under the bubble.
                st.markdown(
                    '<div class="bot-wrap">'
                    + _bot_avatar_html +
                    '<div class="bot-stack">'
                    '<div class="bot-bubble">'
                    '<div class="bot-bubble-head">'
                    '<span class="bot-bubble-name">MediChat AI</span>'
                    '<span class="bot-bubble-spark material-symbols-rounded">auto_awesome</span>'
                    '</div>'
                    + markdown_to_html(content) +
                    '</div>'
                    '</div>'
                    '</div>',
                    unsafe_allow_html=True
                )
                engine_used = msg.get("engine", "")
                msg_sources = msg.get("sources", [])
                is_image_response = "Image Analysis" in msg_sources or engine_used == "groq-vision"
                is_rx_response = "Prescription Reader" in msg_sources or str(engine_used).startswith("rx-reader")
                is_pdf_response = "PDF Report Analysis" in msg_sources

                engine_html = ""
                if str(engine_used).startswith("rx-reader"):
                    engine_html = '<span class="engine-badge engine-vision">Prescription Reader</span>'
                if is_rx_response:
                    source_tags = '<span class="source-tag">📝 Prescription Reader</span>'
                elif is_image_response:
                    source_tags = '<span class="source-tag">📷 Image Analysis</span>'
                elif is_pdf_response:
                    source_tags = '<span class="source-tag">📄 PDF Report Analysis</span>'
                else:
                    source_tags = "".join(['<span class="source-tag">📚 ' + ui_text(s, 50) + '</span>' for s in msg_sources])

                conf_html = ""
                conf_level = msg.get("confidence")
                conf_pct = msg.get("confidence_pct")
                if conf_level and conf_pct and not is_image_response and not is_rx_response and engine_used != "system":
                    conf_level = conf_level if conf_level in ("high", "medium", "low") else "low"
                    conf_label = {"high": "High Confidence", "medium": "Medium Confidence", "low": "Low Confidence"}.get(conf_level, "")
                    conf_color = {"high": "#22c55e", "medium": "#f59e0b", "low": "#ef4444"}.get(conf_level, "#64748b")
                    conf_html = (
                        '<span class="confidence-pill conf-' + conf_level + '">' + conf_label + '</span>'
                        '<span class="confidence-bar"><span class="confidence-fill" style="width:' + str(conf_pct) + '%;background:' + conf_color + ';"></span></span>'
                        '<span class="rag-text">' + str(conf_pct) + '% match</span>'
                    )

                if engine_html or source_tags or conf_html:
                    parts = []
                    if engine_html or source_tags:
                        parts.append('<span class="meta-label meta-label-sources"><span class="material-symbols-rounded">verified_user</span>Sources</span>' + engine_html + source_tags)
                    if conf_html:
                        parts.append('<span class="meta-label meta-label-conf">Confidence</span>' + conf_html)
                    st.markdown('<div class="meta-row">' + "".join(parts) + '</div>', unsafe_allow_html=True)

                if msg_ts:
                    st.markdown('<div class="bot-ts">' + ui_escape(msg_ts) + '</div>', unsafe_allow_html=True)

    if st.session_state.messages:
        # Feedback row removed per user request.
        st.markdown("---")
        st.markdown("**" + L["download_chat"] + "**")

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            if st.button(L["download_chat_btn"], use_container_width=True, key="dl_chat_btn"):
                pdf_bytes = generate_chat_pdf(st.session_state.messages)
                st.download_button(label="Click here to save your PDF", data=pdf_bytes, file_name="MediChat_Conversation_" + datetime.now().strftime("%Y%m%d_%H%M") + ".pdf", mime="application/pdf", use_container_width=True, key="dl_chat_dl")

        with dl_col2:
            if st.button("📋 Doctor Visit Summary", use_container_width=True, key="dl_summary_btn", help="Generates a structured one-page summary you can hand to your GP at your next appointment"):
                with st.spinner("Preparing your doctor visit summary..."):
                    summary_pdf, _summary_text = generate_doctor_visit_summary(st.session_state.messages, st.session_state.patient_name)
                if summary_pdf:
                    st.download_button(
                        label="Save Doctor Visit Summary PDF",
                        data=summary_pdf,
                        file_name="MediChat_DoctorVisitSummary_" + datetime.now().strftime("%Y%m%d_%H%M") + ".pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="dl_summary_dl"
                    )
                else:
                    st.error("Could not generate summary. Please try again or have a longer conversation first.")

        st.markdown(
            '<div style="font-size:0.72rem;color:var(--neutral-500);margin-top:0.5rem;text-align:center;font-style:italic;">'
            'The Doctor Visit Summary is a structured one-pager you can bring to your next GP appointment.'
            '</div>',
            unsafe_allow_html=True
        )


    st.markdown('<div id="chat-input-anchor"></div>', unsafe_allow_html=True)

    # ── Text input + Send (inside form so Enter key submits) ────────────
    user_input = home_user_input
    submit = home_submit or home_vision_analyze
    uploaded_image = home_uploaded_image
    clear = False
    if not home_empty_chat:
        uploaded_image = None
        chat_upload_clicked = False
        chat_voice_clicked = False
        chat_clear_clicked = False
        with st.form("chat_form", clear_on_submit=True):
            user_input = st.text_area(
                "Your message",
                placeholder="Ask anything about your health...",
                label_visibility="collapsed",
                height=100,
                key="chat_input_" + str(st.session_state.chat_input_key),
            )
            # Action row: Upload + Voice + Clear pills on the left, big spacer,
            # round Send ball on the right. Clear lives inside the form so
            # it's visually grouped with the chat box (no orphan button below).
            fc1, fc2, fc3, fc_spacer, fc4 = st.columns([0.56, 0.56, 0.56, 5.2, 0.56], vertical_alignment="center")
            with fc1:
                chat_upload_clicked = st.form_submit_button("Upload", key="chat_upload_btn", icon=":material/attach_file:", use_container_width=True)
            with fc2:
                chat_voice_clicked = st.form_submit_button("Voice", key="chat_voice_btn", icon=":material/mic:", use_container_width=True)
            with fc3:
                chat_clear_clicked = st.form_submit_button("Clear", key="chat_clear_btn", icon=":material/delete:", use_container_width=True)
            with fc4:
                submit = st.form_submit_button(" ", key="chat_send_btn", icon=":material/send:", use_container_width=True, type="primary")
        # If user hit Clear inside the chat form, wire it to the same clear
        # action as the standalone button used by vision/voice panels.
        if chat_clear_clicked:
            clear = True

        if chat_upload_clicked:
            st.session_state.home_show_vision_upload = True
            st.session_state.home_show_voice = False
            st.rerun()
        if chat_voice_clicked:
            st.session_state.home_show_voice = True
            st.session_state.home_show_vision_upload = False
            st.rerun()

        if st.session_state.get("home_show_vision_upload", False):
            st.markdown(
                '<div class="md-vision-panel">'
                '<div class="md-panel-icon material-symbols-rounded">image_search</div>'
                '<div><div class="md-panel-title">Vision Ai upload</div>'
                '<div class="md-panel-subtitle">Upload an X-ray, scan photo, skin image, or blood report PDF for cautious visual/report analysis.</div></div>'
                '</div>',
                unsafe_allow_html=True
            )
            uploaded_image = st.file_uploader(
                "Upload for Vision Ai",
                type=["jpg", "jpeg", "png", "pdf"],
                label_visibility="collapsed",
                key="chat_vision_uploader_" + str(st.session_state.uploader_key),
                help="Upload a medical image or PDF report for Vision Ai analysis.",
            )
            if uploaded_image:
                is_chat_pdf = uploaded_image.name.lower().endswith(".pdf")
                if is_chat_pdf:
                    st.markdown(
                        '<div class="md-file-preview md-home-file-preview">'
                        '<div class="md-file-icon">PDF</div>'
                        '<div style="flex:1;min-width:0;">'
                        '<div class="md-file-name">' + ui_text(uploaded_image.name, 90) + '</div>'
                        '<div class="md-file-help">Ready for Vision Ai report review.</div>'
                        '</div>'
                        '</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.image(uploaded_image, caption="Ready for Vision Ai analysis", use_column_width=True)
            vu1, vu2, vu3 = st.columns([1, 1, 1])
            with vu1:
                home_vision_analyze = st.button("Analyze with Vision Ai", key="chat_vision_analyze", type="primary", use_container_width=True, icon=":material/image_search:")
                submit = submit or home_vision_analyze
                if home_vision_analyze and not uploaded_image:
                    st.error("Please upload an image or PDF first.")
                    submit = False
            with vu2:
                if st.button("Cancel upload", key="chat_vision_cancel", use_container_width=True, icon=":material/close:"):
                    st.session_state.home_show_vision_upload = False
                    st.session_state.uploader_key = st.session_state.get("uploader_key", 0) + 1
                    st.rerun()
            with vu3:
                clear = st.button("Clear chat", use_container_width=True, key="main_clear_btn", icon=":material/delete:")

        if st.session_state.get("home_show_voice", False):
            st.markdown(
                '<div class="md-voice-panel">'
                '<div class="md-panel-icon material-symbols-rounded">mic</div>'
                '<div><div class="md-panel-title">Voice note</div>'
                '<div class="md-panel-subtitle">Record or upload a short voice question. MediChat will transcribe it before answering.</div></div>'
                '</div>',
                unsafe_allow_html=True
            )
            voice_audio = None
            if hasattr(st, "audio_input"):
                voice_audio = st.audio_input("Record your question", key="voice_audio_chat_" + str(st.session_state.voice_audio_key))
            else:
                voice_audio = st.file_uploader(
                    "Upload a voice note",
                    type=["wav", "mp3", "m4a", "ogg", "webm"],
                    key="voice_upload_chat_" + str(st.session_state.voice_audio_key),
                    help="Your Streamlit version does not expose microphone recording, so upload an audio note instead.",
                )
            vc1, vc2, vc3 = st.columns([1, 1, 1])
            with vc1:
                if st.button("Transcribe voice", key="voice_transcribe_chat", type="primary", use_container_width=True, icon=":material/graphic_eq:"):
                    transcript = transcribe_voice_note(voice_audio)
                    if transcript:
                        st.session_state.pending_user_input = transcript
                        st.session_state.home_show_voice = False
                        st.session_state.voice_audio_key = st.session_state.get("voice_audio_key", 0) + 1
                        st.rerun()
                    else:
                        st.error("I could not transcribe that audio. Please try a clearer recording or type the question.")
            with vc2:
                if st.button("Cancel voice", key="voice_cancel_chat", use_container_width=True, icon=":material/close:"):
                    st.session_state.home_show_voice = False
                    st.session_state.voice_audio_key = st.session_state.get("voice_audio_key", 0) + 1
                    st.rerun()
            with vc3:
                clear = st.button("Clear chat", use_container_width=True, key="main_clear_btn_voice", icon=":material/delete:")

        # Plain-state Clear button removed — it's now an inline chip inside
        # the chat composer form (chat_clear_btn). Vision/voice panels still
        # have their own Clear next to Cancel above.

    # ── File preview (shown below button row once attached) ─────────────
    if uploaded_image and not home_empty_chat and not st.session_state.get("home_show_vision_upload", False):
        is_pdf_upload = uploaded_image.name.lower().endswith(".pdf")
        if is_pdf_upload:
            st.markdown(
                '<div class="md-file-preview">'
                '<div class="md-file-icon">PDF</div>'
                '<div style="flex:1;min-width:0;">'
                '<div class="md-file-name">' + ui_text(uploaded_image.name, 90) + '</div>'
                '<div class="md-file-help">Ready for analysis. Ask MediChat what you want to know about this report.</div>'
                '</div>'
                '</div>',
                unsafe_allow_html=True
            )
        else:
            ia, ib, ic = st.columns([1, 2, 1])
            with ib:
                st.image(uploaded_image, caption="Ready for analysis", use_column_width=True)

    if not home_empty_chat:
        st.markdown(
            '<div class="md-home-composer-note">'
            '<span class="material-symbols-rounded md-disclaimer-shield">lock</span>'
            'This is not emergency care. If you feel seriously unwell, seek immediate medical attention.'
            '</div>',
            unsafe_allow_html=True
        )
    st.markdown('<div id="page-bottom-anchor" style="height:1px;"></div>', unsafe_allow_html=True)

    # Enter-to-send (Shift+Enter keeps newline) for both home and in-chat composers.
    import streamlit.components.v1 as _components
    _components.html(
        """
        <script>
            (function () {
                const host = window.parent;
                if (!host || !host.document) return;
                if (host.__medichatEnterSendBound) return;
                host.__medichatEnterSendBound = true;

                function isComposerTextarea(el) {
                    if (!el || el.tagName !== "TEXTAREA") return false;
                    const label = (el.getAttribute("aria-label") || "").trim();
                    return label === "Start a chat" || label === "Your message";
                }

                function findSendButton(textarea) {
                    const form = textarea.closest('[data-testid="stForm"]');
                    if (!form) return null;
                    return form.querySelector(
                        '.st-key-home_send_btn button, .st-key-chat_send_btn button, ' +
                        '[data-testid="stFormSubmitButton"] > button[kind="primaryFormSubmit"], ' +
                        '[data-testid="stFormSubmitButton"] > button[kind="primary"]'
                    );
                }

                host.document.addEventListener("keydown", function (ev) {
                    if (ev.defaultPrevented) return;
                    if (ev.key !== "Enter") return;
                    if (ev.shiftKey || ev.ctrlKey || ev.metaKey || ev.altKey || ev.isComposing) return;
                    const target = ev.target;
                    if (!isComposerTextarea(target)) return;

                    ev.preventDefault();
                    const btn = findSendButton(target);
                    if (btn && !btn.disabled) btn.click();
                }, true);
            })();
        </script>
        """,
        height=0,
    )

    if st.session_state.messages:
        _components.html(
            """
            <script>
                (function() {
                    function scrollToBottom() {
                        try {
                            const doc = window.parent.document;
                            const anchor = doc.getElementById('page-bottom-anchor');
                            if (anchor) {
                                anchor.scrollIntoView({ behavior: 'smooth', block: 'end' });
                                return;
                            }
                            const mainBlock = doc.querySelector('section.main') ||
                                              doc.querySelector('[data-testid="stAppViewContainer"]') ||
                                              doc.querySelector('.main');
                            if (mainBlock) {
                                mainBlock.scrollTo({ top: mainBlock.scrollHeight, behavior: 'smooth' });
                            }
                            window.parent.scrollTo({ top: doc.body.scrollHeight, behavior: 'smooth' });
                        } catch (e) {}
                    }
                    setTimeout(scrollToBottom, 200);
                    setTimeout(scrollToBottom, 600);
                    setTimeout(scrollToBottom, 1200);
                })();
            </script>
            """,
            height=0,
        )

    if clear:
        # Start a fresh chat. The existing conversation stays saved in history.
        st.session_state.last_pdf_name = ""
        st.session_state.emergency_reason = ""
        start_new_chat_session()
        st.rerun()

    pending_user_input = st.session_state.get("pending_user_input", "")
    effective_user_input = pending_user_input if pending_user_input else user_input
    auto_submit = bool(pending_user_input)

    if (submit or auto_submit) and (effective_user_input.strip() or uploaded_image):
        if auto_submit:
            st.session_state.pending_user_input = ""
        st.session_state.qcount += 1
        lang_instruction = LANGUAGES[st.session_state.selected_language]["lang_instruction"]

        if effective_user_input.strip():
            conv_text = " ".join([m.get("content", "") for m in st.session_state.messages if m.get("type") == "text"])
            if is_meta_text(effective_user_input):
                # Pasted docs / test plans / app-feedback — not a symptom report.
                st.session_state.emergency_detected = False
                st.session_state.emergency_reason = ""
                st.session_state.triage_assessment = None
            else:
                is_emerg, reason = detect_emergency(effective_user_input, conv_text)
                if is_emerg:
                    st.session_state.emergency_detected = True
                    st.session_state.emergency_reason = reason
                st.session_state.triage_assessment = assess_triage_tier(effective_user_input, conv_text, st.session_state.patient_memory)

        if uploaded_image:
            is_pdf = uploaded_image.name.lower().endswith(".pdf")
            if is_pdf:
                st.session_state.messages.append({"role": "user", "type": "pdf", "content": (effective_user_input.strip() + " " if effective_user_input.strip() else "") + "[PDF: " + uploaded_image.name + "]", "ts": _msg_now_ts()})
                with st.spinner("Reading your medical report..."):
                    pdf_text = extract_pdf_text(uploaded_image)
                    if not pdf_text:
                        reply = "I had trouble reading that PDF. It might be image-based (scanned) rather than text-based. Could you try uploading it as a JPEG or PNG image instead?"
                        engine_used = "system"
                    else:
                        st.session_state.last_pdf_context = pdf_text
                        st.session_state.last_pdf_name = uploaded_image.name
                        reply, engine_used = medichat_pdf_analysis(effective_user_input, pdf_text, st.session_state.messages, lang_instruction)
                        reply = strip_excessive_disclaimers(reply)
                st.session_state.last_sources = ["PDF Report Analysis"]
                st.session_state.messages.append({"role": "assistant", "type": "text", "content": reply, "sources": st.session_state.last_sources, "confidence": "medium", "confidence_pct": 75, "engine": engine_used, "ts": _msg_now_ts()})
                st.session_state.uploader_key += 1
                st.session_state.home_show_vision_upload = False
            else:
                st.session_state.messages.append({"role": "user", "type": "image", "content": effective_user_input.strip(), "ts": _msg_now_ts()})
                if looks_like_prescription_request(effective_user_input):
                    with st.spinner("Reading prescription handwriting..."):
                        uploaded_image.seek(0)
                        rx_result = read_prescription(
                            uploaded_image.read(),
                            user_note=effective_user_input,
                            lang_instruction=lang_instruction,
                        )
                        reply = strip_excessive_disclaimers(rx_result.get("reading", ""))
                        vision_engine = "rx-reader:" + rx_result.get("model_used", "unknown")
                        conf_map = {"high": 88, "medium": 72, "low": 56}
                        conf_level = rx_result.get("overall_confidence", "medium")
                        conf_pct = conf_map.get(conf_level, 70)
                    st.session_state.last_image_context = (
                        "User uploaded prescription image: " + uploaded_image.name + "\n"
                        "User's prompt: " + (effective_user_input.strip() if effective_user_input.strip() else "(no question provided)") + "\n"
                        "Prescription transcription: " + reply
                    )
                    st.session_state.last_sources = ["Prescription Reader"]
                    st.session_state.messages.append({"role": "assistant", "type": "text", "content": reply, "sources": st.session_state.last_sources, "confidence": conf_level, "confidence_pct": conf_pct, "engine": vision_engine, "ts": _msg_now_ts()})
                else:
                    with st.spinner("Analysing your image..."):
                        uploaded_image.seek(0)
                        reply, vision_engine = medichat_vision(effective_user_input, encode_image(uploaded_image), st.session_state.messages, lang_instruction)
                        reply = strip_excessive_disclaimers(reply)
                    st.session_state.last_image_context = (
                        "User uploaded image: " + uploaded_image.name + "\n"
                        "User's question about image: " + (effective_user_input.strip() if effective_user_input.strip() else "(no question, just the image)") + "\n"
                        "Your visual analysis: " + reply
                    )
                    st.session_state.last_sources = ["Image Analysis"]
                    st.session_state.messages.append({"role": "assistant", "type": "text", "content": reply, "sources": st.session_state.last_sources, "confidence": "medium", "confidence_pct": 75, "engine": vision_engine, "ts": _msg_now_ts()})
                st.session_state.uploader_key += 1
                st.session_state.home_show_vision_upload = False
        else:
            user_msg = {"role": "user", "type": "text", "content": effective_user_input.strip()}
            st.session_state.messages.append(user_msg)
            _t0 = time.time()

            name_for_rag = "" if st.session_state.patient_name == "Guest" else st.session_state.patient_name

            # Build cross-chat summary so the AI knows about the patient's other conversations.
            past_chats_summary = ""
            if st.session_state.is_authenticated and st.session_state.user_email_hash:
                _all_convs = list_conversations(st.session_state.user_email_hash, limit=10)
                _other = [c for c in _all_convs if c["id"] != st.session_state.current_conversation_id]
                if _other:
                    _lines = []
                    for c in _other[:5]:
                        _topic = (c.get("first_user_msg") or c.get("title") or "").strip().replace("\n", " ")
                        _outcome = (c.get("last_assistant_msg") or "").strip().replace("\n", " ")
                        if not _topic:
                            continue
                        _entry = "- Topic: " + _topic[:200]
                        if _outcome:
                            _entry += "\n  Last response summary: " + _outcome[:240]
                        _lines.append(_entry)
                    past_chats_summary = "\n".join(_lines)

            # Ensure _bot_avatar_html is always defined regardless of which
            # history-rendering branch executed above (the elif/pass branch
            # skips the definition when messages were empty at render time).
            if "_bot_avatar_html" not in dir():
                _bot_avatar_uri_fallback = get_brand_logo_data_uri()
                _bot_avatar_html = (
                    '<div class="av av-bot av-bot-image"><img src="' + _bot_avatar_uri_fallback + '" alt="MediChat AI"></div>'
                    if _bot_avatar_uri_fallback
                    else '<div class="av av-bot">M</div>'
                )

            with st.spinner("MediChat is analysing"):
                stream_placeholder = st.empty()
                final_text = ""
                stream_metadata = None
                
                try:
                    for event in medichat_rag_stream(effective_user_input, st.session_state.messages, lang_instruction, name_for_rag, st.session_state.get("last_pdf_context", ""), st.session_state.get("last_image_context", ""), past_chats_summary):
                        kind = event[0]
                        if kind == "chunk":
                            final_text = event[2]
                            stream_placeholder.markdown(
                                '<div class="bot-wrap">' + _bot_avatar_html + '<div class="bot-stack">'
                                '<div class="bot-bubble">'
                                '<div class="bot-bubble-head">'
                                '<span class="bot-bubble-name">MediChat AI</span>'
                                '<span class="bot-bubble-spark material-symbols-rounded">auto_awesome</span>'
                                '</div>' + markdown_to_html(final_text) + '<span class="stream-cursor"></span></div>'
                                '</div>'
                                '</div>', unsafe_allow_html=True
                            )
                        elif kind == "done":
                            final_text = event[1]
                            stream_metadata = event[2]
                    
                    stream_placeholder.empty()
                    
                except Exception as streaming_fault:
                    import traceback as _tb
                    _tb.print_exc()
                    print(f"Streaming crash detail: {type(streaming_fault).__name__}: {str(streaming_fault)}")
                    stream_placeholder.empty()
                    _err_msg = str(streaming_fault) or "Unknown internal error"
                    st.markdown(
                        f'<div class="md-mini-error">MediChat had trouble generating a response. '
                        f'Error: {_err_msg[:300]}. Please try again or check your API keys in Streamlit secrets.</div>',
                        unsafe_allow_html=True,
                    )
                    st.stop()

            final_text = strip_excessive_disclaimers(final_text)

            if stream_metadata is None:
                stream_metadata = {"memory": st.session_state.patient_memory, "sources": [], "confidence": "low", "confidence_pct": 0, "engine": "unknown"}

            memory = stream_metadata["memory"]
            sources = stream_metadata["sources"]
            conf_level = stream_metadata["confidence"]
            conf_pct = stream_metadata["confidence_pct"]
            engine_used = stream_metadata.get("engine", "unknown")
            # Merge fresh extraction with existing profile memory so facts
            # from past chats persist into new ones.
            existing_mem = st.session_state.patient_memory or {}
            merged_mem = {}
            for _key in ("symptoms", "conditions", "medications"):
                _combined = list(dict.fromkeys((existing_mem.get(_key, []) or []) + (memory.get(_key, []) or [])))
                merged_mem[_key] = _combined[:30]
            st.session_state.patient_memory = merged_mem
            memory = merged_mem
            st.session_state.last_sources = sources
            if st.session_state.is_authenticated and st.session_state.user_email_hash:
                persist_profile_state(
                    st.session_state.user_email_hash,
                    patient_memory=memory,
                    name=st.session_state.patient_name or None,
                    language=st.session_state.selected_language,
                )
            _response_time = round(time.time() - _t0, 2)
            st.session_state.response_times.append(_response_time)

            interaction_alerts = check_drug_interactions(final_text, memory)
            if interaction_alerts:
                alert_block = "\n\n---\n\n**Drug Safety Check:**\n"
                for a in interaction_alerts:
                    alert_block += "\n- **" + a["drug"] + "**, given your " + ", ".join(a["conditions"]) + ": " + a["warning"]
                final_text = final_text + alert_block

            _log_entry = {
                "query": effective_user_input.strip(),
                "confidence": conf_level,
                "confidence_pct": conf_pct,
                "sources": sources,
                "response_time": _response_time,
                "language": st.session_state.selected_language,
                "mode": "free_chat",
                "emergency_triggered": st.session_state.emergency_detected,
                "drug_alerts": len(interaction_alerts),
                "engine": engine_used,
            }
            st.session_state.eval_log.append(_log_entry)
            log_query_to_firestore(_log_entry)

            st.session_state.messages.append({"role": "assistant", "type": "text", "content": final_text, "sources": sources, "confidence": conf_level, "confidence_pct": conf_pct, "engine": engine_used, "ts": _msg_now_ts()})

        if st.session_state.is_authenticated and st.session_state.user_email_hash:
            persist_profile_state(
                st.session_state.user_email_hash,
                patient_memory=st.session_state.patient_memory,
                name=st.session_state.patient_name or None,
                language=st.session_state.selected_language,
            )
            new_id = save_conversation(
                st.session_state.user_email_hash,
                st.session_state.current_conversation_id or "",
                st.session_state.messages,
            )
            if new_id:
                st.session_state.current_conversation_id = new_id
        st.session_state.chat_input_key = st.session_state.get("chat_input_key", 0) + 1
        st.rerun()

elif st.session_state.mode == "eval":
    st.markdown(
        '<div style="background:linear-gradient(135deg,#1f2937,#111827);color:white;padding:0.7rem 1.2rem;border-radius:12px;margin-bottom:1rem;display:flex;align-items:center;justify-content:space-between;">'
        '<div style="display:flex;align-items:center;gap:0.5rem;"><span style="font-size:1rem;">🔒</span><span style="font-weight:600;font-size:0.9rem;">Admin Mode. Research & Evaluation Dashboard</span></div>'
        '<div style="font-size:0.75rem;opacity:0.8;">Not visible to patients</div>'
        '</div>',
        unsafe_allow_html=True
    )
    st.markdown("---")
    st.markdown("### 📊 MediChat Analytics Dashboard")
    st.caption("Real-time session analytics for clinical evaluation and research reporting. All patient query text is anonymised.")

    dc1, dc2 = st.columns([3, 2])
    with dc2:
        if FIREBASE_ACTIVE:
            data_source = st.radio(
                "Data source:",
                ["Current Session", "All Patients (Firestore)"],
                horizontal=True,
                key="analytics_data_source",
                label_visibility="collapsed"
            )
        else:
            data_source = "Current Session"
            st.caption("Firestore not connected. Showing current session only.")

    if data_source == "All Patients (Firestore)" and FIREBASE_ACTIVE:
        raw_logs = fetch_all_queries_from_firestore(limit=500)
        logs = [
            {
                "query": " " * d.get("query_word_count", 0),
                "confidence": d.get("confidence", "unknown"),
                "confidence_pct": d.get("confidence_pct", 0),
                "sources": d.get("sources", []),
                "response_time": d.get("response_time", 0),
                "language": d.get("language", "English"),
                "emergency_triggered": d.get("emergency_triggered", False),
                "drug_alerts": d.get("drug_alerts", 0),
            }
            for d in raw_logs
        ]
        st.info("Live data from Firestore, aggregated from all MediChat patients (anonymised). Total records: " + str(len(logs)))
    else:
        logs = st.session_state.eval_log
        if FIREBASE_ACTIVE:
            st.caption("Current session only. Switch to 'All Patients (Firestore)' to see aggregate data.")
    total_queries = len(logs)

    if total_queries == 0:
        st.markdown(
            '<div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:14px;padding:2rem;text-align:center;color:#0369a1;">'
            '<div style="font-size:3rem;margin-bottom:0.8rem;">📊</div>'
            '<div style="font-size:1.1rem;font-weight:600;margin-bottom:0.4rem;">No data yet</div>'
            '<div style="font-size:0.9rem;">Start a chat or symptom assessment in Free Chat mode. Analytics will appear here automatically.</div>'
            '</div>',
            unsafe_allow_html=True
        )
    else:
        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1:
            st.metric("Total Queries", total_queries)
        with mc2:
            avg_conf = sum(l["confidence_pct"] for l in logs) / total_queries
            st.metric("Avg Confidence", str(round(avg_conf, 1)) + "%")
        with mc3:
            avg_time = sum(l["response_time"] for l in logs) / total_queries
            st.metric("Avg Response", str(round(avg_time, 2)) + "s")
        with mc4:
            feedback_data = st.session_state.feedback
            helpful_count = sum(1 for v in feedback_data.values() if v == "helpful")
            not_helpful = sum(1 for v in feedback_data.values() if v == "not_helpful")
            feedback_total = helpful_count + not_helpful
            feedback_pct = round((helpful_count / feedback_total * 100), 1) if feedback_total > 0 else "N/A"
            st.metric("Satisfaction", str(feedback_pct) + ("%" if feedback_total > 0 else ""))

        st.markdown("---")

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("#### Confidence Distribution")
            high = sum(1 for l in logs if l["confidence"] == "high")
            medium = sum(1 for l in logs if l["confidence"] == "medium")
            low = sum(1 for l in logs if l["confidence"] == "low")
            st.markdown(
                '<div style="padding:1rem;background:#f8fafc;border-radius:12px;border:1px solid #e2e8f0;">'
                '<div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:0.6rem;">'
                '<div style="flex:0 0 80px;font-size:0.82rem;color:#166534;font-weight:600;">High</div>'
                '<div style="flex:1;height:18px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(round((high / total_queries) * 100)) + '%;background:#22c55e;"></div>'
                '</div>'
                '<div style="flex:0 0 50px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(high) + '</div>'
                '</div>'
                '<div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:0.6rem;">'
                '<div style="flex:0 0 80px;font-size:0.82rem;color:#92400e;font-weight:600;">Medium</div>'
                '<div style="flex:1;height:18px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(round((medium / total_queries) * 100)) + '%;background:#f59e0b;"></div>'
                '</div>'
                '<div style="flex:0 0 50px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(medium) + '</div>'
                '</div>'
                '<div style="display:flex;align-items:center;gap:0.8rem;">'
                '<div style="flex:0 0 80px;font-size:0.82rem;color:#991b1b;font-weight:600;">Low</div>'
                '<div style="flex:1;height:18px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(round((low / total_queries) * 100)) + '%;background:#ef4444;"></div>'
                '</div>'
                '<div style="flex:0 0 50px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(low) + '</div>'
                '</div>'
                '</div>',
                unsafe_allow_html=True
            )

        with col_b:
            st.markdown("#### Source Usage (RAG Retrieval)")
            pubmed_queries = sum(1 for l in logs if any("PubMed" in s for s in l.get("sources", [])))
            dialog_queries = sum(1 for l in logs if any("Doctor-Patient" in s for s in l.get("sources", [])))
            both = sum(1 for l in logs if len(l.get("sources", [])) > 1)
            st.markdown(
                '<div style="padding:1rem;background:#f8fafc;border-radius:12px;border:1px solid #e2e8f0;">'
                '<div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:0.6rem;">'
                '<div style="flex:0 0 110px;font-size:0.82rem;color:#0369a1;font-weight:600;">PubMed</div>'
                '<div style="flex:1;height:18px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(round((pubmed_queries / total_queries) * 100)) + '%;background:#0ea5e9;"></div>'
                '</div>'
                '<div style="flex:0 0 50px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(pubmed_queries) + '</div>'
                '</div>'
                '<div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:0.6rem;">'
                '<div style="flex:0 0 110px;font-size:0.82rem;color:#7c3aed;font-weight:600;">Doctor-Patient</div>'
                '<div style="flex:1;height:18px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(round((dialog_queries / total_queries) * 100)) + '%;background:#a855f7;"></div>'
                '</div>'
                '<div style="flex:0 0 50px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(dialog_queries) + '</div>'
                '</div>'
                '<div style="display:flex;align-items:center;gap:0.8rem;">'
                '<div style="flex:0 0 110px;font-size:0.82rem;color:#059669;font-weight:600;">Mixed Sources</div>'
                '<div style="flex:1;height:18px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                '<div style="height:100%;width:' + str(round((both / total_queries) * 100)) + '%;background:#10b981;"></div>'
                '</div>'
                '<div style="flex:0 0 50px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(both) + '</div>'
                '</div>'
                '</div>',
                unsafe_allow_html=True
            )

        st.markdown("---")

        st.markdown("#### Safety Metrics")
        sm1, sm2, sm3 = st.columns(3)
        with sm1:
            emergency_fires = sum(1 for l in logs if l.get("emergency_triggered"))
            st.metric("🚨 Emergency Alerts Fired", emergency_fires)
        with sm2:
            drug_warnings = sum(l.get("drug_alerts", 0) for l in logs)
            st.metric("⚠️ Drug Safety Warnings", drug_warnings)
        with sm3:
            mem = st.session_state.patient_memory
            total_extracted = len(mem.get("symptoms", [])) + len(mem.get("conditions", [])) + len(mem.get("medications", []))
            st.metric("🧠 Memory Entries Extracted", total_extracted)

        st.markdown("---")

        lang_counts = {}
        for l in logs:
            lang = l.get("language", "English")
            lang_counts[lang] = lang_counts.get(lang, 0) + 1
        if len(lang_counts) > 0:
            st.markdown("#### Language Distribution")
            lang_html = '<div style="padding:1rem;background:#f8fafc;border-radius:12px;border:1px solid #e2e8f0;">'
            for lang, count in sorted(lang_counts.items(), key=lambda x: x[1], reverse=True):
                pct = round((count / total_queries) * 100)
                lang_html += (
                    '<div style="display:flex;align-items:center;gap:0.8rem;margin-bottom:0.5rem;">'
                    '<div style="flex:0 0 110px;font-size:0.82rem;color:#334155;font-weight:600;">' + lang + '</div>'
                    '<div style="flex:1;height:16px;background:#e5e7eb;border-radius:50px;overflow:hidden;">'
                    '<div style="height:100%;width:' + str(pct) + '%;background:#8b5cf6;"></div>'
                    '</div>'
                    '<div style="flex:0 0 70px;font-size:0.82rem;color:#64748b;text-align:right;">' + str(count) + ' (' + str(pct) + '%)</div>'
                    '</div>'
                )
            lang_html += '</div>'
            st.markdown(lang_html, unsafe_allow_html=True)

        st.markdown("---")

        st.markdown("#### Recent Queries Log (Anonymised)")
        st.caption("For patient privacy, query text is not displayed. Only query length, confidence, timing, and source metadata are shown.")
        log_html = '<div style="padding:1rem;background:#f8fafc;border-radius:12px;border:1px solid #e2e8f0;max-height:300px;overflow-y:auto;">'
        for i, l in enumerate(reversed(logs[-10:]), 1):
            conf_color = {"high": "#22c55e", "medium": "#f59e0b", "low": "#ef4444"}.get(l["confidence"], "#64748b")
            query_len = len(l["query"].split())
            log_html += (
                '<div style="padding:0.6rem 0.8rem;border-bottom:1px solid #e5e7eb;font-size:0.8rem;">'
                '<div style="color:#334155;margin-bottom:0.2rem;font-weight:600;">Query #' + str(total_queries - i + 1) + ', ' + str(query_len) + ' words</div>'
                '<div style="display:flex;gap:0.8rem;font-size:0.7rem;color:#64748b;flex-wrap:wrap;">'
                '<span style="color:' + conf_color + ';font-weight:600;">● ' + str(l["confidence_pct"]) + '% conf</span>'
                '<span>⏱ ' + str(l["response_time"]) + 's</span>'
                '<span>📚 ' + ", ".join(l.get("sources", []) or ["-"]) + '</span>'
                '<span>🌐 ' + l.get("language", "English") + '</span>'
                '</div>'
                '</div>'
            )
        log_html += '</div>'
        st.markdown(log_html, unsafe_allow_html=True)

        st.markdown("---")

        ec1, ec2 = st.columns([1, 1])
        with ec1:
            import io as _io
            try:
                from openpyxl import Workbook as _Workbook
                from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
                from openpyxl.utils import get_column_letter as _col_letter

                wb = _Workbook()
                thin = _Side(border_style="thin", color="CCCCCC")
                cell_border = _Border(top=thin, bottom=thin, left=thin, right=thin)
                header_fill = _Fill(start_color="1F3864", end_color="1F3864", fill_type="solid")
                header_font = _Font(name="Arial", bold=True, color="FFFFFF", size=11)
                label_font = _Font(name="Arial", bold=True, size=11, color="1F3864")
                body_font = _Font(name="Arial", size=10)

                ws1 = wb.active
                ws1.title = "Summary"
                ws1["A1"] = "MediChat Analytics Summary"
                ws1["A1"].font = _Font(name="Arial", bold=True, size=16, color="1F3864")
                ws1.merge_cells("A1:B1")
                ws1["A2"] = "Generated: " + datetime.now().strftime("%B %d, %Y at %I:%M %p")
                ws1["A2"].font = _Font(name="Arial", italic=True, size=10, color="555555")
                ws1.merge_cells("A2:B2")
                ws1["A3"] = "Privacy Note: All patient query text has been anonymised."
                ws1["A3"].font = _Font(name="Arial", italic=True, size=9, color="C2410C")
                ws1.merge_cells("A3:B3")

                summary_rows = [
                    ("Metric", "Value"),
                    ("Total Queries", total_queries),
                    ("Average Confidence (%)", round(avg_conf, 2)),
                    ("Average Response Time (seconds)", round(avg_time, 2)),
                    ("Patient Satisfaction (%)", feedback_pct if feedback_total > 0 else "No feedback yet"),
                    ("", ""),
                    ("Confidence Distribution", ""),
                    ("  High Confidence", high),
                    ("  Medium Confidence", medium),
                    ("  Low Confidence", low),
                    ("", ""),
                    ("Source Usage (RAG Retrieval)", ""),
                    ("  PubMed Research", pubmed_queries),
                    ("  Doctor-Patient Data", dialog_queries),
                    ("  Mixed Sources", both),
                    ("", ""),
                    ("Safety Metrics", ""),
                    ("  Emergency Alerts Fired", emergency_fires),
                    ("  Drug Safety Warnings Raised", drug_warnings),
                    ("  Memory Entries Extracted", total_extracted),
                ]
                for r, (label, value) in enumerate(summary_rows, start=5):
                    ws1.cell(row=r, column=1, value=label).font = header_font if r == 5 else label_font
                    ws1.cell(row=r, column=2, value=value).font = header_font if r == 5 else body_font
                    if r == 5:
                        ws1.cell(row=r, column=1).fill = header_fill
                        ws1.cell(row=r, column=2).fill = header_fill
                    ws1.cell(row=r, column=1).border = cell_border
                    ws1.cell(row=r, column=2).border = cell_border
                ws1.column_dimensions["A"].width = 42
                ws1.column_dimensions["B"].width = 22

                ws2 = wb.create_sheet("Languages")
                ws2["A1"] = "Language Distribution"
                ws2["A1"].font = _Font(name="Arial", bold=True, size=14, color="1F3864")
                ws2.merge_cells("A1:C1")
                ws2["A3"] = "Language"
                ws2["B3"] = "Query Count"
                ws2["C3"] = "Percentage"
                for c in ["A3", "B3", "C3"]:
                    ws2[c].font = header_font
                    ws2[c].fill = header_fill
                    ws2[c].border = cell_border
                row_i = 4
                for lang, count in sorted(lang_counts.items(), key=lambda x: x[1], reverse=True):
                    pct = round((count / total_queries) * 100, 1)
                    ws2.cell(row=row_i, column=1, value=lang).font = body_font
                    ws2.cell(row=row_i, column=2, value=count).font = body_font
                    ws2.cell(row=row_i, column=3, value=str(pct) + "%").font = body_font
                    for col in range(1, 4):
                        ws2.cell(row=row_i, column=col).border = cell_border
                    row_i += 1
                ws2.column_dimensions["A"].width = 18
                ws2.column_dimensions["B"].width = 15
                ws2.column_dimensions["C"].width = 15

                ws3 = wb.create_sheet("Query Log (Anonymised)")
                ws3["A1"] = "Anonymised Query Log"
                ws3["A1"].font = _Font(name="Arial", bold=True, size=14, color="1F3864")
                ws3.merge_cells("A1:H1")
                ws3["A2"] = "All patient query text has been removed. Only metadata is included for privacy compliance."
                ws3["A2"].font = _Font(name="Arial", italic=True, size=9, color="C2410C")
                ws3.merge_cells("A2:H2")

                headers3 = ["Query ID", "Word Count", "Confidence", "Confidence %", "Response Time (s)", "Sources", "Language", "Safety Triggered"]
                for col_i, h in enumerate(headers3, start=1):
                    c = ws3.cell(row=4, column=col_i, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.border = cell_border
                    c.alignment = _Align(horizontal="center")

                for row_i, l in enumerate(logs, start=5):
                    safety = "Yes" if (l.get("emergency_triggered") or l.get("drug_alerts", 0) > 0) else "No"
                    values = [
                        row_i - 4,
                        len(l["query"].split()),
                        l["confidence"].title(),
                        str(l["confidence_pct"]) + "%",
                        l["response_time"],
                        ", ".join(l.get("sources", [])) or "-",
                        l.get("language", "English"),
                        safety,
                    ]
                    for col_i, v in enumerate(values, start=1):
                        c = ws3.cell(row=row_i, column=col_i, value=v)
                        c.font = body_font
                        c.border = cell_border
                        c.alignment = _Align(horizontal="center") if col_i != 6 else _Align(horizontal="left")
                widths3 = [10, 12, 14, 14, 16, 30, 12, 16]
                for col_i, w in enumerate(widths3, start=1):
                    ws3.column_dimensions[_col_letter(col_i)].width = w

                ws4 = wb.create_sheet("Methodology")
                ws4["A1"] = "MediChat Analytics Methodology"
                ws4["A1"].font = _Font(name="Arial", bold=True, size=14, color="1F3864")
                ws4.merge_cells("A1:B1")

                method_notes = [
                    ("Confidence Score", "Calculated from FAISS L2 distance between query embedding and top-3 retrieved documents. Lower distance = higher confidence. < 0.8 = High, 0.8-1.3 = Medium, > 1.3 = Low."),
                    ("Source Classification", "PubMed Research = documents 0-499 in FAISS index (biomedical research papers). Doctor-Patient Data = documents 500-999 (real clinical conversations)."),
                    ("Emergency Detection", "Hybrid system: direct keyword match on 30+ emergency terms, plus 6 symptom-cluster patterns (cardiac, stroke, anaphylaxis, syncope, severe asthma, hyperglycemic crisis)."),
                    ("Drug Safety Warnings", "Hard-coded rules cross-reference 6 drug classes against patient-stated conditions from memory extraction."),
                    ("Response Time", "Measured from query submission to final response display, including FAISS retrieval and LLM inference."),
                    ("Privacy Compliance", "Patient query text is NEVER stored or exported. Only aggregate metadata and word counts are retained for evaluation purposes."),
                    ("Data Retention", "All analytics data is held in browser session state only. Cleared automatically when the patient closes their tab or clicks Reset."),
                ]
                for r, (label, desc) in enumerate(method_notes, start=3):
                    ws4.cell(row=r, column=1, value=label).font = label_font
                    ws4.cell(row=r, column=1).alignment = _Align(vertical="top")
                    ws4.cell(row=r, column=2, value=desc).font = body_font
                    ws4.cell(row=r, column=2).alignment = _Align(wrap_text=True, vertical="top")
                    ws4.row_dimensions[r].height = 60
                ws4.column_dimensions["A"].width = 24
                ws4.column_dimensions["B"].width = 75

                _buffer = _io.BytesIO()
                wb.save(_buffer)
                _buffer.seek(0)

                st.download_button(
                    "📊 Export Analytics (Excel)",
                    data=_buffer.getvalue(),
                    file_name="medichat_analytics_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as _ex:
                st.error("Excel export failed: " + str(_ex))
                st.info("openpyxl library missing. Add 'openpyxl' to requirements.txt and redeploy.")
        with ec2:
            if st.button("🔄 Reset Analytics", use_container_width=True):
                st.session_state.eval_log = []
                st.session_state.response_times = []
                st.rerun()

elif st.session_state.mode == "history":
    # ── Full chat history list ──────────────────────────────────────
    st.markdown('<div class="md-greet-wrap"><div class="md-greet">Your Chats</div>'
                '<div class="md-subgreet">Every conversation you have had with MediChat. Open one to continue, or start a new chat.</div></div>',
                unsafe_allow_html=True)
    if st.button("← Back to home", key="hist_back"):
        st.session_state.mode = "chat"
        st.rerun()
    if st.session_state.is_authenticated and st.session_state.user_email_hash:
        _hist = list_conversations(st.session_state.user_email_hash, limit=200)
        if not _hist:
            st.markdown('<div class="md-rcard" style="text-align:center;color:var(--md-text-3);font-style:italic;padding:1.6rem;">No past chats yet. Start one from the home screen.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="md-history-list">', unsafe_allow_html=True)
            for _h in _hist:
                _ht = (_h.get("title") or "Chat")[:80]
                _hc = _h.get("message_count", 0)
                _hu = _h.get("last_updated")
                try:
                    if _hu and hasattr(_hu, "strftime"):
                        _ago_disp = _hu.strftime("%d %b %Y · %H:%M")
                    else:
                        _ago_disp = ""
                except Exception:
                    _ago_disp = ""
                _preview = (_h.get("first_user_msg") or "")[:120]
                st.markdown(
                    '<div class="md-history-row">'
                    '<div class="md-history-bubble">💬</div>'
                    '<div class="md-history-mid">'
                    '<div class="md-history-title">' + ui_text(_ht, 80) + '</div>'
                    '<div class="md-history-meta">' + str(_hc) + ' messages · ' + ui_text(_ago_disp, 40) + '</div>'
                    + ('<div class="md-history-preview">' + ui_text(_preview, 140) + '</div>' if _preview else '') +
                    '</div></div>',
                    unsafe_allow_html=True
                )
                hc1, hc2 = st.columns([4, 1])
                with hc1:
                    if st.button("Open", key="hist_open_" + _h["id"], use_container_width=True):
                        conv = load_conversation(st.session_state.user_email_hash, _h["id"])
                        if conv is not None:
                            st.session_state.current_conversation_id = _h["id"]
                            st.session_state.messages = conv.get("messages", []) or []
                            st.session_state.qcount = sum(1 for m in st.session_state.messages if m.get("role") == "user")
                            st.session_state.feedback = {}
                            st.session_state.last_sources = []
                            st.session_state.emergency_detected = False
                            st.session_state.mode = "chat"
                            st.rerun()
                with hc2:
                    if st.button("Delete", key="hist_del_" + _h["id"], use_container_width=True):
                        delete_conversation(st.session_state.user_email_hash, _h["id"])
                        st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="md-rcard" style="text-align:center;color:var(--md-text-3);font-style:italic;padding:1.6rem;">Sign in to keep and review your chat history.</div>', unsafe_allow_html=True)

elif st.session_state.mode == "overview":
    # ── Health Overview ─────────────────────────────────────────────
    st.markdown(
        '<div class="md-page-hero md-page-hero-overview">'
        '<div class="md-page-hero-ic"><span class="material-symbols-rounded">monitoring</span></div>'
        '<div class="md-page-hero-text">'
        '<div class="md-page-hero-title">Health Overview</div>'
        '<div class="md-page-hero-sub">A live snapshot of what we know about you and what you have logged today.</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    today = get_daily_metrics()
    water_n = int(today.get("water_glasses", 0) or 0)
    sleep_h = today.get("sleep_hours")
    steps_n = today.get("steps")
    hr_n = today.get("heart_rate_resting")

    # Wearable sync placeholder (no simulated data)
    st.markdown(
        '<div class="md-wearable-card">'
        '<div class="md-wearable-icon">⌚</div>'
        '<div class="md-wearable-body">'
        '<div class="md-wearable-title">Wearable data source: Not connected</div>'
        '<div class="md-wearable-desc">Connect a wearable to track heart rate, steps, and activity automatically. Until then, health overview only uses your manually logged data.</div>'
        '<div class="md-wearable-actions">'
        '<span class="md-wearable-pill">Bluetooth</span>'
        '<span class="md-wearable-pill">Apple Health</span>'
        '<span class="md-wearable-pill">Google Fit</span>'
        '<span class="md-wearable-pill md-wearable-soon">Not connected</span>'
        '</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    ov_c, ov_d = st.columns(2)
    with ov_c:
        sleep_disp = (str(sleep_h) + " hrs") if sleep_h else "-"
        st.markdown(
            '<div class="md-rcard"><div class="md-metric-row" style="border:none;">'
            '<div class="md-metric-icon md-hp-violet">🌙</div>'
            '<div class="md-metric-mid"><div class="md-metric-label">Sleep last night</div>'
            '<div class="md-metric-value">' + ui_text(sleep_disp, 20) + '</div></div></div></div>',
            unsafe_allow_html=True
        )
        with st.form("sleep_form_today", clear_on_submit=True):
            sl_in = st.number_input("Log sleep (hours)", min_value=0.0, max_value=16.0, step=0.5, value=float(sleep_h) if sleep_h else 7.0, key="sleep_input")
            if st.form_submit_button("Save sleep", use_container_width=True):
                update_daily_metric("sleep_hours", float(sl_in))
                st.rerun()
    with ov_d:
        pct_water = int(round((water_n / 8) * 100)) if water_n else 0
        st.markdown(
            '<div class="md-rcard"><div class="md-metric-row" style="border:none;">'
            '<div class="md-metric-icon md-hp-blue">💧</div>'
            '<div class="md-metric-mid"><div class="md-metric-label">Water today</div>'
            '<div class="md-metric-value">' + str(water_n) + ' / 8 glasses</div></div>'
            '<div class="md-metric-status md-status-info">' + str(pct_water) + '%</div></div></div>',
            unsafe_allow_html=True
        )
        wcol_a, wcol_b = st.columns(2)
        with wcol_a:
            if st.button("+1 glass", key="water_inc", use_container_width=True):
                update_daily_metric("water_glasses", water_n + 1)
                st.rerun()
        with wcol_b:
            if st.button("Reset", key="water_reset", use_container_width=True):
                update_daily_metric("water_glasses", 0)
                st.rerun()

    ov_e, ov_f = st.columns(2)
    with ov_e:
        steps_disp = (f"{int(steps_n):,} / 10,000") if steps_n else "-"
        steps_pct = min(100, int(round((int(steps_n) / 10000) * 100))) if steps_n else 0
        st.markdown(
            '<div class="md-rcard"><div class="md-metric-row" style="border:none;">'
            '<div class="md-metric-icon md-hp-green">🚶</div>'
            '<div class="md-metric-mid"><div class="md-metric-label">Steps today</div>'
            '<div class="md-metric-value">' + ui_text(steps_disp, 24) + '</div></div>'
            + ('<div class="md-metric-status md-status-info">' + str(steps_pct) + '%</div>' if steps_n else '')
            + '</div></div>',
            unsafe_allow_html=True
        )
        with st.form("steps_form_today", clear_on_submit=True):
            st_in = st.number_input("Log steps today", min_value=0, max_value=100000, step=500, value=int(steps_n) if steps_n else 0, key="steps_input")
            if st.form_submit_button("Save steps", use_container_width=True):
                update_daily_metric("steps", int(st_in))
                st.rerun()
    with ov_f:
        hr_disp = (str(int(hr_n)) + " BPM") if hr_n else "-"
        _hr_lbl, _hr_cls_p = heart_rate_status(hr_n)
        hr_status_html = ('<div class="md-metric-status ' + (_hr_cls_p or "md-status-info") + '">' + (_hr_lbl or "") + '</div>') if hr_n else ''
        st.markdown(
            '<div class="md-rcard"><div class="md-metric-row" style="border:none;">'
            '<div class="md-metric-icon md-hp-pink">❤</div>'
            '<div class="md-metric-mid"><div class="md-metric-label">Resting heart rate</div>'
            '<div class="md-metric-value">' + ui_text(hr_disp, 16) + '</div></div>'
            + hr_status_html
            + '</div></div>',
            unsafe_allow_html=True
        )
        with st.form("hr_form_today", clear_on_submit=True):
            hr_in = st.number_input("Log resting HR (BPM)", min_value=30, max_value=220, step=1, value=int(hr_n) if hr_n else 70, key="hr_input")
            if st.form_submit_button("Save heart rate", use_container_width=True):
                update_daily_metric("heart_rate_resting", int(hr_in))
                st.rerun()

    # 7-day history table
    st.markdown('<div class="md-smart-head" style="margin-top:1rem;"><div class="md-smart-title">Last 7 days</div></div>', unsafe_allow_html=True)
    history = get_metrics_history(7)
    rows_html = '<div class="md-rcard"><div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0.4rem;font-size:0.78rem;">'
    rows_html += '<div style="font-weight:700;color:var(--md-text-2);">Day</div><div style="font-weight:700;color:var(--md-text-2);">Water</div><div style="font-weight:700;color:var(--md-text-2);">Sleep</div>'
    for d, m in history:
        sl = m.get("sleep_hours")
        rows_html += '<div>' + ui_text(d, 20) + '</div>'
        rows_html += '<div>' + str(int(m.get("water_glasses", 0) or 0)) + ' glasses</div>'
        rows_html += '<div>' + ui_text((str(sl) + " h" if sl else "-"), 20) + '</div>'
    rows_html += '</div></div>'
    st.markdown(rows_html, unsafe_allow_html=True)

elif st.session_state.mode == "medications":
    # ── Medications ─────────────────────────────────────────────────
    st.markdown(
        '<div class="md-page-hero md-page-hero-meds">'
        '<div class="md-page-hero-ic"><span class="material-symbols-rounded">pill</span></div>'
        '<div class="md-page-hero-text">'
        '<div class="md-page-hero-title">Medications</div>'
        '<div class="md-page-hero-sub">Keep track of what you take and when. Stored to your profile so MediChat can reference it during chats.</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    with st.form("add_med_form", clear_on_submit=True):
        st.markdown(
            '<div class="md-form-intro">Add medication</div>'
            '<div class="md-form-sub">Capture dose, timing and notes so MediChat can reference them in future chats.</div>',
            unsafe_allow_html=True
        )
        mc1, mc2 = st.columns(2)
        with mc1:
            m_name = st.text_input("Name", placeholder="e.g. Metformin")
            m_freq = st.selectbox("Frequency", ["Once daily", "Twice daily", "Three times daily", "Four times daily", "As needed", "Weekly"])
        with mc2:
            m_dose = st.text_input("Dose", placeholder="e.g. 500 mg")
            m_time = st.text_input("Time(s) of day", placeholder="e.g. Morning, 8 pm")
        m_notes = st.text_area("Notes (optional)", placeholder="Take with food, etc.", height=80)
        if st.form_submit_button("Save medication", use_container_width=True, type="primary", icon=":material/save:"):
            if add_medication(m_name, m_dose, m_freq, m_time, m_notes):
                st.success("Medication added.")
                st.rerun()
            else:
                st.error("Please enter a name.")

    meds = list_medications()
    if not meds:
        st.markdown('<div class="md-rcard" style="text-align:center;color:var(--md-text-3);font-style:italic;padding:1.6rem;">No medications added yet.</div>', unsafe_allow_html=True)
    else:
        for m in meds:
            mh = '<div class="md-rcard"><div style="display:flex;align-items:flex-start;gap:0.8rem;">'
            mh += '<div class="md-metric-icon md-hp-violet" style="width:42px;height:42px;flex-shrink:0;">💊</div>'
            mh += '<div style="flex:1;min-width:0;">'
            mh += '<div style="font-weight:700;color:var(--md-text-1);font-size:0.98rem;">' + ui_text(m.get("name", ""), 90) + '</div>'
            mh += '<div style="font-size:0.78rem;color:var(--md-text-2);margin-top:0.15rem;">' + ui_text(m.get("dose", "") or "Dose not specified", 40) + ' · ' + ui_text(m.get("frequency", ""), 40) + (" · " + ui_text(m.get("time_of_day"), 40) if m.get("time_of_day") else "") + '</div>'
            if m.get("notes"):
                mh += '<div style="font-size:0.76rem;color:var(--md-text-2);margin-top:0.3rem;font-style:italic;">' + ui_text(m.get("notes"), 260) + '</div>'
            mh += '</div></div></div>'
            st.markdown(mh, unsafe_allow_html=True)
            if st.button("Remove", key="del_med_" + str(m.get("id", "")), use_container_width=False):
                delete_medication(m.get("id"))
                st.rerun()

elif st.session_state.mode == "appointments":
    # ── Appointments ────────────────────────────────────────────────
    st.markdown(
        '<div class="md-page-hero md-page-hero-appts">'
        '<div class="md-page-hero-ic"><span class="material-symbols-rounded">calendar_month</span></div>'
        '<div class="md-page-hero-text">'
        '<div class="md-page-hero-title">Appointments</div>'
        '<div class="md-page-hero-sub">Upcoming visits and reminders. Stored to your profile so MediChat can reference them in chats.</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    with st.form("add_appt_form", clear_on_submit=True):
        st.markdown(
            '<div class="md-form-intro">Add appointment</div>'
            '<div class="md-form-sub">Save upcoming visits, reminders and notes in one place.</div>',
            unsafe_allow_html=True
        )
        ac1, ac2 = st.columns(2)
        with ac1:
            a_title = st.text_input("Title", placeholder="e.g. GP follow-up")
            a_date = st.date_input("Date", min_value=_date.today())
            a_time = st.time_input("Time", value=datetime.now().time().replace(minute=0, second=0, microsecond=0))
        with ac2:
            a_doc = st.text_input("Doctor / clinician", placeholder="e.g. Dr Patel")
            a_loc = st.text_input("Location", placeholder="e.g. Melbourne Health Centre")
        a_notes = st.text_area("Notes (optional)", placeholder="Bring previous lab results, etc.", height=80)
        if st.form_submit_button("Save appointment", use_container_width=True, type="primary", icon=":material/save:"):
            iso = datetime.combine(a_date, a_time).isoformat(timespec="minutes")
            if add_appointment(a_title, iso, a_doc, a_loc, a_notes):
                st.success("Appointment added.")
                st.rerun()
            else:
                st.error("Please enter a title and date.")

    appts = list_appointments()
    if not appts:
        st.markdown('<div class="md-rcard" style="text-align:center;color:var(--md-text-3);font-style:italic;padding:1.6rem;">No appointments scheduled yet.</div>', unsafe_allow_html=True)
    else:
        # Sort by date ascending
        try:
            appts_sorted = sorted(appts, key=lambda a: a.get("date", ""))
        except Exception:
            appts_sorted = appts
        now_iso = datetime.now().isoformat(timespec="minutes")
        for a in appts_sorted:
            past = a.get("date", "") < now_iso
            status_cls = "md-status-warn" if past else "md-status-info"
            status_lbl = "Past" if past else "Upcoming"
            try:
                _dt = datetime.fromisoformat(a.get("date"))
                date_disp = _dt.strftime("%a %d %b %Y · %I:%M %p")
            except Exception:
                date_disp = a.get("date", "")
            ah = '<div class="md-rcard"><div style="display:flex;align-items:flex-start;gap:0.8rem;">'
            ah += '<div class="md-metric-icon md-hp-blue" style="flex-shrink:0;">📅</div>'
            ah += '<div style="flex:1;min-width:0;">'
            ah += '<div style="display:flex;align-items:center;gap:0.5rem;flex-wrap:wrap;"><div style="font-weight:700;color:var(--md-text-1);font-size:0.98rem;">' + ui_text(a.get("title", ""), 90) + '</div>'
            ah += '<span class="md-metric-status ' + status_cls + '">' + status_lbl + '</span></div>'
            ah += '<div style="font-size:0.78rem;color:var(--md-text-2);margin-top:0.15rem;">' + ui_text(date_disp, 60) + '</div>'
            details = []
            if a.get("doctor"):
                details.append("👨‍⚕️ " + ui_text(a["doctor"], 60))
            if a.get("location"):
                details.append("📍 " + ui_text(a["location"], 80))
            if details:
                ah += '<div style="font-size:0.78rem;color:var(--md-text-2);margin-top:0.2rem;">' + " &nbsp; ".join(details) + '</div>'
            if a.get("notes"):
                ah += '<div style="font-size:0.76rem;color:var(--md-text-2);margin-top:0.3rem;font-style:italic;">' + ui_text(a.get("notes"), 260) + '</div>'
            ah += '</div></div></div>'
            st.markdown(ah, unsafe_allow_html=True)
            if st.button("Remove", key="del_appt_" + str(a.get("id", "")), use_container_width=False):
                delete_appointment(a.get("id"))
                st.rerun()

elif st.session_state.mode == "records":
    # ── Health Records ──────────────────────────────────────────────
    st.markdown(
        '<div class="md-page-hero md-page-hero-records">'
        '<div class="md-page-hero-ic"><span class="material-symbols-rounded">medical_information</span></div>'
        '<div class="md-page-hero-text">'
        '<div class="md-page-hero-title">Health Records</div>'
        '<div class="md-page-hero-sub">Upload medical PDFs or images. We extract text and store metadata only. File contents stay on your device unless you choose to keep an Ai summary.</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    with st.form("rec_upload_form", clear_on_submit=True):
        rec_file = st.file_uploader("Drop a medical record here", type=["pdf", "jpg", "jpeg", "png"], key="hr_upload")
        rec_label = st.text_input("Label this record", placeholder="e.g. Blood test - April 2026")
        rec_keep_summary = st.checkbox("Generate and save an Ai summary (recommended)", value=True)
        if st.form_submit_button("Save record", use_container_width=True, type="primary", icon=":material/save:"):
            if not rec_file:
                st.error("Please pick a file.")
            elif not rec_label.strip():
                st.error("Please add a label.")
            else:
                size = len(rec_file.getbuffer())
                summary = ""
                if rec_keep_summary:
                    try:
                        if rec_file.name.lower().endswith(".pdf"):
                            txt = extract_pdf_text(rec_file)
                            if txt:
                                ai_resp, _ = medichat_pdf_analysis("Briefly summarise the key findings of this report in plain language for a patient.", txt[:6000], st.session_state.messages)
                                summary = strip_excessive_disclaimers(ai_resp or "")[:1400]
                    except Exception as _e:
                        print("record summary failed:", _e)
                if add_health_record(rec_label, rec_file.type or rec_file.name.split(".")[-1].upper(), size, summary):
                    st.success("Record saved.")
                    st.rerun()

    records = list_health_records()
    if not records:
        st.markdown('<div class="md-rcard" style="text-align:center;color:var(--md-text-3);font-style:italic;padding:1.6rem;">No records uploaded yet.</div>', unsafe_allow_html=True)
    else:
        records_sorted = sorted(records, key=lambda r: r.get("uploaded_at", ""), reverse=True)
        for r in records_sorted:
            ic = "📄" if "pdf" in (r.get("file_type", "").lower()) else "🖼"
            try:
                kb = round(r.get("size_bytes", 0) / 1024, 1)
                size_lbl = (str(kb) + " KB") if kb < 1024 else (str(round(kb / 1024, 1)) + " MB")
            except Exception:
                size_lbl = ""
            uploaded = r.get("uploaded_at", "")[:16].replace("T", " ")
            rh = '<div class="md-rcard"><div style="display:flex;align-items:flex-start;gap:0.8rem;">'
            rh += '<div class="md-metric-icon md-hp-green" style="width:42px;height:42px;flex-shrink:0;font-size:1.2rem;">' + ic + '</div>'
            rh += '<div style="flex:1;min-width:0;">'
            rh += '<div style="font-weight:700;color:var(--md-text-1);font-size:0.96rem;">' + ui_text(r.get("name", ""), 120) + '</div>'
            rh += '<div style="font-size:0.74rem;color:var(--md-text-3);margin-top:0.15rem;">' + ui_text(r.get("file_type", ""), 30) + ' · ' + ui_text(size_lbl, 20) + ' · uploaded ' + ui_text(uploaded, 30) + '</div>'
            if r.get("summary"):
                rh += '<details style="margin-top:0.4rem;"><summary style="cursor:pointer;font-size:0.78rem;color:var(--md-brand-2);font-weight:600;">View Ai summary</summary><div style="font-size:0.8rem;color:var(--md-text-2);margin-top:0.4rem;line-height:1.5;">' + ui_lines(r.get("summary")) + '</div></details>'
            rh += '</div></div></div>'
            st.markdown(rh, unsafe_allow_html=True)
            if st.button("Remove", key="del_rec_" + str(r.get("id", "")), use_container_width=False):
                delete_health_record(r.get("id"))
                st.rerun()

elif st.session_state.mode == "rx_reader":
    # ── Prescription Reader ────────────────────────────────────────
    rx_uploader_widget_key = "rx_uploader_" + str(st.session_state.get("rx_uploader_key", 0))
    st.markdown(
        '<div class="md-page-hero md-page-hero-rx">'
        '<div class="md-page-hero-ic md-page-hero-ic-rx"><span class="md-rx-glyph">&#8478;</span></div>'
        '<div class="md-page-hero-text">'
        '<div class="md-page-hero-title">Prescription Reader</div>'
        '<div class="md-page-hero-sub">Upload a handwritten prescription photo. MediChat will transcribe text only, with confidence grading and an AU drug-name cross-check.</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    with st.form("rx_reader_form", clear_on_submit=False):
        rx_img = st.file_uploader("Drop your prescription image here", type=["jpg", "jpeg", "png"], key=rx_uploader_widget_key)
        rx_note = st.text_input("Optional context", placeholder="e.g. GP script from today, hard to read medication line")
        rx_btn_a, rx_btn_b = st.columns([3, 1])
        with rx_btn_a:
            rx_submit = st.form_submit_button("Read prescription", use_container_width=True, type="primary", icon=":material/arrow_forward:")
        with rx_btn_b:
            rx_clear = st.form_submit_button("Clear", use_container_width=True)

    if rx_clear:
        reset_prescription_reader_state()

    if rx_submit:
        if not rx_img:
            st.error("Please upload a prescription image first.")
        else:
            with st.spinner("Reading prescription text..."):
                try:
                    rx_img.seek(0)
                    st.session_state.rx_reader_result = read_prescription(
                        rx_img.read(),
                        user_note=rx_note,
                        lang_instruction=LANGUAGES[st.session_state.selected_language]["lang_instruction"],
                    )
                except Exception as e:
                    st.session_state.rx_reader_result = None
                    st.error("Prescription reader failed. Please retry with a clearer image. Error: " + str(e))

    rx_result = st.session_state.get("rx_reader_result")
    if rx_result:
        conf = (rx_result.get("overall_confidence") or "unknown").upper()
        model_used = rx_result.get("model_used", "unknown")
        st.markdown(
            '<div class="md-rcard" style="margin-top:0.55rem;">'
            '<div style="display:flex;gap:0.6rem;flex-wrap:wrap;align-items:center;">'
            '<span class="md-metric-status md-status-info">Model: ' + ui_text(model_used, 40) + '</span>'
            '<span class="md-metric-status md-status-good">Overall confidence: ' + ui_text(conf, 10) + '</span>'
            '</div>'
            '</div>',
            unsafe_allow_html=True
        )
        st.markdown('<div class="md-rcard" style="margin-top:0.65rem;">' + markdown_to_html(rx_result.get("reading", "")) + '</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="md-inline-note" style="margin-top:0.65rem;">'
            'Safety guardrail: this feature is transcription only. It does not prescribe therapy or confirm diagnosis. '
            'Always confirm the script with a pharmacist or the prescriber.'
            '</div>',
            unsafe_allow_html=True
        )
        if st.button("Upload a new prescription", key="rx_reader_reset_after_result", use_container_width=True):
            reset_prescription_reader_state()

elif st.session_state.mode == "privacy":
    # ── Privacy & Consent ─────────────────────────────────────────
    st.markdown(
        '<div class="md-greet-wrap"><div class="md-greet">Privacy & Consent</div>'
        '<div class="md-subgreet">MediChat uses HIPAA-style technical safeguards and is designed to align with the Australian Privacy Principles (APPs) and NDB obligations.</div></div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<div class="md-rcard">'
        '<div style="font-weight:800;color:var(--md-text-1);font-size:0.95rem;margin-bottom:0.45rem;">Alignment with Data Handling Standards</div>'
        '<div style="font-size:0.84rem;color:var(--md-text-2);line-height:1.6;">'
        '<strong>APP-first design:</strong> collection minimisation, purpose limitation, patient access and correction workflows, and secure retention controls.<br>'
        '<strong>NDB-ready response:</strong> breach assessment and notification process aligned with the Australian Notifiable Data Breaches scheme for eligible incidents.<br>'
        '<strong>HIPAA-style safeguards:</strong> role-aware access controls, auditability, and encrypted transit/storage patterns where supported by deployment infrastructure.<br>'
        '<strong>Consent & transparency:</strong> users can continue in Guest mode with lower data retention, or create a profile with explicit consent flow.'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<div class="md-rcard" style="margin-top:0.7rem;">'
        '<div style="font-weight:800;color:var(--md-text-1);font-size:0.95rem;margin-bottom:0.45rem;">Clinical Guardrails</div>'
        '<div style="font-size:0.84rem;color:var(--md-text-2);line-height:1.6;">'
        'MediChat is intentionally restricted from issuing final diagnoses or exact medication dose instructions. '
        'It provides educational guidance, triage cues, and clear escalation advice to licensed clinicians.'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<div class="md-rcard" style="margin-top:0.7rem;">'
        '<div style="font-size:0.84rem;color:var(--md-text-2);line-height:1.6;">'
        'Privacy policy link: <a href="' + ui_escape(PRIVACY_POLICY_URL) + '" target="_blank">Open Privacy Policy (APP + NDB)</a>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

elif st.session_state.mode == "insights":
    # ── Ai Insights ─────────────────────────────────────────────────
    st.markdown(
        '<div class="md-page-hero md-page-hero-insights">'
        '<div class="md-page-hero-ic"><span class="material-symbols-rounded">auto_awesome</span></div>'
        '<div class="md-page-hero-text">'
        '<div class="md-page-hero-title">Ai Insights</div>'
        '<div class="md-page-hero-sub">Personalised observations generated from what you have logged. Refreshes each time you visit.</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )
    insights = []
    mem = st.session_state.patient_memory or {}
    meds = list_medications()
    appts = list_appointments()
    records = list_health_records()
    saved_conversations = []
    if st.session_state.is_authenticated and st.session_state.user_email_hash:
        saved_conversations = list_conversations(st.session_state.user_email_hash, limit=20)
    history = get_metrics_history(7)
    today_m = get_daily_metrics()

    # Hydration
    water_today = int(today_m.get("water_glasses", 0) or 0)
    sleeps = [m.get("sleep_hours") for _, m in history if m.get("sleep_hours") is not None]
    has_profile_data = bool(mem.get("symptoms") or mem.get("conditions") or mem.get("medications"))
    has_logged_metrics = (water_today > 0 or len(sleeps) > 0)
    has_records = bool(records)
    has_chat_history = bool(saved_conversations)

    if not (has_profile_data or has_logged_metrics or has_records or has_chat_history or meds or appts):
        st.markdown(
            '<div class="md-empty-card">'
            '<div class="md-empty-icon"><span class="material-symbols-rounded">lightbulb</span></div>'
            '<div class="md-empty-title">No insights yet</div>'
            '<div class="md-empty-copy">Log a few health details such as water, sleep, steps, medications, or a record, and MediChat will start generating personalised observations here.</div>'
            '</div>',
            unsafe_allow_html=True
        )
        st.stop()

    if water_today < 4:
        insights.append(("💧", "Hydration is low today",
            "You have logged " + str(water_today) + " glasses today. Aim for at least 6-8 glasses across the day, more if active.",
            "warn"))
    elif water_today >= 6:
        insights.append(("💧", "Great hydration today",
            "You are at " + str(water_today) + " glasses. Keep it steady through the evening.", "good"))

    # Sleep
    if len(sleeps) >= 3:
        avg_sleep = round(sum(sleeps) / len(sleeps), 1)
        if avg_sleep < 6:
            insights.append(("😴", "Sleep is running short",
                "Your last " + str(len(sleeps)) + " logged nights average " + str(avg_sleep) + " hours. Most adults need 7-9.", "warn"))
        elif avg_sleep > 9:
            insights.append(("😴", "Long sleep pattern",
                "Average " + str(avg_sleep) + " hours. Long sleep can sometimes signal infection or low mood, worth mentioning to a GP if it persists.", "info"))
        else:
            insights.append(("😴", "Sleep looks healthy",
                "Averaging " + str(avg_sleep) + " hours over your last " + str(len(sleeps)) + " logged nights.", "good"))

    # Medications & conditions cross-check
    if meds and mem.get("conditions"):
        insights.append(("💊", "Profile is well-populated",
            "You have " + str(len(meds)) + " medication(s) and " + str(len(mem.get("conditions"))) + " condition(s) recorded. MediChat will use these in every chat.", "info"))
    elif meds and not mem.get("conditions"):
        insights.append(("💊", "Add the conditions these medications treat",
            "You have " + str(len(meds)) + " medication(s) recorded but no conditions yet. Telling MediChat why you take each one will improve its advice.", "warn"))
    elif not meds and mem.get("conditions"):
        insights.append(("💊", "Any medications to add?",
            "You have conditions recorded (" + ", ".join(mem.get("conditions", [])[:3]) + ") but no medications yet. Add them so MediChat can flag interactions.", "info"))

    # Upcoming appointment
    now_iso = datetime.now().isoformat(timespec="minutes")
    upcoming = [a for a in appts if a.get("date", "") >= now_iso]
    if upcoming:
        upcoming.sort(key=lambda a: a.get("date", ""))
        nx = upcoming[0]
        try:
            _dt = datetime.fromisoformat(nx.get("date"))
            in_days = (_dt.date() - _date.today()).days
            when = "today" if in_days == 0 else ("tomorrow" if in_days == 1 else "in " + str(in_days) + " days")
        except Exception:
            when = "soon"
        insights.append(("📅", "Upcoming appointment " + when,
            (nx.get("title") or "Appointment") + (" with " + nx.get("doctor", "") if nx.get("doctor") else "") + ". Want a Doctor Visit Summary PDF before then?", "info"))

    # Symptom load
    sym_n = len(mem.get("symptoms", []) or [])
    if sym_n >= 5:
        insights.append(("📌", "Several active symptoms",
            "MediChat has " + str(sym_n) + " symptoms on file from your chats. Consider a Symptoms Checker run to see if a pattern emerges.", "warn"))

    if records:
        insights.append(("📄", "Health records available",
            "You have " + str(len(records)) + " uploaded record(s). MediChat can use these to improve context during chat and summaries.", "info"))

    if saved_conversations:
        insights.append(("💬", "Conversation history linked",
            str(len(saved_conversations)) + " saved conversation(s) are available for continuity of care context.", "info"))

    # AI-generated overall insight (one Claude call) if logged in and has data
    if CLAUDE_ACTIVE and (meds or mem.get("conditions") or len(sleeps) >= 2 or records or saved_conversations):
        try:
            data_lines = []
            if mem.get("conditions"):
                data_lines.append("Conditions: " + ", ".join(mem["conditions"][:6]))
            if meds:
                data_lines.append("Medications: " + ", ".join([m["name"] + " " + m.get("dose", "") for m in meds[:6]]))
            if sleeps:
                data_lines.append("Recent sleep avg: " + str(round(sum(sleeps)/len(sleeps), 1)) + " h over " + str(len(sleeps)) + " nights")
            data_lines.append("Hydration today: " + str(water_today) + "/8 glasses")
            blob = "\n".join(data_lines)
            resp = anthropic_client.messages.create(
                model=CLAUDE_MODEL, max_tokens=200, temperature=0.4,
                system="You are a careful AI health companion. Read the patient's logged data and give ONE clear, kind, evidence-aware insight in 2 short sentences. Avoid alarming language. End with a concrete suggestion. No disclaimer.",
                messages=[{"role": "user", "content": blob}],
            )
            ai_text = (resp.content[0].text or "").strip()
            if ai_text:
                insights.append(("✨", "Personalised observation", ai_text, "info"))
        except Exception as _e:
            print("AI insights failed:", _e)

    if not insights:
        st.markdown('<div class="md-rcard" style="text-align:center;color:var(--md-text-3);padding:1.6rem;">'
                    'Add more health information to generate insights.</div>',
                    unsafe_allow_html=True)
    else:
        for icon, title, body, kind in insights:
            badge_cls = {"good": "md-status-good", "warn": "md-status-warn", "info": "md-status-info"}.get(kind, "md-status-info")
            ih = '<div class="md-rcard"><div style="display:flex;align-items:flex-start;gap:0.8rem;">'
            ih += '<div class="md-metric-icon md-hp-violet" style="width:42px;height:42px;flex-shrink:0;font-size:1.2rem;">' + icon + '</div>'
            ih += '<div style="flex:1;min-width:0;">'
            ih += '<div style="display:flex;align-items:center;gap:0.5rem;flex-wrap:wrap;"><div style="font-weight:700;color:var(--md-text-1);font-size:0.95rem;">' + ui_text(title, 100) + '</div>'
            ih += '<span class="md-metric-status ' + badge_cls + '">' + ui_text(kind.upper(), 12) + '</span></div>'
            ih += '<div style="font-size:0.85rem;color:var(--md-text-2);margin-top:0.3rem;line-height:1.5;">' + ui_lines(body) + '</div>'
            ih += '</div></div></div>'
            st.markdown(ih, unsafe_allow_html=True)

else:
    if st.session_state.assessment_complete and st.session_state.assessment_parsed:
        parsed = st.session_state.assessment_parsed
        data = st.session_state.assessment_data
        urgency = parsed.get("urgency", "See a doctor soon")
        urgency_lower = urgency.lower()
        report_date = datetime.now().strftime("%B %d, %Y at %I:%M %p")

        st.markdown("---")
        st.markdown("### " + L["report_title"])
        st.caption("Generated: " + report_date)

        if "emergency" in urgency_lower or "now" in urgency_lower:
            st.error("URGENCY: " + urgency)
        elif "urgent" in urgency_lower or "today" in urgency_lower:
            st.warning("URGENCY: " + urgency)
        else:
            st.success("URGENCY: " + urgency)

        st.markdown("---")
        st.markdown("#### " + L["symptoms_reported"])
        rc1, rc2 = st.columns(2)
        with rc1:
            st.info("**Main symptom:** " + data.get("main_symptom", ""))
            st.info("**Duration:** " + data.get("duration", ""))
            st.info("**Severity:** " + data.get("severity", ""))
        with rc2:
            st.info("**Pattern:** " + data.get("pattern", ""))
            st.info("**Age:** " + data.get("age", ""))
            st.info("**Sex:** " + data.get("gender", ""))
            st.info("**Conditions:** " + data.get("known_conditions", "Not specified"))
            st.info("**Medications:** " + data.get("current_medications", "Not specified"))

        other = data.get("other_symptoms", "")
        if other and other.lower() not in ["no", "none", "n/a", "no other symptoms"]:
            st.info("**Other symptoms:** " + other)
        red_flags = data.get("red_flags", "")
        if red_flags:
            st.info("**Red flags:** " + red_flags)

        st.markdown("---")
        st.markdown("#### " + L["possible_conditions"])
        for c in parsed.get("conditions", []):
            if c.strip():
                st.markdown("- " + c.strip())

        st.markdown("---")
        st.markdown("#### " + L["what_to_do"])
        for i, s in enumerate(parsed.get("next_steps", []), 1):
            if s.strip():
                st.markdown("**" + str(i) + ".** " + s.strip())

        st.markdown("---")
        st.markdown("#### " + L["summary"])
        summary = parsed.get("summary", "")
        if summary:
            st.markdown('<div style="background:#f0fdfa;border:1px solid #99f6e4;border-radius:12px;padding:1rem;font-size:0.92rem;color:#134e4a;line-height:1.6;">' + summary + "</div>", unsafe_allow_html=True)

        if parsed.get("safety"):
            st.markdown('<div class="md-inline-note" style="margin-top:0.6rem;">' + ui_text(parsed.get("safety"), 300) + '</div>', unsafe_allow_html=True)

        st.markdown("---")
        st.warning(L["disclaimer_short"])

        st.markdown("---")
        st.markdown("**" + L["download_chat"] + "**")
        pdf_bytes = generate_assessment_pdf(parsed, data, report_date)
        st.download_button(label=L["download_assess_btn"], data=pdf_bytes, file_name="MediChat_Assessment_" + datetime.now().strftime("%Y%m%d_%H%M") + ".pdf", mime="application/pdf", use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        br1, br2 = st.columns(2)
        with br1:
            if st.button(L["new_assessment"], use_container_width=True):
                st.session_state.assessment_stage = 0
                st.session_state.assessment_data = {}
                st.session_state.assessment_complete = False
                st.session_state.assessment_report = None
                st.session_state.assessment_parsed = None
                st.rerun()
        with br2:
            if st.button(L["switch_chat"], use_container_width=True):
                st.session_state.mode = "chat"
                st.rerun()

    else:
        L = LANGUAGES[st.session_state.selected_language]
        stage = st.session_state.assessment_stage
        total = len(ASSESSMENT_STAGES)
        progress = int((stage / total) * 100)

        st.markdown(
            '<div class="assessment-card">'
            '<div class="assessment-card-head">'
            '<div class="md-page-hero-ic"><span class="material-symbols-rounded">stethoscope</span></div>'
            '<div class="assessment-head-text">'
            '<div class="assessment-title">' + L["symptom_title"] + '</div>'
            '<div class="assessment-subtitle">' + L["symptom_subtitle"] + '</div>'
            '</div>'
            '</div>'
            '<div class="progress-label"><span>Step ' + str(stage + 1) + ' of ' + str(total) + '</span><span>' + str(progress) + '% complete</span></div>'
            '<div class="progress-bar-wrap"><div class="progress-bar-fill" style="width:' + str(progress) + '%;"></div></div>'
            '</div>',
            unsafe_allow_html=True
        )

        if st.session_state.assessment_data:
            with st.expander(L["answers_so_far"], expanded=False):
                for k, v in st.session_state.assessment_data.items():
                    st.markdown("**" + k.replace("_", " ").title() + ":** " + str(v))

        if stage < total:
            current = ASSESSMENT_STAGES[stage]
            st.markdown('<div class="question-bubble">' + current["question"] + '</div>', unsafe_allow_html=True)
            if current["hint"]:
                st.caption(current["hint"])

            if current["options"]:
                st.markdown("**" + L["quick_select"] + "**")
                num_cols = min(len(current["options"]), 3)
                ocols = st.columns(num_cols)
                for i, opt in enumerate(current["options"]):
                    with ocols[i % num_cols]:
                        if st.button(opt, key="opt_" + str(stage) + "_" + str(i), use_container_width=True):
                            st.session_state.assessment_data[current["key"]] = opt
                            if current["key"] in ("main_symptom", "red_flags"):
                                is_emerg, detected_reason = detect_emergency(opt)
                                if is_emerg:
                                    st.session_state.emergency_detected = True
                                    st.session_state.emergency_reason = detected_reason
                            st.session_state.assessment_stage += 1
                            if st.session_state.assessment_stage >= total:
                                lang_instruction = LANGUAGES[st.session_state.selected_language]["lang_instruction"]
                                with st.spinner("Compiling patient clinical evaluation index..."):
                                    try:
                                        report = generate_assessment_report(st.session_state.assessment_data, lang_instruction)
                                        st.session_state.assessment_report = report
                                        st.session_state.assessment_parsed = parse_report(report)
                                        st.session_state.assessment_complete = True
                                    except Exception as report_exception:
                                        st.error("Report assembly interrupted. Re-verifying parameter states.")
                                        print(f"Structural Generation Error: {str(report_exception)}")
                            st.rerun()

            with st.form(key="assessment_form_" + str(stage), clear_on_submit=True):
                typed = st.text_input("", placeholder="Or type your own answer here...", label_visibility="collapsed")
                ac1, ac2 = st.columns([3, 1])
                with ac1:
                    next_btn = st.form_submit_button(L["next"], use_container_width=True, type="primary", icon=":material/arrow_forward:")
                with ac2:
                    cancel_btn = st.form_submit_button(L["cancel"], use_container_width=True)

            if next_btn and typed.strip():
                # Force localized atomic update to block state race conditions
                st.session_state.assessment_data[current["key"]] = typed.strip()
                
                if current["key"] in ("main_symptom", "red_flags"):
                    is_emerg, detected_reason = detect_emergency(typed)
                    if is_emerg:
                        st.session_state.emergency_detected = True
                        st.session_state.emergency_reason = detected_reason
                        
                st.session_state.assessment_stage += 1
                
                if st.session_state.assessment_stage >= total:
                    lang_instruction = LANGUAGES[st.session_state.selected_language]["lang_instruction"]
                    with st.spinner("Compiling patient clinical evaluation index..."):
                        try:
                            report = generate_assessment_report(st.session_state.assessment_data, lang_instruction)
                            st.session_state.assessment_report = report
                            st.session_state.assessment_parsed = parse_report(report)
                            st.session_state.assessment_complete = True
                        except Exception as report_exception:
                            st.error("Report assembly interrupted. Re-verifying parameter states.")
                            print(f"Structural Generation Error: {str(report_exception)}")
                            
                st.rerun()

            if cancel_btn:
                st.session_state.assessment_stage = 0
                st.session_state.assessment_data = {}
                st.session_state.mode = "chat"
                st.rerun()
